"""Clade-filtered SNP-table previews (render_filtered_window + selections).

The tree viewer lets a user click a clade and open the group's SNP table
subset to that clade. The subset must keep exactly what the clade uses: the
selected sample rows (plus the structural root/MQ/annotation rows) and only
the locus columns where a selected sample CALLS something different from the
reference row.

Values decide the column choice, not colours — this is load-bearing. In a
real vSNP3 table every populated call cell has a fill, because the
top-priority rule paints reference-matching calls near-white (`cellIs equal
B$2`), so "has a colour" would keep every column. And missing data ('-'/'N')
is not a call: a low-coverage sample must not drag thousands of columns into
the view. The workbook builder here reproduces the real rule stack, B$2 and
all, so the tests fail on exactly the tables vSNP3 writes.

Matching is deliberately forgiving about spelling: tree tips and table labels
may carry `_zc.vcf` suffixes or different metadata decorations of the same
sample, and both sides are reduced to the canonical on-disk stem before
comparing. What it must NOT be forgiving about is a selection that matches
nothing — that renders as an error, never as an empty grid or a silently
unfiltered table.

Run directly:  python test_xlsx_filter.py
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

import xlsx_html


def assert_eq(actual, expected, label):
    if actual != expected:
        raise AssertionError(f"{label}: expected {expected!r}, got {actual!r}")
    print(f"  OK  {label} = {actual!r}")


def assert_true(condition, label):
    if not condition:
        raise AssertionError(f"{label}: expected truthy, got falsy")
    print(f"  OK  {label}")


def _fill(argb: str) -> PatternFill:
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def _contains_rule(text: str, argb: str) -> Rule:
    return Rule(
        type="containsText", operator="containsText", text=text,
        dxf=DifferentialStyle(fill=_fill(argb)),
        formula=[f'NOT(ISERROR(SEARCH("{text}",B3)))'])


def make_group_table(path: Path, sample_labels: list[str], cols: int,
                     call_at, locus_at=None) -> None:
    """A workbook shaped like a real vSNP3 group table.

    Row 1 is loci, row 2 the `root` reference row (all 'C'), then sample
    rows, then MQ and annotation. ``call_at(sample_index, col)`` supplies
    each sample cell's call ('C' = matches reference, 'A'/'R'/... = SNP,
    '-' = missing). The conditional formatting mirrors vSNP3's stack: the
    reference-match rule FIRST (`equal B$2`, near-white — so every populated
    call cell ends up with a fill), per-base colours after it, missing-data
    colours last, over the sample rows only.

    ``locus_at(col)`` names the position headers. vsnp3 writes them as
    `CHROM:POS`, and CHROM is whatever the reference FASTA calls the contig —
    an accession for a bacterial reference, an isolate name for an influenza
    one. Both are exercised.
    """
    locus_at = locus_at or (lambda c: f"MTBC0:{1000 + c}")
    wb = openpyxl.Workbook()
    ws = wb.active
    for c in range(2, cols + 1):
        ws.cell(row=1, column=c, value=locus_at(c))
    ws.cell(row=2, column=1, value="root")
    for c in range(2, cols + 1):
        ws.cell(row=2, column=c, value="C")
    for i, label in enumerate(sample_labels, start=1):
        r = i + 2
        ws.cell(row=r, column=1, value=label)
        for c in range(2, cols + 1):
            ws.cell(row=r, column=c, value=call_at(i, c))
    ws.cell(row=len(sample_labels) + 3, column=1, value="MQ")
    for c in range(2, cols + 1):
        ws.cell(row=len(sample_labels) + 3, column=c, value=60)
    ws.cell(row=len(sample_labels) + 4, column=1, value="annotation")
    last_col = openpyxl.utils.get_column_letter(cols)
    sample_rng = f"B3:{last_col}{len(sample_labels) + 2}"
    ws.conditional_formatting.add(sample_rng, CellIsRule(
        operator="equal", formula=["B$2"], fill=_fill("FFFDFEFE")))
    for text, argb in (("A", "FF58FA82"), ("G", "FFF7FE2E"),
                       ("C", "FF0000FF"), ("T", "FFFF0000"),
                       ("R", "FFE2CFDD"), ("N", "FFE2CFDD"),
                       ("-", "FFE2CFDD")):
        ws.conditional_formatting.add(sample_rng, _contains_rule(text, argb))
    wb.save(path)
    wb.close()


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="xlsx_filter_"))
    try:
        cols, nsamples = 12, 6

        # S1: SNP at col 2 ('A') and col 3 ('R' — an ambiguous call is still
        # a call). S2: matches reference at col 2, SNP at col 3, MISSING at
        # col 5. S3..S6: SNPs at columns 4+ only.
        def calls(i, c):
            if i == 1:
                return {2: "A", 3: "R"}.get(c, "C")
            if i == 2:
                return {3: "A", 5: "-"}.get(c, "C")
            return "A" if c >= 4 else "C"

        book = tmp / "group.xlsx"
        make_group_table(
            book, [f"S{i}_meta_country" for i in range(1, nsamples + 1)],
            cols, calls)
        total_rows, total_cols = nsamples + 4, cols
        bams = {f"S{i}" for i in range(1, nsamples + 1)}

        print("\n[a clade keeps its samples and only its SNP columns]")
        w = xlsx_html.render_filtered_window(
            book, total_rows, total_cols, None, "p", bams, set(),
            ["S1_meta_country", "S2_zc.vcf"],   # one decorated, one suffixed
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        f = w["filter"]
        assert_eq(f["matched"], 2, "both selected samples matched")
        assert_eq(f["total_samples"], nsamples, "sample total counted")
        assert_eq(f["locus_shown"], 2, "only the clade's SNP columns kept")
        assert_eq(w["loci"], {"2": "MTBC0:1002", "3": "MTBC0:1003"},
                  "cols 2-3 kept: S2's missing '-' at col 5 is not a SNP")
        assert_eq(w["shown_cols"], 3, "kept columns = names + 2 SNP columns")
        assert_eq(w["shown_rows"], 6, "header + root + 2 samples + MQ + annotation")
        assert_eq(w["row_samples"], ["", "", "S1", "S2", "", ""],
                  "IGV stems per kept row (structural rows blank)")
        variant_cells = sum(r.count("xlsx-variant") for r in w["rows"])
        assert_eq(variant_cells, 4,
                  "every populated call cell in a kept sample row stays an IGV target")
        assert_true("#fdfefe" in w["style_css"],
                    "the reference-match near-white resolved while streaming "
                    "(S2's col-2 'C' equals root)")
        page = xlsx_html.compose_page(w)
        assert_true("Filtered to a tree clade" in page, "the page says it is filtered")
        assert_true("2 of 6 samples" in page, "…with the sample count")
        assert_true("Show the full table" in page, "…and offers the way back")
        assert_true("more than a browser can lay out" not in page,
                    "the too-big notice is not shown for a deliberate subset")

        print("\n[differently-decorated spellings of one sample still match]")
        # Table says S5_meta_country; the tree says S5_lab_variantX. Both
        # resolve to the known stem S5 (here known via the VCF database set).
        w = xlsx_html.render_filtered_window(
            book, total_rows, total_cols, None, "p", set(), {"S5"},
            ["S5_lab_variantX"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        assert_eq(w["filter"]["matched"], 1, "decorated tip matched decorated row")
        assert_eq(w["filter"]["locus_shown"], cols - 3,
                  "S5's SNP columns (4+) kept, reference-match columns dropped")
        assert_eq(w["row_samples"], ["", "", "S5", "", ""],
                  "row resolved to the canonical stem")

        print("\n[no recognisable reference row: over-keep, and SAY so]")
        # The severe case. The column filter needs the `root` row as its
        # baseline. When it cannot find one, the old code compared the selected
        # samples against EACH OTHER — which drops every position where the clade
        # agrees internally but differs from the reference, i.e. exactly the
        # clade-defining SNPs. On a real 4-sample clade, relabelling that one row
        # lost 30 of 85 positions, all of them monomorphic within the clade,
        # while the banner went on saying the hidden positions had no SNP.
        noref = tmp / "noref.xlsx"
        make_group_table(
            noref, [f"S{i}_meta" for i in range(1, 4)], 6,
            # Every selected sample carries the SAME variant call at col 2 (a
            # clade-defining SNP) and matches the reference elsewhere.
            call_at=lambda i, c: "A" if c == 2 else "C")
        # Rename the reference row so it is not recognised as one.
        wb = openpyxl.load_workbook(noref)
        wb.active.cell(row=2, column=1, value="MTBC0ref")
        wb.save(noref)
        wb.close()
        w = xlsx_html.render_filtered_window(
            noref, 7, 6, None, "p", {"S1", "S2", "S3"}, set(),
            ["S1_meta", "S2_meta"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        f = w["filter"]
        assert_true(not f["reference_found"], "the missing reference row is noticed")
        assert_true("MTBC0:1002" in w["loci"].values(),
                    "the clade-defining SNP survives (it is monomorphic in-clade)")
        page = xlsx_html.compose_page(w)
        assert_true("no <code>root</code> reference row" in page,
                    "the page says there is no reference row")
        assert_true("matches the reference are hidden" not in page,
                    "…and does NOT claim positions were hidden for having no SNP")

        print("\n[selected samples that match no row are disclosed]")
        # Tips that resolve to nothing take their positions with them, so this is
        # missing DATA, not just missing names. It used to be computed and never
        # rendered, leaving the page reading like a deliberately small clade.
        w = xlsx_html.render_filtered_window(
            book, total_rows, total_cols, None, "p", bams, set(),
            ["S1_meta_country", "NOT_IN_TABLE_A", "NOT_IN_TABLE_B"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        assert_eq(w["filter"]["matched"], 1, "one of three selected tips matched")
        assert_eq(w["filter"]["unmatched"], 2, "the other two are counted")
        page = xlsx_html.compose_page(w)
        assert_true("2 of the 3 selected samples were not found" in page,
                    "the page states how many selected samples are missing")

        print("\n[a name that cannot be told from another sample is excluded]")
        # _canonical_stem maps a label onto the longest KNOWN stem that prefixes
        # it. With only `19-1234` on disk, the row `19-1234_2_Bovine_USA` — a
        # different isolate — resolves to `19-1234` as well. Keeping it added its
        # private SNPs as columns AND labelled it `19-1234`, so a click opened the
        # wrong sample's alignment in IGV.
        alias = tmp / "alias.xlsx"
        make_group_table(
            alias, ["19-1234_Bovine_USA", "19-1234_2_Bovine_USA", "20-9999_Bovine_USA"],
            6, call_at=lambda i, c: "A" if c == i + 1 else "C")
        w = xlsx_html.render_filtered_window(
            alias, 7, 6, None, "p", {"19-1234"}, set(),
            ["19-1234_Bovine_USA"],          # ONE tip selected
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        f = w["filter"]
        assert_eq(f["matched"], 1, "only the selected isolate matched")
        assert_eq(f["dropped_ambiguous"], 1, "the look-alike row was excluded")
        assert_eq(w["filter"]["locus_shown"], 1,
                  "…and the intruder's private SNP is not a column")
        # The surviving row does NOT claim the ambiguous stem. Deliberately
        # conservative: a stem two different labels resolve to is not a reliable
        # identity for either of them, so it is not used as an IGV target at all
        # — including for the exactly-matched row. The cost is one lost link in a
        # rare naming situation; the alternative is a cell that opens a different
        # specimen's reads while looking entirely normal.
        assert_true("19-1234" not in [s for s in w["row_samples"] if s],
                    "no row carries the ambiguous stem as its IGV target")
        assert_true(sum(r.count("xlsx-variant") for r in w["rows"]) == 0,
                    "…so no cell in it is clickable")
        assert_true(sum(r.count("xlsx-igv-none") for r in w["rows"]) > 0,
                    "…and the cells say why (tooltip: no data for this name)")
        page = xlsx_html.compose_page(w)
        assert_true("could not be told apart" in page, "the exclusion is disclosed")

        print("\n[the annotation row is structural even when named differently]")
        # vsnp3 writes "no annotations" when the group has no GBK. Read as a
        # sample, it was dropped from every filtered view — losing the gene and
        # amino-acid annotation for the positions being inspected.
        annot = tmp / "annot.xlsx"
        make_group_table(annot, ["S1", "S2"], 6, call_at=lambda i, c: "A")
        wb = openpyxl.load_workbook(annot)
        wb.active.cell(row=6, column=1, value="no annotations")
        wb.save(annot)
        wb.close()
        w = xlsx_html.render_filtered_window(
            annot, 6, 6, None, "p", {"S1", "S2"}, set(), ["S1"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        assert_eq(w["filter"]["total_samples"], 2,
                  "'no annotations' is not counted as a sample")
        assert_eq(w["row_samples"], ["", "", "S1", "", ""],
                  "…and the annotation row is kept in the filtered view")

        print("\n[a selection that matches nothing is an error, not a table]")
        try:
            xlsx_html.render_filtered_window(
                book, total_rows, total_cols, None, "p", bams, set(),
                ["NOT_A_SAMPLE"],
                xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
            raise AssertionError("expected FilterMatchError")
        except xlsx_html.FilterMatchError as e:
            assert_true("appear in this table" in str(e),
                        "FilterMatchError explains the mismatch")
        try:
            xlsx_html.render_filtered_window(
                book, total_rows, total_cols, None, "p", bams, set(),
                ["root"],   # the outgroup tip is not a sample
                xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
            raise AssertionError("expected FilterMatchError")
        except xlsx_html.FilterMatchError:
            assert_true(True, "a root-only selection has no sample names")

        # The reference decides what a position header LOOKS like, and for a
        # while only one shape was recognised. `NC_002945.4:12345` is a
        # bacterial accession; an influenza reference names each segment after
        # the isolate it came from, so its headers carry '/', '(' and ')'. Not
        # one column of an HPAI table matched, the sheet was classified "not a
        # variant table", and the clade filter handed back the WHOLE table —
        # or, on a big one, the "too large to display, open the tree and click
        # a branch" page, which is what the user had just done. Reported from
        # the Ames HPC on an H5N1 dairy-cattle group.
        print("\n[the position header's contig name is the reference's, not an accession]")
        for label, locus_at in (
            ("influenza segment named after its isolate",
             lambda c: f"A/mallard/Netherlands/12/2022(H5N1)_PB2:{100 + c}"),
            ("a bare contig name",
             lambda c: f"NC_002945.4:{1000 + c}"),
            ("a contig name carrying its own colon",
             lambda c: f"gi|12345|ref|NC_002945.4:{1000 + c}"),
        ):
            named = tmp / f"named_{abs(hash(label))}.xlsx"
            make_group_table(named, [f"S{i}_meta_country" for i in range(1, 5)],
                             8, calls, locus_at=locus_at)
            w = xlsx_html.render_filtered_window(
                named, 4 + 4, 8, None, "p", {f"S{i}" for i in range(1, 5)}, set(),
                ["S1_meta_country", "S2_meta_country"],
                xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
            assert_true(not w["filter"].get("ignored"),
                        f"{label}: recognised as a SNP table")
            assert_eq(w["filter"]["shown_samples"], 2, f"{label}: rows filtered")
            assert_true(w["filter"]["locus_shown"] < w["filter"]["locus_total"],
                        f"{label}: positions filtered "
                        f"({w['filter']['locus_shown']} of {w['filter']['locus_total']})")

        print("\n[a sheet with no locus columns ignores the filter, loudly]")
        plain = tmp / "plain.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, 6):
            for c in range(1, 5):
                ws.cell(row=r, column=c, value=f"v{r}.{c}")
        wb.save(plain)
        wb.close()
        w = xlsx_html.render_filtered_window(
            plain, 5, 4, None, "p", set(), set(), ["v2.1"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        assert_true(w["filter"].get("ignored"), "filter marked ignored")
        assert_eq(w["shown_rows"], 5, "the sheet renders unfiltered")
        page = xlsx_html.compose_page(w)
        assert_true("Clade filter not applied" in page,
                    "the page says the filter did not apply")

        # A filtered window that had to give something up is NOT shown as a
        # table. It is reported as partial, and the endpoint replaces it with the
        # "this clade is still too large" page. That matters more than a banner:
        # the truncation always takes the column TAIL, and in cascade order the
        # tail is the private SNPs — on a real 149-sample clade that silently
        # dropped 3,478 of its 10,000 matching positions while the banner's
        # leading clause still read "positions with no SNP … are hidden".
        print("\n[a clade too large to filter by position is reported partial]")
        saved = xlsx_html.FILTER_BUFFER_MAX_CELLS
        xlsx_html.FILTER_BUFFER_MAX_CELLS = 10   # force the fallback
        try:
            w = xlsx_html.render_filtered_window(
                book, total_rows, total_cols, None, "p", bams, set(),
                ["S3_meta_country", "S4_meta_country"],
                xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        finally:
            xlsx_html.FILTER_BUFFER_MAX_CELLS = saved
        f = w["filter"]
        assert_true(not f["column_filter"], "the per-position filter was skipped")
        assert_eq(f["matched"], 2, "rows were still filtered")
        assert_true(xlsx_html.window_is_partial(w),
                    "the window reports itself partial, so no table is served")

        print("\n[a budget-truncated clade is reported partial too]")
        wide = tmp / "wide.xlsx"
        make_group_table(
            wide, [f"S{i}" for i in range(1, 4)], 400,
            call_at=lambda i, c: "A")   # every position is a SNP for everyone
        w = xlsx_html.render_filtered_window(
            wide, 3 + 4, 400, None, "p", {"S1", "S2", "S3"}, set(),
            ["S1", "S2"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS,
            max_table_bytes=4096)
        assert_true(w["filter"]["truncated_cols"] > 0, "columns were dropped")
        rendered = sum(len(r) for r in w["rows"])
        assert_true(rendered < 4096 * 1.6,
                    f"output ({rendered} B) respects the byte budget")
        assert_true(xlsx_html.window_is_partial(w), "…and it reports partial")

        print("\n[a clade past the row cap is reported partial]")
        w = xlsx_html.render_filtered_window(
            book, total_rows, total_cols, None, "p", bams, set(),
            [f"S{i}_meta_country" for i in range(1, 5)],
            xlsx_html.DEFAULT_MAX_CELLS, max_rows=10)
        assert_eq(w["filter"]["shown_samples"], 2,
                  "max_rows=10 leaves room for 2 sample rows")
        assert_eq(w["filter"]["truncated_rows"], 2, "the rest are counted")
        assert_true(xlsx_html.window_is_partial(w), "…and it reports partial")

        print("\n[selection tokens: deterministic, stored, recoverable]")
        os.environ["VSNP_GUI_PREVIEW_CACHE_DIR"] = str(tmp / "cache")
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from app import main as app_main
        t1 = app_main._snp_selection_token(["b", "a"])
        t2 = app_main._snp_selection_token(["a", "b", "b"])
        assert_eq(t1, t2, "token ignores order and duplicates")
        app_main._store_snp_selection(t1, {"samples": ["a", "b"], "source": "x.tre"})
        assert_eq(app_main._load_snp_selection(t1)["samples"], ["a", "b"],
                  "round trip from memory")
        app_main._SNP_SELECTIONS.clear()
        assert_eq(app_main._load_snp_selection(t1)["samples"], ["a", "b"],
                  "round trip from disk after a restart")
        assert_eq(app_main._load_snp_selection("no-such-token"), None,
                  "unknown token is None, not an error")
        assert_eq(app_main._load_snp_selection("../../etc/passwd"), None,
                  "a token is a hex string, nothing else is looked up")

        print("\nAll xlsx filter tests passed.")
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
        os.environ.pop("VSNP_GUI_PREVIEW_CACHE_DIR", None)


if __name__ == "__main__":
    sys.exit(main())
