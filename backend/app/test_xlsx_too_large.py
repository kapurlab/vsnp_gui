"""A table too large to show in full says so — it does not show a slice of itself.

A vSNP3 cascade table on a big group is ~1,000 samples x 10,000 positions: ten
million cells, which no browser lays out. The old behaviour rendered the first
1,000 x 1,000 of it behind a "showing the first ..." banner. That is a table
which looks complete enough to read and is not, and reading a SNP table with an
unannounced 90% of the positions missing is a wrong answer, not a slow one.

So: if the whole sheet cannot be shown, none of it is. The preview answers with
the two routes that do work — the file in Excel/LibreOffice, or the group's tree
with a clade selected — and it answers in well under a second, because the
verdict comes from the sheet's declared dimensions and never opens a cell.

Measured on this machine (2026-08-13): ~34 bytes of HTML per rendered cell, so
the 1,000,000-cell ceiling is a ~34 MB document — already the outer limit of
what is worth handing a browser. 225 of 2,316 real tables here are above it.

Run directly:  python test_xlsx_too_large.py
"""

from __future__ import annotations

import shutil
import sys
import tempfile
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

import xlsx_html


def assert_eq(actual, expected, label):
    if actual != expected:
        raise AssertionError(f"{label}: expected {expected!r}, got {actual!r}")
    print(f"  OK  {label} = {actual!r}")


def assert_true(condition, label):
    if not condition:
        raise AssertionError(f"{label}: expected truthy, got falsy")
    print(f"  OK  {label}")


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="xlsx_toobig_"))
    try:
        print("\n[the viewable envelope is decided by CELLS, not by a row cap]")
        cap = xlsx_html.FULL_VIEW_MAX_CELLS
        assert_true(xlsx_html.fits_full_view(1_000, 100),
                    "1,000 x 100 fits")
        assert_true(xlsx_html.fits_full_view(1, cap), "exactly at the ceiling fits")
        assert_true(not xlsx_html.fits_full_view(1, cap + 1),
                    "one cell over the ceiling does not")
        assert_true(not xlsx_html.fits_full_view(1_048, 10_001),
                    "the real 1,048 x 10,001 cascade table does not fit")
        # A tall narrow sheet is not "too large" — 5,000 x 12 is 60,000 cells and
        # a browser handles it fine. The old 1,000-row policy cut it at row 1,000.
        assert_true(xlsx_html.fits_full_view(5_000, 12),
                    "a tall narrow sheet fits (no row cap in the verdict)")
        assert_true(not xlsx_html.fits_full_view(0, 0),
                    "an empty/unreadable extent is not claimed to fit")

        print("\n[the verdict is cheap: no cells are read to reach it]")
        big = tmp / "big.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Declared dimension is what read-only mode reports; write two far-apart
        # cells so the sheet is genuinely large without taking all day to build.
        ws.cell(row=1, column=1, value="corner")
        ws.cell(row=1500, column=900, value="far corner")
        wb.save(big)
        wb.close()
        t0 = time.time()
        rows, cols = xlsx_html.sheet_extent(big)
        verdict = xlsx_html.fits_full_view(rows, cols)
        dt = time.time() - t0
        assert_eq((rows, cols), (1500, 900), "extent read")
        assert_true(not verdict, "1,500 x 900 = 1.35M cells is over the ceiling")
        assert_true(dt < 1.0, f"the verdict took {dt:.3f}s (no cell was read)")

        print("\n[the page names the size and offers both routes]")
        page = xlsx_html.too_large_page(
            filename="La1_cascade1_table1-2026-08-04.xlsx",
            total_rows=1048, total_cols=10001,
            tree_url="../../../?view=tree&project=p&path=%2Fx%2FLa1.tre",
            tree_name="La1_2026-08-04_11-03-06.tre")
        assert_true("1,048" in page and "10,001" in page,
                    "the real dimensions are stated")
        assert_true("10,480,048" in page or "10.5 million" in page,
                    "the cell count is stated in words a reader can weigh")
        assert_true("Download" in page, "the spreadsheet can be downloaded")
        assert_true("Excel" in page and ("LibreOffice" in page or "OpenOffice" in page),
                    "…and it says what to open it in")
        assert_true("view=tree" in page, "the group's tree is linked")
        assert_true("La1_2026-08-04_11-03-06.tre" in page, "…and named")
        assert_true("clade" in page.lower(), "…with what to do there")
        # It must not read as a failure: nothing is broken, the file is simply big.
        assert_true("error" not in page.lower() and "failed" not in page.lower(),
                    "the page does not present this as an error")
        assert_true("<table" not in page,
                    "no partial table is rendered alongside the message")

        # Arriving here BY clicking a branch, and being told to click a branch,
        # is a loop — and it is the one an unrecognised position header put a
        # user in on a real HPAI table: the sheet was not recognised as a SNP
        # table, so the filter never applied, so the whole table came back, so
        # it was too large, so the page advised opening the tree.
        print("\n[a filter that could not apply does not send you back to the tree]")
        looped = xlsx_html.too_large_page(
            filename="x.xlsx", total_rows=2000, total_cols=8000,
            tree_url="../../../?view=tree&project=p&path=t.tre", tree_name="t.tre",
            filter_ignored="this sheet has no locus columns")
        assert_true("view=tree" not in looped,
                    "the Open-the-tree button is gone — that is where this came from")
        assert_true("no locus columns" in looped,
                    "…and the page says WHY the clade filter did not apply")
        assert_true("whole table, not your clade" in looped,
                    "…and that the size quoted is the whole table, not the clade")
        assert_true("Download" in looped, "the download route survives")

        print("\n[with no tree beside the table, the page still stands on its own]")
        page2 = xlsx_html.too_large_page(
            filename="x.xlsx", total_rows=2000, total_cols=8000,
            tree_url=None, tree_name=None)
        assert_true("Download" in page2, "the download route survives")
        assert_true("view=tree" not in page2, "no dead tree link is offered")
        assert_true("tree" in page2.lower(),
                    "the tree route is still explained (the group's tree is in Results)")

        print("\n[a sheet that DOES fit renders in full — no truncation banner]")
        fitting = tmp / "fits.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for c in range(2, 40):
            ws.cell(row=1, column=c, value=f"MTBC0:{1000 + c}")
        for r in range(2, 60):
            ws.cell(row=r, column=1, value=f"S{r - 1}")
            for c in range(2, 40):
                ws.cell(row=r, column=c, value="A" if (r + c) % 3 else "C")
        wb.save(fitting)
        wb.close()
        rows, cols = xlsx_html.sheet_extent(fitting)
        assert_true(xlsx_html.fits_full_view(rows, cols), "the small sheet fits")
        w = xlsx_html.render_window(
            fitting, rows, cols, None, "p", None, None,
            xlsx_html.FULL_VIEW_MAX_CELLS, rows)
        assert_eq((w["shown_rows"], w["shown_cols"]), (rows, cols),
                  "every row and column is rendered")
        page3 = xlsx_html.compose_page(w)
        assert_true("Showing the first" not in page3, "no truncation banner")

        print("\n[a tall narrow sheet now renders past row 1,000]")
        tall = tmp / "tall.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for c in range(2, 8):
            ws.cell(row=1, column=c, value=f"MTBC0:{1000 + c}")
        for r in range(2, 1300):
            ws.cell(row=r, column=1, value=f"S{r - 1}")
            for c in range(2, 8):
                ws.cell(row=r, column=c, value="A")
        wb.save(tall)
        wb.close()
        rows, cols = xlsx_html.sheet_extent(tall)
        assert_true(rows > 1000, f"the sheet is {rows} rows")
        assert_true(xlsx_html.fits_full_view(rows, cols), "…and still fits by cells")
        w = xlsx_html.render_window(
            tall, rows, cols, None, "p", None, None,
            xlsx_html.FULL_VIEW_MAX_CELLS, rows)
        assert_eq(w["shown_rows"], rows, "all rows rendered, past the old 1,000 cap")

        print("\n[what counts as 'partial' differs for a filtered window]")
        full = {"shown_rows": 50, "shown_cols": 20, "total_rows": 50, "total_cols": 20}
        assert_true(not xlsx_html.window_is_partial(full), "a complete window is not partial")
        assert_true(xlsx_html.window_is_partial(
            {**full, "shown_cols": 19}), "a column short of the sheet is partial")
        assert_true(xlsx_html.window_is_partial(
            {**full, "shown_rows": 49}), "a row short of the sheet is partial")
        # A filtered window shows fewer rows BY DESIGN — that is not partial.
        filt_ok = {**full, "shown_rows": 6, "filter": {
            "matched": 4, "truncated_cols": 0, "truncated_rows": 0,
            "column_filter": True, "ignored": None}}
        assert_true(not xlsx_html.window_is_partial(filt_ok),
                    "a clade window showing only its clade is complete")
        assert_true(xlsx_html.window_is_partial(
            {**filt_ok, "filter": {**filt_ok["filter"], "truncated_cols": 12}}),
            "…but dropping matching positions makes it partial")
        assert_true(xlsx_html.window_is_partial(
            {**filt_ok, "filter": {**filt_ok["filter"], "truncated_rows": 3}}),
            "…as does dropping matching samples")
        assert_true(xlsx_html.window_is_partial(
            {**filt_ok, "filter": {**filt_ok["filter"], "column_filter": False}}),
            "…as does skipping the per-position filter entirely")
        # A non-variant sheet ignores the filter, so fall back to the geometry.
        assert_true(not xlsx_html.window_is_partial(
            {**full, "filter": {"ignored": "not a SNP table"}}),
            "an ignored filter falls back to the geometric check")

        print("\n[the tree beside a table is found, labeled one preferred]")
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from app.main import _sibling_tree
        group = tmp / "GroupA"
        group.mkdir()
        (group / "GroupA_cascade1_table-2026.xlsx").write_text("")
        assert_eq(_sibling_tree(group / "GroupA_cascade1_table-2026.xlsx"), None,
                  "no tree beside the table")
        (group / "GroupA_2026.tre").write_text("(a,b);")
        assert_eq(_sibling_tree(group / "GroupA_cascade1_table-2026.xlsx").name,
                  "GroupA_2026.tre", "the plain tree is found")
        (group / "GroupA_2026_labeled.tre").write_text("(a,b);")
        assert_eq(_sibling_tree(group / "GroupA_cascade1_table-2026.xlsx").name,
                  "GroupA_2026_labeled.tre",
                  "the re-labeled tree is preferred (descriptive tip names)")

        print("\nAll too-large-table tests passed.")
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
