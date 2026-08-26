"""The SNP-table view aids: the identical-position toggle and colour marks.

Both aids answer requests that came straight from reading clade views. A
clade's table keeps every position where any member differs from the
reference — which includes the positions the whole clade shares, informative
about the clade but noise when the question is what differs WITHIN it. The
toggle hides exactly those (exact-match across the shown sample rows, applied
by the page's own JS). The colour marks are the table's version of the tree
viewer's "Colour it": click a sample name or a position header and it stays
marked while you scan.

What the backend owes these features, and what is tested here, is the page
scaffolding: the controls appear only where they can mean something (a variant
table with sample rows), the toggle is disabled — with the reason on it — when
only one sample is shown, and the full-fidelity renderer now ships the same
row/column metadata the streaming renderer always had, so the aids work on
small tables too. The full path only ships that metadata when no cells are
merged, because the aids resolve cells by DOM position and a merge shifts
every index after it.

Run directly:  python test_xlsx_view_controls.py
"""

from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

import xlsx_html
from test_xlsx_filter import assert_eq, assert_true, make_group_table


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="xlsx_view_controls_"))
    try:
        cols, nsamples = 8, 4

        # S1/S2 share the same 'A' at col 2 (identical within any clade that
        # holds both); S1 alone carries 'G' at col 3.
        def calls(i, c):
            if c == 2:
                return "A" if i in (1, 2) else "C"
            if c == 3:
                return "G" if i == 1 else "C"
            return "C"

        book = tmp / "group.xlsx"
        make_group_table(book, [f"S{i}_meta" for i in range(1, nsamples + 1)],
                         cols, calls)
        total_rows, total_cols = nsamples + 4, cols
        bams = {f"S{i}" for i in range(1, nsamples + 1)}

        print("\n[a clade view carries both aids]")
        w = xlsx_html.render_filtered_window(
            book, total_rows, total_cols, None, "p", bams, set(),
            ["S1_meta", "S2_meta"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        page = xlsx_html.compose_page(w)
        assert_true('id="xlsxInvariant"' in page, "the toggle is on the page")
        assert_true('id="xlsxInvariant" disabled' not in page,
                    "…enabled: two samples are shown")
        assert_true("Hide identical positions" in page, "…and labelled plainly")
        assert_true('id="xlsxInvariantNote"' in page,
                    "…with a slot to report what it hid")
        assert_true("colour" in page and 'id="xlsxHlClear"' in page,
                    "the colour-mark hint and its clear link are on the page")
        assert_true('"1f77b4"' not in page.split("PALETTE")[0],
                    "sanity: the palette lives in the script, not the markup")

        print("\n[one sample: the toggle is present, disabled, and says why]")
        w1 = xlsx_html.render_filtered_window(
            book, total_rows, total_cols, None, "p", bams, set(),
            ["S1_meta"],
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        page1 = xlsx_html.compose_page(w1)
        assert_true('id="xlsxInvariant" disabled' in page1,
                    "single-sample view disables the toggle")
        assert_true("trivially identical" in page1, "…with the reason on it")

        print("\n[a sheet that is not a variant table offers no aids]")
        plain = tmp / "plain.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, 6):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value=f"v{r}{c}")
        wb.save(plain)
        wb.close()
        wp = xlsx_html.render_window(
            plain, 5, 3, None, "p", set(), set(),
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        pagep = xlsx_html.compose_page(wp)
        assert_true('id="xlsxInvariant"' not in pagep,
                    "no toggle without locus columns")
        assert_true('id="xlsxHlClear"' not in pagep, "no colour controls either")

        print("\n[the full-fidelity path ships the same metadata and aids]")
        pagef = xlsx_html.xlsx_to_html(book, project="p",
                                       samples_with_bams=bams,
                                       samples_with_vcfs=set())
        assert_true('id="xlsxTable"' in pagef and 'id="xlsxBody"' in pagef,
                    "small tables now carry the ids the page JS resolves by")
        assert_true('id="xlsxInvariant"' in pagef, "…and the toggle")
        # The loci/samples the JS reads must match the sheet: 7 locus columns
        # (2..8), and one stem per <tr> with structural rows blank.
        loci = json.loads(pagef.split("var LOCI = ", 1)[1].split(";", 1)[0])
        assert_eq(sorted(int(k) for k in loci), list(range(2, cols + 1)),
                  "every locus column is in LOCI")
        samples = json.loads(pagef.split("var SAMPLES = ", 1)[1].split(";", 1)[0])
        assert_eq(samples, ["", "", "S1_meta", "S2_meta", "S3_meta", "S4_meta",
                            "", ""],
                  "one stem per row, structural rows blank")
        assert_true(pagef.count("xlsx-igv-cell") > 0,
                    "the per-cell anchors are still how this path launches IGV")

        print("\n[no project: the full path stays aid-free]")
        pagenp = xlsx_html.xlsx_to_html(book)
        assert_true('id="xlsxInvariant"' not in pagenp,
                    "no project context, no variant table, no aids")
        assert_true("var LOCI = {};" in pagenp, "…and empty metadata")

        print("\n[a merged cell disables the aids rather than mismapping them]")
        merged = tmp / "merged.xlsx"
        shutil.copy(book, merged)
        wbm = openpyxl.load_workbook(merged)
        wbm.active.merged_cells.add("A7:B7")   # merge inside the MQ row
        wbm.save(merged)
        wbm.close()
        pagem = xlsx_html.xlsx_to_html(merged, project="p",
                                       samples_with_bams=bams,
                                       samples_with_vcfs=set())
        assert_true("var LOCI = {};" in pagem,
                    "merged sheet ships no cell metadata (DOM positions shift)")
        assert_true('id="xlsxInvariant"' not in pagem, "…and no controls")
        assert_true("S1_meta" in pagem, "…while the table itself still renders")

        print("\nAll xlsx view-control tests passed.")
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
