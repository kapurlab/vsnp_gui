"""Step 1 stats scanner — the engine behind the Step 1 Results pane.

Reads every ``<sample>/*_stats.xlsx`` under a step1 directory and emits one
row dict per sample (newest run wins), exactly the shape the old embedded
pandas scan produced. Two properties make it usable on 8000+ sample projects
where the old scan took 15+ minutes per request and re-parsed everything on
every visit:

  * **A persistent per-file cache** (``<step1>/.qc_stats_cache.json``), keyed
    by relative path + (mtime_ns, size). A revisit parses only new/changed
    workbooks; everything else is served from the cache. The cache lives in
    the project so every user of a shared project benefits from whoever
    scanned it first. Writes are atomic (tmp + rename) and best-effort — a
    read-only project just re-parses.
  * **Parallel parsing** for the cache misses (ProcessPoolExecutor), because
    the dominant cost is openpyxl unzipping/parsing one workbook per sample.

Runs as a SUBPROCESS of the backend (same interpreter: it needs openpyxl,
which ships with the vsnp3 env's pandas), streaming progress lines on stdout:

    P <done> <total>       ... repeated as files are handled
    DONE <n_rows>          ... scan finished; rows written to --out

The backend turns those into the progress the Results pane polls. Stats are
read with openpyxl (read_only) instead of pandas — same header row + first
data row, several times faster, and empty cells arrive as None (pandas gave
NaN, which json-encoded as the invalid token NaN).
"""
from __future__ import annotations

import argparse
import glob
import gzip
import json
import os
import re
import sys
import tempfile
import warnings
from concurrent.futures import ProcessPoolExecutor

# openpyxl warns about vsnp3's workbooks ("no default style", …) — once per
# file, thousands of times per project. Silence them in this process and in
# every worker it forks.
warnings.filterwarnings("ignore")

CACHE_BASENAME = ".qc_stats_cache.json"
_CACHE_VERSION = 2


def _read_stats_row(path: str):
    """Header row + first data row of the first worksheet, as a dict.
    Equivalent to pandas ``read_excel(f).iloc[0].to_dict()`` for these
    single-record sheets. Returns None when unreadable/empty."""
    from openpyxl import load_workbook  # deferred: workers import it once

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        values = next(it, None)
    except Exception:
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if not header or values is None:
        return None
    row = {}
    for k, v in zip(header, values):
        if k is None:
            continue
        row[str(k)] = v
    return row or None


def _run_date(stats_path: str, row: dict) -> str:
    """Authoritative run date: run_metadata.json's started_at (provenance),
    falling back to the xlsx 'date' column, the timestamp embedded in the
    stats filename, and finally the file mtime — same ladder as always."""
    d = os.path.dirname(stats_path)
    try:
        with open(os.path.join(d, "run_metadata.json")) as fh:
            meta = json.load(fh)
        ts = meta.get("started_at") or meta.get("finished_at")
        if ts:
            return str(ts)
    except Exception:
        pass
    dt = row.get("date")
    if dt not in (None, ""):
        return str(dt)
    m = re.search(r"(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})", os.path.basename(stats_path))
    if m:
        return m.group(1) + "T" + m.group(2).replace("-", ":")
    try:
        import datetime

        return datetime.datetime.fromtimestamp(os.path.getmtime(stats_path)).isoformat()
    except Exception:
        return ""


def _sample_fastqs(d: str):
    """The sample's own reads: top-level fastq.gz, else the legacy zips/
    folder pre-GUI pipelines parked the originals in (read-type display
    only — dispatch never runs a sample whose reads aren't at top level)."""
    fqs = [
        x
        for x in glob.glob(os.path.join(d, "*.fastq.gz"))
        if "_unmapped_" not in os.path.basename(x)
    ]
    if not fqs:
        fqs = glob.glob(os.path.join(d, "zips", "*.fastq.gz"))
    return fqs


def _read_type(d: str) -> str:
    """paired / single / ont for the sample dir, cached in
    .provenance/read_type. The ONT sniff (read a little of the fastq and look
    at read lengths) runs only on a cache miss for single-end samples."""
    mk = os.path.join(d, ".provenance", "read_type")
    marker = ""
    try:
        marker = open(mk).read().strip()
    except Exception:
        pass
    fqs = _sample_fastqs(d)
    r2 = [x for x in fqs if re.search(r"(_R2[_.]|_2\.)", os.path.basename(x))]
    if marker in ("paired", "single", "ont"):
        # Heal one historical wrong answer: legacy samples (reads in zips/)
        # were marked 'single' when this check couldn't see zips/ at all.
        if marker == "single" and r2:
            marker = "paired"
            _write_marker(mk, marker)
        return marker
    if r2:
        rt = "paired"
    else:
        rt = "single"
        tgt = fqs[0] if fqs else None
        try:
            if tgt:
                n = 0
                s = 0
                with gzip.open(tgt, "rt") as gh:
                    for i, line in enumerate(gh):
                        if i >= 1600:
                            break
                        if i % 4 == 1:
                            n += 1
                            s += len(line.rstrip())
                if n and s / float(n) > 600:
                    rt = "ont"
        except Exception:
            pass
    _write_marker(mk, rt)
    return rt


def _write_marker(mk: str, value: str) -> None:
    try:
        os.makedirs(os.path.dirname(mk), exist_ok=True)
        with open(mk, "w") as fh:
            fh.write(value)
    except Exception:
        pass


def _parse_one(stats_path: str):
    """Worker: parse one stats workbook into its finished row dict (with
    _file/_sample/read_type/_run_date attached), or None if unreadable."""
    row = _read_stats_row(stats_path)
    if row is None:
        return None
    row["_file"] = stats_path
    sample = row.get("sample") or os.path.basename(stats_path).split("_")[0]
    row["_sample"] = sample
    row["read_type"] = _read_type(os.path.dirname(stats_path))
    row["_run_date"] = _run_date(stats_path, row)
    return row


def _load_cache(path: str) -> dict:
    try:
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        if data.get("version") == _CACHE_VERSION and isinstance(data.get("files"), dict):
            return data["files"]
    except Exception:
        pass
    return {}


def _save_cache(path: str, files: dict) -> None:
    """Atomic, best-effort. A project we can't write to just stays uncached —
    correctness never depends on the cache existing."""
    tmp = None
    try:
        fd, tmp = tempfile.mkstemp(
            prefix=os.path.basename(path) + ".", dir=os.path.dirname(path)
        )
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            json.dump({"version": _CACHE_VERSION, "files": files}, fh, default=str)
        os.replace(tmp, path)
        tmp = None
        try:
            os.chmod(path, 0o664)  # share the cache with the project's group
        except OSError:
            pass
    except Exception:
        if tmp:
            try:
                os.unlink(tmp)
            except Exception:
                pass


def scan(step1_dir: str, cache_path: str, workers: int, include_direct: bool = False):
    """Yields ('progress', done, total) tuples, then ('rows', rows, stats)."""
    patterns = [os.path.join(step1_dir, "*", "*_stats.xlsx")]
    if include_direct:
        patterns.insert(0, os.path.join(step1_dir, "*_stats.xlsx"))
    files = []
    for pat in patterns:
        files.extend(glob.glob(pat))
    files = sorted(set(files))
    total = len(files)

    cached = _load_cache(cache_path) if cache_path else {}
    fresh: dict = {}
    to_parse = []
    rows_by_file: dict = {}
    done = 0
    for f in files:
        rel = os.path.relpath(f, step1_dir)
        try:
            st = os.stat(f)
            sig = [st.st_mtime_ns, st.st_size]
        except OSError:
            continue
        hit = cached.get(rel)
        if hit and hit.get("sig") == sig and isinstance(hit.get("row"), dict):
            row = dict(hit["row"])
            row["_file"] = f  # cache stores relative identity; path is per-mount
            fresh[rel] = {"sig": sig, "row": hit["row"]}
            rows_by_file[f] = row
            done += 1
        else:
            to_parse.append((f, rel, sig))
    yield ("progress", done, total)

    if to_parse:
        workers = max(1, min(workers, len(to_parse)))
        with ProcessPoolExecutor(max_workers=workers) as pool:
            for (f, rel, sig), row in zip(
                to_parse, pool.map(_parse_one, [t[0] for t in to_parse], chunksize=8)
            ):
                done += 1
                if row is not None:
                    rows_by_file[f] = row
                    stored = dict(row)
                    stored.pop("_file", None)
                    fresh[rel] = {"sig": sig, "row": stored}
                if done % 20 == 0 or done == total:
                    yield ("progress", done, total)

    if cache_path:
        _save_cache(cache_path, fresh)

    # Newest run per sample, same rule as always: highest _run_date wins.
    latest: dict = {}
    for row in rows_by_file.values():
        sample = row.get("_sample")
        rd = row.get("_run_date", "") or ""
        if sample not in latest or rd > (latest[sample].get("_run_date", "") or ""):
            latest[sample] = row
    rows = list(latest.values())
    yield ("rows", rows, {"files": total, "parsed": len(to_parse), "cache_hits": total - len(to_parse)})


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("step1_dir")
    ap.add_argument("--cache", default="", help="cache file (default <step1>/.qc_stats_cache.json; '-' disables)")
    ap.add_argument("--out", required=True, help="write the result JSON here")
    ap.add_argument("--workers", type=int, default=0)
    ap.add_argument("--direct", action="store_true", help="also scan *_stats.xlsx directly in step1_dir (post-hoc folders)")
    args = ap.parse_args()

    step1_dir = os.path.abspath(args.step1_dir)
    if not os.path.isdir(step1_dir):
        print(f"not a directory: {step1_dir}", file=sys.stderr)
        return 2
    cache_path = args.cache or os.path.join(step1_dir, CACHE_BASENAME)
    if cache_path == "-":
        cache_path = ""
    workers = args.workers or int(os.environ.get("VSNP_GUI_QC_WORKERS", "0") or 0)
    if workers < 1:
        workers = min(8, max(2, (os.cpu_count() or 2) // 2))

    result = None
    for item in scan(step1_dir, cache_path, workers, include_direct=args.direct):
        if item[0] == "progress":
            print(f"P {item[1]} {item[2]}", flush=True)
        else:
            result = item
    rows, stats = result[1], result[2]
    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump({"rows": rows, **stats}, fh, default=str)
    print(f"DONE {len(rows)}", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
