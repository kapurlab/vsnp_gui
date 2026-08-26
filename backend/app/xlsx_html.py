"""xlsx → self-contained HTML page preserving vSNP3 cascade-table formatting.

Pure function: takes a `Path` to an xlsx file, returns an HTML string
ready to serve as `text/html`. Uses openpyxl (already in the vsnp3 env
as pandas' xlsx engine) — no new dependencies.

When the optional `project` argument is provided AND the rendered sheet
looks like a vSNP3 variant-alignment table (row 1 has `contig:pos`
headers, column 1 has sample names — i.e. `name-All_cascade*` and
`name-All_sorted_*` outputs, or any future table with the same shape),
variant cells (those with a colored fill) gain a hover-revealed pair of
IGV launch links: "this" (single sample) and "all" (every sample in the
table). The cell itself stays non-clickable to prevent gratuitous IGV
launches while scanning.

Preserved formatting:
  - cell fill color
  - font: bold, italic, color, size, family
  - text alignment (h + v)
  - cell borders (collapsed; single 1px outline if any side has a style)
  - merged cells (rowspan / colspan)
  - column widths
  - frozen panes (top row / left columns become `position: sticky`)
  - number formats (basic — ints, floats, percents, dates render as
    displayed value when possible; fallback to raw value)

What we don't try to preserve:
  - rich text within a single cell (the rendered text uses the cell value;
    inline color runs are lost — vSNP3 doesn't use them)
  - charts, images, conditional formatting rules (we read the *resolved*
    fill colors from openpyxl's `data_only=True` workbook load — that
    gets the cached final colors for most spreadsheets we'll see, but
    purely-rule-based conditional formatting without a cached result
    will appear plain)
  - hyperlinks (could be added; not needed for cascade tables)
"""
from __future__ import annotations

import html
import json
import os
import re
from copy import copy
from pathlib import Path
from urllib.parse import quote

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import CellRange


_DEFAULT_FONT_SIZE = 11
_DEFAULT_FONT_COLOR = "FF000000"

# A position header in a vSNP3 table. vsnp3 builds these as
# `abs_pos = CHROM + ':' + POS` (vsnp3_group_on_defining_snps.py), so the part
# before the colon is whatever the reference FASTA calls the contig — and the
# VCF spec is the only constraint on it: no whitespace.
#
# This used to be `[A-Za-z0-9_.\-]+`, which fits a bacterial accession
# (`NC_002945.4:12345`) and NOTHING ELSE. An influenza reference names each
# segment after the isolate it came from — `A/mallard/.../2024(H5N1)_PB2:1234`
# — so not one column of an HPAI table was recognised as a locus, the sheet was
# classified "not a variant table", and the clade filter silently handed back
# the whole table. On a big one it came back as "too large to display — open
# the tree and click a branch", which is what the user had just done.
_LOCUS_RE = re.compile(r"^\S+:\d+$")

# --- Very large sheets -------------------------------------------------------
#
# vSNP3 cascade tables on a big group are enormous: a real one here is 1,047
# samples x 10,001 positions = 10.4 MILLION populated cells in a 35 MB file.
# openpyxl's normal (read_write) load builds a full object graph with per-cell
# style objects, which for that file took 47 seconds and 3.6 GB of RAM before a
# single row was rendered — on a memory-capped HPC session that is the
# "Internal Server Error after ten minutes" users were hitting.
#
# Above STREAM_ABOVE_CELLS we switch to openpyxl's read-only streaming mode
# (the same file opens in 0.09 s), and we render a bounded window of the sheet.
# The window is not a preference — no browser can lay out ten million table
# cells, and each variant cell additionally carries an IGV launch anchor of
# ~450 bytes, so the full grid would be a multi-gigabyte page. The page says
# plainly how much it is showing and links the xlsx for the rest.
# The streaming renderer now reproduces the full renderer's output exactly on
# these files (same colours, same IGV links — verified cell-for-cell), so the
# threshold is set low: the full path's only remaining advantages are merged
# cells and CF rules that reference other cells, neither of which vSNP3 emits,
# and it is the path with no size bound at all.
STREAM_ABOVE_CELLS = 20_000
# The window held on the server. Cells are cheap enough now (~25 bytes, see the
# style palette and delegated IGV handler) that a 1,000 x 1,000 window is a
# ~25 MB document — held server-side and streamed to the page in row batches,
# never sent in one piece.
DEFAULT_MAX_CELLS = 1_000_000
DEFAULT_MAX_ROWS = 1_000
# Rows in the first response. The rest arrive as the user scrolls.
DEFAULT_INITIAL_ROWS = 200
# …and a cap on the RENDERED BYTES, which is what actually decides whether a
# page opens and how much disk a cached preview costs. A cell budget alone
# bounds neither: bytes per cell range from ~30 (a plain cell) to ~450 (a
# variant cell carrying an IGV launch anchor), so a 480 KB sheet of mostly
# variant cells rendered to 48 MB while a 35 MB sheet rendered to 1.4 MB.
DEFAULT_MAX_TABLE_BYTES = 64 * 1024 * 1024
# The largest sheet that is shown AT ALL.
#
# Above this the preview does not render a slice of the table — it says the table
# is too large and offers the two routes that work (the file in a spreadsheet
# application, or the group's tree with a clade selected). Showing the leading
# corner of a SNP table behind a "showing the first ..." banner produces
# something that reads as complete and is not: a table missing 90% of its
# positions, unannounced in every way that matters to the person reading the
# calls. A slow answer is recoverable; a plausible wrong one is not.
#
# Measured here (2026-08-13): ~34 bytes of HTML per rendered cell, so a
# 1,000,000-cell sheet is a ~34 MB document — the outer edge of what is worth
# handing a browser, and about 10% of the 2,316 real tables on this machine are
# above it. Note this is a CELL budget with no row cap: a tall narrow sheet
# (5,000 x 12) is not "too large" and renders in full.
FULL_VIEW_MAX_CELLS = 1_000_000
# Clade-filtered previews (render_filtered_window) must see every column of a
# kept row before they can decide which SNP columns the clade actually uses,
# so kept rows are buffered at full sheet width during the single streaming
# pass. This caps that buffer: above it (selection rows x total columns) the
# per-column SNP filter is skipped and only rows are filtered — a selection
# that big is the too-big table again, and buffering it would recreate the
# memory blow-up the streaming renderer exists to avoid. 2M cells of small
# fragment strings is roughly 150-200 MB transient, measured tolerable here.
FILTER_BUFFER_MAX_CELLS = 2_000_000
_NON_SAMPLE_LABELS = {
    "root", "mq", "annotation", "position not annotated",
    # vsnp3 names the annotation row "no annotations" when the group has no GBK
    # (vsnp3_fasta_to_snps_table.py). Without this it was read as a SAMPLE:
    # dropped from every filtered view — losing the gene / amino-acid annotation
    # for the very positions being inspected — and counted in the denominator.
    "no annotations",
    "n:p207l, orf1ab",  # vSNP3 sometimes carries annotation hints into col 1; skip
}


class FilterMatchError(ValueError):
    """A clade selection matched no sample row of the table at all.

    Raised instead of rendering a table of nothing-but-header-rows: when the
    tree tips and the table labels disagree this loudly, the user needs the
    message, not an empty grid."""


def _strip_vcf_suffix(s: str) -> str:
    """Reduce a variant-table sample label to the stem used by Step 1.

    vSNP3 variant-alignment tables put names like
    `hCoV-19-deer-USA-IA-201788-2020-EPI_zc.vcf` in column 1. Step 1 sample
    directories are named after the bare stem.
    """
    out = s.strip()
    for suffix in (".vcf.gz", ".vcf", "_zc"):
        if out.lower().endswith(suffix):
            out = out[: -len(suffix)]
    return out


def _canonical_stem(label: str, *stem_sets) -> str:
    """Map a variant-table label back to the bare sample stem used on disk.

    vSNP3 step2 relabels samples with descriptive metadata
    (`ERR930304_Sterne_A_Br_75_..._Denmark`), while step1 dirs and
    `vcf_database/` files are named after the bare id (`ERR930304`). Imported
    projects therefore show long labels in the cascade table that never match
    the on-disk stems — blanking "this" and tripping a false "no Step 1
    outputs". Resolve the label to the longest known stem ``S`` such that
    ``label == S`` or ``label`` starts with ``S + "_"``; fall back to the label
    unchanged when nothing matches (or when no stem sets were supplied).

    Underscores and dashes are treated as the same character in the LAST
    resort, and that clause is the difference between influenza tables having
    IGV links and not having them. Step 1 deliberately dashes a sample prefix
    when it stages reads (`_sanitized_sample_and_name`) — without that, vSNP3
    splits the sample name at the first underscore and silently merges every
    `Mg_280`, `Mg_281`, … into one sample called `Mg`. So a specimen submitted
    as `26G02488-004_DUCK_2026-02-11_Elkhart-20-p11_IN` is on disk as
    `26G02488-004-DUCK-2026-02-11-Elkhart-20-p11-IN`, all dashes, while a table
    or tree label that came from anywhere else — an imported VCF, an older
    run — still carries the underscores. Neither the equality test nor the
    prefix test can bridge that, so the label resolved to itself, matched no
    BAM and no VCF, and every cell in the row was rendered unclickable with no
    hover and no explanation.

    M. bovis and Brucella sample names (`SRR33643035`, `19-1234`) contain no
    underscores at all, so the dashing is a no-op and label == folder. That,
    and not the multi-segment reference, is why those tables always worked.

    The normalised comparison is tried only after the exact and prefix rules
    have both failed, so it can never change an answer that was already
    right. It also cannot alias two genuinely distinct samples: a pair
    differing only in `_` versus `-` cannot both exist as Step 1 samples,
    because staging would have collapsed them to the same folder."""
    known: set[str] = set()
    for s in stem_sets:
        if s:
            known |= s
    if not known or label in known:
        return label
    best: str | None = None
    for s in known:
        if label.startswith(s + "_") and (best is None or len(s) > len(best)):
            best = s
    if best is not None:
        return best
    flat = _flatten_sep(label)
    for s in known:
        fs = _flatten_sep(s)
        if flat == fs or flat.startswith(fs + "-"):
            if best is None or len(s) > len(best):
                best = s
    if best is not None:
        return best

    # Last rule, and the reverse of every rule above: the on-disk name EXTENDS
    # the label's leading id rather than being extended by it.
    #
    # Step 2 relabels a sample with metadata after its id
    # (`24-029315-007_GWTE_2024-09-26_AH0238161_AK`) while Step 1 staged it
    # under a name carrying a suffix of its own
    # (`24-029315-007-original`). Neither string is a prefix of the other, so
    # nothing above can match them, and every row of such a table rendered
    # unclickable — no link, no hover, and no banner either, because nothing
    # resolved far enough to be called ambiguous. Measured on a real HPAI
    # cascade table: 0 of 8 sample rows matched.
    #
    # A separator after the id is required, so `24-029315-007` cannot claim
    # `24-029315-0071-original`. More than one candidate means the id does not
    # identify a sample, and no link is better than the wrong one.
    head = label.split("_", 1)[0]
    if head and head != label:
        cands = [s for s in known
                 if s == head or s.startswith(head + "-") or s.startswith(head + "_")]
        if len(cands) == 1:
            return cands[0]
    return label


def _flatten_sep(s: str) -> str:
    """Underscores and dashes as one character, for stem comparison only."""
    return s.replace("_", "-")


def _selection_key(label: str, *stem_sets) -> str:
    """Normalise a tree tip / table row label for clade-selection matching.

    Tree tips and table rows both come out of the same vSNP3 step 2 run, but
    they are not always spelled identically: either side may carry a
    `_zc.vcf` suffix, and either side may be decorated with metadata
    (`DRR184883_Vietnam_Asia_L4`) — the GUI's re-labeled trees decorate from
    a different source than vsnp3 decorates its tables. Reducing both sides
    to the canonical on-disk stem (when the stem is known from step1 BAMs or
    the imported VCF database) makes differently-decorated spellings of the
    same sample compare equal; unknown stems fall back to the suffix-stripped
    label itself."""
    stem = _strip_vcf_suffix(str(label).strip())
    return _canonical_stem(stem, *stem_sets)


def fits_full_view(total_rows: int, total_cols: int,
                   max_cells: int = FULL_VIEW_MAX_CELLS) -> bool:
    """Can this sheet be shown in its entirety?

    Answered from the declared dimensions alone, so the caller can decide
    before paying to open anything — on the 35 MB cascade table the verdict
    costs 0.13 s where rendering the old truncated window cost 42.

    A zero/unknown extent returns False: nothing is claimed to fit on the
    strength of an extent that could not be read.
    """
    if total_rows < 1 or total_cols < 1:
        return False
    return total_rows * total_cols <= max_cells


def too_large_page(filename: str, total_rows: int, total_cols: int,
                   tree_url: str | None = None, tree_name: str | None = None,
                   clade_samples: int | None = None,
                   download_url: str | None = None,
                   clade_positions: int | None = None,
                   filter_ignored: str | None = None) -> str:
    """The page shown instead of a table that cannot be shown in full.

    Two routes, because both are real: the file itself in a spreadsheet
    application, and the group's tree with one clade selected (which is what
    makes a table this size legible in the browser at all).

    ``clade_samples`` set means this was ALREADY filtered to a clade and is
    still too large — the advice becomes "select a smaller clade", not "select
    a clade".

    ``filter_ignored`` set means a clade WAS selected and the filter could not
    be applied to this sheet at all. The tree route is then removed rather than
    reworded: it is the page's job not to send someone back to do the thing
    they have just done. That loop is exactly what an unrecognised position
    header produced — click a branch, get "open the tree and click a branch".
    """
    cells = total_rows * total_cols
    positions = max(0, total_cols - 1)      # column 1 holds the sample names
    if cells >= 1_000_000:
        size_words = f"{cells / 1_000_000:.1f} million cells"
    else:
        size_words = f"{cells:,} cells"

    if filter_ignored:
        heading = "This table is too large to display"
        lede = (
            f"<strong>{total_rows:,} rows × {total_cols:,} columns</strong> — "
            f"{size_words}, more than a browser can lay out. A clade was selected, "
            f"but the filter could not be applied to this sheet: "
            f"{html.escape(filter_ignored)}. What you are being told the size of is "
            "therefore the whole table, not your clade."
        )
        tree_lead = "Selecting a clade cannot narrow this sheet"
        tree_body = (
            "The clade filter works by keeping only the position columns a clade "
            "uses, and this sheet has none it can recognise. Open the file in a "
            "spreadsheet application instead."
        )
    elif clade_samples is not None:
        # The clade's OWN matching positions when the filter got far enough to
        # count them; the sheet's position count otherwise (which is the case
        # where the selection was too large to filter by position at all).
        shown_positions = clade_positions if clade_positions else positions
        cells = max(1, clade_samples) * max(1, shown_positions)
        size_words = (f"{cells / 1_000_000:.1f} million cells" if cells >= 1_000_000
                      else f"{cells:,} cells")
        heading = "This clade is still too large to display"
        lede = (
            f"The selected clade has <strong>{clade_samples:,} samples</strong> across "
            f"<strong>{shown_positions:,} positions</strong> — {size_words} once laid "
            "out, more than a browser can render."
        )
        tree_lead = "Select a smaller clade"
        tree_body = (
            "Go back to the tree and click a branch further out. A clade of a few "
            "dozen samples typically opens in under a second, showing only the "
            "positions where those samples differ."
        )
    else:
        heading = "This table is too large to display"
        lede = (
            f"<strong>{total_rows:,} rows × {total_cols:,} columns</strong> — "
            f"{size_words}, and {cells:,} exactly. A browser cannot lay out a table "
            "of that size, and showing only its first corner would read as the whole "
            "table while leaving most of the positions out of view."
        )
        tree_lead = "View one clade at a time"
        tree_body = (
            "Open the group's tree and click a branch. The table then opens showing "
            "only that clade's samples and only the positions where they differ — "
            "which is usually the comparison you were after."
        )

    dl = html.escape(download_url or "?download=1", quote=True)
    tree_block = ""
    if filter_ignored:
        # No "Open the tree" button: the tree is where this request came from.
        tree_block = ""
    elif tree_url:
        tree_block = (
            f'<a class="tl-btn" href="{html.escape(tree_url, quote=True)}" '
            f'target="_blank" rel="noopener">Open the tree</a>'
            + (f'<div class="tl-note">{html.escape(tree_name)}</div>' if tree_name else "")
        )
    else:
        tree_block = ('<div class="tl-note">The group\'s tree files (<code>.tre</code>) '
                      'are listed beside this table in Step 2 Results.</div>')

    return _TOO_LARGE_TEMPLATE.format(
        title=html.escape(filename),
        filename=html.escape(filename),
        heading=html.escape(heading),
        lede=lede,
        dims=f"{total_rows:,} × {total_cols:,}",
        download_url=dl,
        tree_lead=html.escape(tree_lead),
        tree_body=html.escape(tree_body),
        tree_block=tree_block,
    )


def window_is_partial(window: dict) -> bool:
    """Does this rendered window show less than it set out to show?

    For a clade-filtered window, showing fewer rows than the sheet has is the
    whole point — so what counts as partial there is the filter having had to
    give something up: columns or rows dropped for budget, or the per-position
    filter skipped entirely because the selection was too large to buffer.

    Callers use this to replace a partial table with the too-large page, so it
    has to be evaluated on cached windows too, not just freshly rendered ones.
    """
    filt = window.get("filter")
    if filt and not filt.get("ignored"):
        return bool(filt.get("truncated_cols") or filt.get("truncated_rows")
                    or not filt.get("column_filter"))
    return (window["shown_rows"] < window["total_rows"]
            or window["shown_cols"] < window["total_cols"])


def message_page(heading: str, body: str, filename: str = "",
                 back_hint: str = "") -> str:
    """A plain, styled page for a preview that cannot be produced.

    These endpoints are opened as PAGES — a user clicks a button and a tab
    navigates to them — so an error has to arrive as something a person can
    read. FastAPI's default handler answers HTTPException with JSON, which in a
    fresh tab is a line of raw `{"detail": …}` on white with no way back; the
    two cases that reach it (an expired clade selection, and a selection that
    matches no row in this table) are precisely the ones needing an explanation.
    """
    return _MESSAGE_TEMPLATE.format(
        title=html.escape(filename or heading),
        bar=(f'<div class="tl-bar">{html.escape(filename)}</div>' if filename else ""),
        heading=html.escape(heading),
        body=html.escape(body),
        hint=(f'<div class="tl-note">{html.escape(back_hint)}</div>'
              if back_hint else ""),
    )


def sheet_extent(xlsx_path: Path) -> tuple[int, int]:
    """(rows, cols) of the active sheet, without loading its cells.

    Read-only mode reads only the sheet's declared dimension, so this is
    effectively free even on a 35 MB workbook — it is what lets the caller
    decide between the full-fidelity and streaming renderers before paying
    for either.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        return int(ws.max_row or 0), int(ws.max_column or 0)
    finally:
        wb.close()


def _sheet_layout(xlsx_path: Path) -> dict:
    """Column widths and frozen panes, straight from the sheet XML.

    Read-only worksheets expose neither `column_dimensions` nor
    `freeze_panes`, but both live in the sheet XML's header — before
    `<sheetData>` — so they can be read from the first few KB of the entry
    without touching the millions of cells that follow. Returns
    {"widths": {col_index: width_units}, "default_width": float|None,
     "freeze_row": int, "freeze_col": int}; empty/zero values on any problem,
    since layout is cosmetic and must never break a preview.
    """
    out = {"widths": {}, "default_width": None, "freeze_row": 0, "freeze_col": 0}
    try:
        import zipfile
        from xml.etree import ElementTree as ET

        with zipfile.ZipFile(xlsx_path) as zf:
            name = next((n for n in zf.namelist()
                         if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")),
                        None)
            if not name:
                return out
            with zf.open(name) as fh:
                for event, el in ET.iterparse(fh, events=("start", "end")):
                    tag = el.tag.split("}")[-1]
                    if event == "start" and tag == "sheetData":
                        break                      # everything we want precedes the cells
                    if event != "end":
                        continue
                    if tag == "sheetFormatPr":
                        w = el.get("defaultColWidth")
                        if w:
                            out["default_width"] = float(w)
                    elif tag == "col":
                        w = el.get("width")
                        if w:
                            lo = int(el.get("min", 1))
                            hi = int(el.get("max", lo))
                            # Guard against the "col span to 16384" idiom that
                            # would otherwise build a dict of every column.
                            for ci in range(lo, min(hi, lo + 20_000) + 1):
                                out["widths"][ci] = float(w)
                    elif tag == "pane":
                        out["freeze_row"] = int(float(el.get("ySplit") or 0))
                        out["freeze_col"] = int(float(el.get("xSplit") or 0))
                    el.clear()
    except Exception:
        return {"widths": {}, "default_width": None, "freeze_row": 0, "freeze_col": 0}
    return out


class _XmlCfRule:
    """A conditional-formatting rule parsed straight from the sheet XML.

    Deliberately not an openpyxl Rule: it only carries the attributes
    _cf_rule_matches() reads, so the streaming renderer can evaluate rules
    without the full (read-write) workbook load that exposes them normally.
    """
    __slots__ = ("type", "operator", "text", "formula", "dxfId", "priority")

    def __init__(self, el):
        self.type = el.get("type")
        self.operator = el.get("operator")
        self.text = el.get("text")
        try:
            self.dxfId = int(el.get("dxfId")) if el.get("dxfId") is not None else None
        except ValueError:
            self.dxfId = None
        try:
            self.priority = int(el.get("priority") or 0)
        except ValueError:
            self.priority = 0
        self.formula = [
            (child.text or "") for child in el
            if child.tag.split("}")[-1] == "formula"
        ]


class _NoRefSheet:
    """Stand-in worksheet for CF evaluation while streaming.

    _resolve_cf_value() falls back to reading another cell for rules whose
    comparison value is a cell reference. Read-only sheets have no random
    access, so those rules simply don't resolve here.
    """
    def cell(self, *_args, **_kwargs):
        raise LookupError("no random access while streaming")


class _ValueCell:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class _CapturedRowsSheet:
    """A worksheet shim that answers cell refs from rows captured in passing.

    The colour-deciding rule in a real vSNP3 table is `cellIs equal B$2` —
    "this call matches the reference row" — an ANCHOR-ROW reference, not a
    literal. A read-only pass cannot look back at row 2 on demand, but it
    doesn't have to: row 2 streams past before any row whose rule needs it,
    so the renderer hands the rows that CF formulas name (by absolute row)
    to this shim as they go by. Refs into rows that were not captured raise,
    which _resolve_cf_value already treats as "rule doesn't resolve" — the
    behaviour all streamed CF evaluation had before this shim existed.
    """
    def __init__(self):
        self._rows: dict[int, dict[int, object]] = {}

    def capture(self, row_idx: int, row) -> None:
        vals: dict[int, object] = {}
        for col_idx, cell in enumerate(row, start=1):
            v = getattr(cell, "value", None)
            if v is not None:
                vals[col_idx] = v
        self._rows[row_idx] = vals

    def cell(self, row: int, column: int):
        vals = self._rows.get(row)
        if vals is None:
            raise LookupError("row not captured while streaming")
        return _ValueCell(vals.get(column))


def _cf_absolute_ref_rows(cf_ranges: list) -> set[int]:
    """Row numbers that CF formulas pin with an absolute row (e.g. B$2).

    These are the rows worth capturing during a streaming pass; relative-row
    references can point anywhere and stay unresolvable, as before."""
    rows: set[int] = set()
    for _cr, rules in cf_ranges:
        for rule in rules:
            if rule.type != "cellIs" or not rule.formula:
                continue
            parsed = _parse_single_ref(str(rule.formula[0]).strip())
            if parsed is None:
                continue
            _col_abs, _col_letter, row_abs, row_num = parsed
            if row_abs and row_num >= 1:
                rows.add(row_num)
    return rows


def _cf_from_sheet_xml(xlsx_path: Path) -> list[tuple[object, list]]:
    """Conditional-formatting ranges + rules, read from the sheet XML.

    This exists because EVERY colour in a vSNP3 SNP table comes from
    conditional formatting — the cells carry no direct fill at all — and
    openpyxl's read-only worksheets do not expose `conditional_formatting`.
    Without this the streaming renderer produced a perfectly laid out,
    completely colourless table, and no IGV links (which are gated on a cell
    having a background).

    In the OOXML schema `conditionalFormatting` elements follow `sheetData`,
    so they sit near the END of the sheet part. The stream is decompressed
    once, keeping only a rolling tail, and the blocks are parsed out of that —
    no cell objects are built.

    Returns [(CellRange, [rules]), ...]; empty on any problem, because a
    preview without colour is bad but a preview that raises is worse.
    """
    TAIL = 8 * 1024 * 1024
    try:
        import zipfile
        from xml.etree import ElementTree as ET

        with zipfile.ZipFile(xlsx_path) as zf:
            name = next((n for n in zf.namelist()
                         if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")),
                        None)
            if not name:
                return []
            tail = b""
            with zf.open(name) as fh:
                while True:
                    chunk = fh.read(4 * 1024 * 1024)
                    if not chunk:
                        break
                    tail = (tail + chunk)[-TAIL:]
        text = tail.decode("utf-8", errors="replace")
        out: list[tuple[object, list]] = []
        for m in re.finditer(r"<conditionalFormatting\b.*?</conditionalFormatting>",
                             text, re.S):
            block = m.group(0)
            try:
                el = ET.fromstring(block)
            except ET.ParseError:
                continue
            sqref = el.get("sqref") or ""
            rules = [_XmlCfRule(child) for child in el
                     if child.tag.split("}")[-1] == "cfRule"]
            if not rules:
                continue
            for part in str(sqref).split():
                try:
                    out.append((CellRange(part), rules))
                except Exception:
                    continue
        return out
    except Exception:
        return []


def _detect_variant_table(ws) -> dict | None:
    """Return variant-alignment-table metadata if `ws` matches the shape.

    Matches vSNP3 `name-All_cascade*_table-*.xlsx` and
    `name-All_sorted_table-*.xlsx` outputs (and any future table with the
    same shape — detection is pattern-based, not filename-based).

    Heuristic: row 1 has ≥2 cells matching `<contig>:<pos>`, and column 1
    (rows 2+) has sample-like labels. Returns:

        {
          "positions": {col_idx: "contig:pos", ...},
          "samples":   {row_idx: sample_stem, ...},  # in row order
        }

    or None when the sheet doesn't look like a variant-alignment table — in which
    case the renderer skips the IGV-link injection entirely.
    """
    if ws.max_row < 2 or ws.max_column < 2:
        return None
    positions: dict[int, str] = {}
    for cell in ws[1]:
        v = cell.value
        if v is None:
            continue
        s = str(v).strip()
        if _LOCUS_RE.match(s):
            positions[cell.column] = s
    if len(positions) < 2:
        return None
    samples: dict[int, str] = {}
    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=1).value
        if v is None:
            continue
        raw = str(v).strip()
        if not raw or raw.lower() in _NON_SAMPLE_LABELS:
            continue
        stem = _strip_vcf_suffix(raw)
        if not stem:
            continue
        samples[row_idx] = stem
    if not samples:
        return None
    return {"positions": positions, "samples": samples}


def _igv_cell_html(
    value: str,
    project: str,
    this_stem: str,
    locus: str,
    this_loadable: bool = True,
    this_calls_only: bool = False,
) -> str:
    """Wrap a variant cell's value in a full-cell IGV-launch anchor.

    The whole cell is clickable — clicking anywhere in a colored variant
    cell opens the row's sample in IGV at this locus. (Previously a small
    "↗ this" text link in the corner was the only target; selecting the
    SNP cell itself is more direct.) The anchor fills the cell via
    ``display:block`` and inherits the cell's color/alignment so the
    nucleotide letter still reads normally.

    Same additive single-window behavior as before: the anchor uses
    ``target="vsnp_igv"`` (a named window) and an onclick that posts to the
    existing IGV tab when one is open, so variant-after-variant clicks build
    up the cohort in one IGV view rather than spawning N tabs. Modifier-click
    (cmd / ctrl / shift / middle button) bypasses the handler so users can
    still force a fresh tab.

    The URL is constructed relative to the current preview path so it
    survives the OOD proxy prefix: the preview is served at
    ``/api/projects/{p}/preview-xlsx``, so ``../../../`` climbs back out
    to the SPA root regardless of the OOD rnode prefix in front of it
    (``/rnode/host/port/api/projects/p/...``).

    ``this_loadable``: false only when the sample has NEITHER a BAM nor an
    imported VCF — the cell renders plain (not clickable) with an
    explanatory tooltip, since clicking would land in an empty IGV.
    ``this_calls_only``: true when the sample has an imported VCF but no
    BAM; the cell stays clickable (calls-only IGV anchored to the project
    reference) and is italicized with a tooltip noting the limitation.
    """
    if not this_loadable:
        return (
            '<span class="xlsx-igv-disabled" '
            f'title="No data for {html.escape(this_stem)} — neither Step 1 BAM '
            f'nor imported VCF in step2/vcf_database/.">{value}</span>'
        )
    enc_proj = quote(project, safe="")
    enc_locus = quote(locus, safe="")
    onclick = (
        "if(event.metaKey||event.ctrlKey||event.shiftKey||event.button===1)return true;"
        "window.__vsnpLaunchIgv(this.href);return false;"
    )
    this_track = f"{enc_proj}:{quote(this_stem, safe='')}"
    this_href = f"../../../?view=igv&tracks={this_track}&locus={enc_locus}"
    if this_calls_only:
        cls = "xlsx-igv-cell xlsx-igv-calls-only"
        title = (
            f"Open {html.escape(this_stem)} in IGV at {html.escape(locus)} "
            "(calls only — imported VCF, no BAM to load)"
        )
    else:
        cls = "xlsx-igv-cell"
        title = f"Open {html.escape(this_stem)} in IGV at {html.escape(locus)}"
    return (
        f'<a class="{cls}" href="{html.escape(this_href, quote=True)}" '
        f'target="vsnp_igv" rel="noopener" onclick="{onclick}" '
        f'title="{title}">{value}</a>'
    )


def _rgb_hex(color) -> str | None:
    """Extract a `#rrggbb` string from openpyxl Color, or None if absent/transparent."""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return None
    # ARGB form like 'FFRRGGBB' → take last 6 chars
    if len(rgb) == 8:
        if rgb[:2] == "00":  # fully transparent
            return None
        return f"#{rgb[-6:].lower()}"
    if len(rgb) == 6:
        return f"#{rgb.lower()}"
    return None


def _cell_rotation_class(cell) -> str:
    """Map openpyxl's text_rotation (Excel encoding) to one of our CSS classes.

    Excel OOXML stores rotation as:
      - 0          → horizontal
      - 1..90      → that many degrees counter-clockwise from horizontal
                     (90 = reads bottom-to-top)
      - 91..180    → (rotation - 90) degrees clockwise from horizontal
                     (180 = reads top-to-bottom)
      - 255        → vertically stacked characters
    vSNP3 cascade tables use 90 on the variant-position header row and 180
    on the annotation row. We map both to vertical writing-mode + transform
    so the cell becomes a narrow tall column.
    """
    rot = (cell.alignment.text_rotation or 0) if cell.alignment else 0
    if rot == 0:
        return ""
    if 1 <= rot <= 90:
        return "xlsx-rot-up"
    if 91 <= rot <= 180:
        return "xlsx-rot-down"
    if rot == 255:
        return "xlsx-rot-stacked"
    return ""


def _cell_inline_style(cell) -> str:
    parts: list[str] = []

    # Fill (background)
    fill = cell.fill
    if fill and getattr(fill, "patternType", None) in ("solid", "darkVertical", "darkHorizontal"):
        bg = _rgb_hex(fill.fgColor)
        if bg:
            parts.append(f"background-color: {bg}")

    # Font
    font = cell.font
    if font:
        if font.bold:
            parts.append("font-weight: 600")
        if font.italic:
            parts.append("font-style: italic")
        fc = _rgb_hex(font.color)
        if fc and fc != "#000000":
            parts.append(f"color: {fc}")
        if font.size and font.size != _DEFAULT_FONT_SIZE:
            parts.append(f"font-size: {float(font.size):.1f}px")
        if font.name and font.name not in ("Calibri", "Arial"):
            parts.append(f"font-family: '{font.name}', sans-serif")

    # Alignment
    align = cell.alignment
    if align:
        if align.horizontal:
            parts.append(f"text-align: {align.horizontal}")
        if align.vertical:
            vmap = {"center": "middle", "top": "top", "bottom": "bottom"}
            parts.append(f"vertical-align: {vmap.get(align.vertical, align.vertical)}")
        if align.wrap_text:
            parts.append("white-space: pre-wrap")

    # Borders — collapse to a single hint when any side has a style
    border = cell.border
    if border:
        for side in ("top", "right", "bottom", "left"):
            b = getattr(border, side, None)
            if b and b.style:
                bc = _rgb_hex(b.color) or "#888"
                parts.append(f"border-{side}: 1px solid {bc}")

    return "; ".join(parts)


def _format_cell_value(cell) -> str:
    """Render a cell value as HTML-safe text, honoring the number format when straightforward."""
    v = cell.value
    if v is None:
        return ""
    nf = (cell.number_format or "General")
    # Percent
    if isinstance(v, (int, float)) and "%" in nf:
        # heuristic: '0.00%' style means multiply by 100; vSNP3 usually stores
        # the displayed value already, so be conservative — only multiply
        # when value is in [0, 1].
        try:
            if 0 <= float(v) <= 1:
                return html.escape(f"{float(v) * 100:.1f}%")
        except (TypeError, ValueError):
            pass
    # Integer-like
    if isinstance(v, float) and v.is_integer():
        return html.escape(str(int(v)))
    # Non-integer float with no specific number format → vsnp3's cascade-table
    # MQ row produces values like 59.5833333333 / 59.9090909091 because it's
    # averaging integer-bounded scores across samples. Python's `str()` gives
    # full precision repr which is unreadable in the preview AND produces
    # variable-width cells (59.58 vs 60 vs 59.5) that look ragged. If the
    # cell author didn't specify a format, round to a whole number — the
    # MQ-average use case is what produces these values and 1-point
    # differences (60 vs 59) aren't biologically meaningful.
    if isinstance(v, float) and nf in ("General", "@"):
        return html.escape(str(int(round(v))))
    return html.escape(str(v))


def _parse_single_ref(formula: str) -> tuple[bool, str, bool, int] | None:
    """Parse a single-cell ref like B$2 / $A2 / A1 / $C$5.

    Returns (col_absolute, col_letter, row_absolute, row_num) or None if not a single ref."""
    s = formula.strip()
    m = re.fullmatch(r"(\$?)([A-Z]+)(\$?)(\d+)", s)
    if not m:
        return None
    col_abs_marker, col_letter, row_abs_marker, row_num = m.groups()
    return (col_abs_marker == "$", col_letter, row_abs_marker == "$", int(row_num))


def _resolve_cf_value(formula: str, ws, anchor_row: int, anchor_col: int, cell_row: int, cell_col: int):
    """Resolve a CF rule's formula to a Python literal for comparison.

    Handles: number literal, quoted string, single cell ref (relative or absolute
    parts honoured). Returns the resolved value or None when we can't evaluate."""
    s = formula.strip()
    # Number
    try:
        if "." in s or "e" in s.lower():
            return float(s)
        return int(s)
    except ValueError:
        pass
    # Quoted string
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        return s[1:-1]
    # Single cell ref
    parsed = _parse_single_ref(s)
    if parsed is None:
        return None
    col_abs, col_letter, row_abs, row_num = parsed
    ref_col = column_index_from_string(col_letter)
    ref_row = row_num
    if not col_abs:
        ref_col += (cell_col - anchor_col)
    if not row_abs:
        ref_row += (cell_row - anchor_row)
    if ref_row < 1 or ref_col < 1:
        return None
    try:
        return ws.cell(row=ref_row, column=ref_col).value
    except Exception:
        return None


def _cf_rule_matches(rule, cell, ws, anchor_row: int, anchor_col: int) -> bool:
    """Return True if a conditional-formatting rule applies to a cell."""
    cv = cell.value
    if rule.type == "containsText":
        target = rule.text
        if target is None or cv is None:
            return False
        return str(target) in str(cv)
    if rule.type == "notContainsText":
        target = rule.text
        if target is None:
            return False
        return str(target) not in str(cv or "")
    if rule.type == "cellIs":
        if not rule.formula:
            return False
        rhs = _resolve_cf_value(rule.formula[0], ws, anchor_row, anchor_col, cell.row, cell.column)
        if rhs is None:
            return False
        op = rule.operator
        try:
            if op == "equal":
                return cv == rhs or str(cv) == str(rhs)
            if op == "notEqual":
                return cv != rhs and str(cv) != str(rhs)
            lhs_f = float(cv) if cv is not None else None
            rhs_f = float(rhs)
            if lhs_f is None:
                return False
            return {
                "lessThan": lhs_f < rhs_f,
                "lessThanOrEqual": lhs_f <= rhs_f,
                "greaterThan": lhs_f > rhs_f,
                "greaterThanOrEqual": lhs_f >= rhs_f,
            }.get(op, False)
        except (TypeError, ValueError):
            return False
    return False


def _dxf_style_fragments(dxf) -> list[str]:
    """Translate an openpyxl differential format (dxf) into CSS fragments."""
    out: list[str] = []
    if dxf is None:
        return out
    fill = getattr(dxf, "fill", None)
    if fill:
        # CF dxf fills typically expose the color via bgColor; fall back to fgColor.
        for color_attr in ("bgColor", "fgColor"):
            color = getattr(fill, color_attr, None)
            hexc = _rgb_hex(color)
            if hexc:
                out.append(f"background-color: {hexc}")
                break
    font = getattr(dxf, "font", None)
    if font:
        fc = _rgb_hex(getattr(font, "color", None))
        if fc and fc != "#000000":
            out.append(f"color: {fc}")
        # openpyxl's dxf font uses .b for bold (not .bold)
        if getattr(font, "b", None) is True:
            out.append("font-weight: 600")
        if getattr(font, "i", None) is True:
            out.append("font-style: italic")
    return out


def _build_cf_extras(ws, dxfs) -> dict[str, list[str]]:
    """Walk every conditional-formatting range × cell, applying first-match rule.

    Returns a dict mapping cell coordinate → list of CSS style fragments to
    append to that cell's inline style.
    """
    extras: dict[str, list[str]] = {}
    try:
        cf_rules = ws.conditional_formatting._cf_rules
    except Exception:
        return extras
    for cf_range, rules in cf_rules.items():
        sqref = cf_range.sqref if hasattr(cf_range, "sqref") else str(cf_range)
        # sqref may be space-separated multi-range ("A1:A5 B1:B5"). Split.
        for r_str in str(sqref).split():
            try:
                cr = CellRange(r_str)
            except Exception:
                continue
            anchor_row, anchor_col = cr.min_row, cr.min_col
            for r in range(cr.min_row, cr.max_row + 1):
                for c in range(cr.min_col, cr.max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is None or cell.value == "":
                        continue
                    # First-match-wins (Excel's default with priority-ordered rules).
                    for rule in rules:
                        if _cf_rule_matches(rule, cell, ws, anchor_row, anchor_col):
                            dxf_id = getattr(rule, "dxfId", None)
                            if dxf_id is None or dxf_id < 0 or dxf_id >= len(dxfs):
                                break
                            frags = _dxf_style_fragments(dxfs[dxf_id])
                            if frags:
                                extras.setdefault(cell.coordinate, []).extend(frags)
                            break
    return extras


def _cf_dxf_for(cell, row_idx: int, col_idx: int,
                cf_ranges: list, dxfs: list, cf_sheet):
    """The differential format a cell picks up from conditional formatting.

    First match wins, mirroring _build_cf_extras — but evaluated per cell as the
    sheet streams past rather than by walking every CF range up front, because
    the ranges in these files span the whole table and materialising them would
    defeat the point of streaming.

    Returns the dxf itself rather than styling for one output format. In a real
    vSNP3 table NOTHING is coloured by a static fill: every colour on screen
    comes from these rules, so both the HTML preview and the xlsx export of a
    clade have to resolve them, and they must resolve them identically.
    """
    if not cf_ranges:
        return None
    for cr, rules in cf_ranges:
        if not (cr.min_row <= row_idx <= cr.max_row
                and cr.min_col <= col_idx <= cr.max_col):
            continue
        for rule in sorted(rules, key=lambda r: r.priority or 0):
            try:
                if not _cf_rule_matches(rule, cell, cf_sheet, cr.min_row, cr.min_col):
                    continue
            except Exception:
                continue
            if rule.dxfId is None or rule.dxfId >= len(dxfs):
                continue
            dxf = dxfs[rule.dxfId]
            if dxf is not None:
                return dxf
    return None


def _cf_fragments_for(cell, row_idx: int, col_idx: int,
                      cf_ranges: list, dxfs: list, cf_sheet) -> list[str]:
    """CSS fragments a cell picks up from conditional formatting."""
    dxf = _cf_dxf_for(cell, row_idx, col_idx, cf_ranges, dxfs, cf_sheet)
    return _dxf_style_fragments(dxf) if dxf is not None else []


def ambiguous_stems(row_labels) -> dict:
    """Stems that more than one distinct row label resolves to.

    A row whose label is not itself an on-disk sample name was matched by
    canonicalisation — its label merely STARTS WITH a known stem. When two
    labels reach the same stem that way, neither cell can be given an IGV link:
    a click would open one specimen's alignment while claiming to be the
    other's. Influenza labels make this common, because they are underscore
    joins led by a submission id, so `26G02776-002_DUCK_2026-02-17_Elkhart-21_IN`
    and `26G02776-007_DUCK_2026-02-17_Elkhart-21_IN` both extend `26G02776`.

    A row that IS the stem exactly is unaffected: it is that sample, and it
    keeps its link.

    @param row_labels iterable of (raw_label, resolved_stem)
    @returns {stem: sorted labels that collide on it}
    """
    reached: dict = {}
    for raw, stem in row_labels:
        if not stem:
            continue
        if _strip_vcf_suffix(str(raw).strip()) == stem:
            continue                      # exact: this row really is that sample
        reached.setdefault(stem, set()).add(str(raw).strip())
    return {k: sorted(v) for k, v in reached.items() if len(v) > 1}


def withheld_note(ambiguous: dict, rows: int) -> str:
    """Why IGV links were withheld, in the user's terms.

    Withholding them SILENTLY is the failure this exists to prevent. A table
    where clicking a SNP does nothing and nothing even hovers is
    indistinguishable from a broken build, and was reported as one.
    """
    if not ambiguous or not rows:
        return ""
    stems = sorted(ambiguous)
    shown = ", ".join(f"<code>{html.escape(s)}</code>" for s in stems[:3])
    if len(stems) > 3:
        shown += f", and {len(stems) - 3} more"
    return (
        f'<strong>IGV links are off for {rows:,} row(s).</strong> Their names '
        f'each extend the same Step 1 sample folder ({shown}), so a click could '
        'not tell which specimen you meant and would risk opening the wrong '
        'alignment. Open those samples from the Step 1 results pane instead.'
    )


def _delink(cells):
    """Strip the IGV affordance from a rendered row's cells."""
    return [
        c.replace("xlsx-variant xlsx-igv-calls-only", "xlsx-igv-none")
         .replace("xlsx-variant", "xlsx-igv-none")
        for c in cells
    ]


def render_window(
    xlsx_path: Path,
    total_rows: int,
    total_cols: int,
    title: str | None,
    project: str | None,
    samples_with_bams: set[str] | None,
    samples_with_vcfs: set[str] | None,
    max_cells: int,
    max_rows: int,
    max_table_bytes: int = None,
) -> str:
    """Render a bounded window of a very large sheet in one streaming pass.

    Same per-cell appearance and same IGV behaviour as the full renderer — it
    reuses the identical style/value/link helpers, which all work on
    read-only cells. What it cannot do is random access, so the
    variant-table detection (row 1 = loci, column 1 = samples) is built up
    during the pass instead of probed up front.

    Conditional-formatting overlays are skipped: the rules are not readable
    in read-only mode. vSNP3 cascade tables colour their variant cells with
    direct fills, which ARE read, so what a user actually looks at survives.
    """
    if max_table_bytes is None:
        max_table_bytes = DEFAULT_MAX_TABLE_BYTES
    render_rows = min(total_rows, max_rows)
    render_cols = max(1, min(total_cols, max_cells // max(1, render_rows)))

    layout = _sheet_layout(xlsx_path)
    default_width_units = layout["default_width"] or 8.43
    colgroup_parts = [
        f'<col style="width: '
        f'{max(int(round(layout["widths"].get(ci, default_width_units) * 7)), 16)}px">'
        for ci in range(1, render_cols + 1)
    ]
    freeze_row, freeze_col = layout["freeze_row"], layout["freeze_col"]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        sheet_title = ws.title or "Sheet1"
        # Conditional formatting is where ALL of a vSNP3 table's colour lives,
        # and read-only sheets don't expose it — so it comes from the XML.
        dxfs = []
        try:
            dxfs = list(wb._differential_styles.styles)
        except Exception:
            dxfs = []
        cf_ranges = _cf_from_sheet_xml(xlsx_path) if dxfs else []
        # vSNP3's top-priority rule compares each call against the reference
        # row (`equal B$2`). Capture the rows such rules pin so the rule
        # resolves here the way it does in Excel — without it, reference-
        # matching calls fell through to the per-base colour rules and the
        # streamed table painted every cell bright instead of near-white.
        cf_sheet = _CapturedRowsSheet()
        cf_capture_rows = _cf_absolute_ref_rows(cf_ranges)
        positions: dict[int, str] = {}
        # Rows are buffered as per-cell lists rather than one flat string, so
        # the column count can be trimmed after the fact — see the byte budget
        # below, which is only measurable once real cells have been rendered.
        rows_cells: list[list[str]] = []
        effective_cols = render_cols
        bytes_so_far = 0
        # Style palette: identical inline styles repeat across tens of
        # thousands of cells in these tables, so each distinct one becomes a
        # class emitted once in a <style> block. Together with dropping the
        # per-cell IGV anchor (see below) this takes a cell from ~450 bytes to
        # ~25, which is the difference between showing 24 columns and 1,000.
        style_classes: dict[str, str] = {}
        row_samples: list[str] = []
        # (label, stem) per rendered row, kept so look-alike names can be found
        # once the whole window is known — see ambiguous_stems().
        row_labels: list[tuple[str, str]] = []

        # Row/column indices are tracked by position, not read off the cell:
        # a read-only sheet yields `EmptyCell` for gaps, and those carry no
        # coordinates at all (nor styles) — reading cell.row on one is an
        # AttributeError, not a missing value.
        for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=render_rows,
                             min_col=1, max_col=render_cols), start=1):
            if not row:
                continue
            if row_idx in cf_capture_rows:
                cf_sheet.capture(row_idx, row)
            # Row 1 carries the locus headers; capture them before any data row
            # needs them (they arrive first, so a single pass is enough).
            if row_idx == 1 and project:
                for col_idx, cell in enumerate(row, start=1):
                    v = getattr(cell, "value", None)
                    if v is not None and _LOCUS_RE.match(str(v).strip()):
                        positions[col_idx] = str(v).strip()

            # Column 1 of a data row is the sample label.
            row_stem = ""
            row_label = ""
            if project and positions and row_idx > 1:
                first = getattr(row[0], "value", None)
                raw = str(first).strip() if first is not None else ""
                if raw and raw.lower() not in _NON_SAMPLE_LABELS:
                    row_label = raw
                    stem = _strip_vcf_suffix(raw)
                    if stem:
                        row_stem = _canonical_stem(
                            stem, samples_with_bams, samples_with_vcfs)

            row_labels.append((row_label, row_stem))
            row_samples.append(row_stem)
            cells: list[str] = []
            for col_idx, cell in enumerate(row, start=1):
                if col_idx > effective_cols:
                    break
                if not hasattr(cell, "column"):     # EmptyCell: no value, no style
                    cells.append("<td></td>")
                    continue
                inline = _cell_inline_style(cell)
                # CF wins over direct cell styles for the properties it sets,
                # so its fragments go last (later declarations override).
                cf_frags = _cf_fragments_for(
                    cell, row_idx, col_idx, cf_ranges, dxfs, cf_sheet)
                if cf_frags:
                    inline = "; ".join([p for p in (inline,) if p] + cf_frags)
                classes = []
                if inline:
                    cls = style_classes.get(inline)
                    if cls is None:
                        cls = f"k{len(style_classes)}"
                        style_classes[inline] = cls
                    classes.append(cls)
                if row_idx <= freeze_row:
                    classes.append("xlsx-sticky-top")
                if col_idx <= freeze_col:
                    classes.append("xlsx-sticky-left")
                rot_class = _cell_rotation_class(cell)
                if rot_class:
                    classes.append(rot_class)
                value = _format_cell_value(cell)
                # A variant cell is marked with a class and NOTHING else. The
                # sample is already in the row's first cell and the locus in
                # the header row, so a delegated click handler can work both
                # out from the cell's position — which removes an anchor
                # carrying a duplicated href, onclick, title and target from
                # every one of tens of thousands of cells.
                if (row_stem and col_idx in positions
                        and "background-color" in inline):
                    has_bam = samples_with_bams is None or row_stem in samples_with_bams
                    has_vcf = (samples_with_vcfs is not None
                               and row_stem in samples_with_vcfs)
                    loadable = (
                        samples_with_bams is None and samples_with_vcfs is None
                    ) or has_bam or has_vcf
                    if not loadable:
                        classes.append("xlsx-igv-none")
                    elif (not has_bam) and has_vcf:
                        classes.append("xlsx-variant")
                        classes.append("xlsx-igv-calls-only")
                    else:
                        classes.append("xlsx-variant")
                attrs = f' class="{" ".join(classes)}"' if classes else ""
                cells.append(f"<td{attrs}>{value}</td>")
            rows_cells.append(cells)
            bytes_so_far += sum(len(c) for c in cells)

            # Bound the OUTPUT, not just the cell count. Bytes per cell vary by
            # more than tenfold — a plain cell is ~30 bytes, a coloured variant
            # cell carrying an IGV anchor ~450 — so a cell budget bounds
            # nothing useful: a 480 KB sheet rendered to 48 MB, which is a page
            # no browser enjoys and a cache entry larger than the spreadsheet
            # that produced it.
            #
            # Enforced continuously rather than predicted once. A single
            # sample row is a bad estimator here: the first rows of a cascade
            # table are headers and annotations with no colour at all, so
            # estimating from them left the budget 5x overshot by the time the
            # dense rows arrived. After each row, project the finished size
            # from what has actually been emitted and narrow the window if it
            # is heading over — already-buffered rows are trimmed to match, so
            # the result is the same as if the width had been right all along.
            if bytes_so_far and effective_cols > 1:
                projected = bytes_so_far / len(rows_cells) * render_rows
                if projected > max_table_bytes:
                    scale = max_table_bytes / projected
                    narrowed = max(1, int(effective_cols * scale))
                    if narrowed < effective_cols:
                        effective_cols = narrowed
                        rows_cells = [r[:effective_cols] for r in rows_cells]
                        bytes_so_far = sum(len(c) for r in rows_cells for c in r)
    finally:
        wb.close()

    # Trim to the width the byte budget allows (a no-op when nothing was over).
    if effective_cols < render_cols:
        rows_cells = [r[:effective_cols] for r in rows_cells]
        colgroup_parts = colgroup_parts[:effective_cols]
    render_cols = min(render_cols, effective_cols)
    render_rows = len(rows_cells)

    # Look-alike names, resolved once the whole window is known.
    #
    # The clade-filtered renderer has always refused to link these, on the
    # grounds that a click would open the wrong specimen's alignment. This view
    # linked them anyway — the same table, the same rows, opposite answers, and
    # the permissive one is the dangerous one. Now they agree, and both say why.
    ambiguous = ambiguous_stems(row_labels[:render_rows])
    withheld = 0
    if ambiguous:
        for i, (raw, stem) in enumerate(row_labels[:render_rows]):
            if not stem or stem not in ambiguous:
                continue
            if _strip_vcf_suffix(str(raw).strip()) == stem:
                continue
            rows_cells[i] = _delink(rows_cells[i])
            withheld += 1

    style_css = "".join(f".{cls}{{{style}}}" for style, cls in style_classes.items())
    return {
        "igv_withheld": withheld,
        "igv_withheld_stems": ambiguous,
        "title": title or xlsx_path.name,
        "filename": xlsx_path.name,
        "sheet": sheet_title,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "shown_rows": render_rows,
        "shown_cols": render_cols,
        "style_css": style_css,
        "colgroup": "".join(colgroup_parts),
        # One entry per rendered row, so the page can ship a prefix and fetch
        # the rest as the user scrolls instead of sending everything at once.
        "rows": ["<tr>" + "".join(r) + "</tr>" for r in rows_cells],
        # Per-row sample stem and per-column locus, so the delegated IGV click
        # handler can resolve a cell without any per-cell markup.
        "row_samples": row_samples,
        "loci": {str(k): v for k, v in positions.items()},
        "project": project or "",
    }


def render_filtered_window(
    xlsx_path: Path,
    total_rows: int,
    total_cols: int,
    title: str | None,
    project: str | None,
    samples_with_bams: set[str] | None,
    samples_with_vcfs: set[str] | None,
    selection: "list[str] | set[str]",
    max_cells: int,
    max_rows: int,
    max_table_bytes: int = None,
) -> dict:
    """Render a SNP table subset to a tree clade, in one streaming pass.

    ``selection`` is the clade's tip names as read off the tree. Kept rows are
    the header, the structural rows (root / MQ / annotation), and every sample
    row whose label resolves to a selected tip (see _selection_key). Kept
    columns are column 1 plus every locus column with at least one COLOURED
    cell among the kept sample rows — colour is what marks a call as a SNP in
    these tables, so "columns with a SNP in this clade" and "columns a user
    sees highlighted" are the same set by construction.

    Single pass by design: a full streaming pass over a 35 MB cascade table
    costs ~30 s in XML parsing alone, so a look-ahead pass to pick columns
    first would double the dominant cost. Instead kept rows are buffered at
    full sheet width and the column choice is applied afterwards. The buffer
    is bounded by FILTER_BUFFER_MAX_CELLS; selections too large for it fall
    back to row filtering only (disclosed on the page), because a clade that
    big is the unfiltered table again for all practical purposes.

    Returns the same window dict as render_window, plus a ``filter`` block
    for the page banner. Raises FilterMatchError when nothing matched.
    """
    if max_table_bytes is None:
        max_table_bytes = DEFAULT_MAX_TABLE_BYTES

    raw_sel = {_strip_vcf_suffix(str(s).strip()) for s in selection
               if str(s).strip()}
    raw_sel = {s for s in raw_sel if s and s.lower() not in _NON_SAMPLE_LABELS}
    sel_keys = {_selection_key(s, samples_with_bams, samples_with_vcfs)
                for s in raw_sel}
    if not sel_keys:
        raise FilterMatchError("The clade selection contains no sample names.")

    # Row caps: samples up to the window bound, structural rows generously but
    # not unboundedly (a vSNP3 table has ~4; a sheet where thousands of rows
    # have no sample label is not a SNP table and must not be buffered whole).
    max_sample_rows = max(1, max_rows - 8)
    max_structural_rows = 50

    # Decide up front whether the full-width buffer fits. Kept rows can't
    # exceed the selection size (plus structural rows), so this is a real
    # bound, not an estimate.
    est_kept_rows = min(len(sel_keys), max_sample_rows) + max_structural_rows
    column_filter = est_kept_rows * total_cols <= FILTER_BUFFER_MAX_CELLS
    if column_filter:
        buffer_width = total_cols
    else:
        buffer_width = max(1, min(total_cols, max_cells // max(1, est_kept_rows)))

    layout = _sheet_layout(xlsx_path)
    default_width_units = layout["default_width"] or 8.43
    freeze_row, freeze_col = layout["freeze_row"], layout["freeze_col"]

    not_variant_table = False
    positions: dict[int, str] = {}
    # A locus column belongs to the clade when at least one selected sample
    # CALLS something different from the reference row there. Values, not
    # colours, decide this: in a real vSNP3 table every populated cell has a
    # fill (reference matches are painted near-white), so "has a colour"
    # would keep every column. Missing data ('-'/'N'/empty) is not a call —
    # a low-coverage sample must not drag thousands of columns into the view.
    col_snp = bytearray(total_cols + 1)
    col_called = bytearray(total_cols + 1)   # a selected sample called ANYTHING here
    # Columns contributed by rows that matched only through canonicalisation, kept
    # per row so that dropping such a row in the ambiguity pass also withdraws
    # its columns. Without this a look-alike row was excluded from the table while
    # its private SNP stayed on as a column of the clade's own. Only non-exact
    # matches are tracked, because only they can be dropped; exact matches go
    # straight into the bitmaps above.
    nonexact_cols: dict[int, tuple[set, set]] = {}
    root_vals: dict[int, str] | None = None  # the reference row's calls
    kept_frags: list[list[str]] = []     # per kept row, one fragment per column
    kept_rows: list[int] = []            # per kept row, its row number in the SHEET
                                         # — what the xlsx export of this clade
                                         # copies, so the download is the same
                                         # subset the page is showing and not a
                                         # second, differently-filtered answer
    kept_samples: list[str] = []         # per kept row, IGV stem ('' = structural)
    kept_meta: list[dict] = []           # per kept row: how it matched, for the
                                         # ambiguity pass below
    key_stems: dict[str, set] = {}       # canonical key -> the distinct label
                                         # stems that resolved to it
    style_classes: dict[str, str] = {}
    matched_rows = 0                     # sample rows that matched the clade
    shown_sample_rows = 0                # ...and were actually buffered
    total_sample_rows = 0                # all sample rows in the sheet
    structural_kept = 0
    bytes_so_far = 0

    def _call_of(cell) -> str:
        v = getattr(cell, "value", None)
        s = str(v).strip() if v is not None else ""
        return "" if s.upper() in ("", "-", "N") else s

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        sheet_title = ws.title or "Sheet1"
        dxfs = []
        try:
            dxfs = list(wb._differential_styles.styles)
        except Exception:
            dxfs = []
        cf_ranges = _cf_from_sheet_xml(xlsx_path) if dxfs else []
        # Same reference-row capture as render_window: the `equal B$2` rule
        # needs row 2 to resolve, and row 2 streams past before any row that
        # asks for it.
        cf_sheet = _CapturedRowsSheet()
        cf_capture_rows = _cf_absolute_ref_rows(cf_ranges)

        for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=total_rows,
                             min_col=1, max_col=total_cols), start=1):
            if row_idx in cf_capture_rows:
                cf_sheet.capture(row_idx, row)
            if row_idx == 1:
                for col_idx, cell in enumerate(row, start=1):
                    v = getattr(cell, "value", None)
                    if v is not None and _LOCUS_RE.match(str(v).strip()):
                        positions[col_idx] = str(v).strip()
                if len(positions) < 2:
                    # Not a variant table — a clade filter is meaningless here.
                    not_variant_table = True
                    break
                keep, row_stem, is_matched_sample = True, "", False
                meta = {"sample": False}
            else:
                first = getattr(row[0], "value", None) if row else None
                raw = str(first).strip() if first is not None else ""
                if not raw or raw.lower() in _NON_SAMPLE_LABELS:
                    if root_vals is None and raw.lower() == "root":
                        # The reference row — the baseline every sample's
                        # calls are compared against for the column choice.
                        root_vals = {}
                        for ci, c in enumerate(row, start=1):
                            if ci in positions:
                                call = _call_of(c)
                                if call:
                                    root_vals[ci] = call
                    keep = structural_kept < max_structural_rows
                    if keep:
                        structural_kept += 1
                    row_stem, is_matched_sample = "", False
                    meta = {"sample": False}
                else:
                    total_sample_rows += 1
                    stem = _strip_vcf_suffix(raw)
                    key = _canonical_stem(
                        stem, samples_with_bams, samples_with_vcfs)
                    key_stems.setdefault(key, set()).add(stem)
                    # EXACT beats canonical. Canonicalisation exists so a tree
                    # tip and a table row that decorate the same sample
                    # differently still match, but its "longest known stem that
                    # prefixes this label" rule can also alias one specimen onto
                    # another: with only `19-1234` known on disk, the row
                    # `19-1234_2_Bovine_USA` (a different isolate) resolves to
                    # `19-1234` too. Recording how each row matched lets the
                    # ambiguity pass after the loop throw those back out.
                    exact = stem in raw_sel
                    is_matched_sample = exact or key in sel_keys
                    if is_matched_sample:
                        matched_rows += 1
                    keep = is_matched_sample and shown_sample_rows < max_sample_rows
                    if keep:
                        shown_sample_rows += 1
                        row_stem = key
                    else:
                        row_stem = ""
                    meta = {"sample": True, "exact": exact, "key": key, "stem": stem}
            if not keep:
                continue

            cells: list[str] = []
            for col_idx, cell in enumerate(row, start=1):
                if col_idx > buffer_width:
                    break
                if not hasattr(cell, "column"):     # EmptyCell: no value, no style
                    cells.append("<td></td>")
                    continue
                inline = _cell_inline_style(cell)
                cf_frags = _cf_fragments_for(
                    cell, row_idx, col_idx, cf_ranges, dxfs, cf_sheet)
                if cf_frags:
                    inline = "; ".join([p for p in (inline,) if p] + cf_frags)
                colored = "background-color" in inline
                if is_matched_sample and row_stem and col_idx in positions:
                    call = _call_of(cell)
                    if call:
                        # Any call at all keeps the column when there is no
                        # reference to compare against (see below) — over-keeping
                        # is the recoverable direction.
                        base = (root_vals or {}).get(col_idx)
                        is_snp = base is not None and call != base
                        if meta.get("exact"):
                            col_called[col_idx] = 1
                            if is_snp:
                                col_snp[col_idx] = 1
                        else:
                            slot = len(kept_frags)   # the index this row will take
                            snp_set, called_set = nonexact_cols.setdefault(
                                slot, (set(), set()))
                            called_set.add(col_idx)
                            if is_snp:
                                snp_set.add(col_idx)
                classes = []
                if inline:
                    cls = style_classes.get(inline)
                    if cls is None:
                        cls = f"k{len(style_classes)}"
                        style_classes[inline] = cls
                    classes.append(cls)
                if row_idx <= freeze_row:
                    classes.append("xlsx-sticky-top")
                if col_idx <= freeze_col:
                    classes.append("xlsx-sticky-left")
                rot_class = _cell_rotation_class(cell)
                if rot_class:
                    classes.append(rot_class)
                value = _format_cell_value(cell)
                if row_stem and col_idx in positions and colored:
                    has_bam = samples_with_bams is None or row_stem in samples_with_bams
                    has_vcf = (samples_with_vcfs is not None
                               and row_stem in samples_with_vcfs)
                    loadable = (
                        samples_with_bams is None and samples_with_vcfs is None
                    ) or has_bam or has_vcf
                    if not loadable:
                        classes.append("xlsx-igv-none")
                    elif (not has_bam) and has_vcf:
                        classes.append("xlsx-variant")
                        classes.append("xlsx-igv-calls-only")
                    else:
                        classes.append("xlsx-variant")
                attrs = f' class="{" ".join(classes)}"' if classes else ""
                cells.append(f"<td{attrs}>{value}</td>")
            kept_frags.append(cells)
            kept_rows.append(row_idx)
            kept_samples.append(row_stem)
            kept_meta.append(meta)
            bytes_so_far += sum(len(c) for c in cells)

            # Row-filter-only mode has no post-pass column choice, so it must
            # bound its output the way render_window does: project the final
            # size from what has been emitted and narrow the buffered width.
            if not column_filter and bytes_so_far and buffer_width > 1:
                projected = bytes_so_far / len(kept_frags) * est_kept_rows
                if projected > max_table_bytes:
                    scale = max_table_bytes / projected
                    narrowed = max(1, int(buffer_width * scale))
                    if narrowed < buffer_width:
                        buffer_width = narrowed
                        kept_frags = [r[:buffer_width] for r in kept_frags]
                        bytes_so_far = sum(
                            len(c) for r in kept_frags for c in r)
    finally:
        wb.close()

    if not_variant_table:
        win = render_window(
            xlsx_path, total_rows, total_cols, title, project,
            samples_with_bams, samples_with_vcfs, max_cells, max_rows,
            max_table_bytes)
        win["filter"] = {
            "ignored": ("this sheet has no locus columns, so it is not a "
                        "SNP table and the clade filter does not apply"),
        }
        return win

    # Ambiguity pass. A row that matched ONLY through canonicalisation, onto a
    # key that more than one distinct label resolves to, is not this clade's
    # row — it is a different specimen whose name happens to extend a known
    # stem. Keeping it would add its private SNPs as columns and, worse, label
    # it with the other sample's stem, so a click would open the WRONG sample
    # in IGV. Dropped, and counted so the page can say so.
    ambiguous_keys = {k for k, stems in key_stems.items() if len(stems) > 1}
    dropped_ambiguous = 0
    surviving_slots = set(range(len(kept_frags)))
    if ambiguous_keys:
        keep_idx = []
        for i, meta in enumerate(kept_meta):
            if (meta.get("sample") and not meta.get("exact")
                    and meta.get("key") in ambiguous_keys):
                dropped_ambiguous += 1
                continue
            keep_idx.append(i)
        if dropped_ambiguous:
            surviving_slots = set(keep_idx)
            kept_frags = [kept_frags[i] for i in keep_idx]
            kept_rows = [kept_rows[i] for i in keep_idx]
            kept_samples = [kept_samples[i] for i in keep_idx]
            kept_meta = [kept_meta[i] for i in keep_idx]
            matched_rows -= dropped_ambiguous
            shown_sample_rows -= dropped_ambiguous
    # Fold in the columns of the non-exact rows that SURVIVED. A dropped row's
    # columns are simply never added, so a look-alike cannot leave its private
    # SNP behind as a column of this clade.
    for slot, (snp_set, called_set) in nonexact_cols.items():
        if slot not in surviving_slots:
            continue
        for c in snp_set:
            col_snp[c] = 1
        for c in called_set:
            col_called[c] = 1
    # A row kept on an ambiguous key must not carry that key as its IGV target
    # either: better a cell that is not clickable than one that opens another
    # sample's alignment. The cells were rendered before the ambiguity was known
    # (it takes the whole sheet to establish), so the classes are rewritten here
    # as well — a cell left looking clickable would resolve to a name with no
    # data behind it, which is a dead end dressed as a link.
    # Counted, not just done: a table where every SNP cell has quietly stopped
    # responding reads as a broken build, and was reported as one. The page says
    # how many rows lost their links and which folder names collided.
    igv_withheld = 0
    withheld_stems: dict = {}
    for i, meta in enumerate(kept_meta):
        if meta.get("sample") and meta.get("key") in ambiguous_keys:
            kept_samples[i] = meta.get("stem") or kept_samples[i]
            kept_frags[i] = _delink(kept_frags[i])
            igv_withheld += 1
            key = str(meta.get("key"))
            withheld_stems[key] = sorted(key_stems.get(meta.get("key"), ()))

    if matched_rows == 0:
        raise FilterMatchError(
            f"None of the {len(sel_keys)} selected samples appear in this "
            f"table (it has {total_sample_rows} sample rows). The tree and "
            "the table may not belong to the same Step 2 group."
        )

    # No reference row means no baseline, and the honest response is to keep
    # every position the clade has a call at rather than invent one.
    #
    # The previous fallback compared the selected samples against EACH OTHER,
    # which drops any position where the clade agrees internally but differs
    # from the reference — that is, precisely the clade-defining SNPs. On a real
    # 4-sample clade, relabelling the reference row alone lost 30 of 85
    # positions, every one of them monomorphic within the clade, while the page
    # went on stating that the hidden positions had no SNP in these samples.
    reference_found = root_vals is not None
    if not reference_found:
        col_snp = col_called

    locus_total = len(positions)
    truncated_cols = 0
    if column_filter:
        kept_cols = [c for c in range(1, total_cols + 1)
                     if c == 1 or c not in positions or col_snp[c]]
        locus_matched = sum(1 for c in kept_cols if c in positions)
        # Cell budget, then byte budget — both trim the column tail, never
        # column 1. A filtered table is nearly always far inside both.
        allowed = max(1, max_cells // max(1, len(kept_frags)))
        if len(kept_cols) > allowed:
            truncated_cols += len(kept_cols) - allowed
            kept_cols = kept_cols[:allowed]
        rows_cells = [[(r[c - 1] if c - 1 < len(r) else "<td></td>")
                       for c in kept_cols] for r in kept_frags]
        total_bytes = sum(len(c) for r in rows_cells for c in r)
        if total_bytes > max_table_bytes:
            scale = max_table_bytes / total_bytes
            narrowed = max(1, int(len(kept_cols) * scale))
            if narrowed < len(kept_cols):
                truncated_cols += len(kept_cols) - narrowed
                kept_cols = kept_cols[:narrowed]
                rows_cells = [r[:narrowed] for r in rows_cells]
    else:
        kept_cols = list(range(1, buffer_width + 1))
        locus_matched = sum(1 for c in kept_cols if c in positions)
        rows_cells = [[(r[c - 1] if c - 1 < len(r) else "<td></td>")
                       for c in kept_cols] for r in kept_frags]

    locus_shown = sum(1 for c in kept_cols if c in positions)
    colgroup_parts = [
        f'<col style="width: '
        f'{max(int(round(layout["widths"].get(ci, default_width_units) * 7)), 16)}px">'
        for ci in kept_cols
    ]
    loci = {str(new_idx): positions[orig]
            for new_idx, orig in enumerate(kept_cols, start=1)
            if orig in positions}
    style_css = "".join(f".{cls}{{{style}}}" for style, cls in style_classes.items())
    return {
        "title": title or xlsx_path.name,
        "filename": xlsx_path.name,
        "sheet": sheet_title,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "shown_rows": len(rows_cells),
        "shown_cols": len(kept_cols),
        "style_css": style_css,
        "colgroup": "".join(colgroup_parts),
        "rows": ["<tr>" + "".join(r) + "</tr>" for r in rows_cells],
        "row_samples": kept_samples,
        # The subset in SHEET coordinates. Carried on the window so the xlsx
        # export reuses the decisions made here — which rows matched the clade,
        # which columns hold one of its SNPs — instead of re-deriving them and
        # risking a download that disagrees with the page it came from.
        "kept_rows": kept_rows,
        "kept_cols": kept_cols,
        "loci": loci,
        "project": project or "",
        "igv_withheld": igv_withheld,
        "igv_withheld_stems": withheld_stems,
        "filter": {
            "selected": len(sel_keys),
            "matched": matched_rows,
            "shown_samples": shown_sample_rows,
            "total_samples": total_sample_rows,
            "locus_total": locus_total,
            "locus_matched": locus_matched,
            "locus_shown": locus_shown,
            "column_filter": column_filter,
            "truncated_cols": truncated_cols,
            "truncated_rows": matched_rows - shown_sample_rows,
            # Everything the page has to admit to. Each of these was a way for
            # the old banner to state something untrue.
            "reference_found": reference_found,
            "unmatched": max(0, len(sel_keys) - matched_rows),
            "dropped_ambiguous": dropped_ambiguous,
            "ignored": None,
        },
    }


def write_filtered_xlsx(xlsx_path: Path, dest: Path,
                        rows: "list[int]", cols: "list[int]",
                        sheet_title: str | None = None) -> None:
    """Write the rows x columns of `xlsx_path` named by `rows`/`cols` to `dest`.

    The subset comes from render_filtered_window, so this copies a decision it
    does not make: the spreadsheet a user downloads from a clade view holds
    exactly the rows and columns that view is showing. Before this existed the
    "Download xlsx" link handed back the WHOLE table — every sample in the group
    and every position in it — because the download branch returned the file
    before the clade selection was even looked at. On a cascade table that is a
    35 MB answer to a question about twelve samples.

    Formatting is carried across, and conditional formatting is RESOLVED into
    static fills rather than re-emitted as rules. In these tables every colour on
    screen comes from CF (`equal B$2` whitens a reference match, then one rule per
    base paints the rest), and those formulas are anchored to the sheet they were
    written for: `B$2` means the reference row is row 2 and the block starts at
    B3. Re-emitting them over a subset would need that layout to survive column
    and row removal, and when it did not the file would open MIS-coloured, which
    is worse than plainly baked colour. Baked colour also guarantees the download
    matches the page, since both resolve through _cf_dxf_for.

    Streaming, read_only in and write_only out: these tables reach 10 million
    cells and this runs inside a web request.
    """
    if not rows or not cols:
        raise ValueError("nothing to write: the subset has no rows or no columns")
    row_set = set(rows)
    last_row = max(rows)
    last_col = max(cols)
    # Source column -> position in the output, so a row can be re-emitted in one
    # pass without searching the column list per cell.
    col_order = {c: i for i, c in enumerate(cols)}

    layout = _sheet_layout(xlsx_path)
    default_width = layout["default_width"] or 8.43

    out = openpyxl.Workbook(write_only=True)
    ws_out = out.create_sheet(sheet_title or "Sheet1")
    for i, c in enumerate(cols, start=1):
        ws_out.column_dimensions[get_column_letter(i)].width = (
            layout["widths"].get(c, default_width)
        )
    # Freeze whatever the source froze, translated into the output's coordinates:
    # a pane split is a number of leading rows/columns, and dropping some of them
    # moves it. Counting the kept ones is the translation.
    #
    # Set BEFORE the first append, not after: a write-only worksheet streams to
    # disk, and the pane lives in the sheet header, which is already written by
    # the time the first row goes out. Assigning it afterwards is silently
    # ignored — the file saves clean and opens with no frozen header, which on a
    # SNP table means scrolling right loses the sample names.
    fr = sum(1 for r in rows if r <= layout["freeze_row"])
    fc = sum(1 for c in cols if c <= layout["freeze_col"])
    if fr or fc:
        ws_out.freeze_panes = f"{get_column_letter(fc + 1)}{fr + 1}"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        try:
            dxfs = list(wb._differential_styles.styles)
        except Exception:
            dxfs = []
        cf_ranges = _cf_from_sheet_xml(xlsx_path) if dxfs else []
        # The `equal B$2` rule needs row 2 to resolve, and row 2 has streamed
        # past by the time any sample row asks for it — same capture the HTML
        # renderer does, for the same reason.
        cf_sheet = _CapturedRowsSheet()
        cf_capture_rows = _cf_absolute_ref_rows(cf_ranges)

        for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=last_row,
                             min_col=1, max_col=last_col), start=1):
            if row_idx in cf_capture_rows:
                cf_sheet.capture(row_idx, row)
            if row_idx not in row_set:
                continue
            emit: list = [None] * len(cols)
            for col_idx, cell in enumerate(row, start=1):
                pos = col_order.get(col_idx)
                if pos is None:
                    continue
                if not hasattr(cell, "column"):     # EmptyCell: no value, no style
                    continue
                new_cell = WriteOnlyCell(ws_out, value=cell.value)
                try:
                    new_cell.number_format = cell.number_format or "General"
                    new_cell.font = copy(cell.font)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.border = copy(cell.border)
                    if cell.fill is not None and getattr(
                            cell.fill, "patternType", None):
                        new_cell.fill = copy(cell.fill)
                except (TypeError, ValueError):
                    pass   # an unusual style is not worth losing the value over
                dxf = _cf_dxf_for(cell, row_idx, col_idx,
                                  cf_ranges, dxfs, cf_sheet)
                if dxf is not None:
                    _apply_dxf(new_cell, dxf)
                emit[pos] = new_cell
            ws_out.append(emit)
    finally:
        wb.close()

    # Written beside the destination and renamed in, so a failed or interrupted
    # build never leaves a truncated .xlsx for the browser to download.
    tmp = dest.with_name(dest.name + ".part")
    out.save(tmp)
    os.replace(tmp, dest)


def _apply_dxf(cell, dxf) -> None:
    """Bake a conditional format onto a cell as ordinary styling."""
    fill = getattr(dxf, "fill", None)
    if fill is not None:
        # A CF dxf fill carries its colour on bgColor; static fills read fgColor,
        # so the colour has to be moved across rather than the fill copied.
        rgb = None
        for attr in ("bgColor", "fgColor"):
            colour = getattr(fill, attr, None)
            raw = getattr(colour, "rgb", None)
            if isinstance(raw, str) and len(raw) in (6, 8) and raw[:2] != "00":
                rgb = raw if len(raw) == 8 else f"FF{raw}"
                break
        if rgb:
            cell.fill = PatternFill(fill_type="solid", start_color=rgb,
                                    end_color=rgb)
    font = getattr(dxf, "font", None)
    if font is not None:
        base = copy(cell.font)
        # openpyxl's dxf font uses the OOXML short names (.b/.i), not .bold/.italic.
        if getattr(font, "b", None) is True:
            base.bold = True
        if getattr(font, "i", None) is True:
            base.italic = True
        colour = getattr(font, "color", None)
        if colour is not None and getattr(colour, "rgb", None):
            base.color = copy(colour)
        cell.font = base


def download_link(label: str, href: str | None) -> str:
    """The bar's download anchor.

    A REAL link when the caller supplies the relative href
    (``?path=…&download=1`` resolves against this page's own URL, so it
    survives any proxy prefix), opening in a fresh tab — the exact mechanism
    the Step 2 Results download buttons use, which is the one download path
    proven to work everywhere this suite is served. The JS-assembled same-tab
    navigation this replaces could be dropped silently by the browser: the
    click ran, location.href was set, and no request ever reached the server
    (observed in Safari on tabs the app had opened with noopener — the same
    page's other JS links still worked). A native anchor has no such moving
    parts, and right-click "Download Linked File" works on it too.

    The JS fallback remains only for callers that cannot know their own URL.
    """
    esc = html.escape(label)
    if href:
        return (f'<a href="{html.escape(href, quote=True)}" '
                f'target="_blank" rel="noopener">{esc}</a>')
    return ('<a href="#" onclick="var u=new URL(window.location.href);'
            "u.searchParams.set('download','1');u.searchParams.delete('rows_from');"
            f'window.location.href=u.toString();return false;">{esc}</a>')


def view_controls(loci: dict, row_samples: "list[str]") -> str:
    """The view aids in the page bar: the identical-position toggle and the
    colour-marking hint.

    Both are statements about samples-at-positions, so they appear only on
    variant tables — locus columns recognised and at least one sample row.
    And both are VIEW aids, applied by the page itself: hiding a column or
    colouring a row changes what is emphasised, never what was rendered, so
    the download links keep handing back every position the view was built
    from.

    The toggle exists for the clade view: a clade's table keeps every position
    where any member differs from the reference, which includes the positions
    the whole clade shares — informative about the clade, noise when the
    question is what differs WITHIN it. "Identical" is exact-match across the
    shown sample rows; the JS says so on the tooltip and reports the count it
    hid.

    With one sample every position is trivially identical and the toggle
    would blank the table, so it renders disabled with the reason on its
    tooltip — a control that comes and goes between views reads as broken.
    """
    n = sum(1 for s in (row_samples or []) if s)
    if not loci or n < 1:
        return ""
    if n >= 2:
        title = (
            f"Hide every position where all {n} shown samples have exactly "
            "the same call — including positions where the whole group "
            "shares one difference from the reference. A position stays "
            "whenever any two shown samples disagree, and comes back when "
            "this is unticked. The download links are unaffected."
        )
        box = '<input type="checkbox" id="xlsxInvariant">'
        cls = "xlsx-ctl"
    else:
        title = ("Only one sample row is shown, so every position is "
                 "trivially identical — there is nothing to compare against.")
        box = '<input type="checkbox" id="xlsxInvariant" disabled>'
        cls = "xlsx-ctl xlsx-ctl-off"
    return (
        f'<label class="{cls}" title="{html.escape(title, quote=True)}">'
        f'{box} Hide identical positions</label>'
        '<span class="xlsx-ctl-note" id="xlsxInvariantNote"></span>'
        '<span class="xlsx-ctl-note" title="The colours cycle through the '
        'tree viewer&#39;s palette. Click the same name or header again to '
        'remove its mark.">Click a sample name or position header to colour '
        'it.</span>'
        '<a href="#" id="xlsxHlClear" style="display:none">clear colours</a>'
    )


def compose_page(window: dict, initial_rows: int = DEFAULT_INITIAL_ROWS,
                 download_href: str | None = None,
                 full_href: str | None = None) -> str:
    """Build the preview page from a rendered window.

    Only the first `initial_rows` rows are inlined; the rest are fetched by the
    page as it is scrolled. A 1,000 x 1,000 window is a million cells — fine to
    hold on the server, far too much to hand a browser in one document.

    ``download_href``/``full_href`` are RELATIVE URLs (query-only) built by the
    endpoint, so they re-enter the same route through any proxy prefix. See
    download_link() for why these are real anchors rather than JS navigation.
    """
    rows = window["rows"]
    head = "".join(rows[:initial_rows])
    total_rows, total_cols = window["total_rows"], window["total_cols"]
    shown_rows, shown_cols = window["shown_rows"], window["shown_cols"]
    filt = window.get("filter")

    # The link target is the same preview minus the selection. Same-tab is
    # right here — the destination is a page, not a file.
    if full_href:
        _clear_filter_link = (f'<a href="{html.escape(full_href, quote=True)}">'
                              'Show the full table</a>')
    else:
        _clear_filter_link = (
            '<a href="#" onclick="var u=new URL(window.location.href);'
            "u.searchParams.delete('selection');u.searchParams.delete('rows_from');"
            'window.location.href=u.toString();return false;">Show the full table</a>'
        )

    notice = ""
    if filt and filt.get("ignored"):
        notice += (
            '<div class="xlsx-filter xlsx-filter-warn">'
            f'<strong>Clade filter not applied:</strong> {html.escape(filt["ignored"])}. '
            f'{_clear_filter_link}</div>'
        )
        filt = None
    if filt:
        bits = (f'<strong>Filtered to a tree clade:</strong> '
                f'{filt["shown_samples"]:,} of {filt["total_samples"]:,} samples')
        # What the position count MEANS depends on whether a reference row was
        # found; claiming "no SNP in these samples" without one is the sentence
        # that made a lost clade-defining SNP invisible.
        if filt.get("reference_found", True):
            bits += (f' and {filt["locus_shown"]:,} of {filt["locus_total"]:,} '
                     'SNP positions — positions where every selected sample '
                     'matches the reference are hidden')
        else:
            bits += (f' and {filt["locus_shown"]:,} of {filt["locus_total"]:,} '
                     'positions — this sheet has no <code>root</code> reference '
                     'row, so positions are kept wherever a selected sample has '
                     'a call, NOT by difference from the reference')
        # A tip that resolved to no row takes its positions with it, so an
        # unmatched count is a statement about missing DATA, not just names.
        if filt.get("unmatched"):
            bits += (f'. <strong>{filt["unmatched"]:,} of the '
                     f'{filt["selected"]:,} selected samples were not found in '
                     'this table</strong> and are not represented here')
        if filt.get("dropped_ambiguous"):
            bits += (f'. {filt["dropped_ambiguous"]:,} row(s) whose name could '
                     'not be told apart from another sample were excluded')
        if window.get("kept_rows") and window.get("kept_cols"):
            bits += ('. <em>Download this clade</em> above saves exactly these '
                     'rows and columns, not the whole table')
        notice += f'<div class="xlsx-filter">{bits}. {_clear_filter_link}</div>'
    elif shown_rows < total_rows or shown_cols < total_cols:
        bits = []
        if shown_rows < total_rows:
            bits.append(f"the first {shown_rows:,} of {total_rows:,} rows")
        if shown_cols < total_cols:
            bits.append(f"the first {shown_cols:,} of {total_cols:,} columns")
        notice = (
            '<div class="xlsx-truncated">'
            f'<strong>Showing {" and ".join(bits)}.</strong> '
            f'The full sheet is {total_rows:,} x {total_cols:,} '
            f'({total_rows * total_cols:,} cells), more than a browser can lay '
            'out. Use <em>Download xlsx</em> above for the complete table.'
            '</div>'
        )

    # Appended to whatever else the page has to say, because it is orthogonal to
    # all of it: a table can be filtered, truncated, both or neither and still
    # have had links withheld. Silence here is what turned a naming collision
    # into "clicking a SNP does nothing, not even a hover".
    wnote = withheld_note(window.get("igv_withheld_stems") or {},
                          window.get("igv_withheld") or 0)
    if wnote:
        notice += f'<div class="xlsx-filter xlsx-filter-warn">{wnote}</div>'

    table_html = (f'<table class="xlsx" id="xlsxTable"><colgroup>{window["colgroup"]}'
                  f'</colgroup><tbody id="xlsxBody">{head}</tbody></table>')
    # The download follows the view. Saying so on the link is half the fix: the
    # complaint that started this was "the download does not work", and what had
    # actually happened was that a clade view handed back the whole group's
    # table under a label that promised nothing more specific.
    download_label = ("Download this clade (xlsx)"
                      if window.get("kept_rows") and window.get("kept_cols")
                      else "Download xlsx")
    return _PAGE_TEMPLATE.format(
        download_link=download_link(download_label, download_href),
        title=html.escape(window["title"]),
        filename=html.escape(window["filename"]),
        sheet=html.escape(window["sheet"]),
        rows=total_rows,
        cols=total_cols,
        notice=notice,
        style_css=window["style_css"],
        table=table_html,
        loaded=len(rows[:initial_rows]),
        available=len(rows),
        project=html.escape(window["project"], quote=True),
        loci_json=json.dumps(window["loci"]),
        samples_json=json.dumps(window["row_samples"]),
        controls=view_controls(window.get("loci") or {},
                               window.get("row_samples") or []),
    )


def xlsx_to_html(
    xlsx_path: Path,
    title: str | None = None,
    project: str | None = None,
    samples_with_bams: set[str] | None = None,
    samples_with_vcfs: set[str] | None = None,
    max_cells: int = DEFAULT_MAX_CELLS,
    max_rows: int = DEFAULT_MAX_ROWS,
    download_href: str | None = None,
) -> str:
    """Render the first (active) sheet of an xlsx file as a self-contained HTML page.

    When ``project`` is provided and the sheet looks like a vSNP3 variant-alignment
    table, variant cells (those with a colored fill) become clickable
    IGV-launch targets — clicking anywhere in the cell opens the row's
    sample in the IgvStandalone viewer at the cell's locus (additive
    across clicks).

    ``samples_with_bams`` (optional) is the set of step1 sample names for
    which a BAM exists on disk.

    ``samples_with_vcfs`` (optional) is the set of sample names that have
    an imported VCF in ``step2/vcf_database/`` (no BAM, calls-only IGV).

    A cell is clickable if the sample is in EITHER set; calls-only mode is
    indicated in the tooltip (and italics) when only the VCF is present.
    The cell renders plain/non-clickable only when the sample is in
    neither set (genuinely nothing to load).
    """
    xlsx_path = Path(xlsx_path)
    # Decide the renderer from the sheet's declared size, which read-only mode
    # answers without reading any cells. Anything of ordinary size keeps the
    # full-fidelity path below, byte for byte as before; only sheets that the
    # old path could not open at all take the streaming route.
    try:
        total_rows, total_cols = sheet_extent(xlsx_path)
    except Exception:
        total_rows = total_cols = 0
    if total_rows * total_cols > STREAM_ABOVE_CELLS:
        return compose_page(render_window(
            xlsx_path, total_rows, total_cols, title, project,
            samples_with_bams, samples_with_vcfs, max_cells, max_rows,
        ))

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    ws = wb.active
    dxfs = list(wb._differential_styles.styles) if hasattr(wb, "_differential_styles") else []
    cf_extras = _build_cf_extras(ws, dxfs)
    vtable = _detect_variant_table(ws) if project else None

    # Pre-compute merged-cell map: anchor coordinate → (rowspan, colspan);
    # all other cells in the merge get skipped.
    merge_anchors: dict[str, tuple[int, int]] = {}
    merged_skip: set[str] = set()
    for mr in ws.merged_cells.ranges:
        anchor = ws.cell(row=mr.min_row, column=mr.min_col).coordinate
        merge_anchors[anchor] = (mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                coord = ws.cell(row=r, column=c).coordinate
                if coord != anchor:
                    merged_skip.add(coord)

    # Column widths — openpyxl stores column-range widths under a single
    # anchor key with `.min` / `.max` spanning the range (e.g. one entry on
    # `B` covers cols 2..52). Resolve to a per-column-index map first, then
    # emit the <colgroup>. Excel char-units → px is ~7 for Calibri 11pt.
    col_widths_units: dict[int, float] = {}
    for _letter, dim in ws.column_dimensions.items():
        if dim.width is None:
            continue
        lo = dim.min if dim.min is not None else 1
        hi = dim.max if dim.max is not None else lo
        for ci in range(lo, hi + 1):
            col_widths_units[ci] = float(dim.width)
    default_width_units = (
        float(ws.sheet_format.defaultColWidth)
        if ws.sheet_format and ws.sheet_format.defaultColWidth
        else 8.43  # Excel's documented default
    )
    colgroup_parts: list[str] = []
    for col_idx in range(1, ws.max_column + 1):
        w_units = col_widths_units.get(col_idx, default_width_units)
        width_px = max(int(round(w_units * 7)), 16)  # never collapse a column to invisible
        colgroup_parts.append(f'<col style="width: {width_px}px">')

    # Frozen panes: openpyxl exposes ws.freeze_panes as the top-left of the
    # *unfrozen* area, e.g. "B2" means row 1 + col A are frozen.
    freeze_row, freeze_col = 0, 0
    if ws.freeze_panes:
        try:
            anchor_cell = ws[ws.freeze_panes]
            freeze_row = anchor_cell.row - 1
            freeze_col = anchor_cell.column - 1
        except Exception:
            freeze_row, freeze_col = 0, 0

    rows_html: list[str] = []
    # One entry per emitted <tr>, aligned by construction (appended in the same
    # loop): the sample stem for a sample row, "" for everything else. This is
    # what lets the view aids (colour marks, the identical-position toggle)
    # tell a sample row from root/MQ/annotation on this path too.
    row_samples_full: list[str] = []
    for row in ws.iter_rows():
        if not row:
            continue
        # Per-row height. Excel stores row height in points (1pt ≈ 1.333px).
        row_idx = row[0].row
        row_samples_full.append(
            vtable["samples"].get(row_idx, "") if vtable else "")
        row_dim = ws.row_dimensions.get(row_idx)
        row_style = ""
        if row_dim and row_dim.height:
            row_style = f' style="height: {int(round(row_dim.height * 1.333))}px"'
        rows_html.append(f"<tr{row_style}>")
        for cell in row:
            if cell.coordinate in merged_skip:
                continue
            inline_parts = [_cell_inline_style(cell)] if _cell_inline_style(cell) else []
            cf_frags = cf_extras.get(cell.coordinate)
            if cf_frags:
                # CF wins over direct cell styles for the properties it sets,
                # so put it last (later declarations override).
                inline_parts.extend(cf_frags)
            inline = "; ".join(inline_parts)
            attrs = ""
            if cell.coordinate in merge_anchors:
                rs, cs = merge_anchors[cell.coordinate]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            # Class bundle: sticky pane hints + rotation
            classes = []
            if cell.row <= freeze_row:
                classes.append("xlsx-sticky-top")
            if cell.column <= freeze_col:
                classes.append("xlsx-sticky-left")
            rot_class = _cell_rotation_class(cell)
            if rot_class:
                classes.append(rot_class)
            if classes:
                attrs += f' class="{" ".join(classes)}"'
            value = _format_cell_value(cell)
            # Cascade-table IGV launch: only on variant cells in a detected
            # variant-alignment table when a project context is available. A "variant"
            # here is any cell in the data area with a resolved background
            # color (direct fill or conditional formatting) — that's the
            # visible signal vSNP3 uses to mark a call that differs from root.
            # The whole variant cell becomes a clickable IGV-launch target
            # (cell_inner wraps the value in an anchor); other cells render
            # their value bare.
            cell_inner = value
            if (
                vtable
                and cell.row in vtable["samples"]
                and cell.column in vtable["positions"]
                and any("background-color" in p for p in inline_parts)
            ):
                # Resolve the (possibly metadata-decorated) cascade label back
                # to the bare on-disk stem so loadability + the IGV track id
                # match the BAM/VCF filenames (imported step2-renamed samples).
                row_stem = _canonical_stem(
                    vtable["samples"][cell.row], samples_with_bams, samples_with_vcfs
                )
                has_bam = samples_with_bams is None or row_stem in samples_with_bams
                has_vcf = samples_with_vcfs is not None and row_stem in samples_with_vcfs
                # Loadable if either source is present (default-loadable when
                # caller didn't supply sets).
                this_loadable = (
                    samples_with_bams is None and samples_with_vcfs is None
                ) or has_bam or has_vcf
                this_calls_only = (not has_bam) and has_vcf
                cell_inner = _igv_cell_html(
                    value,
                    project=project,
                    this_stem=row_stem,
                    locus=vtable["positions"][cell.column],
                    this_loadable=this_loadable,
                    this_calls_only=this_calls_only,
                )
                classes.append("xlsx-variant")
                # Re-emit class attr — `attrs` already had it baked in above
                # for sticky/rotation classes, so replace rather than append.
                attrs = re.sub(r'\s+class="[^"]*"', "", attrs)
                attrs += f' class="{" ".join(classes)}"'
            style_attr = f' style="{inline}"' if inline else ""
            rows_html.append(f"<td{attrs}{style_attr}>{cell_inner}</td>")
        rows_html.append("</tr>")

    display_title = html.escape(title or xlsx_path.name)
    colgroup = "<colgroup>" + "".join(colgroup_parts) + "</colgroup>"
    table_html = ('<table class="xlsx" id="xlsxTable">' + colgroup
                  + '<tbody id="xlsxBody">' + "".join(rows_html)
                  + "</tbody></table>")

    # The view aids resolve a cell by its DOM position, which matches sheet
    # coordinates only while nothing is merged — a merge removes cells and
    # shifts every index after it. vSNP3 tables never merge; a sheet that does
    # simply renders without the aids (loci/samples stay empty, so the page
    # offers no controls rather than mismapped ones). IGV launching is
    # unaffected either way: on this path every variant cell carries its own
    # anchor, which the delegated handler deliberately leaves alone.
    aids_ok = vtable is not None and not merged_skip and not merge_anchors
    loci_map = ({str(c): locus for c, locus in vtable["positions"].items()}
                if aids_ok else {})
    samples_list = row_samples_full if aids_ok else []

    # The full-fidelity path styles every cell inline and gives each variant
    # cell its own anchor, so it needs no style palette and has nothing to page
    # through — the lazy-loading and delegated-click machinery simply sits idle.
    page = _PAGE_TEMPLATE.format(
        download_link=download_link("Download xlsx", download_href),
        title=display_title,
        filename=html.escape(xlsx_path.name),
        sheet=html.escape(ws.title or "Sheet1"),
        rows=ws.max_row,
        cols=ws.max_column,
        notice="",
        style_css="",
        table=table_html,
        loaded=0,
        available=0,
        project=html.escape(project or "", quote=True),
        loci_json=json.dumps(loci_map),
        samples_json=json.dumps(samples_list),
        controls=view_controls(loci_map, samples_list),
    )
    return page


_MESSAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<style>
  :root {{
    --ink: #1a2536; --muted: #6a7585; --line: #d4ccb8;
    --bg: #f6f3ec; --panel: #fbf9f3;
  }}
  * {{ box-sizing: border-box; }}
  html, body {{ margin: 0; padding: 0; min-height: 100%; background: var(--bg);
    color: var(--ink); font-family: 'IBM Plex Sans', system-ui, sans-serif;
    line-height: 1.55; }}
  .tl-bar {{ background: var(--panel); border-bottom: 1px solid var(--line);
    padding: 10px 20px; font-size: 13px; color: var(--muted);
    font-family: 'IBM Plex Mono', ui-monospace, monospace; }}
  .tl-wrap {{ max-width: 640px; margin: 0 auto; padding: 48px 24px; }}
  h1 {{ font-size: 19px; font-weight: 650; margin: 0 0 12px; }}
  p {{ font-size: 15px; margin: 0; }}
  .tl-note {{ font-size: 13px; color: var(--muted); margin-top: 18px; }}
  @media (prefers-color-scheme: dark) {{
    :root {{ --ink: #e8edf2; --muted: #98a4b3; --line: #2d3d47;
      --bg: #0f171d; --panel: #17222b; }}
  }}
</style>
</head>
<body>
{bar}
<div class="tl-wrap">
  <h1>{heading}</h1>
  <p>{body}</p>
  {hint}
</div>
</body>
</html>
"""


_TOO_LARGE_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<style>
  :root {{
    --ink: #1a2536; --muted: #6a7585; --line: #d4ccb8;
    --bg: #f6f3ec; --panel: #fbf9f3; --accent: #b85a3e;
  }}
  * {{ box-sizing: border-box; }}
  html, body {{ margin: 0; padding: 0; min-height: 100%; background: var(--bg);
    color: var(--ink); font-family: 'IBM Plex Sans', system-ui, sans-serif;
    line-height: 1.55; }}
  .tl-bar {{ background: var(--panel); border-bottom: 1px solid var(--line);
    padding: 10px 20px; font-size: 13px; color: var(--muted);
    font-family: 'IBM Plex Mono', ui-monospace, monospace; }}
  .tl-wrap {{ max-width: 760px; margin: 0 auto; padding: 40px 24px 64px; }}
  h1 {{ font-size: 21px; font-weight: 650; margin: 0 0 12px; letter-spacing: -0.01em; }}
  .tl-lede {{ font-size: 15px; margin: 0 0 6px; }}
  .tl-dims {{ font-size: 13px; color: var(--muted);
    font-family: 'IBM Plex Mono', ui-monospace, monospace; margin: 0 0 28px; }}
  .tl-route {{ background: var(--panel); border: 1px solid var(--line);
    border-radius: 10px; padding: 18px 20px; margin: 0 0 14px; }}
  .tl-route h2 {{ font-size: 15px; font-weight: 650; margin: 0 0 6px; }}
  .tl-route p {{ font-size: 14px; margin: 0 0 14px; color: #2c3a4d; }}
  .tl-btn {{ display: inline-block; background: var(--accent); color: #fff;
    text-decoration: none; font-size: 14px; font-weight: 600;
    padding: 8px 16px; border-radius: 7px; }}
  .tl-btn:hover {{ filter: brightness(1.07); }}
  .tl-note {{ font-size: 12.5px; color: var(--muted); margin-top: 10px;
    font-family: 'IBM Plex Mono', ui-monospace, monospace; }}
  .tl-note code {{ font-size: 12px; }}
  @media (prefers-color-scheme: dark) {{
    :root {{ --ink: #e8edf2; --muted: #98a4b3; --line: #2d3d47;
      --bg: #0f171d; --panel: #17222b; --accent: #d4744f; }}
    .tl-route p {{ color: #cbd5df; }}
  }}
</style>
</head>
<body>
<div class="tl-bar">{filename}</div>
<div class="tl-wrap">
  <h1>{heading}</h1>
  <p class="tl-lede">{lede}</p>
  <p class="tl-dims">{dims}</p>

  <div class="tl-route">
    <h2>Open it in a spreadsheet application</h2>
    <p>Excel, LibreOffice Calc and OpenOffice Calc all handle sheets this size,
       with the conditional formatting intact.</p>
    <a class="tl-btn" href="{download_url}">Download the spreadsheet</a>
  </div>

  <div class="tl-route">
    <h2>{tree_lead}</h2>
    <p>{tree_body}</p>
    {tree_block}
  </div>
</div>
</body>
</html>
"""


_PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
  :root {{
    --xlsx-border: #d4ccb8;
    --xlsx-header-bg: #fbf9f3;
    --xlsx-text: #1a2536;
    --xlsx-muted: #6a7585;
  }}
  * {{ box-sizing: border-box; }}
  html, body {{ margin: 0; padding: 0; height: 100%; background: #f6f3ec; color: var(--xlsx-text); font-family: 'IBM Plex Sans', system-ui, sans-serif; }}
  .xlsx-bar {{
    position: sticky; top: 0; z-index: 10;
    background: var(--xlsx-header-bg);
    border-bottom: 1px solid var(--xlsx-border);
    padding: 10px 20px;
    display: flex; align-items: baseline; gap: 18px; flex-wrap: wrap;
  }}
  .xlsx-truncated {{
    margin: 0; padding: 10px 20px;
    background: #fff5e0; border-bottom: 1px solid #e6c98a;
    color: #4a3a12; font-size: 13px; line-height: 1.5;
  }}
  /* Clade-filtered view (opened from the tree viewer): green = a deliberate
     subset, distinct from the amber too-big-to-show truncation above. */
  .xlsx-filter {{
    margin: 0; padding: 10px 20px;
    background: #e9f2e7; border-bottom: 1px solid #a9c6a2;
    color: #1e3a1a; font-size: 13px; line-height: 1.5;
  }}
  .xlsx-filter a {{ color: #b85a3e; }}
  .xlsx-filter-warn {{
    background: #fff5e0; border-bottom: 1px solid #e6c98a; color: #4a3a12;
  }}
  /* Written only when a launch actually failed — a click that opens nothing has
     to say so somewhere the user is already looking. */
  .xlsx-igv-note {{
    padding: 10px 20px; background: #fdecea; border-bottom: 1px solid #e6a8a0;
    color: #5a1d16; font-size: 13px; line-height: 1.5;
  }}
  .xlsx-bar .filename {{ font-weight: 600; font-size: 14px; }}
  .xlsx-bar .meta {{ color: var(--xlsx-muted); font-size: 12px; font-family: 'IBM Plex Mono', monospace; }}
  .xlsx-bar a {{ color: #b85a3e; text-decoration: none; font-size: 13px; }}
  .xlsx-bar a:hover {{ text-decoration: underline; }}
  .xlsx-wrap {{ padding: 16px 20px; overflow: auto; }}
  table.xlsx {{
    border-collapse: collapse;
    background: white;
    font-size: 13px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    table-layout: fixed;  /* honor <colgroup> widths */
  }}
  table.xlsx td {{
    padding: 2px 4px;
    border: 1px solid #eee;  /* light default; cell inline styles override */
    white-space: nowrap;
    vertical-align: middle;
    text-align: center;
    overflow: visible;  /* let rotated annotation rows extend past the row when needed */
  }}
  /* Sample-name column (anchored at first column) should left-align its long
     identifiers rather than centering. */
  table.xlsx td:first-child {{ text-align: left; padding-left: 8px; }}
  .xlsx-sticky-top {{ position: sticky; top: 0; z-index: 2; background: var(--xlsx-header-bg); }}
  .xlsx-sticky-left {{ position: sticky; left: 0; z-index: 1; background: var(--xlsx-header-bg); }}
  .xlsx-sticky-top.xlsx-sticky-left {{ z-index: 3; }}
  /* Excel text-rotation: 1..90 = bottom-to-top read; 91..180 = top-to-bottom. */
  .xlsx-rot-up, .xlsx-rot-down, .xlsx-rot-stacked {{
    white-space: nowrap;
    vertical-align: bottom;
    text-align: left;
    padding: 4px 1px;  /* tight horizontal — narrow column is already the rotated text height */
  }}
  .xlsx-rot-up {{ writing-mode: vertical-rl; transform: rotate(180deg); }}
  .xlsx-rot-down {{ writing-mode: vertical-rl; }}
  .xlsx-rot-stacked {{ writing-mode: vertical-rl; text-orientation: upright; }}
  /* IGV launch on variant cells: the whole colored cell is clickable.
     The anchor fills the cell and inherits color/alignment so the
     nucleotide letter reads normally; a hover outline signals the cell
     is a launch target. */
  td.xlsx-variant {{ position: relative; padding: 0; }}
  td.xlsx-variant a.xlsx-igv-cell {{
    display: block;
    width: 100%;
    height: 100%;
    padding: 2px 4px;
    box-sizing: border-box;
    color: inherit;
    text-align: inherit;
    text-decoration: none;
    cursor: pointer;
    min-height: 1em;
  }}
  td.xlsx-variant a.xlsx-igv-cell:hover {{
    outline: 2px solid rgba(15, 22, 33, 0.6);
    outline-offset: -2px;
  }}
  /* Calls-only — sample is loadable but has only the VCF, no BAM; italic
     hints that this click won't produce a reads pile-up. */
  td.xlsx-variant a.xlsx-igv-cell.xlsx-igv-calls-only {{ font-style: italic; }}
  /* Disabled — sample has neither BAM nor imported VCF; render the value
     plain (not clickable). Tooltip explains why. */
  td.xlsx-variant span.xlsx-igv-disabled {{
    display: block;
    padding: 2px 4px;
    cursor: default;
  }}
  /* Cheap variant cells (streaming renderer): the whole <td> is the launch
     target, so it needs no inner anchor at all. The sample comes from the
     row's first cell and the locus from the header row, both resolved by the
     delegated click handler below. */
  td.xlsx-variant {{ cursor: pointer; }}
  td.xlsx-variant:hover {{
    outline: 2px solid rgba(15, 22, 33, 0.6);
    outline-offset: -2px;
  }}
  td.xlsx-igv-calls-only {{ font-style: italic; }}
  td.xlsx-igv-none {{ cursor: default; }}
  #xlsxMore {{
    padding: 10px 20px; font-size: 13px; color: var(--xlsx-muted);
  }}
  #xlsxMore button {{
    font: inherit; padding: 4px 10px; cursor: pointer;
  }}
  /* View aids in the bar: the identical-position toggle and the colour hint. */
  .xlsx-ctl {{ font-size: 12.5px; white-space: nowrap; }}
  .xlsx-ctl input {{ vertical-align: -2px; margin: 0 4px 0 0; }}
  .xlsx-ctl-off {{ color: var(--xlsx-muted); }}
  .xlsx-ctl-note {{ font-size: 12px; color: var(--xlsx-muted); }}
  /* Cells that toggle a colour mark: sample names and position headers. */
  td.xlsx-hl-target {{ cursor: pointer; }}
  td.xlsx-hl-target:hover {{
    outline: 2px dashed rgba(15, 22, 33, 0.45);
    outline-offset: -2px;
  }}
  /* Per-cell styles, deduplicated into classes — the same handful of fills and
     fonts repeat across every cell of these tables. */
  {style_css}
</style>
</head>
<body>
<div class="xlsx-bar">
  <span class="filename">{filename}</span>
  <span class="meta">sheet: {sheet} · {rows} rows × {cols} cols</span>
  {controls}
  <span style="margin-left: auto;">{download_link}</span>
</div>
{notice}
<div id="xlsxIgvNote" class="xlsx-igv-note" style="display:none"></div>
<div class="xlsx-wrap" id="xlsxWrap">
{table}
</div>
<div id="xlsxMore"></div>
<script>
// IGV launcher for cascade-table variant cells. First click opens an IGV tab;
// later clicks reuse it additively (postMessage tells IgvStandalone to add any
// new samples and navigate to the new locus).
//
// Every failure here used to be silent, which is the whole reason "I click a
// SNP and no tab opens" was so hard to pin down. Three of them:
//
//  1. postMessage into a STALE handle. `w.closed` is false for a tab that is
//     still open but no longer running this app — which is routine under Open
//     OnDemand, where the URL carries the compute node and port and both change
//     when the session is recycled. The message went nowhere, we returned as if
//     delivered, and nothing happened at all. The viewer now has to ACKNOWLEDGE
//     it; if no ack arrives we treat the handle as dead and open a fresh tab.
//  2. `window.open` returning null when a pop-up blocker refuses. Never
//     checked, so the click did nothing and said nothing. Now reported.
//  3. The new tab was never focused, so on setups that open it in the
//     background the user's answer is still "nothing happened".
(function() {{
  var ACK_MS = 700;

  window.addEventListener("message", function(ev) {{
    if (ev.origin !== window.location.origin) return;
    if (ev.data && ev.data.type === "vsnpIgvAck") window.__vsnpIgvAck = true;
  }});

  function say(msg) {{
    var el = document.getElementById("xlsxIgvNote");
    if (!el) return;
    el.textContent = msg;
    el.style.display = msg ? "block" : "none";
  }}

  function openFresh(url) {{
    // A NAMED target reuses a tab called "vsnp_igv" anywhere in this browser
    // session, including one left over from an earlier OOD session pointing at
    // a node and port that no longer exist. That reuse is wanted when the tab
    // is live and is why the name is here; the ack check above is what stops it
    // being a dead end when it is not.
    var win = window.open(url, "vsnp_igv");
    window.__vsnpIgvWin = win;
    if (!win) {{
      say("Your browser blocked the IGV window. Allow pop-ups for this site, "
          + "then click the cell again. (Ctrl- or Cmd-click opens it directly.)");
      return;
    }}
    say("");
    try {{ win.focus(); }} catch (e) {{ /* focus is a courtesy, not a requirement */ }}
  }}

  window.__vsnpLaunchIgv = function(url) {{
    var w = window.__vsnpIgvWin;
    if (w && !w.closed) {{
      window.__vsnpIgvAck = false;
      var sent = false;
      try {{
        w.postMessage({{ type: "vsnpIgvLaunch", url: url }}, window.location.origin);
        sent = true;
        w.focus();
      }} catch (e) {{ sent = false; }}
      if (sent) {{
        setTimeout(function() {{
          if (!window.__vsnpIgvAck) {{
            window.__vsnpIgvWin = null;
            openFresh(url);
          }}
        }}, ACK_MS);
        return;
      }}
    }}
    openFresh(url);
  }};
}})();

// ---------------------------------------------------------------------------
// Delegated variant-cell clicks, view aids, and load-as-you-scroll.
//
// The delegation exists because a cascade table is far too big to hand the
// browser at once. Giving every variant cell its own <a href onclick title
// target> cost ~450 bytes a cell, which capped the preview at ~24 columns of a
// 10,001-column table. A cell now carries only a class; the sample is read from
// its row's first cell and the locus from the header row, so one listener on
// the table resolves any cell from its position.
(function() {{
  var LOCI = {loci_json};        // column index (1-based) -> "contig:pos"
  var SAMPLES = {samples_json};  // row index (0-based, incl. header) -> stem
  var PROJECT = "{project}";
  var loaded = {loaded}, available = {available};
  var body = document.getElementById("xlsxBody");
  var more = document.getElementById("xlsxMore");
  var table = document.getElementById("xlsxTable");
  if (!body || !table) return;

  table.addEventListener("click", function(ev) {{
    // Full-fidelity pages give every variant cell its own anchor (and every
    // no-data cell an explanatory span). Those cells answer for themselves;
    // acting here as well would launch IGV twice per click — or once for a
    // sample whose cell was deliberately left unclickable.
    if (ev.target.closest && ev.target.closest("a, .xlsx-igv-disabled")) return;
    var td = ev.target.closest ? ev.target.closest("td") : null;
    if (!td || !td.classList.contains("xlsx-variant")) return;
    var tr = td.parentNode;
    // Row index within the whole table, so SAMPLES lines up after lazy loads.
    var rowIdx = tr.rowIndex;
    var colIdx = td.cellIndex + 1;
    var sample = SAMPLES[rowIdx];
    var locus = LOCI[String(colIdx)];
    if (!sample || !locus) return;
    var url = "../../../?view=igv&tracks=" + encodeURIComponent(PROJECT + ":" + sample)
            + "&locus=" + encodeURIComponent(locus);
    // Modifier-click keeps the browser's own "open in a new tab" behaviour.
    if (ev.metaKey || ev.ctrlKey || ev.shiftKey) {{ window.open(url, "_blank"); return; }}
    window.__vsnpLaunchIgv(url);
  }});

  // ---- colour marks -------------------------------------------------------
  // Click a sample name to mark its row, a position header to mark its
  // column — the table's version of the tree viewer's "Colour it", in the
  // same palette. Marks are nth-child CSS rules rather than edits to cells:
  // rows that lazy-load later pick their marks up on arrival, and a column
  // mark costs one rule instead of one edit per cell. The marks are drawn as
  // lane edges and a tinted name/header cell, never as cell fills — in these
  // tables the fill IS the data (it says which base was called), and a mark
  // that painted over it would cost exactly the information being compared.
  var PALETTE = ["#1f77b4", "#d95f02", "#2ca02c", "#9467bd",
                 "#8c564b", "#e7298a", "#17becf", "#bcbd22"];
  var hlRows = new Map(), hlCols = new Map(), hlNext = 0;
  var hlStyle = document.createElement("style");
  document.head.appendChild(hlStyle);
  var clearLink = document.getElementById("xlsxHlClear");

  function tintOf(hex) {{
    var mix = function(v) {{ return Math.round(255 - (255 - v) * 0.25); }};
    return "rgb(" + mix(parseInt(hex.slice(1, 3), 16))
         + "," + mix(parseInt(hex.slice(3, 5), 16))
         + "," + mix(parseInt(hex.slice(5, 7), 16)) + ")";
  }}

  function renderMarks() {{
    var css = [];
    hlRows.forEach(function(colour, r) {{
      var tr = "#xlsxBody tr:nth-child(" + (r + 1) + ")";
      css.push(tr + " td {{ box-shadow: inset 0 2px 0 0 " + colour
               + ", inset 0 -2px 0 0 " + colour + "; }}");
      css.push(tr + " td:first-child {{ background-color: " + tintOf(colour)
               + "; box-shadow: inset 3px 0 0 0 " + colour + "; }}");
    }});
    hlCols.forEach(function(colour, c) {{
      css.push("#xlsxBody td:nth-child(" + c + ") {{ box-shadow: inset 2px 0 0 0 "
               + colour + ", inset -2px 0 0 0 " + colour + "; }}");
      css.push("#xlsxBody tr:first-child td:nth-child(" + c
               + ") {{ background-color: " + tintOf(colour)
               + "; box-shadow: inset 0 3px 0 0 " + colour + "; }}");
    }});
    hlStyle.textContent = css.join(" ");
    if (clearLink) clearLink.style.display = (hlRows.size || hlCols.size) ? "" : "none";
  }}

  function toggleMark(map, key) {{
    if (map.has(key)) map.delete(key);
    else {{ map.set(key, PALETTE[hlNext++ % PALETTE.length]); }}
    renderMarks();
  }}

  if (clearLink) clearLink.addEventListener("click", function(ev) {{
    ev.preventDefault();
    hlRows.clear(); hlCols.clear(); hlNext = 0;
    renderMarks();
  }});

  table.addEventListener("click", function(ev) {{
    if (ev.target.closest && ev.target.closest("a")) return;
    var td = ev.target.closest ? ev.target.closest("td") : null;
    if (!td) return;
    var tr = td.parentNode;
    if (tr.rowIndex === 0 && LOCI[String(td.cellIndex + 1)]) {{
      toggleMark(hlCols, td.cellIndex + 1);
    }} else if (td.cellIndex === 0 && SAMPLES[tr.rowIndex]) {{
      toggleMark(hlRows, tr.rowIndex);
    }}
  }});

  // The clickable name/header cells say so on hover. Applied to rows as they
  // arrive — markTargets() starts where the previous call stopped.
  var markedTo = 0;
  function markTargets() {{
    var rows = body.rows;
    if (!rows.length) return;
    if (markedTo === 0) {{
      var hdr = rows[0].cells;
      for (var c = 0; c < hdr.length; c++) {{
        if (LOCI[String(c + 1)]) {{
          hdr[c].classList.add("xlsx-hl-target");
          hdr[c].title = "Click to colour this position";
        }}
      }}
    }}
    for (var r = Math.max(markedTo, 1); r < rows.length; r++) {{
      if (SAMPLES[r] && rows[r].cells.length) {{
        rows[r].cells[0].classList.add("xlsx-hl-target");
        rows[r].cells[0].title = "Click to colour this sample";
      }}
    }}
    markedTo = rows.length;
  }}
  markTargets();

  // ---- the identical-position toggle --------------------------------------
  // "Identical" is exact: a position hides only when every sample row shows
  // the very same text in its cell — same base, same ambiguity code, same
  // nothing. Any one sample disagreeing (a different call, a missing '-',
  // anything) keeps the position on screen.
  var invBox = document.getElementById("xlsxInvariant");
  var invNote = document.getElementById("xlsxInvariantNote");
  var invColsMemo = null;   // computed once; see the note at the toggle
  var colgroupEl = table.querySelector("colgroup");
  var allCols = colgroupEl ? Array.prototype.slice.call(colgroupEl.children) : [];
  var hideStyle = document.createElement("style");
  document.head.appendChild(hideStyle);

  function identicalCols() {{
    var rows = body.rows, sampleRows = [];
    for (var r = 0; r < rows.length; r++) if (SAMPLES[r]) sampleRows.push(rows[r]);
    var out = [];
    if (sampleRows.length < 2) return out;   // nothing to compare — hide nothing
    Object.keys(LOCI).forEach(function(k) {{
      var c = parseInt(k, 10) - 1;
      var first = null;
      for (var i = 0; i < sampleRows.length; i++) {{
        var cell = sampleRows[i].cells[c];
        var t = cell ? cell.textContent.trim() : "";
        if (first === null) first = t;
        else if (t !== first) return;
      }}
      out.push(c + 1);
    }});
    return out;
  }}

  function setHiddenCols(cols) {{
    // The <col> elements are physically removed (and restored) so the fixed
    // table layout keeps handing the right width to each cell that remains;
    // hiding only the cells would shift every column onto its neighbour's
    // width. The cells themselves hide by nth-child rule — those count DOM
    // position, so nothing renumbers and the IGV click mapping stays right.
    var hidden = {{}};
    cols.forEach(function(c) {{ hidden[c] = true; }});
    while (colgroupEl.firstChild) colgroupEl.removeChild(colgroupEl.firstChild);
    for (var i = 0; i < allCols.length; i++) {{
      if (!hidden[i + 1]) colgroupEl.appendChild(allCols[i]);
    }}
    var sels = cols.map(function(c) {{ return "#xlsxTable td:nth-child(" + c + ")"; }});
    var rules = [];
    for (var j = 0; j < sels.length; j += 400) {{
      rules.push(sels.slice(j, j + 400).join(",") + " {{ display: none; }}");
    }}
    hideStyle.textContent = rules.join(" ");
  }}

  if (invBox && !invBox.disabled && colgroupEl) {{
    invBox.addEventListener("change", function() {{
      if (!invBox.checked) {{
        setHiddenCols([]);
        invNote.textContent = "";
        return;
      }}
      // Memoized after the first computation: the answer walks every sample
      // cell of every row (~1M textContent reads near the full-view bound),
      // and it cannot change afterwards — loadAllRows() has fetched the whole
      // window and row text is static (colour marks are CSS classes, not text
      // edits). Re-ticking the box is instant instead of a multi-second walk.
      if (invColsMemo !== null) {{
        setHiddenCols(invColsMemo);
        invNote.textContent = invColsMemo.length
          ? "hiding " + invColsMemo.length + " of " + Object.keys(LOCI).length + " positions"
          : "no position is identical across the shown samples";
        return;
      }}
      invNote.textContent = "checking every row…";
      loadAllRows().then(function() {{
        var cols = identicalCols();
        invColsMemo = cols;
        setHiddenCols(cols);
        invNote.textContent = cols.length
          ? "hiding " + cols.length + " of " + Object.keys(LOCI).length + " positions"
          : "no position is identical across the shown samples";
      }}).catch(function() {{
        invBox.checked = false;
        invNote.textContent = "could not load the remaining rows — nothing was hidden";
      }});
    }});
  }}

  // ---- load-as-you-scroll -------------------------------------------------
  var done = loaded >= available, inflight = null;
  function status() {{
    if (!more) return;
    more.innerHTML = done
      ? ""
      : "Showing " + loaded + " of " + available + " rows. "
        + "<button type=\\"button\\" id=\\"xlsxMoreBtn\\">Load more</button>";
    var b = document.getElementById("xlsxMoreBtn");
    if (b) b.addEventListener("click", function() {{ fetchMore().catch(function() {{}}); }});
  }}
  function fetchMore() {{
    if (done) return Promise.resolve();
    if (inflight) return inflight;
    more.textContent = "Loading rows " + (loaded + 1) + "-"
                     + Math.min(loaded + 200, available) + "…";
    var u = new URL(window.location.href);
    u.searchParams.set("rows_from", loaded);
    u.searchParams.set("rows_count", 200);
    inflight = fetch(u.toString(), {{ headers: {{ "Accept": "text/html" }} }})
      .then(function(r) {{ if (!r.ok) throw new Error("HTTP " + r.status); return r.text(); }})
      .then(function(htmlText) {{
        inflight = null;
        if (!htmlText.trim()) {{ done = true; status(); return; }}
        body.insertAdjacentHTML("beforeend", htmlText);
        loaded = body.rows.length;
        if (loaded >= available) done = true;
        markTargets();
        status();
      }})
      .catch(function(e) {{
        inflight = null;
        more.textContent = "Could not load more rows (" + e.message + "). ";
        var b = document.createElement("button");
        b.textContent = "Retry";
        b.addEventListener("click", function() {{ more.textContent = ""; fetchMore().catch(function() {{}}); }});
        more.appendChild(b);
        throw e;
      }});
    return inflight;
  }}
  // Every window row, before the toggle answers: a claim about "all samples"
  // must not be computed from whatever prefix happened to be scrolled in.
  function loadAllRows() {{
    return done ? Promise.resolve() : fetchMore().then(loadAllRows);
  }}

  if (done) {{ if (more) more.remove(); return; }}
  // Load the next batch as the bottom of the table comes into view, and keep
  // the button as the manual fallback (and for browsers without the observer).
  status();
  if (window.IntersectionObserver) {{
    new IntersectionObserver(function(entries) {{
      if (entries.some(function(e) {{ return e.isIntersecting; }})) fetchMore().catch(function() {{}});
    }}, {{ rootMargin: "600px" }}).observe(more);
  }}
}})();
</script>
</body>
</html>
"""
