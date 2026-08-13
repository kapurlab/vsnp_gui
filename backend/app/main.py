from fastapi import FastAPI, HTTPException, UploadFile, File, Query, Request
from fastapi.responses import Response, FileResponse, HTMLResponse
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple
import zipfile
import csv
import io
import socket
import json
import os
import time
import subprocess
import shutil
import gzip
import signal
import sys
import logging
import shlex
import re
import hashlib
import tempfile
import threading

# Serializes Step 1 dispatch so a fast double-click can't start two batches
# that race over the same per-sample dirs. The check ("is a job already
# running?") and the claim (write .step1_job_id) span ~200 lines; without this
# lock both requests pass the check before either claims. uvicorn runs
# single-process here, so a threading.Lock is sufficient.
_STEP1_DISPATCH_LOCK = threading.Lock()

# Per-sample status cache for the Step 1 status endpoint. A sample that has
# written .provenance/exit_code is terminal — its on-disk outputs won't change
# until it is re-run (which rewrites exit_code, bumping the mtime). We key the
# cache on that mtime so a poll over a mostly-finished 1000-sample project does
# a single stat() per completed sample instead of re-globbing every sample's
# whole output tree every 5 seconds (what tripped "Failed to load Step 1
# status"). Maps sample_dir path -> (exit_code_mtime_ns, cached_entry_without_
# in_vcfs_folder). Access is under the GIL from the sync-endpoint threadpool;
# dict get/set are atomic enough that no extra lock is warranted.
_STEP1_STATUS_CACHE: Dict[str, tuple] = {}

from app.config import load_config, save_config, SITE_ROOT, TOOLS_ROOT, DB_ROOT
from app.jobs import JobManager
from app.request_safety import install_request_safety
from app import qc_verdict
from app import provenance_writer
from app.step2_staging import stage_step2_vcfs
from app.projects import (
    create_project,
    list_projects,
    ensure_project_dirs,
    archive_project,
    delete_project,
    update_project_meta,
    resolve_project_dir,
    vcf_db_dir,
    SCOPE_PERSONAL,
    SCOPE_SHARED,
)
from app.refs import (
    list_references,
    get_reference_paths,
    add_reference_path,
    remove_reference_path,
    reference_roots,
    find_gff_for_fasta,
    sanitize_upstream_paths,
)
from app.sra import expand_accessions, expand_accessions_with_mapping, build_download_script, SRAExpansionError, write_crosswalk_tsv
from app.posthoc import list_tools as posthoc_list_tools, get_tool as posthoc_get_tool, tool_status as posthoc_tool_status

app = FastAPI(title="vSNP GUI API")
logger = logging.getLogger("uvicorn.error")

install_request_safety(app)

cfg = load_config()
projects_root = Path(cfg["projects_root"])
projects_root.mkdir(parents=True, exist_ok=True)
job_manager = JobManager(Path(cfg["projects_root"]) / ".jobs")

# One-time per vsnp3 install: drop the upstream author's shipped reference
# locations (/Users/todstuber/..., /home/tstuber/...) from
# dependencies/reference_options_paths.txt — a fresh deployment must start
# with an empty Reference Locations list, not another machine's.
try:
    _removed_upstream = sanitize_upstream_paths(Path(cfg["vsnp3_path"]))
    if _removed_upstream:
        logger.info("Removed upstream-shipped reference locations: %s",
                    ", ".join(_removed_upstream))
except Exception as _exc:  # never block startup on registry hygiene
    logger.warning("reference locations sanitize skipped: %s", _exc)


def _resolve_app_version() -> str:
    """Version of the deployed checkout — the exact string the Diagnostic
    Tools Dashboard shows for this tool (`git describe --tags --always`,
    the same command bdtools runs). Resolved once at startup; empty when
    git or the .git dir is unavailable, in which case the frontend falls
    back to its built-in constant."""
    try:
        out = subprocess.run(
            ["git", "-C", str(Path(__file__).resolve().parents[2]),
             "describe", "--tags", "--always"],
            capture_output=True, text=True, timeout=10,
        )
        return out.stdout.strip() if out.returncode == 0 else ""
    except Exception:
        return ""


APP_VERSION = _resolve_app_version()


def _project_roots(cfg_in: Dict) -> List:
    """Build the list of (scope, root) pairs from config. Personal first
    (so it wins on name collisions), then shared if present."""
    personal = cfg_in.get("projects_root", "").strip()
    shared = cfg_in.get("shared_projects_root", "").strip()
    out = []
    if personal:
        out.append((SCOPE_PERSONAL, Path(personal)))
    if shared:
        out.append((SCOPE_SHARED, Path(shared)))
    return out


def _project_dir_for(cfg_in: Dict, name: str) -> Path:
    """Resolve a project name to its on-disk directory across roots.
    Falls back to the personal root if the project doesn't exist yet
    (e.g. mid-create). Raises ValueError on invalid names."""
    if "/" in name or name.startswith("."):
        raise ValueError("Invalid project name")
    found = resolve_project_dir(_project_roots(cfg_in), name)
    if found is not None:
        return found
    return Path(cfg_in.get("projects_root", "")) / name


def _resolve_qc_thresholds(cfg: Dict, project_dir: Path | None = None) -> Dict:
    """Merge thresholds: module DEFAULTS < user cfg < project.json override."""
    project_layer = None
    if project_dir is not None:
        pj = project_dir / "project.json"
        if pj.is_file():
            try:
                project_layer = json.loads(pj.read_text()).get("qc_thresholds")
            except (json.JSONDecodeError, OSError):
                project_layer = None
    return qc_verdict.merge_thresholds(cfg.get("qc_thresholds"), project_layer)


def _annotate_qc_rows(rows: list, thresholds: Dict) -> list:
    """Attach `_qc_verdict` to each row in place. Returns the same list."""
    for row in rows:
        if isinstance(row, dict):
            row["_qc_verdict"] = qc_verdict.compute_verdict(row, thresholds)
    return rows


def _project_reference(project_dir: Path) -> str:
    """Return the project's locked reference, or "" if not yet set.
    Reads project.json["reference"] only — callers that want the inferred
    value should call the reference_lock endpoint and let it write it back."""
    meta_path = project_dir / "project.json"
    if not meta_path.exists():
        return ""
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return ""
    return (meta.get("reference") or "").strip()


def _project_reference_fasta_and_gff(project_dir: Path, cfg: Dict) -> tuple[str, str]:
    """Resolve a project's reference fasta + GFF paths for IGV consumption.

    Used by step1_files for imported-VCF samples (those that exist only in
    step2/vcf_database/, never went through Step 1, so have no alignment dir
    to crib the reference from).

    Strategy 1 — borrow from any step1 sample that DOES have an alignment
    dir in this project. Cheap when at least one sample was aligned here.

    Strategy 2 — use project.json's "reference" name to look up the dir
    under the configured reference roots. Needed for pure-import projects.

    Returns ("", "") when neither strategy finds a fasta — caller should
    surface that as a friendly error rather than a partial IGV view.
    """
    vsnp3_path = Path(cfg.get("vsnp3_path", ""))
    step1_dir = project_dir / "step1"
    if step1_dir.is_dir():
        for d in sorted(step1_dir.iterdir()):
            if not d.is_dir():
                continue
            fastas = _align_glob(d, "*.fasta")
            if fastas:
                gff = find_gff_for_fasta(fastas[0], vsnp3_path)
                return str(fastas[0]), (str(gff) if gff else "")
    ref_name = _project_reference(project_dir)
    if ref_name:
        for root in reference_roots(vsnp3_path):
            ref_dir = root / ref_name
            if not ref_dir.is_dir():
                continue
            fastas = sorted(ref_dir.glob("*.fasta"))
            if fastas:
                gff = find_gff_for_fasta(fastas[0], vsnp3_path)
                return str(fastas[0]), (str(gff) if gff else "")
    return "", ""


def _path_under_any_project_root(cfg_in: Dict, target: Path) -> bool:
    """Security check for /serve, /download-file, /posthoc/open: target
    must resolve under any configured projects root."""
    target = target.resolve()
    for _scope, root in _project_roots(cfg_in):
        try:
            if str(target).startswith(str(root.resolve()) + "/") or target == root.resolve():
                return True
        except OSError:
            continue
    return False


def _wrapper_pids(script_path: Path) -> List[int]:
    """PIDs of any process whose command line references the wrapper script.

    Used to detect (and, via _terminate_step1_wrapper, kill) an orphaned bash
    wrapper that survived a backend reload, where in-memory JobManager state is
    gone but the batch is still running.
    """
    try:
        result = subprocess.run(
            ["pgrep", "-f", str(script_path)],
            capture_output=True,
            text=True,
            timeout=3,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return []
    if result.returncode != 0:
        return []
    pids: List[int] = []
    for line in result.stdout.split():
        try:
            pids.append(int(line))
        except ValueError:
            continue
    return pids


def _wrapper_process_alive(script_path: Path) -> bool:
    """True if an orphaned wrapper process for the script is still running."""
    return bool(_wrapper_pids(script_path))


def _terminate_step1_wrapper(script_path: Path) -> bool:
    """Kill an orphaned Step 1 batch by signaling its process group(s).

    Restart-resilient counterpart to JobManager.stop_job: after a backend
    reload the in-memory job (and its Popen handle) is gone, so stop_job can't
    find anything to signal. The wrapper is launched with start_new_session=True
    (jobs.py), so its PID == its process-group id and every vsnp3 worker it
    spawned shares that group. We rediscover the live wrapper via pgrep, derive
    the group(s), SIGTERM them, then SIGKILL survivors after a grace period in a
    daemon thread so the request returns immediately. Returns True if at least
    one group was signaled."""
    pgids = set()
    for pid in _wrapper_pids(script_path):
        try:
            pgids.add(os.getpgid(pid))
        except ProcessLookupError:
            continue
    if not pgids:
        return False

    def _signal_all(sig: int) -> None:
        for pgid in pgids:
            try:
                os.killpg(pgid, sig)
            except (ProcessLookupError, PermissionError):
                continue

    _signal_all(signal.SIGTERM)

    def _escalate() -> None:
        # Poll for the grace period; SIGKILL anything still alive.
        deadline = time.monotonic() + 10
        while time.monotonic() < deadline:
            if not _wrapper_process_alive(script_path):
                return
            time.sleep(0.5)
        _signal_all(signal.SIGKILL)

    threading.Thread(target=_escalate, daemon=True).start()
    return True


def _script_bin_dir(cfg: Dict) -> Optional[Path]:
    vsnp3_path = cfg.get("vsnp3_path", "").strip()
    if not vsnp3_path:
        return None
    candidate = Path(vsnp3_path) / "bin"
    return candidate if candidate.is_dir() else None


def _tool_bin_dir(cfg: Dict) -> Optional[Path]:
    bcftools_path = cfg.get("bcftools_path", "").strip()
    if bcftools_path:
        candidate = Path(bcftools_path).expanduser().resolve().parent
        if candidate.is_dir():
            return candidate
    return None


def build_env(cfg: Dict) -> Dict[str, str]:
    current_path = os.environ.get("PATH", "")
    path_parts: List[str] = []
    tool_bin = _tool_bin_dir(cfg)
    script_bin = _script_bin_dir(cfg)
    if tool_bin:
        path_parts.append(str(tool_bin))
    if script_bin:
        path_parts.append(str(script_bin))
    if current_path:
        path_parts.append(current_path)
    return {"PATH": ":".join(path_parts)}


# T-07 provenance helpers ---------------------------------------------------


def _ood_session_id() -> Optional[str]:
    """Best-effort OOD session UUID extraction.

    OOD's batch_connect spawns uvicorn with cwd somewhere under
    `~/ondemand/data/sys/dashboard/batch_connect/sys/vsnp_gui/output/<UUID>/`.
    If the env doesn't expose the UUID directly (it usually doesn't), we
    derive from the cwd path. Returns None if neither source succeeds —
    actor.ood_session_id will be null in the metadata, which is acceptable.
    """
    sid = os.environ.get("OOD_SESSION_ID") or os.environ.get("OOD_BC_SESSION_ID")
    if sid:
        return sid
    try:
        for parent in [Path.cwd()] + list(Path.cwd().parents):
            if parent.parent.name == "output":
                # Looks like .../output/<UUID>/...; UUID is the dir whose parent
                # is named "output".
                return parent.name
    except OSError:
        pass
    return None


def _current_user() -> str:
    """User identity to record in provenance actor block."""
    return os.environ.get("USER") or str(os.getuid()) if hasattr(os, "getuid") else "unknown"


def _is_shared_project(cfg: Dict, project_dir: Path) -> bool:
    """True iff project_dir lies under cfg['shared_projects_root']."""
    shared_root = cfg.get("shared_projects_root", "")
    if not shared_root:
        return False
    try:
        project_dir.resolve().relative_to(Path(shared_root).resolve())
        return True
    except (ValueError, OSError):
        return False


def _step1_sample_names(step1_dir: Path) -> List[str]:
    """Discover step1 sample names: subdirs with at least one *_R1*.fastq.gz
    (or fallback patterns matching the bash script's R1 detection).

    Filters out hidden / underscore-prefixed dirs (the writer's _provenance/
    sibling) so they don't get treated as samples.

    NB: this is the *discovery* function — it accepts any dir that looks
    sample-shaped, including single-end-only. Callers that actually
    dispatch a batch should use `_step1_dispatch_plan()` which applies the
    stricter "paired + non-junk" gate.
    """
    samples = []
    for p in sorted(step1_dir.iterdir()):
        if not p.is_dir() or p.name.startswith(("_", ".")):
            continue
        if any(p.glob("*_R1*.fastq.gz")) or any(p.glob("*_1*.fastq.gz")) or any(p.glob("*.fastq.gz")):
            samples.append(p.name)
    return samples


def _step1_browser_samples(step1_dir: Path) -> List[Dict]:
    """List step1 sample dirs for the inline project sample browser.

    A step1 sample is any non-hidden subdirectory of step1/ (the writer's
    _provenance/ sibling and dot-dirs are excluded). Unlike
    _step1_sample_names(), this does NOT require reads to still be present:
    a command-line / imported project keeps its reads inside step1/<sample>/,
    but a native GUI run may have had them removed after alignment — either
    way the sample dir (with its vSNP outputs) is a real sample and should
    list. `is_pair` is best-effort from whatever reads remain on disk, mirroring
    the R1/R2 detection used elsewhere; it only drives the "R1+R2" badge."""
    out: List[Dict] = []
    if not step1_dir.is_dir():
        return out
    for p in sorted(step1_dir.iterdir()):
        if not p.is_dir() or p.name.startswith(("_", ".")):
            continue
        has_r2 = any(p.glob("*_R2*.fastq.gz")) or any(p.glob("*_2*.fastq.gz"))
        out.append({"sample": p.name, "is_pair": bool(has_r2)})
    return out


# T-46 Phase 1: filter at dispatch time so Step 1 doesn't abort on samples
# that vsnp3 can't actually process (single-end, or suspiciously small
# fastqs that are usually SRA submission errors). Without this, one bad
# sample takes down the whole batch because T-07 provenance dispatch
# requires every sample's inputs hash cleanly.
#
# The floor is config-driven (config.DEFAULTS["step1_min_fastq_bytes"]); this
# constant is the fallback default. 50 KB catches the ~43-47 KB junk seen in
# the LSDV batch while PASSING legitimately small viral/amplicon reads — a
# 1 MB floor (the original) wrongly flagged SARS-CoV-2 amplicon (~200 KB).
_T46_JUNK_FASTQ_BYTES = 50 * 1024  # 50 KB (fallback; see step1_min_fastq_bytes)


def _safe_stat_size(p: Path) -> Optional[int]:
    """File size, or None if the file is missing / a broken symlink. Step1
    sample dirs are symlinks into download/, which can point at sources that
    were since removed (e.g. SRA fastqs deleted, or a Kraken parsed-read source
    cleaned up) — stat() then raises and must not 500 the whole run."""
    try:
        return p.stat().st_size
    except OSError:
        return None


def _align_glob(sample_dir: Path, pattern: str) -> List[Path]:
    """Match a Step-1 output under alignment_<reference>/, falling back to the
    suffix-less alignment/ that pre-GUI runs (vSNP2-era pipelines, external
    Slurm batches) left behind — those samples must read as Complete and their
    VCFs must collect, instead of sitting at 'Not Started' forever. The GUI
    layout wins when both exist (a re-run supersedes the legacy dir). READ side
    only: the re-run cleanup keeps its strict alignment_* glob, so a dispatch
    can never delete a legacy alignment/."""
    return sorted(sample_dir.glob(f"alignment_*/{pattern}")) or sorted(
        sample_dir.glob(f"alignment/{pattern}")
    )


def _legacy_step1_complete(sample_dir: Path) -> bool:
    """vSNP2-era completion: no exit_code sentinel and different inner file
    names, but a zero-coverage VCF under plain alignment/ is the completion
    artifact everything downstream (VCF collection, step2) actually consumes."""
    return bool(
        next(sample_dir.glob("alignment/*_zc.vcf"), None)
        or next(sample_dir.glob("alignment/*_zc.vcf.gz"), None)
    )


def _step1_is_complete(sample_dir: Path) -> bool:
    """True if a sample already finished Step 1 successfully — same signal the
    status endpoint uses: a `.provenance/exit_code` of 0, or (legacy) the
    annotated VCF + de-duplicated BAM both present. Used to skip re-aligning
    samples that are already Complete when a batch is re-run after adding new
    samples (re-running would delete alignment_* and redo the slow alignment)."""
    ec = sample_dir / ".provenance" / "exit_code"
    if ec.exists():
        try:
            return ec.read_text(encoding="utf-8").strip() == "0"
        except OSError:
            return False
    vcf = next(iter(_align_glob(sample_dir, "*_filtered_hapall_annotated.vcf")), None)
    nodup = next(iter(_align_glob(sample_dir, "*_nodup.bam")), None)
    return bool(vcf and nodup) or _legacy_step1_complete(sample_dir)


def _step1_errored(sample_dir: Path) -> bool:
    """True if the sample has a terminal non-zero exit_code sentinel — it ran and
    failed. Used to skip auto-retrying it on a plain Run (Force re-run overrides)."""
    ec = sample_dir / ".provenance" / "exit_code"
    if not ec.exists():
        return False
    try:
        return ec.read_text(encoding="utf-8").strip() not in ("", "0")
    except OSError:
        return False


def _step1_input_issue(sample_dir: Path, min_bytes: int) -> Optional[str]:
    """Return a user-readable reason the sample's INPUT can't be dispatched
    (no fastq / broken link / too small), or None if the input is usable.
    Shared by the dispatcher (to skip) and the status endpoint (to explain a
    sample that's stuck at 'not started')."""
    r1_matches = sorted(sample_dir.glob("*_R1*.fastq.gz")) or sorted(sample_dir.glob("*_1.fastq.gz"))
    r2_matches = sorted(sample_dir.glob("*_R2*.fastq.gz")) or sorted(sample_dir.glob("*_2.fastq.gz"))
    all_fq = sorted(sample_dir.glob("*.fastq.gz"))
    if r1_matches:
        r1 = r1_matches[0]
        r2 = r2_matches[0] if r2_matches else None
    elif all_fq:
        r1 = all_fq[0]
        r2 = None
    else:
        return "no fastq files in the sample directory"
    r1_size = _safe_stat_size(r1)
    r2_size = _safe_stat_size(r2) if r2 is not None else None
    if r1_size is None or (r2 is not None and r2_size is None):
        missing = [m.name for m in (r1, r2) if m is not None and _safe_stat_size(m) is None]
        return f"fastq missing or broken link ({', '.join(missing)}); its source was removed — re-import the reads"
    if r1_size < min_bytes or (r2_size is not None and r2_size < min_bytes):
        layout = "paired" if r2 is not None else "single-end"
        detail = f"R1={r1_size}" + (f", R2={r2_size}" if r2_size is not None else "")
        return f"input too small ({detail} bytes, under the {min_bytes}-byte floor) — likely a truncated/failed download or SRA submission error"
    return None


def _step1_status_reason(status: str, sample_dir: Path, min_bytes: int) -> str:
    """Human-readable explanation for a non-complete sample's state, shown inline
    in the Samples list so 'Not started'/'Error' aren't opaque. Empty string when
    there's nothing to explain (e.g. a genuinely queued sample, or complete)."""
    if status == "error":
        return "Failed during Step 1 — open View log for the error."
    if status == "unknown":
        return "Interrupted before finishing (batch killed / OOM / restart) — re-run to retry."
    if status == "not_started":
        # If the input is unusable it will never run until fixed; say why.
        return _step1_input_issue(sample_dir, min_bytes) or ""
    return ""


def _step1_dispatch_plan(
    step1_dir: Path, min_bytes: int = _T46_JUNK_FASTQ_BYTES, force_rerun: bool = False
) -> tuple[List[str], List[Dict[str, Any]]]:
    """Decide which sample dirs to actually dispatch and which to skip.

    Returns (samples_to_run, skipped) where `skipped` is a list of
    {"sample", "reason", "size_bytes"} dicts surfaced back to the UI so the
    user knows what was excluded and why. Reasons are user-readable
    strings — they end up in an alert in the GUI.

    Skip rules (in order):
      0. Already completed (exit_code 0, or VCF+BAM present) → skip so a
         re-run after adding new samples doesn't re-align finished ones.
         Overridden by `force_rerun` (the GUI's "Force re-run" option).
      1. No fastq files at all in the sample dir → skip.
      2. R1 (or the lone single-end read) is a broken symlink / missing —
         the download source was removed → skip so it doesn't crash the batch.
      3. Read(s) < min_bytes → suspiciously small. Default floor is 50 KB
         (config: step1_min_fastq_bytes), which catches the 43-47 KB SRA
         submission errors while passing legitimately small viral/amplicon
         reads (SARS-CoV-2 amplicon is ~200 KB).

    Both paired-end and single-end inputs are dispatched: vsnp3 accepts a lone
    ``-r1`` (its help: "A single read file can also be supplied to this option")
    and auto-selects nanopore alignment by read length, so single-end Illumina
    and long-read/ONT are first-class. The generated batch script picks the right
    ``-r1``/``-r2`` combination and sets ``--nanopore`` per sample by read length.
    """
    samples: List[str] = []
    skipped: List[Dict[str, Any]] = []
    for p in sorted(step1_dir.iterdir()):
        if not p.is_dir() or p.name.startswith(("_", ".")):
            continue
        if not force_rerun and _step1_is_complete(p):
            skipped.append({
                "sample": p.name,
                "reason": "already completed in a previous run (use Force re-run to re-align)",
                "size_bytes": 0,
            })
            continue
        if not force_rerun and _step1_errored(p):
            # Ran and failed before. Don't silently retry on every plain Run
            # (a sample that reliably fails — e.g. bcftools on some ONT data —
            # would re-hang the batch each time). Force re-run overrides.
            skipped.append({
                "sample": p.name,
                "reason": "errored in a previous run — use Force re-run to retry",
                "size_bytes": 0,
            })
            continue
        issue = _step1_input_issue(p, min_bytes)
        if issue:
            skipped.append({"sample": p.name, "reason": issue, "size_bytes": 0})
            continue
        samples.append(p.name)
    return samples, skipped


def wrap_cmd(cfg: Dict, command: str) -> str:
    # PYTHONWARNINGS suppresses two classes of cosmetic noise from each vsnp3
    # worker:
    #   - SyntaxWarning: belt-and-suspenders for any regex literals we haven't
    #     patched yet (the known sites in step1/step2 are fixed in
    #     deploy/vsnp3-patches/v3.16-kapurlab.patch).
    #   - DeprecationWarning from markupsafe's __version__ access. vsnp3 still
    #     uses the deprecated attribute; harmless but spams logs once per
    #     subprocess. Scoped to the markupsafe module so other deprecation
    #     warnings still surface.
    env_prefix = 'PYTHONWARNINGS="ignore::SyntaxWarning,ignore::DeprecationWarning:markupsafe"'
    path_parts: List[str] = []
    tool_bin = _tool_bin_dir(cfg)
    script_bin = _script_bin_dir(cfg)
    if tool_bin:
        path_parts.append(str(tool_bin))
    if script_bin:
        path_parts.append(str(script_bin))
    if path_parts:
        return f"{env_prefix} PATH=\"{':'.join(path_parts)}:$PATH\" {command}"
    return f"{env_prefix} {command}"


def _load_vcf_label_map(cfg: Dict[str, str], label_style: str) -> Dict[str, str]:
    mapping_csv = _find_vcf_refs_csv(cfg)
    if not mapping_csv:
        return {}
    import re as _re
    def _short_label(name: str) -> str:
        name = name.strip()
        if name.lower().startswith("lineage "):
            parts = name.split()
            if len(parts) > 1 and parts[1].isdigit():
                return f"L{parts[1]}"
        if name.lower().startswith("m. "):
            name = name[3:]
        m = _re.match(r"^(L\\d+)\\b", name, _re.IGNORECASE)
        if m:
            return m.group(1).upper()
        tokens = _re.findall(r"[A-Za-z0-9]+", name)
        if not tokens:
            return name or "REF"
        first = tokens[0]
        if first.lower().startswith("l") and first[1:].isdigit():
            return first.upper()
        return first[:1].upper() + first[1:]
    def _rich_label(name: str) -> str:
        name = name.strip()
        if name.lower().startswith("m. "):
            name = name[3:]
        tokens = _re.findall(r"[A-Za-z0-9]+", name)
        if not tokens:
            return name or "REF"
        return "_".join(tokens)
    out: Dict[str, str] = {}
    with mapping_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if len(row) < 2:
                continue
            label, ident = row[0].strip(), row[1].strip()
            if not label or not ident or "number" in label.lower():
                continue
            friendly = _rich_label(label) if label_style == "rich" else _short_label(label)
            out[ident] = friendly
    return out


# Step 2 run directories are timestamp-named (vsnp3 step2 run stamp), created
# directly under step2/ (e.g. step2/2026-06-05_13-11-21). Older projects nested
# them under step2/runs/<ts>; those are still recognized for reads.
_STEP2_RUN_RE = re.compile(r"^\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}$")

# Dirs under step2/ that are NOT analysis-group output dirs (the VCF database,
# the legacy runs/ wrapper, and the provenance store).
_STEP2_NON_GROUP_DIRS = ("vcf_database", "vcf_source", "runs", "_provenance")


def _step2_run_dirs(step2_dir: Path) -> Dict[str, Path]:
    """Map run_id -> run directory for a project's step2 area.

    Runs now live directly under ``step2/`` as timestamp-named dirs. Projects
    created before the change kept them under ``step2/runs/<ts>``; those are
    still picked up (the new location wins on a name clash)."""
    runs: Dict[str, Path] = {}
    if step2_dir.is_dir():
        for d in step2_dir.iterdir():
            if d.is_dir() and _STEP2_RUN_RE.match(d.name):
                runs[d.name] = d
    legacy = step2_dir / "runs"
    if legacy.is_dir():
        for d in legacy.iterdir():
            if d.is_dir() and d.name not in runs:
                runs[d.name] = d
    return runs


def _write_figtree_groups(step2_dir: Path, vcf_source_dir: Path, cfg: Dict[str, str], label_style: str) -> None:
    if not vcf_source_dir.exists():
        return
    import re as _re
    manifest_path = vcf_source_dir / ".vcf_source_manifest.csv"
    source_map: Dict[str, str] = {}
    if manifest_path.exists():
        with manifest_path.open("r", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                name = (row.get("filename") or "").strip()
                source_type = (row.get("source_type") or "").strip()
                if name:
                    source_map[name] = source_type
    label_map = _load_vcf_label_map(cfg, label_style)
    # Build group file(s)
    rows = []
    for vcf in sorted(vcf_source_dir.glob("*.vcf*")):
        taxon = vcf.name
        source_type = source_map.get(taxon, "reference")
        group = "sample" if source_type == "step1" else "reference"
        color = "#d1495b" if group == "sample" else "#2b6cb0"
        rows.append((taxon, group, color))
    if not rows:
        return
    group_path = step2_dir / "figtree_groups.tsv"
    with group_path.open("w", encoding="utf-8") as handle:
        handle.write("taxon\tgroup\tcolor\n")
        for taxon, group, color in rows:
            handle.write(f"{taxon}\t{group}\t{color}\n")
    # Labeled version
    if label_map:
        labeled_path = step2_dir / "figtree_groups_labeled.tsv"
        with labeled_path.open("w", encoding="utf-8") as handle:
            handle.write("taxon\tgroup\tcolor\n")
            for taxon, group, color in rows:
                labeled = taxon
                for ident, friendly in label_map.items():
                    labeled = _re.sub(rf"\\b{_re.escape(ident)}\\b", friendly, labeled)
                handle.write(f"{labeled}\t{group}\t{color}\n")


def _count_vcfs(folder: Path) -> int:
    """Count *.vcf and *.vcf.gz files at the top level of a DB folder."""
    try:
        return sum(
            1 for p in folder.iterdir()
            if p.is_file() and (p.name.endswith(".vcf") or p.name.endswith(".vcf.gz"))
        )
    except OSError:
        return 0


def _resolved_vcf_db_folders(cfg: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Discover VCF DBs as a 2-level tree under vcf_db_folders_root:
        <root>/<reference>/<db_name>/*.vcf

    Each entry carries:
        path           absolute folder path
        reference      parent dir name (matches vsnp3 reference key)
        name           leaf folder name (display name)
        sample_count   live count of *.vcf / *.vcf.gz at top level
        enabled        bool (user toggle; shared entries always True in V1)
        scope          "shared" | "user"

    Shared entries (auto-discovered) are emitted first; user entries (explicit
    config list) follow. Paths are deduplicated. User entries must declare
    their reference in the config item dict; legacy plain-string entries are
    surfaced with reference="" and the GUI can prompt the user to set one."""
    result: List[Dict[str, Any]] = []
    seen: set = set()

    def _emit(path: Path, reference: str, scope: str, enabled: bool):
        p = str(path.resolve())
        if p in seen:
            return
        result.append({
            "path": p,
            "reference": reference,
            "name": path.name,
            "sample_count": _count_vcfs(path),
            "enabled": enabled,
            "scope": scope,
        })
        seen.add(p)

    disabled_shared = set(cfg.get("disabled_vcf_db_paths", []) or [])
    root = cfg.get("vcf_db_folders_root", "")
    if root:
        root_path = Path(root)
        if root_path.is_dir():
            for ref_dir in sorted(root_path.iterdir()):
                if not ref_dir.is_dir():
                    continue
                for db_dir in sorted(ref_dir.iterdir()):
                    if not db_dir.is_dir():
                        continue
                    resolved = str(db_dir.resolve())
                    _emit(db_dir, reference=ref_dir.name, scope="shared",
                          enabled=resolved not in disabled_shared)

    for entry in cfg.get("vcf_db_folders", []) or []:
        if isinstance(entry, dict):
            raw = entry.get("path")
            enabled = entry.get("enabled", True)
            reference = entry.get("reference", "") or ""
        else:
            raw = str(entry)
            enabled = True
            reference = ""
        if not raw:
            continue
        p = Path(raw).expanduser()
        _emit(p, reference=reference, scope="user", enabled=bool(enabled))
    return result


def _find_vcf_refs_csv(cfg: Dict[str, str]) -> Optional[Path]:
    candidates: List[Path] = []
    for folder in _resolved_vcf_db_folders(cfg):
        path = Path(folder["path"])
        candidates.append(path / "VCF_refs.csv")
        candidates.append(path / "vcf_refs.csv")
        candidates.append(path.parent / "VCF_refs.csv")
        candidates.append(path.parent / "vcf_refs.csv")
    vsnp3_path = Path(cfg.get("vsnp3_path", ""))
    if vsnp3_path:
        candidates.append(vsnp3_path / "VCF_REFS" / "VCF_refs.csv")
        candidates.append(vsnp3_path / "VCF_REFS" / "vcf_refs.csv")
    projects_root = Path(cfg.get("projects_root", ""))
    if projects_root:
        parent = projects_root.parent
        candidates.append(parent / "VCF_REFS" / "VCF_refs.csv")
        candidates.append(parent / "VCF_REFS" / "vcf_refs.csv")
    for c in candidates:
        if c.exists():
            return c
    return None


def _build_tree_label_script(step2_dir: Path, cfg: Dict[str, str], label_style: str) -> Optional[Path]:
    mapping_csv = _find_vcf_refs_csv(cfg)
    if not mapping_csv:
        return None
    script_path = step2_dir / "_label_trees.py"
    script = """\
import csv
import io
import re
from pathlib import Path

mapping_csv = Path("__MAPPING_CSV__")
step2_dir = Path("__STEP2_DIR__")
label_style = "__LABEL_STYLE__"

def short_label(name: str) -> str:
    name = name.strip()
    if name.lower().startswith("lineage "):
        parts = name.split()
        if len(parts) > 1 and parts[1].isdigit():
            return "L{}".format(parts[1])
    if name.lower().startswith("m. "):
        name = name[3:]
    m = re.match(r"^(L\\d+)\\b", name, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    tokens = re.findall(r"[A-Za-z0-9]+", name)
    if not tokens:
        return name or "REF"
    first = tokens[0]
    if first.lower().startswith("l") and first[1:].isdigit():
        return first.upper()
    return first[:1].upper() + first[1:]

def rich_label(name: str) -> str:
    name = name.strip()
    if name.lower().startswith("m. "):
        name = name[3:]
    tokens = re.findall(r"[A-Za-z0-9]+", name)
    if not tokens:
        return name or "REF"
    return "_".join(tokens)

def load_mapping(path: Path):
    out = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if len(row) < 2:
                continue
            label, ident = row[0].strip(), row[1].strip()
            if not label or not ident or "number" in label.lower():
                continue
            if label_style == "rich":
                friendly = rich_label(label)
            else:
                friendly = short_label(label)
            out[ident] = friendly + "_" + ident
    return out

mapping = load_mapping(mapping_csv)
if not mapping:
    raise SystemExit(0)

def load_color_map(step2_dir: Path):
    # Map accession -> color based on source type (sample vs reference)
    db_dir = step2_dir / "vcf_database"
    if not db_dir.exists():
        db_dir = step2_dir / "vcf_source"
    manifest = db_dir / ".vcf_source_manifest.csv"
    out = {}
    if not manifest.exists():
        return out
    with manifest.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = (row.get("filename") or "").strip()
            source_type = (row.get("source_type") or "").strip()
            m = re.search(r"(SRR\\d+|ERR\\d+|DRR\\d+|SRX\\d+|ERX\\d+|DRX\\d+)", name)
            if not m:
                continue
            ident = m.group(1)
            out[ident] = "#d1495b" if source_type == "step1" else "#2b6cb0"
    return out

color_map = load_color_map(step2_dir)

def annotate_newick(text: str, color_map: dict) -> str:
    # Add FigTree-style color annotations before branch length.
    def repl(match):
        label = match.group(2)
        m = re.search(r"(SRR\\d+|ERR\\d+|DRR\\d+|SRX\\d+|ERX\\d+|DRX\\d+)", label)
        if not m:
            return match.group(0)
        ident = m.group(1)
        color = color_map.get(ident)
        if not color:
            return match.group(0)
        return match.group(1) + label + "[&!color=" + color + "]:"
    return re.sub(r"([,(])([^:(),]+):", repl, text)

tree_files = list(step2_dir.rglob("*.tre")) + list(step2_dir.rglob("*.tree")) + list(step2_dir.rglob("*.nwk"))
for path in tree_files:
    if "_labeled" in path.stem:
        continue
    text = path.read_text(encoding="utf-8", errors="ignore")
    for ident, label in mapping.items():
        text = re.sub(rf"\\b" + re.escape(ident) + r"(_zc\\.vcf(?:\\.gz)?)\\b", label + r"\\1", text)
        text = re.sub(rf"\\b" + re.escape(ident) + r"\\b", label, text)
    labeled = path.with_name(path.stem + "_labeled" + path.suffix)
    labeled.write_text(text, encoding="utf-8")
    # Write a NEXUS file with color annotations for FigTree
    annotated = annotate_newick(text, color_map)
    nexus_path = labeled.with_suffix(".nexus")
    nexus_path.write_text(
        \"#NEXUS\\nbegin trees;\\n  tree tree1 = \" + annotated.strip() + \"\\nend;\\n\",
        encoding=\"utf-8\"
    )
"""
    script = (
        script.replace("__MAPPING_CSV__", str(mapping_csv))
        .replace("__STEP2_DIR__", str(step2_dir))
        .replace("__LABEL_STYLE__", label_style)
    )
    script_path.write_text(script, encoding="utf-8")
    return script_path


def conda_python_cmd(cfg: Dict, code: str, args: Optional[List[str]] = None) -> List[str]:
    args = args or []
    return [sys.executable, "-c", code, *args]


class ConfigUpdate(BaseModel):
    vsnp3_path: Optional[str] = None
    projects_root: Optional[str] = None
    shared_projects_root: Optional[str] = None
    saved_project_roots: Optional[List[str]] = None
    bcftools_path: Optional[str] = None
    step1_max_parallel: Optional[int] = None
    sra: Optional[Dict] = None


class ProjectCreate(BaseModel):
    name: str = Field(min_length=1)
    scope: Optional[str] = None  # "personal" (default) or "shared"
    reference: Optional[str] = None


class SraRequest(BaseModel):
    accessions: List[str]
    folder: Optional[str] = None


# FASTQ names the server-side picker offers and link-local accepts. Step 1
# pairs on _1/_2 (or _R1/_R2) of gzipped reads, so uncompressed .fastq is
# deliberately not in this list.
_FASTQ_SUFFIXES = (".fastq.gz", ".fq.gz")


class LinkLocalRequest(BaseModel):
    path: str = ""
    # Multi-select from the server-side file picker. Each entry is an absolute
    # path to a FASTQ (or a directory, treated the same as `path`). Kept
    # separate from `path` so existing single-path callers are unchanged.
    paths: List[str] = []


class ImportVcfRequest(BaseModel):
    source_paths: List[str] = []
    include_step1: bool = False
    reference: Optional[str] = None
    action: Optional[str] = "copy"  # copy | link
    on_conflict: Optional[str] = "skip"  # skip | overwrite | rename
    allow_mismatch: bool = False
    prefix_duplicates: bool = False
    dedupe: bool = False
    allow_fuzzy_match: bool = True
    confirm_large: bool = False


class Step1Request(BaseModel):
    reference: Optional[str] = None
    debug: bool = False
    assemble_unmap: bool = False
    nanopore: bool = False
    force_rerun: bool = False   # re-align even samples already marked Complete


class Step2Request(BaseModel):
    reference: Optional[str] = None
    no_filters: bool = False
    qual_threshold: Optional[int] = 150
    n_threshold: Optional[int] = 50
    mq_threshold: Optional[int] = 56
    all_vcf: bool = True
    label_style: Optional[str] = "short"
    find_new_filters: bool = False
    hash_groups: bool = False
    show_groups: bool = False
    html_tree: bool = False
    dp: bool = False
    density_threshold: Optional[int] = None
    density_window: Optional[int] = None
    bootstrap: int = 0
    # Authoritative exclusion sets the UI sends so the run never silently
    # ignores exclusions a debounced save hadn't flushed. Kept as two tiers so
    # the panel exemption can apply to Step 1 exclusions only (build-list
    # exclusions are explicit per-run choices and are never exempted).
    step1_exclude: Optional[List[str]] = None   # tier B
    build_exclude: Optional[List[str]] = None   # tier C
    # Deprecated merged field (older frontends). Treated as build-list (tier C)
    # so it is never panel-exempted — safe default.
    exclude: Optional[List[str]] = None


class PosthocRunRequest(BaseModel):
    group: str
    tool: str
    scope: Optional[str] = "all"
    # Which step2 run the group belongs to. Groups live under a timestamped run
    # dir (step2/<run_id>/<group>); without this the group can't be located and
    # the run 404s on the current layout. None → resolver picks current/latest.
    run_id: Optional[str] = None


class PosthocScanRequest(BaseModel):
    folders: List[str]


class PosthocResolveRequest(BaseModel):
    samples: List[str]
    roots: Optional[List[str]] = None


class VcfEditRequest(BaseModel):
    sample: str
    locus: str
    new_alt: str
    note: Optional[str] = ""
    reason: Optional[str] = ""
    user: Optional[str] = ""


class VcfLookupRequest(BaseModel):
    sample: str
    locus: str


@app.get("/api/health")
def health():
    return {"status": "ok"}


def _path_is_executable(path_str: str) -> bool:
    p = Path(path_str)
    return p.is_file() and os.access(p, os.X_OK)


@app.get("/api/config")
def get_config():
    cfg = load_config()
    root_dir = Path(__file__).resolve().parent.parent.parent
    cfg["gui_root"] = str(root_dir)
    # Deployed checkout's version (git describe) — what the dashboard shows.
    cfg["app_version"] = APP_VERSION
    shared_root = cfg.get("shared_projects_root", "").strip()
    cfg["_validation"] = {
        "vsnp3_path": Path(cfg.get("vsnp3_path", "")).is_dir() if cfg.get("vsnp3_path", "").strip() else False,
        "projects_root": Path(cfg.get("projects_root", "")).is_dir() if cfg.get("projects_root", "").strip() else False,
        "shared_projects_root": Path(shared_root).is_dir() if shared_root else None,
        "bcftools_path": _path_is_executable(cfg.get("bcftools_path", "")) if cfg.get("bcftools_path", "").strip() else None,
    }
    return cfg


@app.post("/api/config")
def update_config(update: ConfigUpdate):
    cfg = load_config()
    if update.vsnp3_path is not None:
        cfg["vsnp3_path"] = update.vsnp3_path
    if update.projects_root is not None:
        cfg["projects_root"] = update.projects_root
    if update.shared_projects_root is not None:
        cfg["shared_projects_root"] = update.shared_projects_root
    if update.saved_project_roots is not None:
        # Curated, user-managed bookmarks (add/remove in Settings). Store a
        # de-duplicated, order-preserving list of non-empty paths.
        seen, cleaned = set(), []
        for r in update.saved_project_roots:
            r = (r or "").strip()
            if r and r not in seen:
                seen.add(r)
                cleaned.append(r)
        cfg["saved_project_roots"] = cleaned
    if update.bcftools_path is not None:
        cfg["bcftools_path"] = update.bcftools_path
    if update.step1_max_parallel is not None:
        cfg["step1_max_parallel"] = update.step1_max_parallel
    if update.sra is not None:
        cfg["sra"].update(update.sra)
    save_config(cfg)
    return cfg


def _disk_usage_for(p: Path) -> Optional[Dict[str, int]]:
    """Free/total bytes for the filesystem holding `p`, or None if unavailable.

    Shown in the folder picker so a project root is not chosen on a volume that
    cannot hold the run — a Step 1 batch of a few thousand samples is easily
    several TB of intermediates.
    """
    try:
        usage = shutil.disk_usage(p)
    except OSError:
        return None
    return {"total": usage.total, "free": usage.free, "used": usage.used}


@app.get("/api/browse-dirs")
def browse_dirs(path: str = "", include_files: bool = False, exts: str = ""):
    """List sub-directories of `path` for the project-root folder picker.

    Runs as the OOD session user, so the OS filesystem permissions are the only
    limit on what can be browsed (no artificial base restriction). Defaults to
    the user's home when no path is given. Returns the resolved path, its parent
    (null at the filesystem root), the immediate sub-directories, and the free /
    total bytes on the filesystem holding it.

    With ?include_files=1 the response also carries the immediate regular files,
    optionally narrowed to a comma-separated ?exts= suffix list (case
    insensitive, e.g. ".fastq.gz,.fq.gz"). This backs the server-side file
    picker for "Add local FASTQ", so files that already live on the server are
    selected in place rather than round-tripped through a browser upload.
    """
    try:
        p = (Path(path).expanduser() if path.strip() else Path.home()).resolve()
    except (OSError, RuntimeError):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not p.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a directory: {p}")
    suffixes = tuple(
        s.strip().lower() for s in exts.split(",") if s.strip()
    )
    entries: List[Dict[str, str]] = []
    files: List[Dict] = []
    try:
        for child in sorted(p.iterdir(), key=lambda c: c.name.lower()):
            if child.name.startswith("."):
                continue
            try:
                if child.is_dir():
                    entries.append({"name": child.name, "path": str(child)})
                elif include_files and child.is_file():
                    if suffixes and not child.name.lower().endswith(suffixes):
                        continue
                    try:
                        size = child.stat().st_size
                    except OSError:
                        size = 0
                    files.append(
                        {"name": child.name, "path": str(child), "size": size}
                    )
            except OSError:
                continue
    except PermissionError:
        raise HTTPException(status_code=403, detail=f"Permission denied: {p}")
    parent = str(p.parent) if p.parent != p else None
    return {
        "path": str(p),
        "parent": parent,
        "entries": entries,
        "files": files,
        "disk": _disk_usage_for(p),
    }


@app.get("/api/disk-usage")
def disk_usage(path: str = ""):
    """Free/total bytes for the filesystem holding `path` (blank -> home).

    Used by the Settings panel to show headroom for the current Projects root
    without re-listing its contents.
    """
    try:
        p = (Path(path).expanduser() if path.strip() else Path.home()).resolve()
    except (OSError, RuntimeError):
        raise HTTPException(status_code=400, detail="Invalid path")
    return {"path": str(p), "disk": _disk_usage_for(p)}


class VcfDbFolderAction(BaseModel):
    action: str  # "add", "remove", "toggle"
    path: Optional[str] = None
    index: Optional[int] = None
    reference: Optional[str] = None  # required for "add" — which reference this DB targets


@app.get("/api/vcf-db-folders")
def get_vcf_db_folders():
    cfg = load_config()
    return _resolved_vcf_db_folders(cfg)


@app.post("/api/vcf-db-folders")
def update_vcf_db_folders(payload: VcfDbFolderAction):
    """Mutates the user's explicit vcf_db_folders list. Shared (auto-discovered)
    entries are read-only — managed via the filesystem at vcf_db_folders_root."""
    cfg = load_config()
    folders = cfg.get("vcf_db_folders", [])
    target_path = (
        str(Path(payload.path).expanduser().resolve())
        if payload.path
        else None
    )
    if payload.action == "add":
        if not target_path:
            raise HTTPException(status_code=400, detail="path is required for add")
        ref = (payload.reference or "").strip()
        if not ref:
            raise HTTPException(status_code=400, detail="reference is required for add")
        if not any(
            (f.get("path") if isinstance(f, dict) else str(f)) == target_path
            for f in folders
        ):
            folders.append({"path": target_path, "enabled": True, "reference": ref})
    elif payload.action == "remove":
        if target_path:
            folders = [
                f for f in folders
                if (f.get("path") if isinstance(f, dict) else str(f)) != target_path
            ]
        elif payload.index is not None and 0 <= payload.index < len(folders):
            folders.pop(payload.index)
    elif payload.action == "toggle":
        # First check if path matches a user-explicit entry — flip its enabled.
        # Otherwise treat as a shared path: add/remove from disabled_vcf_db_paths.
        idx = None
        if target_path:
            for i, f in enumerate(folders):
                fp = f.get("path") if isinstance(f, dict) else str(f)
                if fp == target_path:
                    idx = i
                    break
        elif payload.index is not None and 0 <= payload.index < len(folders):
            idx = payload.index
        if idx is not None and isinstance(folders[idx], dict):
            folders[idx]["enabled"] = not folders[idx].get("enabled", True)
        elif target_path:
            disabled = list(cfg.get("disabled_vcf_db_paths", []) or [])
            if target_path in disabled:
                disabled = [p for p in disabled if p != target_path]
            else:
                disabled.append(target_path)
            cfg["disabled_vcf_db_paths"] = disabled
    else:
        raise HTTPException(status_code=400, detail="action must be add, remove, or toggle")
    cfg["vcf_db_folders"] = folders
    save_config(cfg)
    return _resolved_vcf_db_folders(cfg)


class RefPathRequest(BaseModel):
    path: str


class RefDownloadRequest(BaseModel):
    accession: str
    output_dir: str
    display_name: Optional[str] = None


@app.get("/api/references")
def references():
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    return list_references(vsnp3_path)


@app.get("/api/references/paths")
def ref_paths():
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    return {"paths": get_reference_paths(vsnp3_path)}


@app.post("/api/references/paths")
def ref_path_add(payload: RefPathRequest):
    p = Path(payload.path).expanduser().resolve()
    if not p.is_dir():
        raise HTTPException(status_code=400, detail=f"Directory not found: {p}")
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    updated = add_reference_path(vsnp3_path, str(p))
    refs = list_references(vsnp3_path)
    return {"paths": updated, "references": refs}


@app.delete("/api/references/paths")
def ref_path_remove(payload: RefPathRequest):
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    updated = remove_reference_path(vsnp3_path, payload.path)
    refs = list_references(vsnp3_path)
    return {"paths": updated, "references": refs}


_REF_DISPLAY_NAME_RE = re.compile(r"^[A-Za-z0-9._-]+$")


@app.post("/api/references/download")
def ref_download(payload: RefDownloadRequest):
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    accession = payload.accession.strip()
    if not accession:
        raise HTTPException(status_code=400, detail="Accession is required")
    output_dir = Path(payload.output_dir).expanduser().resolve()
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        raise HTTPException(status_code=400, detail=f"Cannot create output directory: {e}")

    # Optional display name controls the subdir name (and therefore the entry
    # shown in the Reference dropdown). File stems inside still use the
    # accession — matches the lab's existing convention where dir name and
    # file prefixes can differ (e.g. mtbc0_v1.1/ dir contains H37_*.xlsx).
    display_name = (payload.display_name or "").strip()
    if display_name:
        if len(display_name) > 100:
            raise HTTPException(status_code=400, detail="Display name too long (max 100 chars)")
        if display_name.startswith("."):
            raise HTTPException(status_code=400, detail="Display name cannot start with '.'")
        if not _REF_DISPLAY_NAME_RE.match(display_name):
            raise HTTPException(
                status_code=400,
                detail="Display name may only contain letters, digits, underscore, hyphen, and period",
            )
        subdir_name = display_name
    else:
        subdir_name = accession
    acc_dir = output_dir / subdir_name
    if acc_dir.exists() and any(acc_dir.iterdir()):
        raise HTTPException(
            status_code=409,
            detail=f"Reference directory already exists and is non-empty: {acc_dir}",
        )
    acc_dir.mkdir(parents=True, exist_ok=True)
    # Copy template files from dependencies, renaming "template" to accession
    template_dir = vsnp3_path / "dependencies"
    for tpl in ["template_define_filter.xlsx", "template_remove_from_analysis.xlsx"]:
        src = template_dir / tpl
        dest_name = tpl.replace("template", accession)
        dest = acc_dir / dest_name
        if src.is_file() and not dest.exists():
            shutil.copy2(str(src), str(dest))
    # Build script: download fasta/gbk/gff
    lines = [
        "#!/bin/bash",
        "set -euo pipefail",
        f"cd \"{acc_dir}\"",
        f"vsnp3_download_fasta_gbk_gff_by_acc.py -a {accession} -fbg",
    ]
    script_content = "\n".join(lines) + "\n"
    script_path = acc_dir / "download_ref.sh"
    script_path.write_text(script_content, encoding="utf-8")
    script_path.chmod(0o755)
    # Add parent dir to reference paths
    add_reference_path(vsnp3_path, str(output_dir))
    job_id = job_manager.start_job(
        name="ref_download",
        command=wrap_cmd(cfg, f"bash {shlex.quote(str(script_path))}"),
        cwd=acc_dir,
        env=build_env(cfg)
    )
    return {"job_id": job_id, "accession": accession, "target_dir": str(acc_dir)}


@app.get("/api/references/{ref_name}/files")
def ref_files(ref_name: str):
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    refs = list_references(vsnp3_path)
    ref = next((r for r in refs if r["name"] == ref_name), None)
    if not ref:
        raise HTTPException(status_code=404, detail=f"Reference not found: {ref_name}")
    ref_dir = Path(ref["path"])
    define_filter = list(ref_dir.glob("*define_filter*"))
    remove_from = list(ref_dir.glob("*remove_from_analysis*"))
    files = []
    for f in define_filter:
        files.append({"name": f.name, "path": str(f), "exists": f.exists(), "type": "define_filter"})
    for f in remove_from:
        files.append({"name": f.name, "path": str(f), "exists": f.exists(), "type": "remove_from_analysis"})
    meta_files = [f for f in ref_dir.glob("*meta*xlsx") if not f.name.startswith("~$")]
    for f in meta_files:
        files.append({"name": f.name, "path": str(f), "exists": f.exists(), "type": "metadata"})
    return {"ref_name": ref_name, "ref_path": str(ref_dir), "files": files}


def _read_metadata_xlsx(cfg: Dict, meta_path: Path) -> List[Dict[str, str]]:
    code = (
        "import pandas as pd, json, sys; "
        "df = pd.read_excel(sys.argv[1], header=None, usecols=[0,1], names=['original','display_name']); "
        "df = df.dropna(subset=['original']); "
        "df['original'] = df['original'].astype(str); "
        "df['display_name'] = df['display_name'].astype(str); "
        "print(json.dumps(df.to_dict(orient='records')))"
    )
    result = subprocess.run(conda_python_cmd(cfg, code, [str(meta_path)]), text=True, capture_output=True)
    if result.returncode != 0:
        raise HTTPException(status_code=500, detail=f"Failed to read metadata: {result.stderr.strip()}")
    try:
        return json.loads(result.stdout.strip())
    except json.JSONDecodeError:
        return []


@app.get("/api/references/{ref_name}/metadata")
def ref_get_metadata(ref_name: str):
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    refs = list_references(vsnp3_path)
    ref = next((r for r in refs if r["name"] == ref_name), None)
    if not ref:
        raise HTTPException(status_code=404, detail=f"Reference not found: {ref_name}")
    ref_dir = Path(ref["path"])
    meta_files = [f for f in ref_dir.glob("*meta*xlsx") if not f.name.startswith("~$")]
    if not meta_files:
        return {"rows": [], "filename": None, "exists": False}
    meta_path = meta_files[0]
    rows = _read_metadata_xlsx(cfg, meta_path)
    return {"rows": rows, "filename": meta_path.name, "exists": True}


class MetadataAddRequest(BaseModel):
    rows: List[Dict[str, str]]


@app.post("/api/references/{ref_name}/metadata/add-rows")
def ref_add_metadata_rows(ref_name: str, payload: MetadataAddRequest):
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    refs = list_references(vsnp3_path)
    ref = next((r for r in refs if r["name"] == ref_name), None)
    if not ref:
        raise HTTPException(status_code=404, detail=f"Reference not found: {ref_name}")
    ref_dir = Path(ref["path"])
    if not payload.rows:
        raise HTTPException(status_code=400, detail="No rows provided")
    for row in payload.rows:
        if not str(row.get("original", "")).strip():
            raise HTTPException(status_code=400, detail="Each row requires a non-empty 'original' field")
        if not str(row.get("display_name", "")).strip():
            raise HTTPException(status_code=400, detail="Each row requires a non-empty 'display_name' field")

    meta_files = [f for f in ref_dir.glob("*meta*xlsx") if not f.name.startswith("~$")]
    meta_path = meta_files[0] if meta_files else ref_dir / f"{ref_name}_metadata.xlsx"

    existing_rows: List[Dict[str, str]] = []
    if meta_path.exists():
        existing_rows = _read_metadata_xlsx(cfg, meta_path)

    orig_to_idx = {r["original"]: i for i, r in enumerate(existing_rows)}
    added, updated = 0, 0
    for row in payload.rows:
        orig = str(row["original"]).strip()
        disp = str(row["display_name"]).strip()
        if orig in orig_to_idx:
            existing_rows[orig_to_idx[orig]]["display_name"] = disp
            updated += 1
        else:
            existing_rows.append({"original": orig, "display_name": disp})
            orig_to_idx[orig] = len(existing_rows) - 1
            added += 1

    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, dir=str(ref_dir)) as tf:
        json.dump(existing_rows, tf)
        tmp_json = tf.name
    try:
        code = (
            "import pandas as pd, json, sys; "
            "rows = json.load(open(sys.argv[2])); "
            "df = pd.DataFrame([[r['original'], r['display_name']] for r in rows]); "
            "df.to_excel(sys.argv[1], index=False, header=False)"
        )
        result = subprocess.run(
            conda_python_cmd(cfg, code, [str(meta_path), tmp_json]),
            text=True, capture_output=True
        )
        if result.returncode != 0:
            raise HTTPException(status_code=500, detail=f"Failed to write metadata: {result.stderr.strip()}")
    finally:
        Path(tmp_json).unlink(missing_ok=True)

    return {"filename": meta_path.name, "rows_total": len(existing_rows), "added": added, "updated": updated}


def _backup_ref_file(ref_dir: Path, target: Path) -> tuple[str, str]:
    """Copy an existing reference file aside under .history/ before mutating it.

    Returns (old_sha256, archived_path) — both empty strings if the target
    doesn't exist yet. Mirrors the archiving the Replace (upload-file) flow
    does, so add-group / add-sample edits are equally recoverable.
    """
    if not target.exists():
        return "", ""
    old_sha = _sha256_of_path(target)
    history_dir = ref_dir / ".history"
    history_dir.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())
    archived = history_dir / f"{ts}_{old_sha[:8]}_{target.name}"
    shutil.copy2(target, archived)
    return old_sha, str(archived)


def _current_os_user() -> str:
    try:
        import pwd
        return pwd.getpwuid(os.getuid()).pw_name
    except Exception:
        return os.environ.get("USER", "")


def _ref_dir_or_404(ref_name: str) -> Path:
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    ref = next((r for r in list_references(vsnp3_path) if r["name"] == ref_name), None)
    if not ref:
        raise HTTPException(status_code=404, detail=f"Reference not found: {ref_name}")
    return Path(ref["path"]).resolve()


class DefineFilterAddGroupRequest(BaseModel):
    group: str
    positions: List[str]
    rationale: str


# openpyxl mutator for the defining-SNP filter: appends one column per
# position (header = chrom:pos, row 2 = group name), preserving the rest of
# the sheet's formatting. Detects the contig prefix from existing column
# headers so the user may type a bare position. Rejects malformed or
# duplicate positions. Emits a JSON summary on stdout, "ERR:<code>" on stderr.
_DEFINE_FILTER_ADD_CODE = r"""
import openpyxl, json, sys, re
target = sys.argv[1]
payload = json.load(open(sys.argv[2]))
group = str(payload['group']).strip()
positions = payload['positions']
wb = openpyxl.load_workbook(target)
ws = wb.worksheets[0]
chrom = None
for c in range(2, ws.max_column + 1):
    v = ws.cell(row=1, column=c).value
    if v is not None and ':' in str(v):
        chrom = str(v).split(':', 1)[0].lstrip('#').strip()
        break
existing = set()
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=c).value
    if v is not None and str(v).strip():
        existing.add(str(v).strip().lstrip('#'))
norm = []
for p in positions:
    p = str(p).strip()
    if not p:
        continue
    if ':' not in p:
        if not chrom:
            sys.stderr.write('ERR:no_chrom'); sys.exit(2)
        p = '{}:{}'.format(chrom, p)
    if not re.fullmatch(r'\S+:\d+', p):
        sys.stderr.write('ERR:bad_position:' + p); sys.exit(3)
    if p in existing or p in norm:
        sys.stderr.write('ERR:dup_position:' + p); sys.exit(4)
    norm.append(p)
if not norm:
    sys.stderr.write('ERR:no_positions'); sys.exit(5)
col = ws.max_column
for p in norm:
    col += 1
    ws.cell(row=1, column=col, value=p)
    ws.cell(row=2, column=col, value=group)
wb.save(target)
print(json.dumps({'chrom': chrom, 'positions': norm, 'group': group}))
"""


@app.post("/api/references/{ref_name}/define-filter/add-group")
def ref_define_filter_add_group(ref_name: str, payload: DefineFilterAddGroupRequest):
    """Add a defining-SNP group to a reference's *_define_filter.xlsx.

    A group is one or more absolute positions (chrom:pos) that map to a
    single group name. Each position becomes its own column (header = the
    position, row 2 = the group name) — the format vSNP3 reads. This is a
    permanent edit to a shared reference file, so it follows the same
    backup + audit-log flow as the Replace action.
    """
    cfg = load_config()
    ref_dir = _ref_dir_or_404(ref_name)
    group = (payload.group or "").strip()
    if not group:
        raise HTTPException(status_code=400, detail="Group name is required")
    positions = [str(p).strip() for p in (payload.positions or []) if str(p).strip()]
    if not positions:
        raise HTTPException(status_code=400, detail="At least one position is required")
    if not (payload.rationale or "").strip():
        raise HTTPException(status_code=400, detail="A rationale is required")

    define_files = [
        f for f in ref_dir.glob("*define_filter*.xlsx") if not f.name.startswith("~$")
    ]
    if not define_files:
        raise HTTPException(
            status_code=404,
            detail="No define_filter file found for this reference. Create one from template first.",
        )
    target = define_files[0]

    old_sha, archived = _backup_ref_file(ref_dir, target)
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, dir=str(ref_dir)) as tf:
        json.dump({"group": group, "positions": positions}, tf)
        tmp_json = tf.name
    try:
        result = subprocess.run(
            conda_python_cmd(cfg, _DEFINE_FILTER_ADD_CODE, [str(target), tmp_json]),
            text=True, capture_output=True,
        )
    finally:
        Path(tmp_json).unlink(missing_ok=True)

    if result.returncode != 0:
        err = (result.stderr or "").strip()
        if err.startswith("ERR:no_chrom"):
            detail = "Could not detect a contig prefix — enter positions as chrom:position (e.g. NC_000962:12345)."
        elif err.startswith("ERR:bad_position:"):
            detail = f"Invalid position '{err.split(':',2)[2]}' — use chrom:position with a numeric position."
        elif err.startswith("ERR:dup_position:"):
            detail = f"Position '{err.split(':',2)[2]}' is already in the define_filter file."
        elif err.startswith("ERR:no_positions"):
            detail = "No valid positions provided."
        else:
            detail = f"Failed to add group: {err}"
        raise HTTPException(status_code=400, detail=detail)
    try:
        summary = json.loads(result.stdout.strip())
    except json.JSONDecodeError:
        summary = {"group": group, "positions": positions}

    new_sha = _sha256_of_path(target)
    record = {
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "action": "define_filter_add_group",
        "reference": ref_name,
        "filename": target.name,
        "user": _current_os_user(),
        "rationale": payload.rationale.strip(),
        "group": group,
        "positions": summary.get("positions", positions),
        "old_sha256": old_sha,
        "new_sha256": new_sha,
        "archived_old": archived,
        "target": str(target),
    }
    audit_path = _t39_audit_append(record, ref_dir / ".history")
    return {
        "ok": True,
        "filename": target.name,
        "group": group,
        "positions": summary.get("positions", positions),
        "added": len(summary.get("positions", positions)),
        "archived_old": archived,
        "audit_log": audit_path,
    }


class RemoveSampleAddRequest(BaseModel):
    samples: List[str]
    rationale: str


# openpyxl mutator for remove_from_analysis: appends sample names to the
# single (header-less) column A, skipping any already present. Emits a JSON
# summary {"added": [...], "skipped": [...]} on stdout.
_REMOVE_SAMPLE_ADD_CODE = r"""
import openpyxl, json, sys
target = sys.argv[1]
payload = json.load(open(sys.argv[2]))
samples = payload['samples']
wb = openpyxl.load_workbook(target)
ws = wb.worksheets[0]
existing = set()
last = 0
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is not None and str(v).strip():
        existing.add(str(v).strip())
        last = r
added, skipped = [], []
row = last
for s in samples:
    s = str(s).strip()
    if not s:
        continue
    if s in existing:
        skipped.append(s)
        continue
    row += 1
    ws.cell(row=row, column=1, value=s)
    existing.add(s)
    added.append(s)
wb.save(target)
print(json.dumps({'added': added, 'skipped': skipped}))
"""


@app.post("/api/references/{ref_name}/remove-from-analysis/add-sample")
def ref_remove_add_sample(ref_name: str, payload: RemoveSampleAddRequest):
    """Add sample name(s) to a reference's *_remove_from_analysis.xlsx.

    vSNP3 reads column 1 (no header) as sample names to drop from analysis.
    Permanent edit to a shared reference file → backup + audit-log flow.
    """
    cfg = load_config()
    ref_dir = _ref_dir_or_404(ref_name)
    samples = [str(s).strip() for s in (payload.samples or []) if str(s).strip()]
    if not samples:
        raise HTTPException(status_code=400, detail="At least one sample name is required")
    if not (payload.rationale or "").strip():
        raise HTTPException(status_code=400, detail="A rationale is required")

    remove_files = [
        f for f in ref_dir.glob("*remove_from_analysis*.xlsx") if not f.name.startswith("~$")
    ]
    if not remove_files:
        raise HTTPException(
            status_code=404,
            detail="No remove_from_analysis file found for this reference. Create one from template first.",
        )
    target = remove_files[0]

    old_sha, archived = _backup_ref_file(ref_dir, target)
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, dir=str(ref_dir)) as tf:
        json.dump({"samples": samples}, tf)
        tmp_json = tf.name
    try:
        result = subprocess.run(
            conda_python_cmd(cfg, _REMOVE_SAMPLE_ADD_CODE, [str(target), tmp_json]),
            text=True, capture_output=True,
        )
    finally:
        Path(tmp_json).unlink(missing_ok=True)

    if result.returncode != 0:
        raise HTTPException(status_code=400, detail=f"Failed to add sample: {(result.stderr or '').strip()}")
    try:
        summary = json.loads(result.stdout.strip())
    except json.JSONDecodeError:
        summary = {"added": samples, "skipped": []}

    new_sha = _sha256_of_path(target)
    record = {
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "action": "remove_from_analysis_add_sample",
        "reference": ref_name,
        "filename": target.name,
        "user": _current_os_user(),
        "rationale": payload.rationale.strip(),
        "added": summary.get("added", samples),
        "skipped": summary.get("skipped", []),
        "old_sha256": old_sha,
        "new_sha256": new_sha,
        "archived_old": archived,
        "target": str(target),
    }
    audit_path = _t39_audit_append(record, ref_dir / ".history")
    return {
        "ok": True,
        "filename": target.name,
        "added": summary.get("added", samples),
        "skipped": summary.get("skipped", []),
        "archived_old": archived,
        "audit_log": audit_path,
    }


class RefCreateFileRequest(BaseModel):
    file_type: str  # "define_filter" or "remove_from_analysis"


@app.post("/api/references/{ref_name}/create-file")
def ref_create_file(ref_name: str, payload: RefCreateFileRequest):
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    refs = list_references(vsnp3_path)
    ref = next((r for r in refs if r["name"] == ref_name), None)
    if not ref:
        raise HTTPException(status_code=404, detail=f"Reference not found: {ref_name}")
    ref_dir = Path(ref["path"])
    template_dir = vsnp3_path / "dependencies"
    if payload.file_type == "define_filter":
        template_name = "template_define_filter.xlsx"
        dest_name = f"{ref_name}_define_filter.xlsx"
    elif payload.file_type == "remove_from_analysis":
        template_name = "template_remove_from_analysis.xlsx"
        dest_name = f"{ref_name}_remove_from_analysis.xlsx"
    else:
        raise HTTPException(status_code=400, detail="file_type must be define_filter or remove_from_analysis")
    dest = ref_dir / dest_name
    if dest.exists():
        raise HTTPException(status_code=400, detail=f"File already exists: {dest_name}")
    template = template_dir / template_name
    if template.exists():
        shutil.copy2(template, dest)
    else:
        # Create an empty xlsx if no template exists
        code = (
            "import pandas as pd, sys; "
            "pd.DataFrame().to_excel(sys.argv[1], index=False)"
        )
        cmd_list = conda_python_cmd(cfg, code, [str(dest)])
        result = subprocess.run(cmd_list, text=True, capture_output=True)
        if result.returncode != 0:
            raise HTTPException(status_code=500, detail=f"Failed to create file: {result.stderr.strip()}")
    return {"created": str(dest), "name": dest_name}


def _resolve_ref_file(ref_name: str, filename: str) -> Path:
    """Resolve a reference-dir-relative filename to an absolute path with
    a directory traversal guard. Raises HTTPException on validation
    failures. Used by the reference preview / download endpoints."""
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    refs = list_references(vsnp3_path)
    ref = next((r for r in refs if r["name"] == ref_name), None)
    if not ref:
        raise HTTPException(status_code=404, detail=f"Reference not found: {ref_name}")
    ref_dir = Path(ref["path"]).resolve()
    if "/" in filename or filename.startswith("."):
        raise HTTPException(status_code=400, detail="Invalid filename")
    target = (ref_dir / filename).resolve()
    if not str(target).startswith(str(ref_dir) + "/") and target != ref_dir:
        raise HTTPException(status_code=400, detail="Path not allowed")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    return target


@app.get("/api/references/{ref_name}/preview-xlsx", response_class=HTMLResponse)
def ref_preview_xlsx(ref_name: str, filename: str = Query(...), download: int = 0):
    """Render a reference-dir xlsx as a self-contained HTML page. With
    ?download=1, return the raw xlsx instead — used by the "Download xlsx"
    link inside the preview page (JS handler in xlsx_html appends
    download=1 to the current URL preserving the filename param)."""
    target = _resolve_ref_file(ref_name, filename)
    if target.suffix.lower() not in (".xlsx", ".xlsm"):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm supported")
    if download:
        return FileResponse(
            target,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=target.name,
        )
    from app import xlsx_html
    try:
        html_page = xlsx_html.xlsx_to_html(target)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"xlsx render failed: {type(e).__name__}: {e}")
    return HTMLResponse(content=html_page)


@app.get("/api/references/{ref_name}/download-file")
def ref_download_file(ref_name: str, filename: str = Query(...), inline: int = 0):
    """Serve a file from a reference directory. Default = attachment;
    ?inline=1 lets the browser render in-tab where it can. Used by the
    Reference Editor's Download button."""
    target = _resolve_ref_file(ref_name, filename)
    suffix = target.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif suffix == ".xls":
        media_type = "application/vnd.ms-excel"
    elif suffix == ".csv":
        media_type = "text/csv"
    else:
        media_type = "application/octet-stream"
    headers = {}
    if not inline:
        headers["Content-Disposition"] = f'attachment; filename="{target.name}"'
    return FileResponse(target, media_type=media_type, headers=headers)


# T-39: re-upload reference xlsx to replace in place. Closes the offline-edit
# loop (download → edit locally in Excel/Numbers/LibreOffice → re-upload).
# Intentionally narrow:
#   - Only the three known reference xlsx filenames are accepted (defining
#     filter, remove_from_analysis, and metadata). Other reference files
#     (.fasta, .gbk, .gff, best_reference.txt) are read-only via this endpoint.
#   - 10 MB hard cap; these files are typically < 100 KB.
#   - Old file moved aside to <ref>/.history/ with a timestamp prefix —
#     recoverable without `git`.
#   - Audit log appended to /srv/kapurlab/audit/reference-changes.jsonl
#     (best-effort; fall back to <ref>/.history/_audit.jsonl if the shared
#     audit dir isn't writable).
#   - No approval queue. T-17a will layer the proposal+admin-review flow on
#     top when shipped.
_T39_ALLOWED_REF_FILENAMES = re.compile(r"^[A-Za-z0-9._-]+_(define_filter|remove_from_analysis|metadata)\.xlsx$")
_T39_MAX_UPLOAD_BYTES = 10 * 1024 * 1024
_T39_SHARED_AUDIT_PATH = SITE_ROOT / "audit" / "reference-changes.jsonl"


def _sha256_of_path(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _t39_audit_append(record: dict, fallback_dir: Path) -> str:
    """Append one JSON line to the audit log. Returns the path actually used."""
    line = json.dumps(record, sort_keys=True) + "\n"
    try:
        _T39_SHARED_AUDIT_PATH.parent.mkdir(parents=True, exist_ok=True)
        with _T39_SHARED_AUDIT_PATH.open("a", encoding="utf-8") as fh:
            fh.write(line)
        return str(_T39_SHARED_AUDIT_PATH)
    except (OSError, PermissionError):
        fallback = fallback_dir / "_audit.jsonl"
        fallback.parent.mkdir(parents=True, exist_ok=True)
        with fallback.open("a", encoding="utf-8") as fh:
            fh.write(line)
        return str(fallback)


@app.post("/api/references/{ref_name}/upload-file")
async def ref_upload_file(
    ref_name: str,
    file: UploadFile = File(...),
    rationale: str = Query(..., min_length=1, max_length=4000),
):
    cfg = load_config()
    vsnp3_path = Path(cfg["vsnp3_path"])
    refs = list_references(vsnp3_path)
    ref = next((r for r in refs if r["name"] == ref_name), None)
    if not ref:
        raise HTTPException(status_code=404, detail=f"Reference not found: {ref_name}")
    ref_dir = Path(ref["path"]).resolve()

    # Filename validation: client-provided name only; we ignore any path
    # components and enforce the strict whitelist.
    upload_name = Path(file.filename or "").name
    if not upload_name or not _T39_ALLOWED_REF_FILENAMES.match(upload_name):
        raise HTTPException(
            status_code=400,
            detail="Only *_define_filter.xlsx, *_remove_from_analysis.xlsx, or *_metadata.xlsx may be replaced via this endpoint",
        )

    target = (ref_dir / upload_name).resolve()
    if not str(target).startswith(str(ref_dir) + "/"):
        raise HTTPException(status_code=400, detail="Path not allowed")

    # Spool to a temp file in the same dir (so the os.replace is atomic on
    # the same filesystem) while enforcing the size cap. Read in chunks so
    # we don't load 10 MB into memory.
    history_dir = ref_dir / ".history"
    history_dir.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())
    tmp_path = ref_dir / f".{upload_name}.{ts}.tmp"
    bytes_written = 0
    try:
        with tmp_path.open("wb") as out:
            while True:
                chunk = await file.read(1 << 20)
                if not chunk:
                    break
                bytes_written += len(chunk)
                if bytes_written > _T39_MAX_UPLOAD_BYTES:
                    raise HTTPException(
                        status_code=413,
                        detail=f"File too large (max {_T39_MAX_UPLOAD_BYTES // (1024 * 1024)} MB)",
                    )
                out.write(chunk)
        if bytes_written == 0:
            raise HTTPException(status_code=400, detail="Empty upload")
    except HTTPException:
        try: tmp_path.unlink()
        except OSError: pass
        raise

    # Minimal sanity check that it's actually an xlsx (zip magic bytes).
    with tmp_path.open("rb") as fh:
        magic = fh.read(4)
    if magic[:2] != b"PK":
        try: tmp_path.unlink()
        except OSError: pass
        raise HTTPException(status_code=400, detail="File is not a valid xlsx (missing zip magic)")

    new_sha = _sha256_of_path(tmp_path)
    old_sha = ""
    archived_old_path = ""
    if target.exists():
        old_sha = _sha256_of_path(target)
        # Move existing file aside under .history/ before replacing.
        archived = history_dir / f"{ts}_{old_sha[:8]}_{upload_name}"
        try:
            target.replace(archived)  # atomic rename
            archived_old_path = str(archived)
        except OSError as e:
            try: tmp_path.unlink()
            except OSError: pass
            raise HTTPException(status_code=500, detail=f"Failed to archive previous file: {e}")

    # Atomic install of the new file.
    try:
        tmp_path.replace(target)
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"Failed to install new file: {e}")

    # Best-effort user identification (uvicorn runs under one OS user per OOD session).
    try:
        import pwd
        user = pwd.getpwuid(os.getuid()).pw_name
    except Exception:
        user = os.environ.get("USER", "")

    record = {
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "action": "replace",
        "reference": ref_name,
        "filename": upload_name,
        "user": user,
        "rationale": rationale.strip(),
        "old_sha256": old_sha,
        "new_sha256": new_sha,
        "size_bytes": bytes_written,
        "archived_old": archived_old_path,
        "target": str(target),
    }
    audit_path = _t39_audit_append(record, history_dir)

    return {
        "ok": True,
        "filename": upload_name,
        "new_sha256": new_sha,
        "old_sha256": old_sha,
        "archived_old": archived_old_path,
        "audit_log": audit_path,
        "size_bytes": bytes_written,
    }


@app.get("/api/projects")
def projects():
    cfg = load_config()
    return list_projects(_project_roots(cfg))


@app.post("/api/projects")
def project_create(payload: ProjectCreate):
    cfg = load_config()
    scope = getattr(payload, "scope", None) or SCOPE_PERSONAL
    reference = (payload.reference or "").strip()
    try:
        project_dir = create_project(_project_roots(cfg), payload.name, scope=scope, reference=reference)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    # Return the actual (normalized) directory name so the frontend can
    # update its state if the name differs from what the user typed.
    return {"path": str(project_dir), "name": project_dir.name, "scope": scope, "reference": reference}


@app.post("/api/projects/{project}/archive")
def project_archive(project: str):
    cfg = load_config()
    try:
        target = archive_project(_project_roots(cfg), project)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return {"archived_to": str(target)}


@app.delete("/api/projects/{project}")
def project_delete(project: str):
    cfg = load_config()
    try:
        deleted = delete_project(_project_roots(cfg), project)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if deleted is None:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"deleted": project}


class SetReferenceRequest(BaseModel):
    reference: str


@app.post("/api/projects/{project}/set_reference")
def project_set_reference(project: str, payload: SetReferenceRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    reference = (payload.reference or "").strip()
    meta = update_project_meta(project_dir, {"reference": reference} if reference else {})
    if not reference:
        # Explicitly clearing — remove the key
        meta.pop("reference", None)
        with open(project_dir / "project.json", "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, sort_keys=True)
    return {"reference": reference, "name": project}


@app.post("/api/projects/{project}/link-local")
def project_link_local(project: str, payload: LinkLocalRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    ensure_project_dirs(project_dir)
    raw_paths = [str(payload.path or "")] + list(payload.paths or [])
    raw_paths = [r.strip() for r in raw_paths if r and r.strip()]
    if not raw_paths:
        raise HTTPException(status_code=400, detail="No input path given")

    download_dir = project_dir / "download"
    # Accept either a directory of fastqs, or individual FASTQ files (a Kraken
    # parsed-read file pulled in so it can be re-run through Step 1, or a
    # multi-select from the server-side file picker). Files are symlinked to
    # their real target so the link keeps working even if the original is
    # itself a symlink.
    candidates: List[Path] = []
    missing: List[str] = []
    skipped_not_fastq: List[str] = []
    for raw_path in raw_paths:
        src = Path(raw_path).expanduser()
        if not src.exists():
            print(f"Link-local failed. Raw path: {raw_path!r} Resolved: {src}")
            missing.append(str(src))
            continue
        if src.is_file():
            if src.name.lower().endswith(_FASTQ_SUFFIXES):
                candidates.append(src)
            else:
                skipped_not_fastq.append(src.name)
        else:
            candidates.extend(sorted(src.glob("*.fastq.gz")))
            candidates.extend(sorted(src.glob("*.fq.gz")))

    # A single bad path used to abort the whole request. With multi-select that
    # would throw away every good file in the selection, so only fail when
    # nothing at all resolved.
    if missing and not candidates:
        raise HTTPException(
            status_code=400, detail=f"Input path not found: {missing[0]}"
        )

    count = 0
    already: List[str] = []
    for f in sorted(set(candidates)):
        target = download_dir / f.name
        # is_symlink() as well as exists(): a dangling symlink (source moved or
        # deleted) reports exists() == False but still occupies the name, so
        # symlink_to would raise FileExistsError and abort the whole batch.
        if target.exists() or target.is_symlink():
            already.append(f.name)
            continue
        target.symlink_to(f.resolve())
        count += 1
    return {
        "linked": count,
        "already_present": already,
        "missing": missing,
        "skipped_not_fastq": skipped_not_fastq,
    }


@app.post("/api/projects/{project}/import-vcfs")
def project_import_vcfs(project: str, payload: ImportVcfRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    ensure_project_dirs(project_dir)

    vcfs = []
    source_roots = []
    missing_sources: List[str] = []
    for raw in payload.source_paths or []:
        src = Path((raw or "").strip()).expanduser()
        if not src.exists():
            # Skip a missing source (e.g. a typo'd DB path) instead of aborting
            # the whole import — one bad path must not block the valid databases.
            # Surfaced to the UI as skipped_missing so the user still sees it.
            missing_sources.append(str(src))
            continue
        source_roots.append(src)
        vcfs.extend(list(src.rglob("*_zc.vcf")))
        vcfs.extend(list(src.rglob("*_zc.vcf.gz")))

    if payload.include_step1:
        step1_dir = project_dir / "step1"
        if step1_dir.exists():
            source_roots.append(step1_dir)
            vcfs.extend(list(step1_dir.rglob("*_zc.vcf")))
            vcfs.extend(list(step1_dir.rglob("*_zc.vcf.gz")))

    step1_dir = project_dir / "step1"
    resolved_vcfs = []
    vcf_sample_override = {}
    for vcf in vcfs:
        try:
            if step1_dir.exists() and str(vcf.resolve()).startswith(str(step1_dir.resolve())):
                sample = _sample_from_vcf(vcf)
                sample_dir = step1_dir / sample
                patched = _find_patched_vcf(sample_dir, sample, vcf)
                if patched:
                    resolved_vcfs.append(patched)
                    vcf_sample_override[patched] = sample
                    continue
        except FileNotFoundError:
            pass
        resolved_vcfs.append(vcf)
    vcfs = resolved_vcfs

    # De-duplicate by the real file each VCF resolves to. The project's
    # step2/vcf_database is now the cumulative collection of the same step1
    # _zc.vcf files (kept as symlinks), so including both step1 and vcf_database
    # as sources would otherwise count and import each shared VCF twice — the
    # "Large import (1981)" surprise. Two *different* files that merely share a
    # sample ID resolve to different paths and are both kept here; collapsing
    # those is what the dedupe / prefix-duplicates options handle downstream.
    _seen_targets: set = set()
    _deduped = []
    for v in vcfs:
        try:
            key = v.resolve()
        except OSError:
            key = v
        if key in _seen_targets:
            continue
        _seen_targets.add(key)
        _deduped.append(v)
    vcfs = _deduped

    if not vcfs:
        raise HTTPException(status_code=400, detail="No *_zc.vcf files found in provided sources")
    if len(vcfs) > 500 and not payload.confirm_large:
        raise HTTPException(status_code=400, detail=f"Large import ({len(vcfs)} VCFs). Confirm to continue.")

    alias_map = _reference_alias_map(Path(cfg["vsnp3_path"]))
    detected_refs = _detect_vcf_references(vcfs, alias_map)
    if not payload.reference:
        if len(detected_refs) > 1:
            raise HTTPException(status_code=400, detail=f"Mixed references detected: {', '.join(sorted(detected_refs))}")
        detected_ref = next(iter(detected_refs), "")
        if not detected_ref:
            raise HTTPException(status_code=400, detail="Reference is required (could not detect from VCF headers)")
    else:
        detected_ref = payload.reference

    vcf_source_dir = vcf_db_dir(project_dir / "step2")
    vcf_source_dir.mkdir(parents=True, exist_ok=True)
    action = (payload.action or "copy").lower()
    on_conflict = (payload.on_conflict or "skip").lower()

    imported = 0
    already_present = 0  # already in vcf_source, not re-copied (on_conflict=skip)
    ref_skipped = 0      # excluded: reference mismatch or unknown ref
    dedup_skipped = 0    # excluded: older duplicate sample name (dedupe=true)
    renamed = 0
    mismatched = []
    seen_samples = {}
    manifest_path = vcf_source_dir / ".vcf_source_manifest.csv"
    manifest_exists = manifest_path.exists()
    with manifest_path.open("a", encoding="utf-8") as manifest_handle:
        if not manifest_exists:
            manifest_handle.write("filename,source_type,source_path\n")
        for vcf in vcfs:
            vcf_ref = _detect_vcf_reference(vcf, alias_map)
            if vcf_ref and not _refs_match(vcf_ref, detected_ref, payload.allow_fuzzy_match):
                mismatched.append({"path": str(vcf), "reference": vcf_ref})
                if not payload.allow_mismatch:
                    ref_skipped += 1
                    continue
            if not vcf_ref:
                mismatched.append({"path": str(vcf), "reference": "unknown"})
                if not payload.allow_mismatch:
                    ref_skipped += 1
                    continue
            if payload.dedupe:
                sample = vcf_sample_override.get(vcf, _sample_from_vcf(vcf))
                if sample in seen_samples:
                    prev = seen_samples[sample]
                    if vcf.stat().st_mtime <= prev.stat().st_mtime:
                        dedup_skipped += 1
                        continue
                seen_samples[sample] = vcf
            target = vcf_source_dir / vcf.name
            if target.exists():
                if on_conflict == "skip":
                    already_present += 1
                    continue
                if on_conflict == "rename":
                    if payload.prefix_duplicates:
                        prefix = _source_prefix(vcf, source_roots)
                        target = _unique_target(vcf_source_dir, f"{prefix}__{vcf.name}")
                    else:
                        target = _unique_target(vcf_source_dir, vcf.name)
                    renamed += 1
                else:
                    target.unlink()
            if action == "link":
                target.symlink_to(vcf)
            else:
                shutil.copy2(vcf, target)
            imported += 1
            try:
                vcf_path = vcf.resolve()
            except FileNotFoundError:
                vcf_path = vcf
            source_type = "step1" if step1_dir.exists() and str(vcf_path).startswith(str(step1_dir.resolve())) else "reference"
            manifest_handle.write(f"{target.name},{source_type},{vcf_path}\n")

    mismatch_report = ""
    if mismatched:
        report_path = project_dir / "step2" / "mismatch_report.csv"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("path,reference\n")
            for row in mismatched:
                f.write(f"\"{row['path']}\",\"{row['reference']}\"\n")
        mismatch_report = str(report_path)

    edited_in_source = _edited_samples_in_dir(vcf_source_dir)
    _write_step2_edit_summary(vcf_source_dir.parent, edited_in_source)

    return {
        "imported": imported,
        "already_present": already_present,
        "ref_skipped": ref_skipped,
        "dedup_skipped": dedup_skipped,
        "skipped": ref_skipped + dedup_skipped,  # backward compat
        "renamed": renamed,
        "detected_reference": detected_ref or payload.reference or "",
        "mismatched": len(mismatched),
        "mismatch_report": mismatch_report,
        "total_found": len(vcfs),
        "skipped_missing": missing_sources,
    }


@app.post("/api/projects/{project}/upload")
async def project_upload(project: str, files: List[UploadFile] = File(...)):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    ensure_project_dirs(project_dir)
    download_dir = project_dir / "download"
    saved = 0
    for f in files:
        if not f.filename:
            continue
        target = download_dir / Path(f.filename).name
        with open(target, "wb") as out:
            while True:
                chunk = await f.read(1024 * 1024)
                if not chunk:
                    break
                out.write(chunk)
        saved += 1
    return {"uploaded": saved}


@app.get("/api/projects/{project}/sra-crosswalk")
def project_sra_crosswalk(project: str):
    """Serve <project>/download/sra_crosswalk.tsv as text/plain so the browser
    renders it in-tab. Generated by the SRA download flow when expanding
    sample/study-level inputs (SRS/DRS/SRX/PRJNA) into run accessions
    (SRR/DRR). 404 if no crosswalk file exists for this project (project
    was created before the auto-write landed, OR no SRA downloads have
    been kicked off in this project yet, OR the user uploaded fastqs
    directly without going through the SRA path)."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    crosswalk = project_dir / "download" / "sra_crosswalk.tsv"
    if not crosswalk.is_file():
        raise HTTPException(status_code=404, detail="No SRA crosswalk for this project")
    return FileResponse(crosswalk, media_type="text/plain")


@app.get("/api/projects/{project}/sra-download-report")
def project_sra_download_report(project: str):
    """Parsed outcome of the most recent SRA download into this project.

    The download script writes download/sra_download_report.tsv (one
    ``outcome<TAB>accession`` row per accession, overwritten each run). We parse
    it into buckets so the UI can persistently show which accessions were
    downloaded, skipped because already aligned in Step 1, or failed — the
    skipped ones leave no other on-screen trace (they were never fetched, so
    they don't appear in download/ or the "Ready to run" list).

    Returns empty buckets (not 404) when no download has run yet, so the caller
    can treat "no report" and "report with nothing skipped" uniformly.
    """
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    report = project_dir / "download" / "sra_download_report.tsv"
    buckets = {"downloaded": [], "already_in_step1": [], "failed": []}
    if report.is_file():
        for line in report.read_text(encoding="utf-8").splitlines():
            if not line or line.startswith("#"):
                continue
            outcome, _, acc = line.partition("\t")
            acc = acc.strip()
            if acc and outcome in buckets:
                buckets[outcome].append(acc)
    return buckets


@app.get("/api/projects/{project}/inputs")
def project_inputs(project: str):
    """List files currently in <project>/download/.

    Returns name + size + mtime per entry, plus an aggregate. Used by the
    GUI's "Files in this project" panel under the upload dropzone."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    download_dir = project_dir / "download"
    files: List[Dict] = []
    total_bytes = 0
    if download_dir.is_dir():
        for p in sorted(download_dir.iterdir()):
            if not p.is_file() or p.name.startswith("."):
                continue
            try:
                stat = p.stat()
            except OSError:
                continue
            files.append({
                "name": p.name,
                "size": stat.st_size,
                "mtime": stat.st_mtime,
            })
            total_bytes += stat.st_size
    return {"files": files, "total_bytes": total_bytes, "count": len(files)}


@app.delete("/api/projects/{project}/inputs/{filename}")
def project_input_delete(project: str, filename: str):
    """Delete a single file from <project>/download/.

    Filename is treated as a basename only — any directory traversal attempt
    (`..`, slashes, leading dots) is rejected. Idempotent: deleting a missing
    file returns 404."""
    if not filename or "/" in filename or "\\" in filename or filename.startswith(".") or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    target = (project_dir / "download" / filename).resolve()
    download_dir = (project_dir / "download").resolve()
    # Defense in depth: even if the basename guard above slipped, the resolved
    # target must live inside the project's download dir.
    try:
        target.relative_to(download_dir)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid filename")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    try:
        target.unlink()
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"Delete failed: {e}")
    return {"deleted": filename}


@app.post("/api/projects/{project}/sra/expand")
def sra_expand(project: str, payload: SraRequest):
    _ = project
    try:
        expanded = expand_accessions(payload.accessions, strict=True)
    except SRAExpansionError as e:
        raise HTTPException(status_code=502, detail=f"NCBI eutils unavailable: {e}")
    return {"expanded": expanded}


@app.post("/api/projects/{project}/sra/download")
def sra_download(project: str, payload: SraRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    ensure_project_dirs(project_dir)
    try:
        expanded, mapping = expand_accessions_with_mapping(payload.accessions, strict=True)
    except SRAExpansionError as e:
        # Fail fast — building a download script with the literal unexpanded
        # accession (e.g. SRX*/SRP*) would just produce a guaranteed-fail job.
        # The 502 status is a hint that the failure is upstream (NCBI), not
        # the user's input.
        raise HTTPException(
            status_code=502,
            detail=(
                f"Could not resolve SRA accessions via NCBI eutils: {e}. "
                "This is usually NCBI rate-limiting; wait ~30 s and retry. "
                "If it persists, set NCBI_API_KEY in the backend env to get a "
                "10 req/s allowance instead of 3."
            ),
        )
    download_root = project_dir / "download"
    if payload.folder:
        safe = Path(payload.folder).name
        download_root = download_root / safe
    download_root.mkdir(parents=True, exist_ok=True)
    # Persist the input→runs crosswalk so the user can map any downloaded
    # SRR/DRR back to the SRS/DRS/SRX they originally submitted. Without
    # this the resolution is computed at download time and then discarded,
    # leaving no way to reconcile post-hoc.
    try:
        write_crosswalk_tsv(download_root, mapping)
    except OSError as e:
        logger.warning("Failed to write sra_crosswalk.tsv: %s", e)
    script = build_download_script(
        download_root,
        expanded,
        cfg["sra"]["allow_insecure_https"],
        step1_dir=project_dir / "step1",
        # Always at the top-level download/ (not the optional subfolder) so the
        # report endpoint has one fixed place to read regardless of subfolder.
        report_path=project_dir / "download" / "sra_download_report.tsv",
    )
    script_path = download_root / "download_sra.sh"
    script_path.write_text(script, encoding="utf-8")
    script_path.chmod(0o755)
    job_id = job_manager.start_job(
        name="sra_download",
        command=wrap_cmd(cfg, f"bash {shlex.quote(str(script_path))}"),
        cwd=download_root,
        env=build_env(cfg)
    )
    return {"job_id": job_id}


# Greedy prefix (.+) binds the read marker to the RIGHTMOST _R1/_R2 (or _1/_2),
# so a sample ID that itself ends in _1/_2 (Mg_2_R1 -> Mg-2) isn't mis-split on
# the first such token. Non-greedy (.+?) would latch onto the first _1/_2 and
# let the lane group swallow the real _R1, collapsing Mg_2 back to "Mg".
_FASTQ_SAMPLE_RE = re.compile(r"(.+)(?:_R?[12])(?:_[^./]+)?\.fastq\.gz$")


def _sanitized_sample_and_name(filename: str) -> Tuple[str, str]:
    """Return (sample, on-disk filename) for a FASTQ, dashing the sample prefix.

    vSNP3 derives the sample name from the FASTQ filename by splitting at the
    FIRST '_'. An underscore *inside* the sample prefix therefore collapses every
    such sample to the shared prefix — Mg_280, Mg_281, … all become "Mg",
    silently merging distinct samples into one VCF in Step 2. Replacing the
    prefix's underscores with '-' (Mg_280 -> Mg-280) keeps each sample distinct
    while leaving the _R1/_R2 read indicator and any _001 lane suffix intact.

    Returns the original name unchanged when the prefix has no underscore.
    """
    m = _FASTQ_SAMPLE_RE.match(filename)
    sample = m.group(1) if m else filename.split(".")[0]
    safe = sample.replace("_", "-")
    if safe == sample:
        return sample, filename
    return safe, safe + filename[len(sample):]


@app.post("/api/projects/{project}/step1/setup")
def step1_setup(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    ensure_project_dirs(project_dir)
    download_dir = project_dir / "download"
    step1_dir = project_dir / "step1"

    # Group by sample prefix before _R1/_R2 (scan recursively for subfolders)
    fastqs = list(download_dir.rglob("*.fastq.gz"))
    if not fastqs:
        return {"created": 0, "message": "No FASTQ files found"}

    created = 0
    renamed = 0
    for f in fastqs:
        sample, safe_name = _sanitized_sample_and_name(f.name)
        # Rename underscored stems in place (Mg_280_R1 -> Mg-280_R1) BEFORE
        # staging, so vSNP3 never sees an underscore in the sample prefix.
        # download/ entries may be symlinks (rename moves the link, not the
        # target) or real files — both are safe to rename. Idempotent: a name
        # already dashed, or a re-run, is a no-op.
        if safe_name != f.name:
            new_path = f.with_name(safe_name)
            if new_path.exists():
                # A dashed file is already present (e.g. the project shipped
                # both Mg_280_R1 and Mg-280_R1). Renaming would clobber it or
                # leave this one orphaned, so skip this underscored duplicate —
                # the dashed file is staged on its own pass through `fastqs`.
                logger.warning(
                    "step1_setup: dashed target %s already exists; skipping "
                    "underscored duplicate %s", new_path.name, f.name)
                continue
            f.rename(new_path)
            renamed += 1
            f = new_path
        sample_dir = step1_dir / sample
        sample_dir.mkdir(parents=True, exist_ok=True)
        target = sample_dir / f.name
        if not target.exists():
            # Real COPY, not a symlink: the step1 sample folder must retain the
            # exact reads used for its alignment even if download/ is later moved
            # or deleted. copy2 follows the source (download/ entries may
            # themselves be symlinks) so we copy the actual bytes. The reads stay
            # in download/ too. Cost: ~doubles read storage.
            #
            # Only NEW entries are copied. We deliberately do NOT rewrite existing
            # symlink entries from before this change into copies here — doing so
            # would make the next Setup on a large project copy hundreds of GB
            # synchronously and hang the request. Legacy symlinked samples keep
            # their symlink until re-staged; migrate them separately if needed.
            shutil.copy2(f, target)
            created += 1
    return {"created": created, "renamed": renamed}


@app.post("/api/projects/{project}/step1/run")
def step1_run(project: str, payload: Step1Request):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    step1_dir = project_dir / "step1"
    # Hold the dispatch lock across the whole check→claim so a simultaneous
    # double-click serializes: the first request claims .step1_job_id and starts
    # the job; the second then sees it "running" and gets a 409 (below).
    with _STEP1_DISPATCH_LOCK:
        return _step1_dispatch(project, payload, cfg, project_dir, step1_dir)


def _step1_dispatch(
    project: str, payload: Step1Request, cfg: Dict[str, Any],
    project_dir: Path, step1_dir: Path,
):
    # Refuse to spawn a second batch while a prior step1 job is still
    # running — concurrent batches share the same per-sample dirs and race
    # over the SAM / log / .provenance/exit_code files, producing the
    # FileNotFoundError-on-temp_fastq_seqkit_stats.txt class of failures
    # we hit on the M. sciuri panel. 409 is the right shape; the frontend
    # surfaces it via the same setSraStatus-style error handler.
    prior_job_id_path = step1_dir / ".step1_job_id"
    if prior_job_id_path.exists():
        try:
            prior_id = prior_job_id_path.read_text(encoding="utf-8").strip()
        except OSError:
            prior_id = ""
        if prior_id:
            prior_job = job_manager.get_job(prior_id)
            if prior_job and prior_job.get("status") == "running":
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Step 1 is already running (job {prior_id}). "
                        "Wait for it to finish before clicking Run again."
                    ),
                )
    script_path = step1_dir / "run_step1.sh"

    # Concurrency guard: reject duplicate runs so parallel wrappers don't
    # trample each other's output directories (the wrapper's per-sample
    # cleanup deletes alignment_* on entry, which corrupts any in-flight run).
    job_id_path = step1_dir / ".step1_job_id"
    if job_id_path.exists():
        existing_id = job_id_path.read_text(encoding="utf-8").strip()
        if existing_id:
            existing_job = job_manager.get_job(existing_id)
            if existing_job and existing_job.get("status") == "running":
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Step 1 is already running for this project (job {existing_id}). "
                        "Wait for it to finish before starting a new run."
                    ),
                )
    if script_path.exists() and _wrapper_process_alive(script_path):
        raise HTTPException(
            status_code=409,
            detail=(
                "Step 1 is already running for this project "
                "(a previous wrapper process is still active). "
                "Wait for it to finish before starting a new run."
            ),
        )

    debug_flag = "--debug" if payload.debug else ""
    assemble_unmap_flag = "-assemble_unmap" if payload.assemble_unmap else ""
    nanopore_flag = "--nanopore" if payload.nanopore else ""
    # Auto-populate reference from project.json if not supplied in payload
    if not payload.reference:
        payload.reference = _project_reference(project_dir)
    ref_arg = f"-t {payload.reference}" if payload.reference else ""
    if payload.reference:
        update_project_meta(project_dir, {
            "reference": payload.reference,
            "display_name": f"{project}_{payload.reference}"
        })
    max_parallel = cfg.get("step1_max_parallel", 1) or 1
    try:
        max_parallel = max(1, int(max_parallel))
    except (TypeError, ValueError):
        max_parallel = 1

    # Decide which sample dirs to actually run BEFORE writing the batch script,
    # so the script iterates only the valid samples. This keeps a stale dir
    # (e.g. one whose download source was deleted, leaving broken symlinks)
    # from being run and failing the whole batch — those are surfaced to the
    # user as `skipped_samples` instead.
    samples, skipped_samples = _step1_dispatch_plan(
        step1_dir,
        min_bytes=int(cfg.get("step1_min_fastq_bytes", _T46_JUNK_FASTQ_BYTES) or _T46_JUNK_FASTQ_BYTES),
        force_rerun=bool(getattr(payload, "force_rerun", False)),
    )
    samples_bash = " ".join(shlex.quote(s) for s in samples)
    script_path.write_text(
        "\n".join([
            "#!/bin/bash",
            "set -uo pipefail",
            "FAIL=0",
            f"MAX_PARALLEL={max_parallel}",
            "OVERALL_START=$(date +%s)",
            "OVERALL_LOG=\"step1_run_summary.log\"",
            "echo \"Start: $(date -u +%Y-%m-%dT%H:%M:%SZ)\" >> \"$OVERALL_LOG\"",
            "run_sample() {",
            "  local d=\"$1\"",
            "  if [ ! -d \"$d\" ]; then",
            "    return 0",
            "  fi",
            "  # Skip writer/janitor scaffolding (T-07 _provenance, .git, etc.).",
            "  case \"$d\" in",
            "    _*/|.*/|_*|.*) return 0 ;;",
            "  esac",
            "  echo \"== Running step1 in $d ==\"",
            "  cd \"$d\"",
            "  LOG=run_step1.log",
            "  echo \"== Running step1 in $d ==\" | tee -a \"$LOG\"",
            "  START_TS=$(date +%s)",
            "  echo \"Start: $(date -u +%Y-%m-%dT%H:%M:%SZ)\" | tee -a \"$LOG\"",
            "  if [ \"" + ("1" if payload.debug else "0") + "\" = \"0\" ]; then",
            "    for dir in alignment_*; do",
            "      if [ -d \"$dir\" ]; then",
            "        rm -rf \"$dir\"",
            "      fi",
            "    done",
            "    if [ -d \"unmapped_reads\" ]; then",
            "      rm -rf \"unmapped_reads\"",
            "    fi",
            "    if [ -d \"sourmash\" ]; then",
            "      rm -rf \"sourmash\"",
            "    fi",
            "  fi",
            "  R1=$(ls *_R1*.fastq.gz 2>/dev/null | head -n1 || true)",
            "  R2=$(ls *_R2*.fastq.gz 2>/dev/null | head -n1 || true)",
            "  if [ -z \"$R1\" ]; then R1=$(ls *_1*.fastq.gz 2>/dev/null | head -n1 || true); fi",
            "  if [ -z \"$R2\" ]; then R2=$(ls *_2*.fastq.gz 2>/dev/null | head -n1 || true); fi",
            # No read-tag naming → single-end / long-read: use the lone fastq as
            # R1. vsnp3 accepts -r1 by itself (see dispatch-plan docstring).
            "  if [ -z \"$R1\" ]; then R1=$(ls *.fastq.gz 2>/dev/null | head -n1 || true); fi",
            "  if [ -z \"$R1\" ]; then",
            "    echo \"Missing R1 in $d\" | tee -a \"$LOG\"",
            "    cd ..",
            "    return 0",
            "  fi",
            # Read-type detection. NP starts from the batch-level nanopore
            # override (checkbox); if that's off, auto-detect long reads by the
            # average length of the first ~400 reads and flip on --nanopore when
            # it exceeds 600 bp. This never forces nanopore OFF, so it composes
            # safely with vsnp3's own internal >701 bp auto-detection.
            "  NP=\"" + nanopore_flag + "\"",
            "  if [ -z \"$NP\" ]; then",
            "    AVG=$(zcat \"$R1\" 2>/dev/null | head -n 1600 | awk 'NR%4==2{n++;s+=length($0)} END{if(n>0)printf \"%d\", s/n; else print 0}')",
            "    if [ \"${AVG:-0}\" -gt 600 ]; then",
            "      NP=\"--nanopore\"",
            "      echo \"Auto-detected long reads (avg ${AVG}bp) — using --nanopore\" | tee -a \"$LOG\"",
            "    fi",
            "  fi",
            "  mkdir -p .provenance",
            # Record the read type for the results table: paired (R2 present),
            # ont (nanopore alignment — auto-detected or forced), else single.
            "  if [ -n \"$R2\" ]; then echo paired > .provenance/read_type; elif [ -n \"$NP\" ]; then echo ont > .provenance/read_type; else echo single > .provenance/read_type; fi",
            "  date -u +%s.%N > .provenance/started_at",
            # Build the exact command as a shell variable so we can BOTH record it
            # verbatim to .provenance AND run it (via eval) — guaranteeing the
            # recorded line is literally what executed. This is the per-sample
            # "what ran on the command line" provenance the sample folder shows.
            f"  if [ -n \"$R2\" ]; then",
            f"    RUN_CMD=\"vsnp3_step1.py -r1 \\\"$R1\\\" -r2 \\\"$R2\\\" {ref_arg} {debug_flag} {assemble_unmap_flag} $NP\"",
            "  else",
            f"    RUN_CMD=\"vsnp3_step1.py -r1 \\\"$R1\\\" {ref_arg} {debug_flag} {assemble_unmap_flag} $NP\"",
            "  fi",
            "  PROV_TS=$(date '+%Y-%m-%d %H:%M:%S %z')",
            "  PROV_SAMPLE=$(basename \"$d\")",
            "  PROV_CWD=$(pwd)",
            "  PROV_VSNP3=$(command -v vsnp3_step1.py 2>/dev/null || echo vsnp3_step1.py)",
            "  {",
            "    echo \"# ================================================================\"",
            "    echo \"# vsnp_gui - Step 1 (per-sample alignment)\"",
            "    echo \"# run at:      $PROV_TS\"",
            "    echo \"# sample:      $PROV_SAMPLE\"",
            "    echo \"# working dir: $PROV_CWD\"",
            "    echo \"# tool:        $PROV_VSNP3\"",
            "    echo \"# ----------------------------------------------------------------\"",
            "    echo \"# Command executed (copy/paste to reproduce):\"",
            "    echo \"$RUN_CMD\"",
            "    echo \"\"",
            "  } >> .provenance/vsnp_gui_step1_run_cmd.txt 2>/dev/null || true",
            "  eval \"$RUN_CMD\" >> \"$LOG\" 2>&1",
            "  STATUS=$?",
            "  echo $STATUS > .provenance/exit_code",
            "  date -u +%s.%N > .provenance/finished_at",
            "  if [ \"$STATUS\" -eq 0 ]; then",
            "    END_TS=$(date +%s)",
            "    DURATION=$((END_TS-START_TS))",
            "    echo \"Complete: $(date -u +%Y-%m-%dT%H:%M:%SZ)\" | tee -a \"$LOG\"",
            "    echo \"Duration: ${DURATION}s\" | tee -a \"$LOG\"",
            "  else",
            "    echo \"Error: exit $STATUS\" | tee -a \"$LOG\"",
            "    FAIL=1",
            "  fi",
            "  cd ..",
            "  return $STATUS",
            "}",
            # Rolling worker pool throttled to MAX_PARALLEL. We MUST NOT use
            # `wait -n` here: it only exists in bash 4.3+, and macOS ships bash
            # 3.2 as /bin/bash (what `bash run_step1.sh` resolves to). Under 3.2
            # `wait -n` fails instantly as an invalid option WITHOUT blocking, so
            # the gate is skipped and every sample launches at once — which
            # exhausts RAM/CPU and hard-locks the machine. Instead poll the
            # running-job count (`jobs -r -p`), which is portable to 3.2, and
            # only launch a new sample once a slot frees. The SRA download path
            # already throttles portably via `xargs -P` (see sra.py).
            "pids=()",
            f"SAMPLES=({samples_bash})",
            "for d in \"${SAMPLES[@]}\"; do",
            "  while [ \"$(jobs -r -p | wc -l | tr -d ' ')\" -ge \"$MAX_PARALLEL\" ]; do",
            "    sleep 1",
            "  done",
            "  run_sample \"$d\" &",
            "  pids+=(\"$!\")",
            "done",
            "for p in \"${pids[@]}\"; do",
            "  wait \"$p\" || FAIL=1",
            "done",
            "OVERALL_END=$(date +%s)",
            "OVERALL_DURATION=$((OVERALL_END-OVERALL_START))",
            "echo \"End: $(date -u +%Y-%m-%dT%H:%M:%SZ)\" >> \"$OVERALL_LOG\"",
            "echo \"Duration: ${OVERALL_DURATION}s\" >> \"$OVERALL_LOG\"",
            "exit $FAIL",
        ]),
        encoding="utf-8"
    )
    script_path.chmod(0o755)

    # T-07: provenance dispatch. Pre-create per-sample run_metadata.json with
    # frozen dispatch_state sub-block; bash batch will write sentinel files
    # alongside vsnp3_step1.py invocations; finalize_callback rewrites per-
    # sample metadata as terminal once the batch exits. Skip provenance if
    # no reference is set (legacy/unconfigured runs) — the bash still runs.
    prov_finalize_cb = None
    provenance_warning = ""
    if payload.reference:
        if samples:
            # A capture failure must NOT stop the batch — the record describes
            # the run, it is not a precondition for it (this used to 500 the
            # whole Run click over metadata). Warn loudly and run anyway.
            prov_batch_run_id = None
            try:
                prov_batch_run_id, _sample_run_ids = provenance_writer.dispatch_step1_batch(
                    cfg, project_dir, samples, payload.reference,
                    user=_current_user(),
                    ood_session_id=_ood_session_id(),
                )
            except provenance_writer.DispatchFailed as e:
                provenance_warning = (
                    f"Run-metadata (provenance) was not recorded for this run: {e}"
                )
                logger.warning("Step 1 provenance dispatch failed; batch proceeds "
                               "without run_metadata.json: %s", e)

            if prov_batch_run_id is not None:
                def prov_finalize_cb(job_id, exit_code, started_at, finished_at):
                    provenance_writer.finalize_step1_batch(
                        project_dir, prov_batch_run_id, exit_code, started_at, finished_at,
                    )

    job_id = job_manager.start_job(
        name="step1",
        command=wrap_cmd(cfg, f"bash {shlex.quote(str(script_path))}"),
        cwd=step1_dir,
        env=build_env(cfg),
        finalize_callback=prov_finalize_cb,
    )
    (step1_dir / ".step1_job_id").write_text(job_id, encoding="utf-8")
    # T-46 Phase 1: surface samples auto-excluded from the dispatch so the
    # user can see what didn't run and why — but only the NOTEWORTHY ones.
    # "Already completed" skips are expected on any re-run (that's the whole
    # point of skipping finished work), so listing them just spams the dispatch
    # popup with hundreds of lines; drop them. Junk-size / broken-link /
    # errored / no-fastq skips remain — those are the "why didn't my sample
    # run?" cases the popup exists for.
    noteworthy_skips = [
        s for s in skipped_samples
        if not str(s.get("reason", "")).startswith("already completed")
    ]
    return {
        "job_id": job_id,
        "skipped_samples": noteworthy_skips,
        # Non-empty when the batch started WITHOUT a provenance record (T-07
        # dispatch failed). The analysis is unaffected; UI shows it as a note.
        "provenance_warning": provenance_warning,
    }


@app.get("/api/projects/{project}/step1/status")
def step1_status(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")

    min_bytes = int(cfg.get("step1_min_fastq_bytes", _T46_JUNK_FASTQ_BYTES) or _T46_JUNK_FASTQ_BYTES)
    job_id_path = step1_dir / ".step1_job_id"
    job_id = job_id_path.read_text(encoding="utf-8").strip() if job_id_path.exists() else ""
    job = job_manager.get_job(job_id) if job_id else None
    job_status = job["status"] if job else "unknown"
    # Restart resilience: the batch is one bash job that keeps running (as an
    # orphan) across a backend restart, but the in-memory JobManager state is
    # lost, so `job` is None and job_status would read "unknown" — making live
    # samples render as "unknown" instead of "running". If the wrapper process
    # is still alive on this host, it IS running; report that so the panel keeps
    # showing progress after a reconnect/restart.
    if job_status in ("unknown", ""):
        script_path = step1_dir / "run_step1.sh"
        if script_path.exists() and _wrapper_process_alive(script_path):
            job_status = "running"

    vcfs_dir = vcf_db_dir(project_dir / "step2")
    in_vcfs_folder: set = set()
    if vcfs_dir.exists():
        for vf in vcfs_dir.glob("*_zc.vcf*"):
            stem = vf.name.replace("_zc.vcf.gz", "").replace("_zc.vcf", "")
            in_vcfs_folder.add(stem)

    statuses = []
    for sample_dir in sorted(step1_dir.glob("*")):
        if not sample_dir.is_dir():
            continue
        # Skip writer/janitor scaffolding (e.g. _provenance/) so they don't
        # surface as "Unknown" samples in the GUI.
        if sample_dir.name.startswith(("_", ".")):
            continue
        sample = sample_dir.name
        log_path = sample_dir / "run_step1.log"
        exit_code_path = sample_dir / ".provenance" / "exit_code"

        # Fast path: a sample with an exit_code sentinel is terminal. Serve it
        # from cache (keyed on the sentinel's mtime) so we skip the per-sample
        # globs below on every poll. in_vcfs_folder is layered on fresh since it
        # tracks the separate VCF collection, which the user can change anytime.
        cache_key = str(sample_dir)
        try:
            ec_mtime = exit_code_path.stat().st_mtime_ns
        except OSError:
            ec_mtime = None
        if ec_mtime is not None:
            cached = _STEP1_STATUS_CACHE.get(cache_key)
            if cached and cached[0] == ec_mtime:
                entry = dict(cached[1])
                entry["in_vcfs_folder"] = sample in in_vcfs_folder
                entry["reason"] = _step1_status_reason(entry.get("status", ""), sample_dir, min_bytes)
                statuses.append(entry)
                continue

        vcf = next(iter(_align_glob(sample_dir, "*_filtered_hapall_annotated.vcf")), None)
        nodup = next(iter(_align_glob(sample_dir, "*_nodup.bam")), None)
        # Non-recursive: the per-sample zc VCF is always written under
        # alignment_*/ (or the legacy plain alignment/). A recursive **/*_zc.vcf
        # glob walked each sample's entire output subtree (unmapped_reads/,
        # sourmash/, spoligo/, …) on every poll — the dominant cost that made
        # the endpoint time out at ~1000 samples.
        zc_vcf = next(iter(_align_glob(sample_dir, "*_zc.vcf")), None) or next(
            iter(_align_glob(sample_dir, "*_zc.vcf.gz")), None
        )

        # Status logic (in priority order):
        #   1. .provenance/exit_code present  → authoritative per-sample terminal
        #      state (T-07 sentinel written by the bash batch after vsnp3_step1.py
        #      exits). 0 = success, non-zero = real failure.
        #   2. Outputs (VCF + BAM) present     → complete (legacy projects pre-T-07
        #      sentinels, or sentinel write raced).
        #   3. log_path exists + job running   → still running.
        #   4. log_path exists + job not running → unknown (sample's bash leg died
        #      before writing the sentinel — kill / OOM / batch interrupted).
        #   5. else                            → not_started.
        # The previous heuristic grepped the running log for "Error:" / "Exception"
        # / "Traceback", which false-positives on vsnp3's verbose intermediate
        # output (deprecation warnings, etc) and made every sample flicker into
        # "Error" before transitioning to "Complete" once the VCF landed.
        status = "not_started"
        legacy_complete = False
        exit_code_str = ""
        if exit_code_path.exists():
            try:
                exit_code_str = exit_code_path.read_text(encoding="utf-8").strip()
            except OSError:
                exit_code_str = ""
        if exit_code_str:
            status = "complete" if exit_code_str == "0" else "error"
        elif vcf and nodup:
            status = "complete"
        elif zc_vcf is not None and zc_vcf.parent.name == "alignment":
            # vSNP2-era run (plain alignment/, no sentinel, different inner file
            # names): the zero-coverage VCF is the artifact step2 consumes, so
            # the sample is Complete, not 'Not Started' with a fastq complaint.
            status = "complete"
            legacy_complete = True
        elif log_path.exists():
            status = "running" if job_status == "running" else "unknown"
        entry = {
            "sample": sample,
            "status": status,
            "log_path": str(log_path),
            "has_log": log_path.exists(),
            "has_outputs": bool(vcf and nodup) or legacy_complete,
            "has_zc_vcf": bool(zc_vcf),
            "in_vcfs_folder": sample in in_vcfs_folder,
            "reason": (
                "aligned before this GUI (legacy alignment/ layout)"
                if legacy_complete
                else _step1_status_reason(status, sample_dir, min_bytes)
            ),
        }
        # Cache terminal samples (exit_code written) so later polls take the
        # fast path above. Store everything except in_vcfs_folder and reason,
        # which are re-layered per poll (the former tracks the VCF collection;
        # the latter is cheap and kept fresh alongside it).
        if ec_mtime is not None and status in ("complete", "error"):
            _STEP1_STATUS_CACHE[cache_key] = (
                ec_mtime, {k: v for k, v in entry.items() if k not in ("in_vcfs_folder", "reason")}
            )
        statuses.append(entry)
    return {"job_status": job_status, "samples": statuses}


def _safe_child(parent: Path, name: str) -> Path:
    """Resolve `name` as a direct child of `parent`, rejecting separators /
    traversal. Raises HTTPException(400) on anything suspicious."""
    if not name or "/" in name or "\\" in name or name in (".", ".."):
        raise HTTPException(status_code=400, detail="Invalid name")
    child = (parent / name).resolve()
    if child.parent != parent.resolve():
        raise HTTPException(status_code=400, detail="Invalid path")
    return child


def _step1_batch_running(step1_dir: Path) -> bool:
    job_id_path = step1_dir / ".step1_job_id"
    job_id = job_id_path.read_text(encoding="utf-8").strip() if job_id_path.exists() else ""
    job = job_manager.get_job(job_id) if job_id else None
    if job and job.get("status") == "running":
        return True
    script_path = step1_dir / "run_step1.sh"
    return script_path.exists() and _wrapper_process_alive(script_path)


@app.delete("/api/projects/{project}/step1/samples/{sample}")
def step1_remove_sample(project: str, sample: str):
    """Move one Step 1 sample to the project's Quarantine — recoverable, not a
    hard delete.

    The sample's reads are moved out of download/ (and any real fastqs in the
    step1 dir) into quarantine/<sample>/, so step1_setup no longer re-creates it
    on the next Setup, but nothing irreplaceable is destroyed — from the
    Quarantine panel the user can Restore (reads back to download/) or Delete
    forever. The step1 output dir (regenerable by re-running) is discarded.
    Refuses while a batch is running. Path-guarded to stay inside step1/."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")
    target = _safe_child(step1_dir, sample)
    if not target.is_dir():
        raise HTTPException(status_code=404, detail=f"Sample not found: {sample}")
    if _step1_batch_running(step1_dir):
        raise HTTPException(status_code=409, detail="Step 1 is running — stop it before removing samples.")

    quarantine_dir = project_dir / "quarantine" / sample
    quarantine_dir.mkdir(parents=True, exist_ok=True)

    # Preserve one copy of each read in quarantine and clear it from download/
    # so a quarantined sample no longer shows under "Files in download". The
    # step1 folder now holds real COPIES, and download/ holds its own copy, so
    # for each read we MOVE the download/ copy into quarantine (preferred — that
    # empties download/); if there's no download copy, fall back to the step1
    # copy. The step1 dir (with any remaining copies + outputs) is removed below.
    download_root = project_dir / "download"
    moved: List[str] = []
    for fq in list(target.glob("*.fastq.gz")) + list(target.glob("*.fastq")):
        name = fq.name
        dest = quarantine_dir / name
        src: Optional[Path] = None
        if download_root.is_dir():
            for cand in download_root.rglob(name):
                if cand.is_file():
                    src = cand
                    break
        if src is None and fq.is_file():
            src = fq  # no download copy left — preserve the step1 copy
        if src is not None:
            try:
                shutil.move(str(src), str(dest))
                moved.append(name)
            except OSError:
                pass

    size_bytes = 0
    for m in moved:
        size_bytes += _safe_stat_size(quarantine_dir / m) or 0
    meta = {
        "sample": sample,
        "removed_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "files": moved,
        "size_bytes": size_bytes,
        "reason": _step1_status_reason(
            "error" if _step1_errored(target) else "not_started",
            target,
            int(cfg.get("step1_min_fastq_bytes", _T46_JUNK_FASTQ_BYTES) or _T46_JUNK_FASTQ_BYTES),
        ),
    }
    (quarantine_dir / ".quarantine.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

    try:
        shutil.rmtree(target)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Failed to remove sample: {exc}")
    _STEP1_STATUS_CACHE.pop(str(step1_dir / sample), None)
    return {"quarantined": sample, "files": moved}


@app.get("/api/projects/{project}/quarantine")
def quarantine_list(project: str):
    """List samples currently in the project's Quarantine (removed from Step 1,
    recoverable). Each entry carries when/why it was removed and its read size."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    qroot = project_dir / "quarantine"
    items: List[Dict[str, Any]] = []
    if qroot.is_dir():
        for d in sorted(qroot.iterdir()):
            if not d.is_dir() or d.name.startswith("."):
                continue
            meta: Dict[str, Any] = {}
            mp = d / ".quarantine.json"
            if mp.exists():
                try:
                    meta = json.loads(mp.read_text(encoding="utf-8"))
                except (OSError, json.JSONDecodeError):
                    meta = {}
            fastqs = list(d.glob("*.fastq.gz")) + list(d.glob("*.fastq"))
            items.append({
                "sample": d.name,
                "removed_at": meta.get("removed_at", ""),
                "reason": meta.get("reason", ""),
                "files": [f.name for f in fastqs],
                "size_bytes": meta.get("size_bytes") or sum(_safe_stat_size(f) or 0 for f in fastqs),
                "restorable": bool(fastqs),
            })
    return {"quarantine": items}


@app.post("/api/projects/{project}/quarantine/{sample}/restore")
def quarantine_restore(project: str, sample: str):
    """Restore a quarantined sample: move its reads back to download/ so the next
    Setup re-creates it for Step 1. Skips a read whose name already exists in
    download/ (won't clobber). Removes the quarantine entry when empty."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    qroot = project_dir / "quarantine"
    qdir = _safe_child(qroot, sample)
    if not qdir.is_dir():
        raise HTTPException(status_code=404, detail=f"Not in quarantine: {sample}")
    download_dir = project_dir / "download"
    download_dir.mkdir(parents=True, exist_ok=True)
    restored: List[str] = []
    skipped: List[str] = []
    for fq in list(qdir.glob("*.fastq.gz")) + list(qdir.glob("*.fastq")):
        dest = download_dir / fq.name
        if dest.exists():
            skipped.append(fq.name)
            continue
        try:
            shutil.move(str(fq), str(dest))
            restored.append(fq.name)
        except OSError:
            skipped.append(fq.name)
    # Only tear down the quarantine entry if nothing was left behind.
    if not skipped:
        try:
            shutil.rmtree(qdir)
        except OSError:
            pass
    return {"restored": restored, "skipped": skipped, "sample": sample}


@app.delete("/api/projects/{project}/quarantine/{sample}")
def quarantine_delete(project: str, sample: str):
    """Permanently delete a quarantined sample (its held reads). The only place
    a hard delete happens — Remove just quarantines."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    qroot = project_dir / "quarantine"
    qdir = _safe_child(qroot, sample)
    if not qdir.is_dir():
        raise HTTPException(status_code=404, detail=f"Not in quarantine: {sample}")
    try:
        shutil.rmtree(qdir)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Failed to delete: {exc}")
    return {"deleted": sample}


@app.post("/api/projects/{project}/step1/stop")
def step1_stop(project: str):
    """Stop a running Step 1 batch for a project — restart-resilient.

    The generic /api/jobs/{id}/stop path 404s after a backend reload because
    the in-memory JobManager no longer knows the (orphaned) batch. This
    project-scoped endpoint first tries the normal in-memory stop, then falls
    back to killing the live wrapper's process group directly (see
    _terminate_step1_wrapper). Finished samples keep their outputs; in-flight
    samples are left partial. Returns 409 only if nothing is actually running.
    """
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")

    job_id_path = step1_dir / ".step1_job_id"
    job_id = job_id_path.read_text(encoding="utf-8").strip() if job_id_path.exists() else ""

    # Normal path: the JobManager still holds this run (no backend restart since
    # dispatch) and can signal its tracked Popen process group.
    if job_id and job_manager.stop_job(job_id):
        return {"stopped": True, "method": "job_manager", "job_id": job_id}

    # Fallback path: in-memory state was lost (backend reloaded / reconnected
    # from another machine), but an orphaned wrapper may still be running.
    script_path = step1_dir / "run_step1.sh"
    if script_path.exists() and _terminate_step1_wrapper(script_path):
        return {"stopped": True, "method": "wrapper_group", "job_id": job_id}

    raise HTTPException(status_code=409, detail="No running Step 1 batch to stop")


@app.get("/api/projects/{project}/step1/log")
def step1_log(project: str, sample: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    log_path = project_dir / "step1" / sample / "run_step1.log"
    if not log_path.exists():
        raise HTTPException(status_code=404, detail="Log not found")
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            lines = f.readlines()[-400:]
    except OSError as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"sample": sample, "log": "".join(lines)}


@app.get("/api/projects/{project}/vcfs")
def project_vcfs_list(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    vcfs_dir = vcf_db_dir(project_dir / "step2")
    vcfs_dir.mkdir(parents=True, exist_ok=True)
    vcfs = sorted([*vcfs_dir.glob("*_zc.vcf"), *vcfs_dir.glob("*_zc.vcf.gz")], key=lambda p: p.name)
    samples = []
    for v in vcfs:
        stem = v.name.replace("_zc.vcf.gz", "").replace("_zc.vcf", "")
        samples.append({"filename": v.name, "sample": stem})
    return {"count": len(samples), "path": str(vcfs_dir), "folder_name": vcfs_dir.name, "samples": samples}


class VcfsCollectRequest(BaseModel):
    force_samples: List[str] = []


@app.post("/api/projects/{project}/vcfs/collect")
def project_vcfs_collect(project: str, payload: VcfsCollectRequest):
    """Scan step1/ for passing _zc.vcf files and symlink them into the cumulative
    VCF database (step2/vcf_database/). Accumulates — existing entries (including
    imported/historical VCFs with no step1 source) are preserved. Samples in
    force_samples are included even if they did not pass Step 1."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    step1_dir = project_dir / "step1"
    vcfs_dir = vcf_db_dir(project_dir / "step2")
    vcfs_dir.mkdir(parents=True, exist_ok=True)

    force_set = set(payload.force_samples or [])
    # Samples excluded in Step 1 Results must not be auto-collected into the
    # VCF database (they would otherwise feed a Step 2 build the user meant to drop).
    # An explicit force_samples check still overrides, since that is a deliberate
    # per-sample "add this anyway" action.
    excluded_set = set(_read_step1_exclusions(project_dir / "step2"))
    auto_added: List[str] = []
    force_added: List[str] = []
    already_present: List[str] = []
    no_vcf: List[str] = []
    excluded_skipped: List[str] = []

    for sample_dir in sorted(step1_dir.glob("*")):
        if not sample_dir.is_dir() or sample_dir.name.startswith(("_", ".")):
            continue
        sample = sample_dir.name

        if sample in excluded_set and sample not in force_set:
            excluded_skipped.append(sample)
            continue

        # Determine pass/fail from provenance sentinel, fall back to output presence
        exit_code_path = sample_dir / ".provenance" / "exit_code"
        passed = False
        if exit_code_path.exists():
            try:
                passed = exit_code_path.read_text(encoding="utf-8").strip() == "0"
            except OSError:
                pass
        else:
            vcf_out = next(iter(_align_glob(sample_dir, "*_filtered_hapall_annotated.vcf")), None)
            nodup_out = next(iter(_align_glob(sample_dir, "*_nodup.bam")), None)
            passed = bool(vcf_out and nodup_out) or _legacy_step1_complete(sample_dir)

        if not passed and sample not in force_set:
            continue

        # Find the latest _zc.vcf under alignment_*/ (non-recursive to avoid
        # walking the whole sample tree; falls back to the legacy plain
        # alignment/ for pre-GUI runs). Prefer an edited/patched VCF if one
        # exists, and use the same target-name rule as step2_setup so both
        # writers agree on filenames in the shared database.
        all_candidates = sorted(
            _align_glob(sample_dir, "*_zc.vcf*"), key=lambda p: p.stat().st_mtime
        )
        if not all_candidates:
            if sample in force_set:
                no_vcf.append(sample)
            continue

        source_vcf = all_candidates[-1]
        patched_vcf = _find_patched_vcf(sample_dir, sample, source_vcf)
        chosen_vcf = (patched_vcf or source_vcf).resolve()
        target = vcfs_dir / _target_name_for_vcf(source_vcf, chosen_vcf)

        if target.exists() or target.is_symlink():
            # Already in the database. If it's a legacy (or broken) symlink,
            # replace it with a durable real copy so the cumulative collection
            # is self-contained and survives a step1 cleanup; a real file is
            # left untouched (accumulate, never clobber).
            if target.is_symlink():
                target.unlink()
                shutil.copy2(chosen_vcf, target)
            already_present.append(sample)
            continue

        # Copy (don't symlink) so vcf_database is a standalone, permanent store.
        shutil.copy2(chosen_vcf, target)
        if sample in force_set and not passed:
            force_added.append(sample)
        else:
            auto_added.append(sample)

    total = len(list(vcfs_dir.glob("*_zc.vcf"))) + len(list(vcfs_dir.glob("*_zc.vcf.gz")))
    return {
        "auto_added": auto_added,
        "force_added": force_added,
        "already_present": already_present,
        "no_vcf": no_vcf,
        "excluded_skipped": excluded_skipped,
        "total": total,
    }


@app.post("/api/projects/{project}/step2/setup")
def step2_setup(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    step1_dir = project_dir / "step1"
    step2_dir = vcf_db_dir(project_dir / "step2")
    step2_dir.mkdir(parents=True, exist_ok=True)
    # vcf_database is the cumulative, persistent store — do NOT wipe it, and do
    # NOT delete QC-excluded samples from it. We accumulate: add/refresh every
    # step1 VCF and preserve any VCF whose step1 source is gone (imported /
    # historical runs). QC exclusions are applied at RUN time (step2_run's
    # -remove_by_name), so an excluded sample stays in the DB (part of the total)
    # but is left out of the comparison. The response reports total / comparison /
    # excluded so the UI shows the breakdown honestly (total = comparison + excluded).

    # Excluded count for the setup breakdown = tier A (reference blocklist) ∪
    # tier B (Step 1 exclusions). Both are always filtered out of the comparison
    # at run time. Tier C (build-list) is per-run and shown live in the UI.
    excluded_names: set = set(_read_step1_exclusions(project_dir / "step2")) | set(
        _reference_blocklist_names(cfg, _project_reference(project_dir))
    )

    count = 0
    edited_samples = []
    step1_samples: set[str] = set()
    for sample_dir in sorted(step1_dir.glob("*")):
        if not sample_dir.is_dir() or sample_dir.name.startswith(("_", ".")):
            continue
        sample = sample_dir.name
        step1_samples.add(sample)
        # Every step1 sample goes into the cumulative DB, including QC-excluded
        # ones — they're filtered out at run time, not deleted here. Legacy
        # plain-alignment/ samples (pre-GUI runs) contribute their VCFs too.
        vcf_candidates = sorted(_align_glob(sample_dir, "*_zc.vcf*"), key=lambda p: p.stat().st_mtime)
        if not vcf_candidates:
            continue
        source_vcf = vcf_candidates[-1]
        patched_vcf = _find_patched_vcf(sample_dir, sample, source_vcf)
        chosen_vcf = patched_vcf or source_vcf
        target_name = _target_name_for_vcf(source_vcf, chosen_vcf)
        target = step2_dir / target_name
        if patched_vcf:
            edited_samples.append(sample)
        if target.exists() or target.is_symlink():
            # Upgrade a legacy (or broken) symlink entry to a durable real copy;
            # leave an existing real file in place (accumulate, never clobber).
            if target.is_symlink():
                target.unlink()
                shutil.copy2(chosen_vcf, target)
                count += 1
            continue
        # Copy (don't symlink) so vcf_database is a standalone, permanent store.
        shutil.copy2(chosen_vcf, target)
        count += 1

    # Rebuild the manifest from the FINAL database contents so preserved /
    # imported VCFs (those with no step1 source) are recorded too, not just
    # the freshly linked ones.
    preserved = 0
    excluded = 0
    manifest_path = step2_dir / ".vcf_source_manifest.csv"
    with manifest_path.open("w", encoding="utf-8") as manifest:
        manifest.write("filename,source_type,source_path\n")
        for vcf in sorted([*step2_dir.glob("*_zc.vcf"), *step2_dir.glob("*_zc.vcf.gz")], key=lambda p: p.name):
            stem = vcf.name.replace("_zc.vcf.gz", "").replace("_zc.vcf", "")
            try:
                resolved = vcf.resolve()
            except OSError:
                resolved = vcf
            if stem in step1_samples:
                source_type = "step1"
            else:
                source_type = "imported"
                preserved += 1
            if stem in excluded_names:
                excluded += 1
            manifest.write(f"{vcf.name},{source_type},{resolved}\n")
    _write_step2_edit_summary(step2_dir.parent, edited_samples)
    total = len(list(step2_dir.glob("*_zc.vcf"))) + len(list(step2_dir.glob("*_zc.vcf.gz")))
    # total = every VCF in the cumulative DB; comparison = what Step 2 actually
    # compares (total minus QC-excluded); excluded = QC-excluded but still in DB.
    return {
        "linked": count,
        "total": total,
        "comparison": total - excluded,
        "excluded": excluded,
        "edited": len(set(edited_samples)),
        # kept for back-compat with older callers:
        "skipped_excluded": excluded,
        "preserved": preserved,
    }


# --- Step 2 concurrency control ---------------------------------------------
# Two layers:
#   1. A per-project lock (restart-proof) so a project can never have two Step 2
#      runs against its shared vcf_database at once — the in-memory JobManager
#      guard alone fails open across a backend restart.
#   2. A global cap + queue across projects (VSNP3_STEP2_MAX_CONCURRENT, default
#      1): excess runs wait in status "queued" and start automatically as slots
#      free, rather than all piling onto the CPU at once.

def _step2_max_concurrent() -> int:
    """Max Step 2 runs allowed to run concurrently across all projects. Excess
    runs queue. Configurable via VSNP3_STEP2_MAX_CONCURRENT (default 1). Note
    each run can spawn a large multiprocessing pool + RAxML -T 4, so raise this
    only alongside a lower VSNP3_MAX_CPUS."""
    try:
        n = int(os.environ.get("VSNP3_STEP2_MAX_CONCURRENT", "1"))
    except ValueError:
        n = 1
    return max(1, n)


def _pgid_alive(pgid: int) -> bool:
    """True if a process group still exists (used to detect a Step 2 run that
    was orphaned by a backend restart but is still running on the box)."""
    try:
        os.killpg(pgid, 0)
        return True
    except ProcessLookupError:
        return False
    except PermissionError:
        # Exists but owned by another uid — treat as alive (conservative).
        return True


def _step2_active_job(step2_dir: Path) -> Optional[str]:
    """Return the job id of an active (running or queued) Step 2 run for this
    project, or None. Authoritative when the backend that launched it is still
    up (JobManager knows the job); falls back to an OS process-group liveness
    check via the recorded .step2_pgid so a run orphaned by a backend restart
    still blocks a second dispatch against the same vcf_database."""
    jid_path = step2_dir / ".step2_job_id"
    if not jid_path.exists():
        return None
    try:
        job_id = jid_path.read_text(encoding="utf-8").strip()
    except OSError:
        return None
    if not job_id:
        return None
    job = job_manager.get_job(job_id)
    if job and job.get("status") in ("running", "queued"):
        return job_id
    # Backend no longer tracks it (likely restarted). Is the process tree still
    # alive on this host?
    pgid_path = step2_dir / ".step2_pgid"
    if pgid_path.exists():
        try:
            pgid = int(pgid_path.read_text(encoding="utf-8").strip())
        except (OSError, ValueError):
            pgid = 0
        if pgid and _pgid_alive(pgid):
            return job_id
    return None


@app.post("/api/projects/{project}/step2/run")
def step2_run(project: str, payload: Step2Request):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    # Auto-populate reference: project.json first, then reference_lock inference
    if not payload.reference:
        payload.reference = _project_reference(project_dir)
    lock = reference_lock(project)
    refs = lock["references"]
    if len(refs) > 1:
        # Name the samples behind each reference — "which runs do I split out?"
        # is the question this error creates, so answer it in the error.
        by_ref = lock.get("samples_by_reference", {})
        parts = []
        for r in refs:
            samples = sorted(by_ref.get(r, []))
            shown = ", ".join(samples[:4]) + (f", +{len(samples) - 4} more" if len(samples) > 4 else "")
            parts.append(f"{r} ({shown})" if shown else r)
        raise HTTPException(status_code=400, detail=f"Mixed references detected: {'; '.join(parts)}")
    if not payload.reference:
        if len(refs) == 1:
            payload.reference = refs[0]
        else:
            raise HTTPException(status_code=400, detail="Reference type is required for Step 2")
    if len(refs) == 1 and payload.reference != refs[0]:
        raise HTTPException(status_code=400, detail=f"Reference mismatch: expected {refs[0]}")
    # Store the resolved reference back into project.json
    update_project_meta(project_dir, {"reference": payload.reference})

    step2_dir = project_dir / "step2"
    step2_dir.mkdir(parents=True, exist_ok=True)

    # Concurrency guard: at most one Step 2 run per project (restart-proof —
    # checks the JobManager and, if the backend was restarted, the OS liveness
    # of any orphaned run). Prevents two runs clobbering the same vcf_database.
    job_id_path = step2_dir / ".step2_job_id"
    active_id = _step2_active_job(step2_dir)
    if active_id:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Step 2 is already running or queued for this project (job {active_id}). "
                "Wait for it to finish, or Stop it, before starting a new run."
            ),
        )

    vcf_source_dir = vcf_db_dir(step2_dir)

    # Timestamped run directory — each run gets its own subdirectory directly
    # under step2/ (e.g. step2/2026-06-05_13-11-21) so multiple comparisons
    # accumulate without overwriting previous outputs.
    from datetime import datetime as _dt
    run_ts = _dt.now().strftime("%Y-%m-%d_%H-%M-%S")
    run_dir = step2_dir / run_ts
    run_dir.mkdir(parents=True, exist_ok=True)

    # Effective removal set = Step 1 QC exclusions (remove_from_analysis.xlsx)
    # ∪ Step 2 build-list exclusions (.step2_build_excluded.json) ∪ the
    # authoritative set the UI sends in this request. The payload set is what
    # closes the silent-miss bug: previously the run trusted only the on-disk
    # files, which are written by a debounced best-effort save that Run didn't
    # wait for — so a run could analyze everything while the UI showed N
    # excluded. When the UI sends `exclude`, we also persist it so the on-disk
    # build-exclusions stay consistent for later loads.
    # Tiers (see the exclusion-tiers helpers), unioned with the authoritative
    # sets the UI sends (so a debounced save that hadn't flushed can't cause a
    # silent miss):
    #   A reference blocklist  — always excluded.
    #   B Step 1 exclusions    — EXEMPTED for accessions available from a
    #     reference panel (external reference VCFs, not Step 1 samples).
    #   C build-list exclusions — explicit per-run; never exempted.
    # Computed BEFORE staging so the copy loop can skip what the run drops.
    def _clean(xs):
        return {str(s).strip() for s in (xs or []) if str(s).strip()}
    ref_block = set(_reference_blocklist_names(cfg, payload.reference))
    panel_accessions = _reference_panel_accessions(cfg, payload.reference)
    step1_names = (set(_read_step1_exclusions(step2_dir)) | _clean(payload.step1_exclude)) - panel_accessions
    build_names = (
        set(_read_step2_build_exclusions(step2_dir))
        | _clean(payload.build_exclude)
        | _clean(payload.exclude)  # deprecated merged field -> treat as build (not exempted)
    )
    effective_removals = sorted(ref_block | step1_names | build_names)

    # Copy the VCFs out of the persistent database into this dated run folder and
    # run vsnp3 against the COPIES, never the database itself. vsnp3_step2.py
    # deletes every VCF out of its -wd after ingesting them (they survive only
    # inside the vcf_starting_files zip it writes). If -wd pointed at
    # step2/vcf_database, each run would empty the cumulative collection. The DB
    # is the ongoing store for the project; only the dated folder is disposable.
    #
    # Only the VCFs this run will analyze are staged: vsnp3 parses every VCF in
    # its -wd BEFORE applying -remove_by_name, so staging the whole database
    # made a 10-sample comparison copy AND parse all 9,372 VCFs of the big
    # Ames project. stage_step2_vcfs skips exactly what vsnp3's removal would
    # drop; -remove_by_name is still passed below, so a skipped-vs-removed
    # disagreement can only cost time, never correctness.
    try:
        copied_vcfs, skipped_excluded, staged_vcf_names = stage_step2_vcfs(
            vcf_source_dir, run_dir, effective_removals,
        )
    except OSError as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to stage VCFs into the run folder: {exc}",
        )
    if not copied_vcfs:
        if skipped_excluded:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"All {skipped_excluded} VCFs in step2/vcf_database are excluded "
                    "by the current selection/exclusions — nothing to compare. "
                    "Tick a source or list at least one sample that is in the set."
                ),
            )
        raise HTTPException(
            status_code=400,
            detail=(
                "No VCFs in step2/vcf_database to compare. Run Step 1 and collect "
                "VCFs (or import VCFs) before starting Step 2."
            ),
        )

    remove_arg = ""
    if effective_removals:
        remove_file = run_dir / "remove_by_name.xlsx"
        try:
            import pandas as pd  # vsnp3 env
            pd.DataFrame(effective_removals).to_excel(remove_file, header=False, index=False)
            remove_arg = f" -remove_by_name {shlex.quote(str(remove_file))}"
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Failed to build removal list: {exc}")
    _write_step2_edit_summary(run_dir, _edited_samples_in_dir(vcf_source_dir))
    _write_figtree_groups(run_dir, vcf_source_dir, cfg, payload.label_style or "short")
    # Build Step 2 command with options
    step2_flags = []
    if payload.all_vcf:
        step2_flags.append("-a")
    if payload.no_filters:
        step2_flags.append("-n")
    if payload.find_new_filters:
        step2_flags.append("-i")
    if payload.hash_groups:
        step2_flags.append("-hash")
    if payload.show_groups:
        step2_flags.append("--show_groups")
    if payload.html_tree:
        step2_flags.append("-html_tree")
    if payload.dp:
        step2_flags.append("-dp")
    if payload.qual_threshold is not None and payload.qual_threshold != 150:
        step2_flags.append(f"-w {payload.qual_threshold}")
    if payload.n_threshold is not None and payload.n_threshold != 50:
        step2_flags.append(f"-x {payload.n_threshold}")
    if payload.mq_threshold is not None and payload.mq_threshold != 56:
        step2_flags.append(f"-y {payload.mq_threshold}")
    if payload.density_threshold is not None:
        step2_flags.append(f"--density_threshold {payload.density_threshold}")
    if payload.density_window is not None:
        step2_flags.append(f"--density_window {payload.density_window}")
    flags_str = " ".join(step2_flags)
    # -wd is the dated run folder holding the staged COPIES (see the copy loop
    # above), not step2/vcf_database — vsnp3 removes VCFs from its -wd, so the
    # cumulative database must never be handed to it directly.
    cmd = f"vsnp3_step2.py -wd {shlex.quote(str(run_dir))} {flags_str} -t {payload.reference}{remove_arg}"
    label_style = payload.label_style or "short"
    label_script = _build_tree_label_script(run_dir, cfg, label_style)
    if label_script:
        # The tree-tip re-labeling is cosmetic (FigTree labels/colors). It must
        # NOT decide the run's success: vsnp3 has already written the trees and
        # tables (and, per the vsnp3 per-group isolation patch, completes even
        # when individual groups are skipped). Keep vsnp3's own exit code
        # authoritative — run the label step only if vsnp3 succeeded (&&), but
        # swallow a label-step failure (|| echo …) so a cosmetic hiccup can't
        # mark a valid run "failed". The warning still lands in the job log.
        cmd = (
            f"{cmd} && {{ python {shlex.quote(str(label_script))} "
            f"|| echo 'WARN: tree tip re-labeling failed; trees/tables are valid and complete'; }}"
        )
    step2_env = build_env(cfg)
    if payload.bootstrap and payload.bootstrap > 0:
        step2_env["VSNP3_BOOTSTRAP"] = str(int(payload.bootstrap))

    # T-07: provenance dispatch. Writes pipeline_run record (linking step1
    # samples) and step2 run_metadata.json with frozen dispatch_state.
    # On shared projects, refuses to dispatch if any step1 sample is still
    # running (HTTP 409). On personal projects, warn-and-proceed via
    # consumed_step1_run_ids_complete: false in the pipeline_run record.
    #
    # A capture failure (DispatchFailed) must NOT stop the analysis: the
    # record describes the run, it is not a precondition for it. Failing here
    # used to 500 before the job even started — no Live Logs, nothing to
    # debug — over metadata. Warn loudly, tell the frontend, run anyway.
    prov_step2_run_id = None
    provenance_warning = ""
    try:
        prov_step2_run_id, _prov_pipeline_run_id = provenance_writer.dispatch_step2(
            cfg, project_dir, payload.reference,
            cli_command=cmd, cli_flags=step2_flags,
            user=_current_user(),
            ood_session_id=_ood_session_id(),
            is_shared=_is_shared_project(cfg, project_dir),
            resolved_vcf_db_folders=_resolved_vcf_db_folders(cfg),
            step2_run_dir=run_dir,
            staged_vcf_names=staged_vcf_names,
        )
    except provenance_writer.Step2DispatchBlocked as e:
        raise HTTPException(status_code=409, detail=str(e))
    except provenance_writer.DispatchFailed as e:
        provenance_warning = (
            f"Run-metadata (provenance) was not recorded for this run: {e}"
        )
        logger.warning("Step 2 provenance dispatch failed; run proceeds without "
                       "run_metadata.json: %s", e)

    pgid_path = step2_dir / ".step2_pgid"

    def prov_finalize_cb(job_id, exit_code, started_at, finished_at):
        # Release the per-project lock so a finished/cancelled run frees the
        # project for the next dispatch. Best-effort; the .step2_job_id guard
        # (JobManager status) also stops treating it as active once terminal.
        for p in (pgid_path,):
            try:
                p.unlink()
            except OSError:
                pass
        # No dispatch record to finalize when dispatch itself failed (the run
        # was allowed to proceed provenance-less) — the lock release above
        # must still happen either way.
        if prov_step2_run_id is not None:
            provenance_writer.finalize_step2(
                project_dir, prov_step2_run_id, exit_code, started_at, finished_at,
                step2_run_dir=run_dir,
            )

    def on_start_cb(job_id, pid):
        # Record the OS process-group id (== pid, via start_new_session) so the
        # per-project guard can detect a run orphaned by a backend restart.
        try:
            pgid_path.write_text(str(os.getpgid(pid)), encoding="utf-8")
        except OSError:
            pass

    # Provenance: drop a plain-text record of the exact step2 command into the
    # comparison folder (step2/<timestamp>/.provenance/), so the run folder
    # plainly shows what was executed — the companion to step1's per-sample file.
    provenance_writer.append_run_command(
        run_dir, "vsnp_gui_step2_run_cmd.txt",
        title="vsnp_gui - Step 2 (comparison / phylogeny)",
        command=cmd,
        working_dir=str(run_dir),
        tool="vsnp3_step2.py",
        extra_lines=[("reference", payload.reference)],
    )

    job_id = job_manager.start_job(
        name="step2",
        command=wrap_cmd(cfg, cmd),
        cwd=run_dir,
        env=step2_env,
        finalize_callback=prov_finalize_cb,
        category="step2",
        max_concurrent=_step2_max_concurrent(),
        on_start=on_start_cb,
    )
    # Record the active job and the current run so the frontend can auto-select
    # it (and so the per-project guard survives a page reload / backend restart).
    job_id_path.write_text(job_id, encoding="utf-8")
    (step2_dir / ".current_run").write_text(run_ts, encoding="utf-8")
    # Snapshot the initial status (may already be "queued" if the global cap is
    # full) so the GUI can show Queued vs Running immediately.
    snap = job_manager.get_job(job_id) or {}
    # Report what will actually be excluded/compared so the UI can confirm the
    # run matches intent (and warn if it doesn't) instead of trusting a caption.
    try:
        vcf_total = sum(1 for p in vcf_source_dir.iterdir()
                        if p.name.endswith((".vcf", ".vcf.gz")))
    except OSError:
        vcf_total = 0
    excluded_count = len(effective_removals)
    # How many Step 1 exclusions were overridden because the accession is in a
    # reference panel (so the UI can explain the count).
    panel_exempt_count = len(
        (set(_read_step1_exclusions(step2_dir)) | {str(s).strip() for s in (payload.step1_exclude or []) if str(s).strip()})
        & panel_accessions
    )
    return {
        "job_id": job_id,
        "run_id": run_ts,
        "status": snap.get("status", "running"),
        "excluded_count": excluded_count,
        "blocklist_count": len(ref_block),
        "panel_exempt_count": panel_exempt_count,
        "vcf_total": vcf_total,
        # Exactly what was staged for vsnp3 — not vcf_total minus removal
        # NAMES, which overcounted when stale exclusion names matched nothing.
        "comparison_count": copied_vcfs,
        # Non-empty when the run started WITHOUT a provenance record (T-07
        # dispatch failed). The run itself is unaffected; the UI shows this
        # as a warning note, never as a blocking error.
        "provenance_warning": provenance_warning,
    }


@app.get("/api/jobs")
def jobs_list():
    """Suite-management contract: expose active/queued jobs conservatively.

    The consolidated dashboard checks this endpoint before restart, shutdown,
    or tool updates so it never orphans a vSNP analysis.
    """
    return job_manager.list_jobs()


@app.get("/api/jobs/{job_id}")
def job_status(job_id: str):
    job = job_manager.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@app.get("/api/projects/{project}/step2/active")
def step2_active(project: str):
    """Return the active (running or queued) Step 2 job for this project, if any.

    The Step 2 job id normally lives only in browser state (set at launch), so a
    page reload mid-run loses the Run/Stop UI. This lets the GUI rehydrate that
    state from the server (the source of truth). `controllable` is true when the
    JobManager still owns the job (so Stop works); false for a run orphaned by a
    backend restart (still alive on the box, but not stoppable via the API).
    Returns {"job_id": None} when nothing is active.
    """
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    job_id = _step2_active_job(step2_dir)
    if not job_id:
        return {"job_id": None}
    job = job_manager.get_job(job_id)
    controllable = bool(job and job.get("status") in ("running", "queued"))
    status = job.get("status") if job else "running"
    run_id = ""
    cur = step2_dir / ".current_run"
    if cur.exists():
        try:
            run_id = cur.read_text(encoding="utf-8").strip()
        except OSError:
            run_id = ""
    return {"job_id": job_id, "run_id": run_id, "status": status, "controllable": controllable}


@app.post("/api/jobs/{job_id}/stop")
def job_stop(job_id: str):
    """Cancel a running job. Signals the whole process group (SIGTERM, then
    SIGKILL after a grace period) so a step1 batch and every vsnp3 worker it
    spawned are torn down together. Samples already finished keep their outputs;
    in-flight samples are left partial and report as unknown/error on the next
    status poll. Returns 404 for an unknown job, 409 if it is not running."""
    job = job_manager.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if not job_manager.stop_job(job_id):
        raise HTTPException(status_code=409, detail="Job is not running")
    return {"job_id": job_id, "stopped": True}


@app.get("/api/jobs/{job_id}/events")
def job_events(job_id: str):
    job = job_manager.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    log_path = Path(job["log_path"])
    # T-05: for step1 batch jobs, also tail every per-sample run_step1.log so
    # the GUI shows live bwa/samtools/etc output during the long part of the
    # run, not just the bash-batch coordination noise. Other job types (step2,
    # SRA, genome download) keep the original single-log behavior.
    is_step1 = job.get("name") == "step1"
    step1_dir = Path(job["cwd"]) if is_step1 and job.get("cwd") else None
    batch_prefix = "[batch] " if is_step1 else ""

    def event_stream():
        offsets: Dict[Path, int] = {}
        sample_offsets: Dict[Path, int] = {}

        def discover_step1_samples() -> None:
            if not (is_step1 and step1_dir and step1_dir.is_dir()):
                return
            try:
                children = sorted(step1_dir.iterdir())
            except OSError:
                return
            for child in children:
                if not child.is_dir():
                    continue
                slog = child / "run_step1.log"
                if slog.exists() and slog not in sample_offsets:
                    sample_offsets[slog] = 0

        def emit_new(path: Path, offset_dict: Dict[Path, int], prefix: str):
            try:
                size = path.stat().st_size
            except OSError:
                return
            last = offset_dict.get(path, 0)
            if size <= last:
                return
            try:
                with open(path, "r", encoding="utf-8", errors="replace") as f:
                    f.seek(last)
                    chunk = f.read(size - last)
            except OSError:
                return
            offset_dict[path] = size
            for line in chunk.splitlines():
                yield f"data: {prefix}{line}\n\n"

        def drain_all():
            if log_path.exists():
                yield from emit_new(log_path, offsets, batch_prefix)
            if is_step1:
                discover_step1_samples()
                for slog in sorted(sample_offsets.keys()):
                    sample = slog.parent.name
                    yield from emit_new(slog, sample_offsets, f"[{sample}] ")

        while True:
            yield from drain_all()
            j = job_manager.get_job(job_id)
            if j and j["status"] in {"succeeded", "failed", "cancelled"}:
                # Catch anything written between the last poll and process exit.
                yield from drain_all()
                yield f"data: [job:{j['status']}]\n\n"
                break
            time.sleep(0.5)

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.get("/api/preflight")
def preflight(debug: bool = Query(False)):
    cfg = load_config()
    vsnp3_path = cfg.get("vsnp3_path", "").strip()
    if not vsnp3_path:
        raise HTTPException(status_code=400, detail="vSNP3 path is not set")
    vsnp3_dir = Path(vsnp3_path)
    if not vsnp3_dir.is_dir():
        raise HTTPException(status_code=400, detail=f"vSNP3 path not found: {vsnp3_path}")
    check_code = (
        "import importlib.util, json; "
        "mods=['pandas','Bio']; "
        "missing=[m for m in mods if importlib.util.find_spec(m) is None]; "
        "issues=[]; "
        "print(json.dumps({'missing': missing, 'checked': mods, 'issues': issues}))"
    )
    debug_code = (
        "import importlib.util, json, sys, site; "
        "mods=['pandas','Bio']; "
        "result={m: bool(importlib.util.find_spec(m)) for m in mods}; "
        "print(json.dumps({"
        "'executable': sys.executable, "
        "'site': site.getsitepackages(), "
        "'modules': result"
        "}))"
    )
    code = debug_code if debug else check_code
    cmd_list = conda_python_cmd(cfg, code)
    result = subprocess.run(cmd_list, text=True, capture_output=True)
    if debug:
        return {
            "cmd": " ".join(cmd_list),
            "returncode": result.returncode,
            "stdout": result.stdout,
            "stderr": result.stderr,
        }
    if result.returncode != 0:
        raise HTTPException(status_code=500, detail=f"Preflight failed: {result.stderr.strip()}")
    try:
        return json.loads(result.stdout.strip())
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail=f"Preflight output parse failed: {result.stdout.strip()}")


# ---------------------------------------------------------------------------
# Step 1 Results (QC summary) scan orchestration.
#
# The scan itself lives in app/qc_scan.py and runs as a subprocess: it parses
# every <sample>/*_stats.xlsx (openpyxl, in parallel), keeps a per-file cache
# in <step1>/.qc_stats_cache.json so a revisit parses only what changed, and
# streams "P <done> <total>" progress lines. Here one background thread owns
# the scan per step1 dir, and /qc_summary answers instantly with either the
# finished rows (kept in memory per backend process) or a progress snapshot
# the frontend polls. The previous design — one blocking request that
# pd.read_excel'd every workbook, sequentially, on every visit — took 15+
# minutes on an 8000-sample project and died in the OOD proxy's ~60 s read
# timeout, so the pane showed "Loading..." forever with nothing arriving.
_QC_SCANNER = Path(__file__).resolve().parent / "qc_scan.py"
_QC_STATE_LOCK = threading.Lock()
_QC_SCANS: Dict[str, Dict[str, Any]] = {}  # str(step1_dir) -> scan state
# Finished row sets are tens of MB for the biggest projects; keep only the
# most recently used ones in memory. The on-disk cache makes a re-scan of an
# evicted project a matter of seconds, not minutes.
_QC_KEEP_READY = 4


def _qc_run_scanner(scan_dir: Path, direct: bool = False, progress_cb=None) -> List[Dict[str, Any]]:
    """Run one qc_scan.py subprocess to completion and return its rows.
    progress_cb (if given) receives (done, total) as the scan advances."""
    out_fd, out_path = tempfile.mkstemp(prefix="qc_scan_", suffix=".json")
    os.close(out_fd)
    cmd = [sys.executable, str(_QC_SCANNER), str(scan_dir), "--out", out_path]
    if direct:
        cmd.append("--direct")
    err_file = tempfile.TemporaryFile(mode="w+", encoding="utf-8", errors="replace")
    try:
        # stderr goes to a file, not a pipe: we block reading stdout for the
        # progress stream, and a stderr pipe filling up (openpyxl warns per
        # workbook) would deadlock the scan.
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=err_file, text=True
        )
        for line in proc.stdout:
            parts = line.split()
            if len(parts) == 3 and parts[0] == "P" and progress_cb is not None:
                try:
                    progress_cb(int(parts[1]), int(parts[2]))
                except ValueError:
                    pass
        rc = proc.wait()
        if rc != 0:
            err_file.seek(0)
            tail = "\n".join(err_file.read().strip().splitlines()[-8:])
            raise RuntimeError(tail or f"scanner exited {rc}")
        with open(out_path, encoding="utf-8") as fh:
            return json.load(fh).get("rows", [])
    finally:
        err_file.close()
        try:
            os.unlink(out_path)
        except OSError:
            pass


def _qc_evict_lru() -> None:
    """Drop finished row sets beyond the LRU budget. Call under the lock."""
    ready = [(k, s) for k, s in _QC_SCANS.items() if s.get("status") == "ready"]
    ready.sort(key=lambda kv: kv[1].get("used_at", 0.0))
    for k, _ in ready[: max(0, len(ready) - _QC_KEEP_READY)]:
        _QC_SCANS.pop(k, None)


def _qc_start_scan(step1_dir: Path) -> Dict[str, Any]:
    """Start the background scan for one step1 dir, or join the running one."""
    key = str(step1_dir)
    with _QC_STATE_LOCK:
        state = _QC_SCANS.get(key)
        if state is not None and state["status"] == "scanning":
            return state
        state = {
            "status": "scanning",
            "done": 0,
            "total": None,
            "rows": None,
            "error": "",
            "scanned_at": None,
            "used_at": time.time(),
        }
        _QC_SCANS[key] = state

    def _progress(done: int, total: int) -> None:
        state["done"], state["total"] = done, total

    def _worker() -> None:
        try:
            rows = _qc_run_scanner(step1_dir, progress_cb=_progress)
        except Exception as exc:
            state["error"] = str(exc)
            state["status"] = "error"
            return
        state["rows"] = rows
        state["scanned_at"] = time.strftime("%Y-%m-%dT%H:%M:%S")
        state["status"] = "ready"
        with _QC_STATE_LOCK:
            state["used_at"] = time.time()
            _qc_evict_lru()

    threading.Thread(
        target=_worker, daemon=True, name=f"qc-scan:{step1_dir.parent.name}"
    ).start()
    return state


def _qc_rows_blocking(step1_dir: Path) -> List[Dict[str, Any]]:
    """Rows for endpoints that need the finished set within one request
    (reference_lock, CSV/XLSX downloads). Instant when a scan has already
    completed; otherwise waits on one (fast when the on-disk cache is warm)."""
    key = str(step1_dir)
    state = _QC_SCANS.get(key)
    if state is None or state["status"] == "error":
        state = _qc_start_scan(step1_dir)
    while state["status"] == "scanning":
        time.sleep(0.3)
    if state["status"] == "error":
        detail = state["error"]
        with _QC_STATE_LOCK:
            _QC_SCANS.pop(key, None)  # next request retries
        raise HTTPException(status_code=500, detail=f"QC summary failed: {detail}")
    with _QC_STATE_LOCK:
        state["used_at"] = time.time()
    return state["rows"] or []


def _qc_apply_filters(
    rows: List[Dict[str, Any]],
    start: Optional[str],
    end: Optional[str],
    q: Optional[str],
) -> List[Dict[str, Any]]:
    """Date-range / name filter for the download endpoints — the same rule the
    old embedded scan applied via the QC_START / QC_END / QC_Q env vars."""
    start = (start or "").strip()
    end = (end or "").strip()
    q = (q or "").strip().lower()

    def _keep(row: Dict[str, Any]) -> bool:
        rd = str(row.get("_run_date") or "")[:10]
        if start and (not rd or rd < start):
            return False
        if end and (not rd or rd > end):
            return False
        if q and q not in str(row.get("_sample", "")).lower():
            return False
        return True

    return [r for r in rows if _keep(r)]


def _qc_table(rows: List[Dict[str, Any]]) -> Tuple[List[str], List[List[Any]]]:
    """Rows as a rectangular table for the CSV/XLSX exports: columns in
    first-seen order (what pandas.DataFrame produced), _file dropped,
    _run_date renamed run_date."""
    cols: List[str] = []
    for row in rows:
        for k in row:
            if k != "_file" and k not in cols:
                cols.append(k)
    header = ["run_date" if c == "_run_date" else c for c in cols]
    data = [[row.get(c) for c in cols] for row in rows]
    return header, data


@app.get("/api/projects/{project}/qc_summary")
def qc_summary(project: str, refresh: int = 0):
    """Step 1 Results rows. Non-blocking: while a scan runs this returns
    {"status": "scanning", "done", "total"} for the frontend to poll; when
    finished, {"status": "ready", "rows": [...]}. Pass refresh=1 to rescan a
    project whose rows are already in memory — new/changed stats files are
    parsed, everything unchanged comes from the per-file cache in seconds."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")
    key = str(step1_dir)
    state = _QC_SCANS.get(key)
    if state is None or state["status"] == "error" or (refresh and state["status"] == "ready"):
        if state is not None and state["status"] == "error":
            with _QC_STATE_LOCK:
                if _QC_SCANS.get(key) is state:
                    _QC_SCANS.pop(key, None)
        state = _qc_start_scan(step1_dir)
    if state["status"] == "scanning":
        return {"status": "scanning", "done": state["done"], "total": state["total"]}
    if state["status"] == "error":
        raise HTTPException(status_code=500, detail=f"QC summary failed: {state['error']}")
    with _QC_STATE_LOCK:
        state["used_at"] = time.time()
    # Annotate a copy: verdicts depend on thresholds the user can edit, and
    # the cached rows must stay pristine for the CSV/XLSX exports.
    rows = [dict(r) for r in state["rows"] or []]
    _annotate_qc_rows(rows, _resolve_qc_thresholds(cfg, project_dir))
    return {
        "status": "ready",
        "rows": rows,
        "scanned_at": state["scanned_at"],
        "count": len(rows),
    }


@app.get("/api/projects/{project}/qc_summary.csv")
def qc_summary_csv(
    project: str,
    start: Optional[str] = None,
    end: Optional[str] = None,
    q: Optional[str] = None,
):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")
    rows = _qc_apply_filters(_qc_rows_blocking(step1_dir), start, end, q)
    if not rows:
        return Response(content="", media_type="text/csv")
    header, data = _qc_table(rows)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for row in data:
        writer.writerow(["" if v is None else v for v in row])
    return Response(content=buf.getvalue(), media_type="text/csv")


@app.get("/api/projects/{project}/qc_summary.xlsx")
def qc_summary_xlsx(
    project: str,
    start: Optional[str] = None,
    end: Optional[str] = None,
    q: Optional[str] = None,
):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")
    rows = _qc_apply_filters(_qc_rows_blocking(step1_dir), start, end, q)
    header, data = _qc_table(rows)
    from openpyxl import Workbook  # ships with the env's pandas

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(header)
    for row in data:
        ws.append(
            [v if v is None or isinstance(v, (int, float, str, bool)) else str(v) for v in row]
        )
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    # Kept alongside the project like before; best-effort on read-only trees.
    try:
        (step1_dir / "combined_excelworksheets.xlsx").write_bytes(content)
    except OSError:
        pass
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/posthoc/step1/scan")
def posthoc_step1_scan(payload: PosthocScanRequest):
    """Scan arbitrary Step-1-style folders for stats rows. Uses the same
    scanner (and per-folder .qc_stats_cache.json) as the Results pane, so
    re-loading a folder parses only new/changed workbooks."""
    cfg = load_config()
    folders = [Path(p).expanduser() for p in payload.folders]
    if not folders:
        return []
    rows: List[Dict[str, Any]] = []
    for folder in folders:
        if not folder.is_dir():
            continue
        try:
            folder_rows = _qc_run_scanner(folder, direct=True)
        except (RuntimeError, OSError, json.JSONDecodeError) as exc:
            raise HTTPException(status_code=500, detail=f"Post-hoc scan failed for {folder}: {exc}")
        for row in folder_rows:
            f = str(row.get("_file") or "")
            sample_dir = os.path.dirname(f)
            step1_dir = os.path.dirname(sample_dir)
            row["_sample_dir"] = sample_dir
            row["_step1_dir"] = step1_dir
            row["_project"] = os.path.basename(os.path.dirname(step1_dir))
            sample = row.get("_sample") or ""
            edits_dir = Path(sample_dir) / "vcf_edits"
            patched = ""
            if edits_dir.is_dir():
                candidates = sorted(
                    (c for c in edits_dir.glob("*.vcf*") if not c.name.endswith(".tbi")),
                    key=lambda p: p.stat().st_mtime,
                )
                if candidates:
                    patched = str(candidates[-1])
            edit_log = edits_dir / f"{sample}_patchlog.jsonl"
            row["_patched_vcf"] = patched
            row["_edit_log"] = str(edit_log) if edit_log.exists() else ""
            row["_edited"] = bool(patched) and edit_log.exists()
            rows.append(row)
    # Newest run per (project, sample) across every folder scanned.
    latest: Dict[tuple, Dict[str, Any]] = {}
    for row in rows:
        key = (row.get("_project"), row.get("_sample"))
        rd = row.get("_run_date", "") or ""
        if key not in latest or rd > (latest[key].get("_run_date", "") or ""):
            latest[key] = row
    # Post-hoc rows aggregate across many projects, so per-project overrides
    # don't apply cleanly here — annotate with cfg-resolved thresholds only.
    return _annotate_qc_rows(list(latest.values()), _resolve_qc_thresholds(cfg))


@app.post("/api/posthoc/step1/resolve_samples")
def posthoc_resolve_samples(payload: PosthocResolveRequest):
    cfg = load_config()
    samples = [s.strip() for s in payload.samples if s and s.strip()]
    if not samples:
        raise HTTPException(status_code=400, detail="Samples are required")
    roots = payload.roots or []
    if roots:
        root_dirs = [Path(r).expanduser() for r in roots]
    else:
        projects_root = Path(cfg["projects_root"])
        root_dirs = list(projects_root.glob("*/step1"))
    found = []
    missing = []
    for sample in samples:
        hit = None
        for root in root_dirs:
            candidate = root / sample
            if candidate.is_dir():
                hit = candidate
                break
        if hit:
            step1_dir = hit.parent
            found.append({
                "sample": sample,
                "sample_dir": str(hit),
                "step1_dir": str(step1_dir),
                "project": step1_dir.parent.name
            })
        else:
            missing.append(sample)
    return {"found": found, "missing": missing}


@app.get("/api/posthoc/tools")
def posthoc_tools():
    cfg = load_config()
    tool_bin = str(_tool_bin_dir(cfg) or "")
    tools = []
    for tool in posthoc_list_tools():
        status = posthoc_tool_status(tool, tool_bin)
        tools.append({
            "id": tool.tool_id,
            "label": tool.label,
            "description": tool.description,
            "requires": tool.requires,
            "outputs": tool.outputs,
            "available": status["available"],
            "missing": status["missing"],
            "requirements": status["requirements"],
        })
    return tools


@app.post("/api/projects/{project}/posthoc/run")
def posthoc_run(project: str, payload: PosthocRunRequest):
    cfg = load_config()
    tool = posthoc_get_tool(payload.tool)
    if not tool:
        raise HTTPException(status_code=404, detail="Unknown posthoc tool")
    status = posthoc_tool_status(tool, str(_tool_bin_dir(cfg) or ""))
    if not status["available"]:
        raise HTTPException(status_code=400, detail=f"Missing dependencies: {', '.join(status['missing'])}")
    project_dir = Path(cfg["projects_root"]) / project
    step2_dir = project_dir / "step2"
    # Groups live under a timestamped run dir (step2/<run_id>/<group>), so
    # resolve the run the group belongs to instead of assuming step2/<group>.
    output_dir = _resolve_step2_output_dir(step2_dir, payload.run_id)
    group_dir = output_dir / payload.group
    if not group_dir.exists():
        raise HTTPException(status_code=404, detail=f"Group not found: {payload.group}")
    posthoc_dir = group_dir / "posthoc"
    posthoc_dir.mkdir(parents=True, exist_ok=True)
    lock_path = _posthoc_lock_path(group_dir, tool.tool_id)
    _posthoc_clear_stale_lock(lock_path)
    if lock_path.exists():
        raise HTTPException(status_code=409, detail="Posthoc job already running for this group")
    stats_path = posthoc_dir / "stats.json"
    scope = (payload.scope or "all").lower()
    if tool.tool_id == "snp_analysis":
        cmd = _posthoc_snp_analysis_command(
            group_dir,
            payload.group,
            posthoc_dir,
            str(_tool_bin_dir(cfg) or ""),
            scope,
        )
    else:
        cmd = _posthoc_stub_command(cfg, stats_path, payload.group, tool.tool_id)
    backend_root = Path(__file__).resolve().parent.parent
    job_id = job_manager.start_job(
        name=f"posthoc:{tool.tool_id}:{project}:{payload.group}",
        command=cmd,
        cwd=backend_root,
    )
    lock_path.write_text(job_id, encoding="utf-8")
    return {"job_id": job_id, "group": payload.group, "tool": tool.tool_id, "outputs": tool.outputs}


@app.get("/api/projects/{project}/posthoc/status")
def posthoc_status(project: str, group: str, tool: str = "snp_analysis", run_id: Optional[str] = Query(None)):
    cfg = load_config()
    tool_obj = posthoc_get_tool(tool)
    if not tool_obj:
        raise HTTPException(status_code=404, detail="Unknown posthoc tool")
    step2_dir = Path(cfg["projects_root"]) / project / "step2"
    output_dir = _resolve_step2_output_dir(step2_dir, run_id)
    group_dir = output_dir / group
    if not group_dir.exists():
        raise HTTPException(status_code=404, detail=f"Group not found: {group}")
    posthoc_dir = group_dir / "posthoc"
    lock_path = _posthoc_lock_path(group_dir, tool_obj.tool_id)
    _posthoc_clear_stale_lock(lock_path)
    running = lock_path.exists()
    outputs = []
    for rel in tool_obj.outputs:
        path = group_dir / rel
        outputs.append({"path": str(path), "exists": path.exists()})
    return {"running": running, "outputs": outputs}


@app.get("/api/projects/{project}/reference_lock")
def reference_lock(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")
    # Same rows the Results pane shows (newest run per sample), served from the
    # shared scan state/cache — this endpoint used to re-read every stats
    # workbook in a second full pass, doubling the pane's load time.
    raw_refs: Dict[str, List[str]] = {}
    for row in _qc_rows_blocking(step1_dir):
        ref = str(row.get("Reference") or "")
        ref = ref.replace(" Forced", "").replace(" by Best Reference", "").strip()
        if ref:
            raw_refs.setdefault(ref, []).append(row.get("_sample"))
    # The stats sheet records whatever `-r` the run was given. A GUI run records
    # the reference NAME, but a command-line run imported into the project may
    # have recorded a full FASTA path from another machine — which then surfaced
    # verbatim in "Mixed references detected" as a path that doesn't exist here.
    # Normalize path-like values to the reference they actually are (same alias
    # machinery the VCF importer uses: NC_006932-NC_006933.fasta →
    # Brucella_abortus1); a plain name passes through untouched.
    alias_map = _reference_alias_map(Path(cfg["vsnp3_path"]))
    by_ref: Dict[str, List[str]] = {}
    for raw, samples in raw_refs.items():
        looks_pathy = ("/" in raw or "\\" in raw
                       or raw.lower().endswith((".fasta", ".fa", ".fna", ".fas")))
        name = _normalize_reference(raw, alias_map) if looks_pathy else raw
        by_ref.setdefault(name or raw, []).extend(s for s in samples if s)
    refs = sorted(by_ref)
    if len(refs) == 1:
        update_project_meta(project_dir, {
            "reference": refs[0],
            "display_name": f"{project}_{refs[0]}"
        })
    return {"references": refs, "samples_by_reference": by_ref}


class ExcludeRequest(BaseModel):
    samples: List[str]


@app.get("/api/projects/{project}/step2/vcf_count")
def step2_vcf_count(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    vcf_source_dir = vcf_db_dir(project_dir / "step2")
    if not vcf_source_dir.exists():
        return {"count": 0}
    vcfs = list(vcf_source_dir.glob("*.vcf")) + list(vcf_source_dir.glob("*.vcf.gz"))
    edited_samples = set()
    for vcf in vcfs:
        if _vcf_is_edited(vcf):
            edited_samples.add(_sample_from_vcf(vcf))
    # Comparison breakdown, same rule as step2/setup: excluded = tier A (reference
    # blocklist) ∪ tier B (Step 1 Results exclusions); those stay in the DB but
    # are dropped from the Step 2 comparison at run time. comparison = total −
    # excluded. Surfaced so the Build panel shows what will actually be compared,
    # not just the raw stored count.
    reference = _project_reference(project_dir)
    excluded_names = set(_read_step1_exclusions(project_dir / "step2")) | set(
        _reference_blocklist_names(cfg, reference)
    )
    comparison_stems: List[str] = []
    excluded = 0
    for vcf in vcfs:
        stem = vcf.name.replace("_zc.vcf.gz", "").replace("_zc.vcf", "").replace(".vcf.gz", "").replace(".vcf", "")
        if stem in excluded_names:
            excluded += 1
        else:
            comparison_stems.append(stem)
    total = len(vcfs)
    # Composition of the comparison set by source database. Each comparison-set
    # VCF is assigned to the FIRST enabled panel whose accession list contains
    # it; whatever no panel claims is attributed to this project's own
    # vcf_database samples. Priority bucketing keeps the buckets mutually
    # exclusive so the counts always sum to the comparison total, letting the
    # user see how many VCFs each database in use contributes.
    panels = _reference_panels_by_name(cfg, reference)
    panel_counts = {name: 0 for name, _ in panels}
    own = 0
    # Duplicate = a comparison sample whose ID (filename stem) is available from
    # more than one selected source — this project's own Step 1 samples and/or
    # one or more reference panels. Since vcf_database is a flat directory the
    # sample is physically present only once; this counts the cross-source
    # identity overlap so the user can see, at a glance, how many of the
    # comparison samples the reference databases share with each other or with
    # the project. It also explains why a panel's bucket above can be smaller
    # than the panel's raw size (shared IDs are attributed to one bucket only).
    step1_dir = project_dir / "step1"
    step1_set = set(_step1_sample_names(step1_dir)) if step1_dir.is_dir() else set()
    duplicates = 0
    for stem in comparison_stems:
        sources = sum(1 for _, accs in panels if stem in accs)
        if stem in step1_set:
            sources += 1
        if sources >= 2:
            duplicates += 1
        for name, accs in panels:
            if stem in accs:
                panel_counts[name] += 1
                break
        else:
            own += 1
    composition = [{"name": "vcf_database", "count": own}] + [
        {"name": name, "count": panel_counts[name]} for name, _ in panels
    ]
    return {
        "count": total,
        "path": str(vcf_source_dir),
        "edited_count": len(edited_samples),
        "total": total,
        "excluded": excluded,
        "comparison": total - excluded,
        "reference": reference or "",
        "composition": composition,
        "duplicates": duplicates,
    }


@app.get("/api/projects/{project}/step2/vcf_database/samples")
def step2_vcf_database_samples(project: str):
    """Return all sample names in the VCF database directory.

    The on-disk ``*_zc.vcf(.gz)`` files are the source of truth for what the
    collection contains: command-line / imported projects (and any VCFs copied
    straight into vcf_database/) may not be recorded in the manifest. Older
    versions read only the manifest, so those VCFs were invisible in the Step 2
    build list even though they sat in the database (and matched the card's
    "VCF DB" count). We now enumerate every VCF on disk and enrich it with the
    manifest's source_type/source_path when an entry exists — so the full
    collection lists exactly as if the GUI had produced it, while GUI-imported
    samples keep their provenance badge."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    vcf_source_dir = vcf_db_dir(project_dir / "step2")
    if not vcf_source_dir.exists():
        return []
    # Manifest (when present) carries provenance metadata keyed by filename.
    meta_by_filename: Dict[str, Dict] = {}
    manifest_path = vcf_source_dir / ".vcf_source_manifest.csv"
    if manifest_path.exists():
        with manifest_path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                fn = row.get("filename", "").strip()
                if fn and fn not in meta_by_filename:
                    meta_by_filename[fn] = {
                        "source_type": row.get("source_type", ""),
                        "source_path": row.get("source_path", ""),
                    }
    samples = []
    seen: set = set()
    for vcf in sorted(vcf_source_dir.glob("*_zc.vcf")) + sorted(vcf_source_dir.glob("*_zc.vcf.gz")):
        fn = vcf.name
        if fn in seen:
            continue
        seen.add(fn)
        meta = meta_by_filename.get(fn, {})
        samples.append({
            "filename": fn,
            "sample": fn.replace("_zc.vcf.gz", "").replace("_zc.vcf", ""),
            "source_type": meta.get("source_type", ""),
            "source_path": meta.get("source_path", str(vcf)),
        })
    samples.sort(key=lambda x: x["sample"].lower())
    return samples


@app.get("/api/projects/{project}/step2/runs")
def step2_runs_list(project: str):
    """List all timestamped step2 runs newest-first. Falls back to a synthetic
    'legacy' entry when group dirs sit directly under step2/ with no runs."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    if not step2_dir.exists():
        return []
    run_dirs = _step2_run_dirs(step2_dir)
    results = []
    if run_dirs:
        for run_id in sorted(run_dirs.keys(), reverse=True):
            run_entry = run_dirs[run_id]
            meta_path = run_entry / "run_metadata.json"
            started_at = None
            status = "unknown"
            reference = ""
            if meta_path.exists():
                try:
                    meta = json.loads(meta_path.read_text(encoding="utf-8"))
                    started_at = meta.get("started_at")
                    status = meta.get("status", "running")
                    reference = (
                        meta.get("dispatch_state", {})
                        .get("reference", {})
                        .get("name", "")
                    )
                except (json.JSONDecodeError, OSError):
                    pass
            else:
                # run_metadata.json missing → job is still running or very new
                status = "running"
            group_count = sum(1 for d in run_entry.iterdir() if d.is_dir() and not d.name.startswith("_"))
            results.append({
                "run_id": run_id,
                "started_at": started_at,
                "status": status,
                "reference": reference,
                "group_count": group_count,
            })
    else:
        # Legacy flat layout: group dirs directly under step2/ (no run dirs).
        def _is_group(d):
            return d.is_dir() and d.name not in _STEP2_NON_GROUP_DIRS and not d.name.startswith(".")
        groups = [d for d in step2_dir.iterdir() if _is_group(d)]
        if groups:
            results.append({
                "run_id": "legacy",
                "started_at": None,
                "status": "ok",
                "reference": "",
                "group_count": len(groups),
            })
    return results


@app.post("/api/projects/{project}/qc_exclude")
def qc_exclude(project: str, payload: ExcludeRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    step2_dir.mkdir(parents=True, exist_ok=True)
    # Tier B (Step 1 exclusions). Stored in step2/.step1_excluded.json — NOT
    # remove_from_analysis.xlsx, which is the reference-level blocklist (tier A).
    names = _write_step1_exclusions(step2_dir, payload.samples)
    return {"count": len(names)}


@app.get("/api/projects/{project}/qc_exclude")
def qc_exclude_get(project: str):
    """Return the persisted Step 1 exclusion set (tier B) so the GUI can hydrate
    the QC table / build list on load. Empty list = no exclusions."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    return {"samples": _read_step1_exclusions(project_dir / "step2")}


@app.get("/api/projects/{project}/step2/blocklist")
def step2_blocklist_get(project: str):
    """Tier A: the reference-level permanent blocklist for this project's
    reference. Names here are always excluded and locked in the build list."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    ref = _project_reference(project_dir) or ""
    return {"reference": ref, "samples": _reference_blocklist_names(cfg, ref)}


@app.get("/api/projects/{project}/step2/panel-accessions")
def step2_panel_accessions_get(project: str):
    """Accessions available from enabled reference VCF-db panels for this
    project's reference. The build list marks these so the user sees that a
    Step 1 exclusion of the same accession is overridden (the panel keeps it)."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    ref = _project_reference(project_dir) or ""
    return {"reference": ref, "samples": sorted(_reference_panel_accessions(cfg, ref))}


@app.get("/api/projects/{project}/step2/panels")
def step2_panels_get(project: str):
    """Every reference VCF-db panel that applies to this project's reference —
    checked AND unchecked — each with its full sample list.

    The Step 2 setup pane drives the comparison off these checkboxes: a checked
    panel's VCFs are added to the set, and an unchecked panel's samples are left
    OUT of the comparison even when an earlier build already copied them into
    vcf_database (the cumulative store is never pruned). Deciding what to leave
    out needs the sample list of a panel the user just unchecked, so — unlike
    _reference_panel_accessions / _reference_panels_by_name — this deliberately
    does not filter on `enabled`; the flag is returned instead."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    ref = _project_reference(project_dir) or ""
    panels = []
    if ref:
        for folder in _resolved_vcf_db_folders(cfg):
            fref = folder.get("reference", "") or ""
            # A folder tagged for another reference doesn't apply; an untagged
            # (legacy) folder is treated as applicable, same as the run-time
            # helpers above.
            if fref and not _refs_match(fref, ref, True):
                continue
            p = Path(folder.get("path", ""))
            if not p.is_dir():
                continue
            samples = sorted(
                f.name.replace("_zc.vcf.gz", "").replace("_zc.vcf", "")
                for f in list(p.glob("*_zc.vcf")) + list(p.glob("*_zc.vcf.gz"))
            )
            panels.append({
                "name": folder.get("name", p.name),
                "path": str(p),
                "reference": fref,
                "scope": folder.get("scope", ""),
                "enabled": bool(folder.get("enabled")),
                "sample_count": len(samples),
                "samples": samples,
            })
    return {"reference": ref, "panels": panels}


# --- Exclusion tiers --------------------------------------------------------
# Step 2 filters samples out of the comparison at run time via vsnp3
# -remove_by_name (vcf_database stays the cumulative store). Three independent
# tiers, each with its own store, unioned at run time:
#   A. Reference blocklist  — <ref>_remove_from_analysis.xlsx in the reference
#      dir. A permanent, per-reference dependency file: any name in it is NEVER
#      included in an analysis. Read-only here (edited via the reference editor);
#      shown locked in the build list.
#   B. Step 1 exclusions    — step2/.step1_excluded.json, set by the Step 1
#      Results "exclude" checkboxes. (Historically written to the project's
#      remove_from_analysis.xlsx, which collided in name/meaning with tier A —
#      migrated away below.)
#   C. Step 2 build-list    — step2/.step2_build_excluded.json, set by the
#      "Exclude" checkboxes in the Step 2 build list.

def _step1_exclusions_path(step2_dir: Path) -> Path:
    return step2_dir / ".step1_excluded.json"


def _read_step1_exclusions(step2_dir: Path) -> List[str]:
    p = _step1_exclusions_path(step2_dir)
    if not p.exists():
        # One-time migration from the legacy project-level remove_from_analysis
        # .xlsx (tier B used to live there, colliding with the tier-A reference
        # file of the same name). Move it into the JSON store and drop the xlsx.
        legacy = step2_dir / "remove_from_analysis.xlsx"
        if legacy.exists():
            names = sorted(set(_read_remove_xlsx_names(legacy)))
            try:
                if names:
                    p.write_text(json.dumps(names), encoding="utf-8")
                legacy.unlink()
            except OSError:
                pass
            return names
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return [str(s).strip() for s in data if str(s).strip()]
    except (json.JSONDecodeError, OSError):
        pass
    return []


def _write_step1_exclusions(step2_dir: Path, samples) -> List[str]:
    names = sorted({str(s).strip() for s in (samples or []) if str(s).strip()})
    p = _step1_exclusions_path(step2_dir)
    try:
        if names:
            p.write_text(json.dumps(names), encoding="utf-8")
        else:
            p.unlink(missing_ok=True)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Failed to write Step 1 exclusions: {exc}")
    return names


def _reference_blocklist_names(cfg: Dict, reference: Optional[str]) -> List[str]:
    """Tier A: names in the reference's *_remove_from_analysis.xlsx — a permanent
    per-reference blocklist, never included in any analysis. Read-only here."""
    if not reference:
        return []
    # Search EVERY registered reference root, not just the single configured
    # root: a reference can live in any of them (e.g. a shared /srv set the user
    # registered while the GUI's primary root holds only supplemental refs). The
    # configured root is included first; reference_roots() adds the rest from
    # reference_options_paths.txt. De-duped by resolved path.
    roots: List[Path] = []
    seen_roots: set = set()
    configured = str(cfg.get("vsnp3_reference_options_root", "") or "").strip()
    candidates = ([Path(configured)] if configured else []) + list(
        reference_roots(Path(str(cfg.get("vsnp3_path", "") or "")))
    )
    for r in candidates:
        try:
            key = str(r.resolve())
        except OSError:
            key = str(r)
        if key in seen_roots:
            continue
        seen_roots.add(key)
        roots.append(r)
    names: set = set()
    for root in roots:
        ref_dir = root / reference
        if not ref_dir.is_dir():
            continue
        for f in ref_dir.glob("*remove_from_analysis*.xlsx"):
            if f.name.startswith("~$"):
                continue
            names.update(_read_remove_xlsx_names(f))
    return sorted(names)


def _reference_panel_accessions(cfg: Dict, reference: Optional[str]) -> set:
    """Sample names available from ENABLED reference VCF-db panels for this
    reference. These are EXEMPT from Step 1 name exclusions: a curated panel VCF
    is an external reference file, not a Step 1 sample, so a Step 1 QC exclusion
    of the same accession must not drop it from the comparison. Re-derived at run
    time from the enabled db folders, so it survives later Step 1 re-exclusions.
    (Tier A blocklist and explicit build-list exclusions still apply.)"""
    if not reference:
        return set()
    names: set = set()
    for folder in _resolved_vcf_db_folders(cfg):
        if not folder.get("enabled"):
            continue
        fref = folder.get("reference", "") or ""
        # A folder tagged for another reference doesn't apply; an untagged
        # (legacy) folder the user enabled is treated as applicable.
        if fref and not _refs_match(fref, reference, True):
            continue
        p = Path(folder.get("path", ""))
        if not p.is_dir():
            continue
        for f in list(p.glob("*_zc.vcf")) + list(p.glob("*_zc.vcf.gz")):
            names.add(f.name.replace("_zc.vcf.gz", "").replace("_zc.vcf", ""))
    return names


def _reference_panels_by_name(cfg: Dict, reference: Optional[str]) -> List[tuple]:
    """Per-panel accession sets for the ENABLED reference VCF-db panels matching
    this reference, as an ordered list of (panel_name, {accession, ...}).

    Unlike _reference_panel_accessions (which merges every panel into one set),
    this keeps each panel separate so the Build panel can report how many of the
    comparison set's VCFs are accounted for by each database being used. The
    accession lists are read live from the panel folders — not from the project
    manifest, which loses origin-DB attribution once Collect copies VCFs into
    vcf_database and rebuilds the manifest from the destination folder."""
    panels: List[tuple] = []
    if not reference:
        return panels
    for folder in _resolved_vcf_db_folders(cfg):
        if not folder.get("enabled"):
            continue
        fref = folder.get("reference", "") or ""
        if fref and not _refs_match(fref, reference, True):
            continue
        p = Path(folder.get("path", ""))
        if not p.is_dir():
            continue
        accs = {
            f.name.replace("_zc.vcf.gz", "").replace("_zc.vcf", "")
            for f in list(p.glob("*_zc.vcf")) + list(p.glob("*_zc.vcf.gz"))
        }
        if accs:
            panels.append((folder.get("name", p.name), accs))
    return panels


# --- Step 2 build-list exclusions (tier C) ----------------------------------

def _step2_build_exclusions_path(step2_dir: Path) -> Path:
    return step2_dir / ".step2_build_excluded.json"


def _read_step2_build_exclusions(step2_dir: Path) -> List[str]:
    p = _step2_build_exclusions_path(step2_dir)
    if not p.exists():
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return [str(s).strip() for s in data if str(s).strip()]
    except (json.JSONDecodeError, OSError):
        pass
    return []


def _read_remove_xlsx_names(path: Path) -> List[str]:
    """Read sample names from a header-less single-column remove xlsx."""
    if not path.exists():
        return []
    try:
        import pandas as pd  # vsnp3 env
        df = pd.read_excel(path, header=None)
        return [
            str(s).strip()
            for s in df.iloc[:, 0].tolist()
            if str(s).strip() and str(s).strip().lower() != "nan"
        ]
    except Exception:
        return []


@app.get("/api/projects/{project}/step2/build-exclusions")
def step2_build_exclusions_get(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    return {"samples": _read_step2_build_exclusions(project_dir / "step2")}


@app.post("/api/projects/{project}/step2/build-exclusions")
def step2_build_exclusions_set(project: str, payload: ExcludeRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    step2_dir.mkdir(parents=True, exist_ok=True)
    samples = sorted({str(s).strip() for s in (payload.samples or []) if str(s).strip()})
    p = _step2_build_exclusions_path(step2_dir)
    if not samples:
        p.unlink(missing_ok=True)
        return {"samples": [], "count": 0}
    p.write_text(json.dumps(samples), encoding="utf-8")
    return {"samples": samples, "count": len(samples)}


def _parse_step2_groupings(html_text: str) -> tuple[Dict[str, List[str]], int]:
    """Parse the 'Groupings with N listed' table from a vSNP3 step2 summary.

    Each data row is `<td>SampleName</td>` followed by one `<td>` per group
    that sample belongs to. Returns ({group_name: [sample names]}, n_samples).
    The sample name is taken verbatim from the HTML — it usually carries
    metadata (e.g. a state code), which is what the Step 2 group search box
    matches against.

    Splits on ``<tr`` rather than matching ``<tr>...</tr>`` because vSNP3's
    summary uses a malformed ``<tr>`` (instead of ``</tr>``) to close its
    header row; splitting tolerates that.
    """
    import html as _html
    idx = html_text.find("Groupings with")
    if idx == -1:
        return {}, 0
    tbl_start = html_text.find("<table", idx)
    tbl_end = html_text.find("</table>", tbl_start) if tbl_start != -1 else -1
    if tbl_start == -1 or tbl_end == -1:
        return {}, 0
    section = html_text[tbl_start:tbl_end]
    groups: Dict[str, List[str]] = {}
    samples_seen: set = set()
    for chunk in re.split(r"<tr\b", section, flags=re.IGNORECASE):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", chunk, re.DOTALL | re.IGNORECASE)
        cells = [_html.unescape(re.sub(r"<[^>]+>", "", c)).strip() for c in cells]
        cells = [c for c in cells if c]
        if len(cells) < 2:
            continue
        sample = cells[0]
        samples_seen.add(sample)
        for g in cells[1:]:
            groups.setdefault(g, []).append(sample)
    return groups, len(samples_seen)


@app.get("/api/projects/{project}/step2/groupings")
def step2_groupings(project: str, run_id: Optional[str] = Query(None)):
    """Return {group_name: [sample names]} parsed from the run's
    vSNP_step2_summary-*.html. Powers the Step 2 Results group search."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    if not step2_dir.exists():
        return {"groups": {}, "summary_html": None, "sample_count": 0}
    output_dir = _resolve_step2_output_dir(step2_dir, run_id)
    summaries = sorted(output_dir.glob("vSNP_step2_summary-*.html"))
    if not summaries:
        return {"groups": {}, "summary_html": None, "sample_count": 0}
    summary = summaries[-1]
    try:
        html_text = summary.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Failed to read summary: {exc}")
    groups, sample_count = _parse_step2_groupings(html_text)
    # The summary's Groupings table carries RAW accessions (e.g. ERR1462610),
    # but the tree leaves the user sees are lineage-labeled (e.g. Caprae_ERR1462610)
    # via the vcf_refs label map. So a search for "Caprae" found nothing. Append
    # each sample's friendly label (and the "Label_accession" form) to the
    # group's member list as extra SEARCH TOKENS so the search matches what's on
    # the tree. (These lists are used only by the group-search box.)
    label_map = _load_vcf_label_map(cfg, "short")
    if label_map:
        for members in groups.values():
            tokens: List[str] = []
            for s in members:
                for ident, friendly in label_map.items():
                    if ident and friendly and ident in s:
                        tokens.append(friendly)
                        tokens.append(f"{friendly}_{s}")
            for t in tokens:
                if t not in members:
                    members.append(t)
    return {"groups": groups, "summary_html": summary.name, "sample_count": sample_count}


@app.post("/api/projects/{project}/step2/clear")
def step2_clear(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    vcf_source_dir = vcf_db_dir(step2_dir)
    if vcf_source_dir.exists():
        shutil.rmtree(vcf_source_dir)
    vcf_source_dir.mkdir(parents=True, exist_ok=True)
    edited_summary = step2_dir / "edited_samples.json"
    if edited_summary.exists():
        edited_summary.unlink()
    return {"cleared": True}


def _resolve_sample_dir(step1_dir: Path, sample: str) -> Optional[Path]:
    """Resolve sample name to its step1 subdirectory.

    Tries exact match first, then falls back to matching directories
    whose name starts with ``sample_`` (e.g. sample ``13-1941-6``
    matches directory ``13-1941-6_S4_L001``).
    """
    exact = step1_dir / sample
    if exact.is_dir():
        return exact
    candidates = sorted(
        d for d in step1_dir.iterdir()
        if d.is_dir() and d.name.startswith(f"{sample}_")
    )
    return candidates[0] if candidates else None


def _resolve_kraken_sample_dir(kraken_dir: Path, sample: str) -> Optional[Path]:
    """Resolve a step1 sample name to its Kraken output subdirectory.

    Unlike step1, Kraken's dir name can be either LONGER or SHORTER than the
    step1 sample name depending on which read-tag/lane suffix each tool
    stripped (and whether Kraken was run from its own GUI on the raw download
    fastqs, before step1 existed). So match in both directions:

      1. exact: kraken/<sample>
      2. kraken dir longer:  <sample>_*           (e.g. sample 13-1941-6
         matches kraken dir 13-1941-6_S4_L001)
      3. kraken dir shorter: sample == <dir>_*    (e.g. step1 sample
         13-1941-6_S4_L001 matches kraken dir 13-1941-6)

    For the shorter case, prefer the longest-matching dir name so we don't
    match ``S1`` to a sample named ``S10_...``.
    """
    exact = kraken_dir / sample
    if exact.is_dir():
        return exact
    try:
        dirs = [d for d in kraken_dir.iterdir() if d.is_dir() and not d.name.startswith(".")]
    except (OSError, PermissionError):
        return None
    # Kraken dir longer than the sample.
    longer = sorted(d for d in dirs if d.name.startswith(f"{sample}_"))
    if longer:
        return longer[0]
    # Kraken dir shorter than the sample (run on the bare/raw fastq name).
    shorter = [d for d in dirs if sample.startswith(f"{d.name}_")]
    if shorter:
        return max(shorter, key=lambda d: len(d.name))
    return None


@app.get("/api/projects/{project}/step1/files")
def step1_files(project: str, sample: str = Query(...)):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    sample_dir = _resolve_sample_dir(step1_dir, sample) if step1_dir.is_dir() else None
    if not sample_dir:
        # Imported-VCF case: sample lives only in step2/vcf_database/ (no Step 1
        # alignment, so no BAM). Still useful in IGV as a calls-only track —
        # anchor it to the project's reference so the user can compare the
        # variant positions against the local cohort.
        vcf_source_dir = vcf_db_dir(project_dir / "step2")
        imported_vcf = None
        if vcf_source_dir.is_dir():
            for suffix in ("_zc.vcf", "_zc.vcf.gz", ".vcf", ".vcf.gz"):
                cand = vcf_source_dir / f"{sample}{suffix}"
                if cand.exists():
                    imported_vcf = cand
                    break
        if imported_vcf is not None:
            ref_fasta, ref_gff = _project_reference_fasta_and_gff(project_dir, cfg)
            if not ref_fasta:
                # No way to anchor the VCF — surface as a clearly-distinct error.
                raise HTTPException(status_code=404, detail="imported_vcf_no_reference")
            return {
                "stats": "",
                "bam": "",
                "alignment_dir": "",
                "reference_fasta": ref_fasta,
                "reference_gff": ref_gff,
                "annotated_vcf": "",  # imports lack the rich ID-column annotation
                "sample_dir": str(vcf_source_dir),
                "source_vcf": str(imported_vcf),
                "patched_vcf": "",
                "edit_log": "",
                "edited": False,
                "kind": "imported_vcf",
            }
        raise HTTPException(status_code=404, detail="no_step1")
    stats_files = sorted(sample_dir.glob(f"{sample}_*_stats.xlsx"), key=lambda p: p.stat().st_mtime)
    stats_path = str(stats_files[-1]) if stats_files else ""
    bam_files = sorted(sample_dir.glob(f"**/{sample}_nodup.bam"), key=lambda p: p.stat().st_mtime)
    bam_path = str(bam_files[-1]) if bam_files else ""
    align_dir = str(bam_files[-1].parent) if bam_files else ""
    ref_fasta = ""
    ref_gff = ""
    annotated_vcf = ""
    if align_dir:
        fasta_files = sorted(Path(align_dir).glob("*.fasta"))
        if fasta_files:
            ref_fasta = str(fasta_files[0])
            # GFF lives in the reference options dir, not the alignment dir.
            vsnp3_path = Path(cfg.get("vsnp3_path", ""))
            gff_path = find_gff_for_fasta(fasta_files[0], vsnp3_path)
            if gff_path:
                ref_gff = str(gff_path)
        # The *_filtered_hapall_annotated.vcf holds per-variant annotation
        # in the ID column (gene/product/codon/AA change/mutation_type).
        # Surface its path so igv.js can render it as a variant track and
        # show the annotation on hover.
        ann_candidates = sorted(
            Path(align_dir).glob(f"{sample}_filtered_hapall_annotated.vcf*"),
            key=lambda p: p.stat().st_mtime,
        )
        if ann_candidates:
            annotated_vcf = str(ann_candidates[-1])
    vcf_candidates = sorted(sample_dir.glob(f"**/{sample}*zc.vcf*"), key=lambda p: p.stat().st_mtime)
    source_vcf = vcf_candidates[-1] if vcf_candidates else None
    patched_vcf = _find_patched_vcf(sample_dir, sample, source_vcf)
    edit_log = _edit_log_path(sample_dir, sample)
    return {
        "stats": stats_path,
        "bam": bam_path,
        "alignment_dir": align_dir,
        "reference_fasta": ref_fasta,
        "reference_gff": ref_gff,
        "annotated_vcf": annotated_vcf,
        "sample_dir": str(sample_dir),
        "source_vcf": str(source_vcf) if source_vcf else "",
        "patched_vcf": str(patched_vcf) if patched_vcf else "",
        "edit_log": str(edit_log) if edit_log.exists() else "",
        "edited": bool(patched_vcf) and edit_log.exists()
    }


def _latest_step1_stats(project: str, sample: str) -> Path:
    """Discover the latest *_stats.xlsx for a sample. Shared by the download
    and preview endpoints so they stay in sync on the resolution rules."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    sample_dir = _resolve_sample_dir(step1_dir, sample) if step1_dir.is_dir() else None
    if not sample_dir:
        raise HTTPException(status_code=404, detail="Sample not found")
    stats_files = sorted(sample_dir.glob(f"{sample}_*_stats.xlsx"), key=lambda p: p.stat().st_mtime)
    if not stats_files:
        raise HTTPException(status_code=404, detail="Stats file not found")
    return stats_files[-1]


@app.get("/api/projects/{project}/step1/samples/{sample}/stats/download")
def step1_sample_stats_download(project: str, sample: str):
    target = _latest_step1_stats(project, sample)
    return FileResponse(
        target,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{sample}_stats.xlsx",
    )


@app.get("/api/projects/{project}/step1/samples/{sample}/stats/preview", response_class=HTMLResponse)
def step1_sample_stats_preview(project: str, sample: str, download: int = 0):
    """Render the latest stats xlsx for a sample as a formatted HTML preview
    (cell colors + conditional formatting preserved via openpyxl). With
    ?download=1, returns the raw xlsx — used by the "Download xlsx" link
    inside the preview page so users can still grab the file for offline
    spreadsheet work without leaving the preview."""
    target = _latest_step1_stats(project, sample)
    if download:
        return FileResponse(
            target,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{sample}_stats.xlsx",
        )
    from app import xlsx_html
    try:
        html_page = xlsx_html.xlsx_to_html(target)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"xlsx render failed: {type(e).__name__}: {e}")
    return HTMLResponse(content=html_page)


@app.get("/api/projects/{project}/step1/samples")
def step1_samples(project: str):
    """List step1 sample directories for the inline project sample browser.

    Native GUI projects stage reads in download/ and get a step1/<sample>/
    dir when Step 1 runs; command-line / imported projects keep their reads
    inside step1/<sample>/ and never populate download/. Surfacing the step1
    dirs directly makes both layouts list the same samples under the project,
    so a command-line run appears exactly as if the GUI had produced it. The
    frontend merges this with any download-only (not-yet-run) FASTQ groups."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    step1_dir = project_dir / "step1"
    return {"samples": _step1_browser_samples(step1_dir)}


@app.get("/api/projects/{project}/step1/samples/{sample}/files")
def step1_sample_files(project: str, sample: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    sample_dir = _resolve_sample_dir(step1_dir, sample) if step1_dir.is_dir() else None
    if not sample_dir:
        raise HTTPException(status_code=404, detail="Sample not found")
    base = sample_dir.resolve()
    entries = []
    for path in sorted(base.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(".~lock"):
            continue
        try:
            rel = path.relative_to(base).as_posix()
            stat = path.stat()
        except (OSError, ValueError):
            continue
        entries.append({
            "name": path.name,
            "relpath": rel,
            "path": str(path),
            "size": stat.st_size,
            "mtime": stat.st_mtime,
            "type": path.suffix.lstrip(".").lower() or "file",
        })
    return {
        "project": project,
        "sample": sample,
        "sample_dir": str(base),
        "files": entries,
    }


@app.get("/api/projects/{project}/kraken/samples/{sample}/files")
def kraken_sample_files(project: str, sample: str):
    """List Kraken ID Parse outputs for a sample (cross-tool visibility).

    Kraken ID Parse GUI writes its per-sample results into
    ``<project>/kraken/<sample>/`` — the same project directory vSNP uses.
    Since both tools share /srv/kapurlab/projects, those files already sit
    on disk next to our step1 output; this endpoint simply surfaces them so
    a user looking at a sample in vSNP can see (and download, via the
    existing /download-file route) whatever Kraken produced for it.

    Returns ``present: false`` (not 404) when no Kraken run exists for the
    sample, so the UI can show a calm "no Kraken run yet" instead of an error.
    """
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    kraken_dir = project_dir / "kraken"
    # Kraken strips _R1/_R2 / _1/_2 read tags, so its sample dir name can be
    # longer OR shorter than the step1 sample name (especially when Kraken was
    # run from its own GUI on the raw download fastqs before step1). Match in
    # both directions so those earlier runs still surface here.
    sample_dir = _resolve_kraken_sample_dir(kraken_dir, sample) if kraken_dir.is_dir() else None
    if not sample_dir:
        return {
            "project": project,
            "sample": sample,
            "present": False,
            "sample_dir": "",
            "files": [],
        }
    base = sample_dir.resolve()
    entries = []
    for path in sorted(base.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith(".~lock"):
            continue
        try:
            rel = path.relative_to(base).as_posix()
            stat = path.stat()
        except (OSError, ValueError):
            continue
        ext = path.suffix.lower()
        entries.append({
            "name": path.name,
            "relpath": rel,
            "path": str(path),
            "size": stat.st_size,
            "mtime": stat.st_mtime,
            "type": path.suffix.lstrip(".").lower() or "file",
            # Browser-renderable (open in a tab) vs download-only. Krona/report
            # HTML, coverage PDFs, preview PNGs etc. open inline.
            "openable": ext in (".html", ".htm", ".pdf", ".png", ".jpg", ".jpeg",
                                 ".svg", ".txt", ".log", ".csv", ".json"),
        })
    # Surface the most useful artifacts first: report/krona HTML, then the rest.
    def _rank(e):
        n = e["name"].lower()
        if n.endswith("_krona.html") or n == "report.html":
            return 0
        if n.endswith(".html") or n.endswith(".pdf"):
            return 1
        return 2
    entries.sort(key=lambda e: (_rank(e), e["relpath"]))
    return {
        "project": project,
        "sample": sample,
        "present": True,
        "sample_dir": str(base),
        "files": entries,
    }


@app.get("/api/projects/{project}/kraken/samples/{sample}/krona")
def kraken_sample_krona(project: str, sample: str):
    """Open the interactive Krona chart for a sample's Kraken run.

    Resolves the sample's Kraken output dir (same matching rules as
    ``kraken_sample_files``) and serves its ``*_krona.html`` inline so it
    renders in a browser tab. The Krona graph is produced in every Kraken
    mode, so any sample with a Kraken dir has one. Returns 404 when no run
    exists or the run produced no Krona file (e.g. an interrupted run).
    """
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    kraken_dir = project_dir / "kraken"
    sample_dir = _resolve_kraken_sample_dir(kraken_dir, sample) if kraken_dir.is_dir() else None
    if not sample_dir:
        raise HTTPException(status_code=404, detail="No Kraken run for this sample")
    krona = next(
        (p for p in sorted(sample_dir.rglob("*_krona.html")) if p.is_file()),
        None,
    )
    if krona is None:
        raise HTTPException(status_code=404, detail="No Krona graph found for this sample")
    return FileResponse(krona, media_type="text/html")


@app.get("/api/projects/{project}/kraken/samples")
def kraken_samples(project: str):
    """List the sample names that have a Kraken output dir under
    <project>/kraken/. Lets the UI flag which download samples already have
    Kraken results without a per-sample request each."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    kraken_dir = project_dir / "kraken"
    names: List[str] = []
    if kraken_dir.is_dir():
        try:
            names = sorted(
                d.name for d in kraken_dir.iterdir()
                if d.is_dir() and not d.name.startswith(".")
            )
        except (OSError, PermissionError):
            names = []
    return {"project": project, "samples": names}


# ---------------------------------------------------------------------------
# Run Kraken ID Parse on a sample (cross-tool invocation).
#
# vSNP and Kraken ID Parse share /srv/kapurlab/projects, so we can kick off the
# Kraken pipeline directly from the vSNP Step 1 view and have it write into the
# same <project>/kraken/<sample>/ directory the results panel already reads.
#
# Crucially this uses the SHARED Kraken install — its own conda env and bin/
# scripts in the sibling kraken_id_parse_gui checkout — NOT the vSNP env (the
# vsnp3 env has neither kraken2, krona, nor SPAdes). The Kraken tool itself is
# lab-shared functionality; only the project data lives per-user/shared.
# TOOLS_ROOT (config.py) is the dir containing the sibling checkouts, resolved
# from BDTOOLS_TOOLS_ROOT / SITE_ROOT / this checkout's own parent — never a
# baked-in site path, so a fresh deployment finds its OWN Kraken install.
# ---------------------------------------------------------------------------
_KRAKEN_GUI_ROOT = TOOLS_ROOT / "kraken_id_parse_gui"

# Shared taxon search-name list, owned by the Kraken ID Parse repo. Both GUIs
# read and append to this same file so the preset list stays in sync.
_KRAKEN_TAXA_YAML = _KRAKEN_GUI_ROOT / "config" / "taxa.yaml"

_KRAKEN_TAXA_HEADER = (
    "# Kraken ID Parse — taxon search names\n"
    "#\n"
    "# Single source of truth for the \"Target Taxon\" presets shown in BOTH the\n"
    "# Kraken ID Parse GUI and the vSNP GUI. Each entry is a taxonomy\n"
    "# classification name passed verbatim to the read parser (-t) and must match\n"
    "# the name exactly as Kraken reports it. Plain YAML sequence; add by hand or\n"
    "# via the \"Add search name\" control in either GUI.\n"
)


def _read_kraken_taxa() -> List[str]:
    """Read the shared taxon list. Dependency-free flat-YAML-sequence parser
    (``- name`` per line) so it works whether or not PyYAML is installed."""
    taxa: List[str] = []
    try:
        text = _KRAKEN_TAXA_YAML.read_text(encoding="utf-8")
    except (OSError, FileNotFoundError):
        return taxa
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("- "):
            line = line[2:].strip()
        elif line == "-":
            continue
        if line and line[0] not in "\"'" and " #" in line:
            line = line.split(" #", 1)[0].strip()
        if len(line) >= 2 and line[0] in "\"'" and line[-1] == line[0]:
            line = line[1:-1]
        if line:
            taxa.append(line)
    return taxa


def _write_kraken_taxa(taxa: List[str]) -> None:
    """Rewrite the shared taxa.yaml as a flat YAML sequence, preserving header."""
    _KRAKEN_TAXA_YAML.parent.mkdir(parents=True, exist_ok=True)
    lines = [_KRAKEN_TAXA_HEADER]
    for name in taxa:
        name = name.strip()
        if not name:
            continue
        if name[0] in "-?:[]{}#&*!|>'\"%@`" or ": " in name or name.endswith(":"):
            esc = name.replace("\\", "\\\\").replace('"', '\\"')
            lines.append(f'- "{esc}"')
        else:
            lines.append(f"- {name}")
    _KRAKEN_TAXA_YAML.write_text("\n".join(lines) + "\n", encoding="utf-8")

def _kraken_gui_config() -> Dict[str, Any]:
    """Read the Kraken ID Parse GUI's own user config so the vSNP GUI inherits
    whatever DB paths the user already set THERE (that GUI has a Settings panel
    with kraken_db / blast_db fields; the vSNP GUI does not). Mirrors that tool's
    config-dir logic — XDG_CONFIG_HOME, else ~/.config/kraken_id_parse_gui.

    Without this, a local `bdtools` install — where VSNP_GUI_SITE_ROOT points at
    the per-user vsnp3-site tree that has no databases/ dir — falls through to a
    non-existent default and Kraken dies with "does not contain ... taxo.k2d",
    even though the Kraken GUI runs fine off its configured /srv DB.
    """
    xdg = os.environ.get("XDG_CONFIG_HOME", "").strip()
    base = Path(xdg) if xdg else (Path.home() / ".config")
    path = base / "kraken_id_parse_gui" / "config.json"
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}
    return data if isinstance(data, dict) else {}

def _existing_site_kraken_db() -> str:
    """The reference install's conventional Kraken2 DB — but only when it is
    really there. hash.k2d is what kraken2 itself requires, so its presence is
    the honest test. Everywhere else return "" and let the caller error with
    instructions instead of pointing at a directory this site never had."""
    cand = DB_ROOT / "kraken2" / "k2_standard_08gb"
    return str(cand) if (cand / "hash.k2d").is_file() else ""


def _existing_site_blast_db() -> str:
    """Site BLAST DB prefix (files <prefix>.n*), existence-guarded like the
    Kraken default. "" means: let the pipeline fall back to its own default
    ("nt", NCBI-remote) rather than a dead local path."""
    cand = DB_ROOT / "blast" / "ref_prok_rep_genomes"
    try:
        has_files = any(cand.parent.glob(cand.name + ".*"))
    except OSError:
        has_files = False
    return str(cand) if has_files else ""


# Read-tag stripping identical to Kraken ID Parse's own pairing logic, so the
# sample name and R1/R2 selection match what that tool would pick on its own.
_KRAKEN_READ_TAG_RE = re.compile(r'(?:_R([12])(?:_\d+)?|_([12]))\.fastq\.gz$', re.IGNORECASE)


def _kraken_strip_read_tag(filename: str):
    m = _KRAKEN_READ_TAG_RE.search(filename)
    if m:
        return filename[:m.start()], (m.group(1) or m.group(2))
    return filename[:-len(".fastq.gz")], None


def _find_sample_fastqs(download_dir: Path, sample: str):
    """Return (r1, r2) Paths for a sample from <project>/download/.

    Matches the sample's read files by stripping Illumina/SRA read tags, with
    the same prefix fallback used elsewhere (sample ``13-1941-6`` matches
    ``13-1941-6_S4_L001_R1_001.fastq.gz``). r2 is None for single-end.
    """
    if not download_dir.is_dir():
        return None, None
    try:
        all_fq = sorted(download_dir.glob("*.fastq.gz"))
    except OSError:
        return None, None
    r1 = r2 = single = None
    for fq in all_fq:
        base, tag = _kraken_strip_read_tag(fq.name)
        if base != sample and not base.startswith(f"{sample}"):
            continue
        # Require the bare sample name to match the file's base (exact) or be a
        # prefix up to a lane/index separator, to avoid matching "S1" to "S10".
        if base != sample and not base.startswith(f"{sample}_"):
            continue
        if tag == "1":
            r1 = r1 or fq
        elif tag == "2":
            r2 = r2 or fq
        else:
            single = single or fq
    if r1:
        return r1, r2
    if single:
        return single, None
    return None, None


def _resolve_step1_sample_dir(step1_dir: Path, sample: str) -> Optional[Path]:
    """Find a sample's step1 output dir, for sourcing its FASTQs.

    The normal GUI flow keeps the inputs in ``<project>/download/`` and step1
    only holds symlinks back to them. But data run *outside* the GUI (vsnp3 on
    the command line, then dropped into a project) lands as step1 output with
    the real FASTQs inside ``step1/<sample>/`` and an empty ``download/``. This
    locates that folder so Kraken can still find the reads. Dir name usually
    equals the sample; fall back to the same prefix matching used elsewhere.
    """
    if not step1_dir.is_dir():
        return None
    exact = step1_dir / sample
    if exact.is_dir():
        return exact
    for name in _step1_sample_names(step1_dir):
        if name == sample or name.startswith(f"{sample}_") or sample.startswith(f"{name}_"):
            return step1_dir / name
    return None


def _dash_delimited_import_name(filename: str) -> str:
    """Rename a parsed read file so vSNP treats it as a NEW, distinct sample.

    vSNP keys a sample on the text left of the first underscore, so
    ``<sample>_<taxon>_R1.fastq.gz`` would collapse back to ``<sample>`` and
    overwrite the original run. Replacing every underscore in the
    sample-identifying stem with a dash (keeping only the ``_R1``/``_R2``
    read-tag underscore) yields a unique name like
    ``<sample>-<taxon-words>_R1.fastq.gz``.
    """
    m = _KRAKEN_READ_TAG_RE.search(filename)
    if not m:
        if filename.endswith(".fastq.gz"):
            stem = filename[: -len(".fastq.gz")]
            return stem.replace("_", "-") + ".fastq.gz"
        return filename.replace("_", "-")
    stem = filename[: m.start()]
    tag = filename[m.start():]  # keep the read tag (e.g. "_R1.fastq.gz") verbatim
    return stem.replace("_", "-") + tag


def _import_parsed_reads(run_dir: Path, download_dir: Path) -> List[str]:
    """Copy a Kraken run's taxon-parsed reads into the project's inputs.

    parse_reads.py writes ``<token>_<taxon>_R1.fastq.gz`` / ``_R2`` at the top
    level of ``run_dir``. Each is copied into ``<project>/download/`` under a
    dash-delimited name (see _dash_delimited_import_name) so vSNP sees a brand
    new sample and a re-run does not overwrite the original sample's results.
    Returns the destination filenames created.

    Match on the read-tag suffix rather than the taxon string: the canonical
    taxon the tool stamps into the filename can differ from the user-entered
    search taxon (e.g. "Mycobacterium tuberculosis" → "..._complex"), so a
    taxon-prefixed glob silently imports nothing. Every ``*_R1``/``*_R2`` (or
    ``_1``/``_2``) ``.fastq.gz`` at the top level is a parsed output — the
    ``kraken/`` subdir holds reports, not reads, and the glob is non-recursive."""
    imported: List[str] = []
    try:
        download_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        return imported
    for src in sorted(run_dir.glob("*.fastq.gz")):
        if not src.is_file() or not _KRAKEN_READ_TAG_RE.search(src.name):
            continue
        dest = download_dir / _dash_delimited_import_name(src.name)
        try:
            shutil.copy2(src, dest)
            imported.append(dest.name)
        except OSError:
            continue
    return imported


@app.get("/api/kraken/taxa")
def kraken_taxa():
    """Return the shared taxon search names (kraken repo's config/taxa.yaml)."""
    return {"taxa": _read_kraken_taxa()}


class KrakenTaxonPayload(BaseModel):
    name: str


@app.post("/api/kraken/taxa")
def kraken_taxa_add(payload: KrakenTaxonPayload):
    """Append a new taxon search name to the shared taxa.yaml (no duplicates)."""
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="A taxon name is required.")
    taxa = _read_kraken_taxa()
    if any(name.lower() == t.lower() for t in taxa):
        return {"taxa": taxa, "added": False}
    taxa.append(name)
    try:
        _write_kraken_taxa(taxa)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Could not save taxon list: {exc}")
    return {"taxa": taxa, "added": True}


@app.get("/api/kraken/dbs")
def kraken_dbs():
    """Kraken2 databases available to a vSNP-launched run, and which one a run
    would use right now. The Kraken ID Parse tool owns DB configuration (its
    Settings remember every DB added there); this endpoint just mirrors that
    tool's list so the vSNP Kraken dialog can show it and switch per-run."""
    cfg = load_config()
    _kgui = _kraken_gui_config()
    current = cfg.get("kraken_db", "") \
        or str(_kgui.get("kraken_db", "") or "").strip() \
        or _existing_site_kraken_db()
    options: List[str] = []
    for p in [current, str(_kgui.get("kraken_db", "") or "").strip(),
              *(_kgui.get("saved_kraken_dbs") or []),
              _existing_site_kraken_db()]:
        p = str(p or "").strip()
        if p and p not in options:
            options.append(p)
    return {"current": current, "databases": options}


def _resolve_kraken_runtime() -> Dict[str, str]:
    """Locate the shared Kraken install's python + bin. Falls back to a
    per-user miniforge env only if the shared env is absent (mirrors the OOD
    launch script's resolution). Returns {python, gui_root, env_bin}."""
    shared_env = _KRAKEN_GUI_ROOT / "env"
    personal_env = Path.home() / "miniforge3" / "envs" / "kraken_id_parse"
    if (shared_env / "bin" / "python").exists():
        env_dir = shared_env
    elif (personal_env / "bin" / "python").exists():
        env_dir = personal_env
    else:
        env_dir = None
    python = str((env_dir / "bin" / "python")) if env_dir else sys.executable
    return {
        "python": python,
        "gui_root": str(_KRAKEN_GUI_ROOT),
        "env_bin": str(env_dir / "bin") if env_dir else "",
    }


class KrakenRunRequest(BaseModel):
    sample: str
    # "full": classify + parse reads + assemble + BLAST identification (needs a
    # taxon). "parse_only": classify + parse target reads, then stop — skips
    # assembly/BLAST/coverage (needs a taxon). "kraken_only": Kraken2 + Krona
    # graph only (no taxon needed).
    mode: str = "full"
    taxon: Optional[str] = None
    kraken_db: Optional[str] = None
    blast_db: Optional[str] = None


@app.post("/api/projects/{project}/kraken/run")
def kraken_run(project: str, payload: KrakenRunRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")

    if not _KRAKEN_GUI_ROOT.is_dir():
        raise HTTPException(
            status_code=503,
            detail=f"Kraken ID Parse is not installed at {_KRAKEN_GUI_ROOT}.",
        )
    script = _KRAKEN_GUI_ROOT / "bin" / "kraken_id_parse.py"
    if not script.is_file():
        raise HTTPException(status_code=503, detail=f"Kraken pipeline script missing: {script}")

    mode = (payload.mode or "full").strip()
    kraken_only = mode == "kraken_only"
    parse_only = mode == "parse_only"
    taxon = (payload.taxon or "").strip()
    if not kraken_only and not taxon:
        raise HTTPException(status_code=400, detail="A target taxon is required for this run mode.")

    sample = (payload.sample or "").strip()
    if not sample or "/" in sample or sample.startswith("."):
        raise HTTPException(status_code=400, detail="Invalid sample name")

    r1, r2 = _find_sample_fastqs(project_dir / "download", sample)
    if r1 is None:
        # Data run outside the GUI (command-line vsnp3, then placed in a
        # project) has an empty download/ and keeps its FASTQs inside
        # step1/<sample>/. Fall back to that so Kraken can still run on it.
        step1_sample_dir = _resolve_step1_sample_dir(project_dir / "step1", sample)
        if step1_sample_dir:
            r1, r2 = _find_sample_fastqs(step1_sample_dir, sample)
    if r1 is None:
        raise HTTPException(
            status_code=404,
            detail=f"No FASTQ files found for sample {sample!r} in the project's download/ or step1/{sample}/ folder.",
        )

    # Name the output dir EXACTLY as the Kraken ID Parse tool would when run
    # from its own GUI on these same FASTQs (read-tag stripped from R1). This
    # keeps <project>/kraken/<dir> identical no matter which GUI launched the
    # run, so the Kraken GUI — which lists samples by that stripped name —
    # finds the results instead of showing "No Kraken results yet".
    kraken_sample = _kraken_strip_read_tag(r1.name)[0]

    # Kraken/BLAST DB resolution, in order:
    #   1. request override (payload)
    #   2. vsnp GUI's own config (if a user set one here)
    #   3. the Kraken ID Parse GUI's config — inherit whatever DB that tool is
    #      already using, so a working Kraken GUI means a working vSNP Kraken run
    #      (the vSNP GUI has no DB field of its own).
    #   4. the site database root — ONLY if that DB actually exists on disk. A
    #      new deployment must resolve to "" (and get a pointed error below),
    #      never to another site's layout.
    _kgui = _kraken_gui_config()
    kraken_db = (payload.kraken_db or "").strip() or cfg.get("kraken_db", "") \
        or str(_kgui.get("kraken_db", "") or "").strip() \
        or _existing_site_kraken_db()
    blast_db = (payload.blast_db or "").strip() or cfg.get("blast_db", "") \
        or str(_kgui.get("blast_db", "") or "").strip() \
        or _existing_site_blast_db()
    if not kraken_db:
        raise HTTPException(
            status_code=400,
            detail=(
                "No Kraken2 database is configured. Set one in the Kraken ID "
                "Parse tool's Settings — vSNP runs Kraken with the database "
                "configured there."
            ),
        )

    run_dir = project_dir / "kraken" / kraken_sample
    # Guard against two runs racing on the same output dir (they'd clobber each
    # other's temp/output folders). Track the last job id in a sentinel file,
    # the same pattern step1 uses with .step1_job_id.
    job_id_file = run_dir / ".kraken_job_id"
    if job_id_file.exists():
        prior = job_manager.get_job(job_id_file.read_text(encoding="utf-8").strip())
        if prior and prior.get("status") == "running":
            raise HTTPException(
                status_code=409,
                detail=f"A Kraken run is already in progress for {sample}.",
            )
    run_dir.mkdir(parents=True, exist_ok=True)

    rt = _resolve_kraken_runtime()
    parts = [shlex.quote(rt["python"]), "-u", shlex.quote(str(script)), "-r1", shlex.quote(str(r1))]
    if r2 is not None:
        parts += ["-r2", shlex.quote(str(r2))]
    if taxon:
        parts += ["-t", shlex.quote(taxon)]
    if kraken_db:
        parts += ["-k", shlex.quote(kraken_db)]
    if kraken_only:
        parts.append("--kraken-only")
    elif parse_only:
        parts.append("--no-blast")
    elif blast_db:
        parts += ["-b", shlex.quote(blast_db)]

    # Prepend the Kraken env bin so kraken2/krona/spades/seqkit resolve, and set
    # PYTHONPATH so the bin/ scripts' local imports work — exactly as the Kraken
    # GUI launches them.
    path_prefix = f"{rt['env_bin']}:" if rt["env_bin"] else ""
    command = (
        f'PYTHONUNBUFFERED=1 PYTHONPATH={shlex.quote(rt["gui_root"] + "/bin")} '
        f'PATH="{path_prefix}$PATH" {" ".join(parts)}'
    )

    if kraken_only:
        label = "Kraken-only (Krona)"
    elif parse_only:
        label = f"{taxon} (parse only)"
    else:
        label = taxon or "identification"

    # When the run extracts target reads (full or parse-only — anything with a
    # taxon), auto-import them into the project's inputs so the user can re-run
    # them through vSNP without manual copying. Renamed to a dash-delimited
    # sample so vSNP treats them as a distinct sample (see _import_parsed_reads).
    download_dir = project_dir / "download"
    output_prefix = taxon.replace(" ", "_") if taxon else ""

    def _on_kraken_done(jid, exit_code, started, finished,
                        _run_dir=run_dir, _dl=download_dir,
                        _prefix=output_prefix, _kraken_only=kraken_only):
        if exit_code != 0 or _kraken_only or not _prefix:
            return
        imported = _import_parsed_reads(_run_dir, _dl)
        if imported:
            logger.info(
                "kraken auto-import: copied %d parsed read file(s) into %s: %s",
                len(imported), _dl, ", ".join(imported),
            )

    job_id = job_manager.start_job(
        name="kraken",
        command=command,
        cwd=run_dir,
        env=build_env(cfg),
        finalize_callback=_on_kraken_done,
    )
    job_id_file.write_text(job_id, encoding="utf-8")
    return {"job_id": job_id, "run_dir": str(run_dir), "sample": kraken_sample, "mode": mode if mode in ("kraken_only", "parse_only", "full") else "full", "label": label}


@app.get("/api/projects/{project}/step1/edits")
def step1_edits(project: str):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    if not step1_dir.exists():
        raise HTTPException(status_code=404, detail="Step1 directory not found")
    edits = {}
    for sample_dir in sorted(step1_dir.glob("*")):
        if not sample_dir.is_dir():
            continue
        sample = sample_dir.name
        # Non-recursive on purpose: a **/ glob here walked every sample's whole
        # output subtree per request — the same class of cost that made the
        # status endpoint time out at ~1000 samples. The zc VCF only ever
        # lives under alignment_*/ (or the legacy plain alignment/).
        vcf_candidates = sorted(
            _align_glob(sample_dir, f"{sample}*zc.vcf*"), key=lambda p: p.stat().st_mtime
        )
        source_vcf = vcf_candidates[-1] if vcf_candidates else None
        patched_vcf = _find_patched_vcf(sample_dir, sample, source_vcf)
        edit_log = _edit_log_path(sample_dir, sample)
        if patched_vcf or edit_log.exists():
            edits[sample] = {
                "patched_vcf": str(patched_vcf) if patched_vcf else "",
                "edit_log": str(edit_log) if edit_log.exists() else "",
                "edited": bool(patched_vcf) and edit_log.exists()
            }
    return edits




def _resolve_step2_output_dir(step2_dir: Path, run_id: Optional[str]) -> Path:
    """Resolve which directory to read step2 outputs from.

    Priority:
    1. Explicit run_id → step2/{run_id}/ (or legacy step2/runs/{run_id}/)
    2. .current_run sentinel → that run's dir
    3. Latest run by directory name (lexicographic = chronological)
    4. Legacy flat layout: step2/ itself
    """
    run_dirs = _step2_run_dirs(step2_dir)
    if run_id and run_id != "legacy":
        if run_id in run_dirs:
            return run_dirs[run_id]
    if run_id == "legacy":
        return step2_dir
    current_file = step2_dir / ".current_run"
    if current_file.exists():
        current_ts = current_file.read_text(encoding="utf-8").strip()
        if current_ts in run_dirs:
            return run_dirs[current_ts]
    if run_dirs:
        latest = sorted(run_dirs.keys(), reverse=True)[0]
        return run_dirs[latest]
    return step2_dir  # legacy flat


@app.get("/api/projects/{project}/step2_outputs")
def step2_outputs(project: str, run_id: Optional[str] = Query(None)):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    if not step2_dir.exists():
        raise HTTPException(status_code=404, detail="Step2 directory not found")
    output_dir = _resolve_step2_output_dir(step2_dir, run_id)

    def _safe_name(value: str) -> str:
        return "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in value)

    def _find_group_fasta(group_dir: Path) -> Optional[Path]:
        for pattern in ("*.fasta", "*.fa", "*.fna"):
            matches = sorted(group_dir.glob(pattern))
            if matches:
                return matches[-1]
        return None

    def _count_fasta_sequences(fasta_path: Optional[Path]) -> int:
        if not fasta_path or not fasta_path.exists():
            return 0
        count = 0
        with fasta_path.open("r", encoding="utf-8", errors="ignore") as handle:
            for line in handle:
                if line.startswith(">"):
                    count += 1
        return count

    top = []
    html_files = sorted(output_dir.glob("*.html"), key=lambda p: p.stat().st_mtime)
    if html_files:
        latest_html = html_files[-1]
        top.append({
            "label": latest_html.name,
            "path": str(latest_html),
            "type": "html",
            "download_name": f"{_safe_name(project)}__{latest_html.name}",
        })
    for f in output_dir.glob("*.zip"):
        top.append({
            "label": f.name,
            "path": str(f),
            "type": "zip",
            "download_name": f"{_safe_name(project)}__{f.name}",
        })
    mismatch_report = output_dir / "mismatch_report.csv"
    if mismatch_report.exists():
        top.append({
            "label": mismatch_report.name,
            "path": str(mismatch_report),
            "type": "csv",
            "download_name": f"{_safe_name(project)}__{mismatch_report.name}",
        })
    top.sort(key=lambda x: x["label"])

    groups = []
    for d in sorted(output_dir.iterdir()):
        if not d.is_dir():
            continue
        if d.name in _STEP2_NON_GROUP_DIRS or _STEP2_RUN_RE.match(d.name):
            continue
        if d.name.startswith("."):
            continue
        fasta_path = _find_group_fasta(d)
        fasta_count = _count_fasta_sequences(fasta_path)
        # If a *_labeled.tre exists for a base group tree, hide the unlabeled
        # sibling — labeled has the lineage prefix prepended to each leaf
        # (e.g. L4_ERR2704709_zc.vcf), which is what makes the tree useful
        # for placing run samples in context. Unlabeled file is still on
        # disk for anyone going via the filesystem.
        labeled_bases = {
            f.name.removesuffix("_labeled.tre")
            for f in d.iterdir()
            if f.is_file() and f.name.endswith("_labeled.tre")
        }
        files = []
        for f in sorted(d.iterdir()):
            if f.is_file():
                if f.name.endswith(".tre") and not f.name.endswith("_labeled.tre"):
                    if f.name.removesuffix(".tre") in labeled_bases:
                        continue
                ext = f.suffix.lstrip(".")
                files.append({
                    "label": f.name,
                    "path": str(f),
                    "type": ext or "file",
                    "download_name": f"{_safe_name(project)}__{_safe_name(d.name)}__{f.name}",
                })
            elif f.is_dir() and f.name == "posthoc":
                for pf in sorted(f.iterdir()):
                    if pf.is_file():
                        ext = pf.suffix.lstrip(".")
                        files.append({
                            "label": f"posthoc/{pf.name}",
                            "path": str(pf),
                            "type": ext or "file",
                            "download_name": f"{_safe_name(project)}__{_safe_name(d.name)}__posthoc__{pf.name}",
                        })
        if files:
            groups.append({
                "name": d.name,
                "files": files,
                "posthoc_possible": fasta_count >= 3,
                "posthoc_reason": "" if fasta_count >= 3 else "Requires a FASTA with at least 3 sequences",
                "posthoc_sequence_count": fasta_count,
            })
    return {"top": top, "groups": groups, "run_id": output_dir.name if output_dir != step2_dir else "legacy"}


@app.get("/api/projects/{project}/step2/trees")
def step2_trees(project: str, run_id: Optional[str] = Query(None)):
    """List the latest .tre file per group under step2/."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step2_dir = project_dir / "step2"
    if not step2_dir.exists():
        raise HTTPException(status_code=404, detail="Step2 directory not found")
    output_dir = _resolve_step2_output_dir(step2_dir, run_id)
    trees = []
    for d in sorted(output_dir.iterdir()):
        if not d.is_dir():
            continue
        if d.name in _STEP2_NON_GROUP_DIRS or _STEP2_RUN_RE.match(d.name) or d.name.startswith("."):
            continue
        tre_files = sorted(d.glob("*.tre"), key=lambda p: p.stat().st_mtime)
        if not tre_files:
            continue
        latest = tre_files[-1]
        trees.append({
            "group": d.name,
            "name": latest.name,
            "path": str(latest),
            "size": latest.stat().st_size,
            "mtime": latest.stat().st_mtime,
        })
    return {"trees": trees}


@app.post("/api/bootstrap")
def bootstrap():
    cfg = load_config()
    root_dir = Path(__file__).resolve().parent.parent.parent
    script_path = root_dir / "scripts" / "bootstrap_vsnp3.sh"
    if not script_path.exists():
        raise HTTPException(status_code=404, detail="Bootstrap script not found")
    job_id = job_manager.start_job(
        name="bootstrap",
        command=f"bash {script_path}",
        cwd=root_dir,
        env=build_env(cfg)
    )
    return {"job_id": job_id}


@app.get("/api/projects/{project}/preview-xlsx", response_class=HTMLResponse)
def preview_xlsx(project: str, path: str = Query(...), download: int = 0):
    """Render an xlsx file as a self-contained HTML page (formatting preserved
    via openpyxl). With ?download=1, returns the raw xlsx (for the "Download
    xlsx" link inside the preview page)."""
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    target = Path(path).resolve()
    if not str(target).startswith(str(project_dir.resolve())):
        raise HTTPException(status_code=400, detail="Path not allowed")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    if target.suffix.lower() not in (".xlsx", ".xlsm"):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm supported")
    if download:
        return FileResponse(
            target,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=target.name,
        )
    # Build sets of samples loadable in IGV so the cascade-table render
    # can correctly enable / grey out the "↗ this" affordance per row:
    #   - samples_with_bams: have a Step 1 BAM (full IGV: reads + calls)
    #   - samples_with_vcfs: have an imported VCF in step2/vcf_database/
    #     (calls-only IGV, anchored to the project reference)
    # A sample qualifies for "↗ this" if it's in either set.
    samples_with_bams: set[str] = set()
    samples_with_vcfs: set[str] = set()
    step1_dir = project_dir / "step1"
    if step1_dir.is_dir():
        for d in step1_dir.iterdir():
            if not d.is_dir():
                continue
            if any(d.glob(f"**/{d.name}_nodup.bam")):
                samples_with_bams.add(d.name)
            # _resolve_sample_dir accepts a bare sample name even when the
            # step1 dir carries lane suffixes (e.g. dir `13-1941-6_S4_L001`
            # resolves from input `13-1941-6`). Mirror that fallback here so
            # cascade stems match.
            if "_" in d.name:
                prefix = d.name.split("_")[0]
                if prefix and any(d.glob(f"**/{prefix}_nodup.bam")):
                    samples_with_bams.add(prefix)
    vcf_source_dir = vcf_db_dir(project_dir / "step2")
    if vcf_source_dir.is_dir():
        for f in vcf_source_dir.iterdir():
            if not f.is_file():
                continue
            name = f.name
            # Strip the standard vSNP3 suffixes back to the stem.
            for suffix in ("_zc.vcf.gz", "_zc.vcf", ".vcf.gz", ".vcf"):
                if name.endswith(suffix):
                    samples_with_vcfs.add(name[: -len(suffix)])
                    break

    from app import xlsx_html
    try:
        html_page = xlsx_html.xlsx_to_html(
            target,
            project=project,
            samples_with_bams=samples_with_bams,
            samples_with_vcfs=samples_with_vcfs,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"xlsx render failed: {type(e).__name__}: {e}")
    return HTMLResponse(content=html_page)


@app.get("/api/projects/{project}/download-file")
def download_file(project: str, path: str = Query(...), inline: int = 0, download_name: Optional[str] = Query(None)):
    """Serve a file from within a project directory.

    Default (no ?inline) sets `Content-Disposition: attachment` so the browser
    downloads. With ?inline=1, omits the attachment disposition so the browser
    renders the file in-tab when it can (html, fasta, vcf, png, pdf, …) —
    used by the "View" buttons on the step2 results panel. With ?download_name=…,
    overrides the attachment filename (server-side enforcement of the project-
    qualified naming convention).
    """
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    target = Path(path).resolve()
    if not str(target).startswith(str(project_dir.resolve())):
        raise HTTPException(status_code=400, detail="Path not allowed")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    media_type = "application/octet-stream"
    suffix = target.suffix.lower()
    # Map suffix -> MIME so browsers render inline (?inline=1 path) for
    # everything the frontend's fileViewMode() flags as viewable. Without
    # this, e.g. JSON returns octet-stream and browsers force-download.
    if suffix in (".html", ".htm"):
        media_type = "text/html"
    elif suffix == ".csv":
        media_type = "text/csv"
    elif suffix in (".xlsx", ".xls", ".xlsm"):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif suffix == ".pdf":
        media_type = "application/pdf"
    elif suffix == ".json":
        media_type = "application/json"
    elif suffix in (".jsonl", ".ndjson"):
        # NDJSON / JSON-lines — not valid as a single JSON document, so
        # render as plain text (browser shows the lines; doesn't try to
        # parse a tree). Used by Mg***_patchlog.jsonl from VCF edits.
        media_type = "text/plain"
    elif suffix == ".svg":
        media_type = "image/svg+xml"
    elif suffix == ".png":
        media_type = "image/png"
    elif suffix in (".jpg", ".jpeg"):
        media_type = "image/jpeg"
    elif suffix == ".gif":
        media_type = "image/gif"
    elif suffix == ".webp":
        media_type = "image/webp"
    elif suffix in (
        ".tre", ".nwk", ".nexus", ".nex",
        ".fasta", ".fa", ".fna",
        ".vcf", ".txt", ".tsv", ".log",
        ".yaml", ".yml", ".md",
    ):
        media_type = "text/plain"
    if inline:
        # Defensive: if we didn't recognise the suffix and would otherwise
        # serve application/octet-stream, force text/plain so the browser
        # shows the file in-tab instead of triggering a download. This is
        # the safe default for "View" buttons — research data files often
        # have idiosyncratic extensions (e.g. `.amr.tsv`, `.patchlog.jsonl`)
        # that we don't want to enumerate exhaustively.
        if media_type == "application/octet-stream":
            media_type = "text/plain"
        return FileResponse(target, media_type=media_type)
    filename = target.name
    if download_name:
        safe_name = Path(download_name).name
        if safe_name:
            filename = safe_name
    return FileResponse(target, media_type=media_type, filename=filename)


@app.get("/api/download-file")
def download_file_global(path: str = Query(...)):
    """Download any file from within any configured projects root."""
    cfg = load_config()
    target = Path(path).resolve()
    if not _path_under_any_project_root(cfg, target):
        raise HTTPException(status_code=400, detail="Path not allowed")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(target, media_type="application/octet-stream", filename=target.name)


_IGV_SERVE_MEDIA_TYPES = {
    ".bam": "application/octet-stream",
    ".bai": "application/octet-stream",
    ".cram": "application/octet-stream",
    ".crai": "application/octet-stream",
    ".vcf": "text/plain",
    ".gz": "application/gzip",
    ".tbi": "application/octet-stream",
    ".fasta": "text/plain",
    ".fa": "text/plain",
    ".fai": "text/plain",
    ".gff": "text/plain",
    ".gff3": "text/plain",
    ".bed": "text/plain",
    ".gbk": "text/plain",
}


def _range_response(target: Path, request: Request, media_type: str):
    file_size = target.stat().st_size
    range_header = request.headers.get("range") or request.headers.get("Range")
    if not range_header:
        def _full():
            with open(target, "rb") as f:
                while True:
                    chunk = f.read(1024 * 1024)
                    if not chunk:
                        break
                    yield chunk
        headers = {
            "Accept-Ranges": "bytes",
            "Content-Length": str(file_size),
        }
        return StreamingResponse(_full(), media_type=media_type, headers=headers)
    if not range_header.startswith("bytes="):
        raise HTTPException(status_code=400, detail="Invalid Range header")
    try:
        spec = range_header[len("bytes="):].split(",", 1)[0].strip()
        start_s, end_s = spec.split("-", 1)
        start = int(start_s) if start_s else 0
        end = int(end_s) if end_s else file_size - 1
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid Range header")
    if start < 0 or end >= file_size or start > end:
        return Response(status_code=416, headers={"Content-Range": f"bytes */{file_size}"})
    length = end - start + 1

    def _slice():
        with open(target, "rb") as f:
            f.seek(start)
            remaining = length
            while remaining > 0:
                chunk = f.read(min(1024 * 1024, remaining))
                if not chunk:
                    break
                remaining -= len(chunk)
                yield chunk

    headers = {
        "Accept-Ranges": "bytes",
        "Content-Range": f"bytes {start}-{end}/{file_size}",
        "Content-Length": str(length),
    }
    return StreamingResponse(_slice(), status_code=206, media_type=media_type, headers=headers)


def _under(child: Path, parent: Path) -> bool:
    """Is `child` inside `parent`? Compared at path boundaries.

    A bare startswith is wrong twice over: "/data/proj2" startswith "/data/proj",
    so a sibling directory passes; and a root of "" would match everything.
    """
    parent_s = os.path.normpath(str(parent))
    child_s = os.path.normpath(str(child))
    if not parent_s or parent_s in (".", os.sep):
        return False
    return child_s == parent_s or child_s.startswith(parent_s.rstrip(os.sep) + os.sep)


def _serve_path_allowed(requested: Path, roots) -> bool:
    """May this path be served? True when it is inside one of `roots`.

    Both sides are compared in BOTH their as-given and fully-resolved forms,
    because a reference root can be a directory of SYMLINKS rather than of files.
    A local install with no shared reference set builds exactly that: each
    reference under <site>/refs/vsnp3/reference_options is a symlink into
    ~/.local/share/bdtools/vsnp3-refs/. Resolving only the target then lands it
    outside every (unresolved) root, and a GFF the UI itself just handed out came
    back 400 "Path not allowed" — which igv.js turns into a dead viewer.

    Not just a laxer check: the as-given path is normalized first, so ".." cannot
    walk out of a root, and every accepted path is still confined to a root the
    operator configured.
    """
    candidates = {Path(os.path.normpath(str(requested)))}
    try:
        candidates.add(requested.resolve())
    except OSError:
        pass
    for root in roots:
        if not str(root).strip():
            continue
        forms = {Path(os.path.normpath(str(root)))}
        try:
            forms.add(root.resolve())
        except OSError:
            pass
        for cand in candidates:
            for form in forms:
                if _under(cand, form):
                    return True
    return False


@app.get("/api/projects/{project}/serve")
def serve_project_file(project: str, request: Request, path: str = Query(...)):
    """Serve a file from within the project directory with HTTP byte-range support.

    Used by igv.js to stream BAM/BAI/FASTA/FAI without forcing a full
    download. Also permits paths under any configured reference options
    root (so igv.js can fetch the reference GFF annotation track — the
    GFF lives next to the source fasta in the reference dir, not in the
    project's alignment dir).
    """
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    roots = [project_dir]
    roots.extend(reference_roots(Path(cfg.get("vsnp3_path", ""))))
    # The configured reference root is where the UI gets its GFF paths from
    # (find_gff_for_fasta walks it), so this endpoint must be willing to serve
    # from it. It is NOT always listed in reference_options_paths.txt: a local
    # install that finds every reference already available elsewhere deliberately
    # drops its own managed dir from that file.
    configured_refs = str(cfg.get("vsnp3_reference_options_root", "") or "").strip()
    if configured_refs:
        roots.append(Path(configured_refs))
    if not _serve_path_allowed(Path(path), roots):
        raise HTTPException(status_code=400, detail="Path not allowed")
    target = Path(path).resolve()
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    suffix = target.suffix.lower()
    media_type = _IGV_SERVE_MEDIA_TYPES.get(suffix, "application/octet-stream")
    return _range_response(target, request, media_type)


@app.post("/api/projects/{project}/vcf_edit")
def vcf_edit(project: str, payload: VcfEditRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    step1_dir = project_dir / "step1"
    sample_dir = _resolve_sample_dir(step1_dir, payload.sample) if step1_dir.is_dir() else None
    if not sample_dir:
        raise HTTPException(status_code=404, detail="Sample not found")

    bcftools = _resolve_bcftools(cfg)
    if not bcftools or not Path(bcftools).exists():
        raise HTTPException(status_code=400, detail="bcftools not configured or not found")

    # Find source VCF (prefer *_zc.vcf[.gz])
    source_vcf = _find_source_vcf(sample_dir, payload.sample)
    if not source_vcf:
        raise HTTPException(status_code=404, detail="Source VCF not found")

    edits_dir = sample_dir / "vcf_edits"
    edits_dir.mkdir(parents=True, exist_ok=True)
    expected_patched = _expected_patched_vcf_path(edits_dir, source_vcf)
    legacy_patched = edits_dir / f"{payload.sample}_patched.vcf.gz"
    if legacy_patched.exists() and not expected_patched.exists():
        patched_vcf = legacy_patched
    else:
        patched_vcf = expected_patched
    if ":" not in payload.locus:
        raise HTTPException(status_code=400, detail="Locus must be in contig:pos format")
    contig, pos_str = payload.locus.split(":", 1)
    contig = contig.strip()
    if not contig:
        raise HTTPException(status_code=400, detail="Contig is required")
    try:
        pos = int(pos_str.replace(",", ""))
    except ValueError:
        raise HTTPException(status_code=400, detail="Position must be an integer")
    new_alt = payload.new_alt.strip().upper()
    if not new_alt or not all(c in "ACGTN" for c in new_alt):
        raise HTTPException(status_code=400, detail="ALT must be A/C/G/T/N (no commas)")
    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Reason is required")

    base_vcf = patched_vcf if patched_vcf.exists() else (legacy_patched if legacy_patched.exists() else source_vcf)
    if base_vcf and base_vcf != source_vcf:
        if not _scan_vcf_for_locus(base_vcf, contig, pos):
            base_vcf = source_vcf

    tbi_path = patched_vcf.with_suffix(patched_vcf.suffix + ".tbi")
    if tbi_path.exists():
        tbi_path.unlink()
    tmp_vcf = edits_dir / f".{payload.sample}_edit_{int(time.time())}.vcf"
    try:
        edit_meta = _rewrite_vcf_with_alt(base_vcf, tmp_vcf, contig, pos, new_alt)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    compress = subprocess.run([bcftools, "view", "-Oz", "-o", str(patched_vcf), str(tmp_vcf)], text=True, capture_output=True)
    try:
        tmp_vcf.unlink()
    except OSError:
        pass
    if compress.returncode != 0:
        raise HTTPException(status_code=500, detail=f"VCF compression failed: {compress.stderr.strip()}")

    # Index patched VCF
    idx = subprocess.run([bcftools, "index", "-t", str(patched_vcf)], text=True, capture_output=True)
    if idx.returncode != 0:
        raise HTTPException(status_code=500, detail=f"Index failed: {idx.stderr.strip()}")

    edit_meta = edit_meta or {}

    log_path = _edit_log_path(sample_dir, payload.sample)
    entry = {
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "user": payload.user or "",
        "project": project,
        "sample": payload.sample,
        "reference": contig,
        "locus": f"{contig}:{pos}",
        "original": {
            "ref": edit_meta.get("old_ref", ""),
            "alt": edit_meta.get("old_alt", ""),
            "dp": edit_meta.get("old_dp", None),
            "ad": edit_meta.get("old_ad", None),
        },
        "updated": {
            "ref": edit_meta.get("old_ref", ""),
            "alt": new_alt,
            "note": payload.note or "",
            "reason": reason
        },
        "source_vcf": str(source_vcf),
        "base_vcf": str(base_vcf),
        "patched_vcf": str(patched_vcf)
    }
    with log_path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(entry) + "\n")

    return {
        "patched_vcf": str(patched_vcf),
        "log": str(log_path),
        "entry": entry
    }


@app.post("/api/projects/{project}/vcf_lookup")
def vcf_lookup(project: str, payload: VcfLookupRequest):
    cfg = load_config()
    project_dir = _project_dir_for(cfg, project)
    sample = (payload.sample or "").strip()
    if not sample:
        raise HTTPException(status_code=400, detail="Sample is required")
    step1_dir = project_dir / "step1"
    sample_dir = _resolve_sample_dir(step1_dir, sample) if step1_dir.is_dir() else None
    if not sample_dir:
        raise HTTPException(status_code=404, detail="Sample not found")
    if ":" not in payload.locus:
        raise HTTPException(status_code=400, detail="Locus must be in contig:pos format")
    contig, pos_str = payload.locus.split(":", 1)
    contig = contig.strip()
    if not contig:
        raise HTTPException(status_code=400, detail="Contig is required")
    try:
        pos = int(pos_str.replace(",", ""))
    except ValueError:
        raise HTTPException(status_code=400, detail="Position must be an integer")
    source_vcf = _find_source_vcf(sample_dir, sample)
    if not source_vcf:
        raise HTTPException(status_code=404, detail="Source VCF not found")
    edits_dir = sample_dir / "vcf_edits"
    patched_vcf = _find_patched_vcf(sample_dir, sample, source_vcf)
    base_vcf = patched_vcf or source_vcf
    record = _scan_vcf_for_locus(base_vcf, contig, pos)
    if not record and patched_vcf and source_vcf and patched_vcf != source_vcf:
        record = _scan_vcf_for_locus(source_vcf, contig, pos)
    if not record:
        logger.warning(
            "VCF lookup miss contig=%s pos=%s base=%s fallback=%s",
            contig,
            pos,
            base_vcf,
            source_vcf
        )
        raise HTTPException(status_code=400, detail="Record not found at locus")
    return {
        "ref": record.get("ref", ""),
        "alt": record.get("alt", ""),
        "dp": record.get("dp", None),
        "ad": record.get("ad", None),
        "base_vcf": record.get("path", str(base_vcf))
    }


def _edit_log_path(sample_dir: Path, sample: str) -> Path:
    return sample_dir / "vcf_edits" / f"{sample}_patchlog.jsonl"


def _expected_patched_vcf_path(edits_dir: Path, source_vcf: Path) -> Path:
    suffixes = source_vcf.suffixes
    if suffixes[-2:] == [".vcf", ".gz"]:
        return edits_dir / source_vcf.name
    if suffixes and suffixes[-1] == ".vcf":
        return edits_dir / f"{source_vcf.name}.gz"
    return edits_dir / f"{source_vcf.name}.gz"


def _find_patched_vcf(sample_dir: Path, sample: str, source_vcf: Optional[Path]) -> Optional[Path]:
    edits_dir = sample_dir / "vcf_edits"
    if not edits_dir.exists():
        return None
    legacy = edits_dir / f"{sample}_patched.vcf.gz"
    if source_vcf:
        expected = _expected_patched_vcf_path(edits_dir, source_vcf)
        if expected.exists():
            return expected
    if legacy.exists():
        return legacy
    candidates = sorted(
        [p for p in edits_dir.glob("*.vcf*") if p.suffix != ".tbi"],
        key=lambda p: p.stat().st_mtime
    )
    return candidates[-1] if candidates else None


def _target_name_for_vcf(source_vcf: Path, chosen_vcf: Path) -> str:
    source_name = source_vcf.name
    chosen_suffixes = chosen_vcf.suffixes
    if chosen_suffixes[-2:] == [".vcf", ".gz"]:
        if source_name.endswith(".gz"):
            return source_name
        return f"{source_name}.gz"
    return source_name


def _vcf_is_edited(path: Path) -> bool:
    try:
        resolved = path.resolve()
    except FileNotFoundError:
        return False
    return "vcf_edits" in resolved.parts


def _edited_samples_in_dir(vcf_source_dir: Path) -> List[str]:
    edited = set()
    if not vcf_source_dir.exists():
        return []
    vcfs = list(vcf_source_dir.glob("*.vcf")) + list(vcf_source_dir.glob("*.vcf.gz"))
    for vcf in vcfs:
        if _vcf_is_edited(vcf):
            edited.add(_sample_from_vcf(vcf))
    return sorted(edited)


def _write_step2_edit_summary(step2_dir: Path, edited_samples: List[str]) -> None:
    payload = {
        "edited_samples": sorted(set(edited_samples)),
        "edited_count": len(set(edited_samples)),
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    }
    try:
        (step2_dir / "edited_samples.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    except OSError:
        pass


def _scan_vcf_for_locus(path: Path, contig: str, pos: int) -> Optional[Dict[str, object]]:
    if not path or not path.exists():
        return None
    opener = gzip.open if str(path).endswith(".gz") else open
    try:
        with opener(path, "rt", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                if not line or line[0] == "#":
                    continue
                parts = line.rstrip("\n").split("\t")
                if len(parts) < 8:
                    continue
                if parts[0] != contig:
                    continue
                try:
                    lpos = int(parts[1])
                except ValueError:
                    continue
                if lpos != pos:
                    continue
                ref = parts[3]
                alt = parts[4].split(",")[0] if parts[4] else ""
                dp = None
                ad = None
                for field in parts[7].split(";"):
                    if field.startswith("DP="):
                        try:
                            dp = int(field.split("=", 1)[1])
                        except ValueError:
                            dp = None
                        break
                if len(parts) >= 10:
                    fmt = parts[8].split(":")
                    sample = parts[9].split(":")
                    if "AD" in fmt:
                        try:
                            idx = fmt.index("AD")
                            ad_val = sample[idx] if idx < len(sample) else ""
                            ad = [int(x) for x in ad_val.split(",") if x != ""] if ad_val else None
                        except Exception:
                            ad = None
                return {
                    "ref": ref,
                    "alt": alt,
                    "dp": dp,
                    "ad": ad,
                    "path": str(path)
                }
    except OSError:
        return None
    return None


def _rewrite_vcf_with_alt(base_vcf: Path, out_vcf: Path, contig: str, pos: int, new_alt: str) -> Dict[str, object]:
    opener = gzip.open if str(base_vcf).endswith(".gz") else open
    found = False
    old_ref = ""
    old_alt = ""
    old_dp = None
    old_ad = None
    with opener(base_vcf, "rt", encoding="utf-8", errors="ignore") as src, out_vcf.open("w", encoding="utf-8") as dst:
        for line in src:
            if not line or line[0] == "#":
                dst.write(line)
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 5:
                dst.write(line)
                continue
            try:
                lpos = int(parts[1])
            except ValueError:
                dst.write(line)
                continue
            if parts[0] == contig and lpos == pos:
                alt_field = parts[4]
                if not alt_field:
                    raise ValueError("Record has no ALT")
                if "," in alt_field:
                    raise ValueError("Record must have exactly one ALT")
                old_ref = parts[3]
                old_alt = alt_field
                if old_alt == new_alt:
                    raise ValueError("ALT is unchanged")
                parts[4] = new_alt
                if len(parts) > 7:
                    for field in parts[7].split(";"):
                        if field.startswith("DP="):
                            try:
                                old_dp = int(field.split("=", 1)[1])
                            except ValueError:
                                old_dp = None
                            break
                if len(parts) > 9:
                    fmt = parts[8].split(":")
                    sample = parts[9].split(":")
                    if "AD" in fmt:
                        try:
                            idx = fmt.index("AD")
                            ad_val = sample[idx] if idx < len(sample) else ""
                            old_ad = [int(x) for x in ad_val.split(",") if x != ""] if ad_val else None
                        except Exception:
                            old_ad = None
                found = True
                line = "\t".join(parts) + "\n"
            dst.write(line)
    if not found:
        raise ValueError("Record not found at locus")
    return {"old_ref": old_ref, "old_alt": old_alt, "old_dp": old_dp, "old_ad": old_ad}


def _scan_vcf_first_record(path: Optional[Path]) -> Optional[Dict[str, object]]:
    if not path or not Path(path).exists():
        return None
    opener = gzip.open if str(path).endswith(".gz") else open
    try:
        with opener(path, "rt", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                if not line or line[0] == "#":
                    continue
                parts = line.rstrip("\n").split("\t")
                if len(parts) < 2:
                    continue
                return {"contig": parts[0], "pos": parts[1]}
    except OSError:
        return None
    return None


def _find_source_vcf(sample_dir: Path, sample: str) -> Optional[Path]:
    candidates = []
    for vcf in sample_dir.glob(f"**/{sample}*zc.vcf*"):
        if vcf.suffix == ".tbi":
            continue
        if "vcf_edits" in vcf.parts:
            continue
        candidates.append(vcf)
    if not candidates:
        return None
    return sorted(candidates, key=lambda p: p.stat().st_mtime)[-1]


def _resolve_bcftools(cfg: Dict) -> str:
    # bcftools ships with the vsnp3 conda env the backend runs in, so resolve it
    # from PATH first — no explicit path config needed (or shown in the UI). An
    # explicit bcftools_path in config.json is still honored as an override for
    # unusual installs.
    found = shutil.which("bcftools")
    if found:
        return found
    return cfg.get("bcftools_path", "").strip()
    subprocess.run(["open", str(target)])


def _detect_vcf_references(vcfs: List[Path], alias_map: Dict[str, str]) -> set:
    refs = set()
    for vcf in vcfs:
        ref = _detect_vcf_reference(vcf, alias_map)
        if ref:
            refs.add(ref)
    return {r for r in refs if r}


def _detect_vcf_reference(vcf: Path, alias_map: Dict[str, str]) -> str:
    try:
        opener = gzip.open if vcf.suffix == ".gz" else open
        with opener(vcf, "rt", encoding="utf-8", errors="ignore") as f:
            for _ in range(200):
                line = f.readline()
                if not line:
                    break
                if line.startswith("##reference="):
                    ref = line.split("=", 1)[1].strip()
                    return _normalize_reference(ref, alias_map)
    except Exception as e:
        print(f"[vcf-ref] Failed to read {vcf}: {e}")
        return ""
    return ""


def _normalize_reference(ref: str, alias_map: Dict[str, str]) -> str:
    ref = ref.replace("file://", "").strip()
    lower_ref = ref.lower()
    for name in alias_map.values():
        if f"/{name.lower()}/" in lower_ref:
            return name
    ref_path = Path(ref)
    candidate = ref_path.stem
    if candidate in alias_map:
        return alias_map[candidate]
    return candidate.replace("_", " ").strip().replace(" ", "_")


def _reference_alias_map(vsnp3_path: Path) -> Dict[str, str]:
    """Map every plausible FASTA stem -> reference directory name.

    vsnp3_step1.py copies the reference FASTA into each sample's alignment
    dir and renames it, stripping NCBI version suffixes — so a reference
    FASTA shipped as `NZ_LS483305.1.fasta` becomes `NZ_LS483305.fasta` in
    the sample dirs, and the resulting VCF's `##reference=` points at the
    renamed file. To survive that rename, index both the original stem
    and the version-stripped form (and a fully-extensionless variant for
    multi-dot filenames like `MTBC0_v1.1.fasta`).
    """
    import re
    aliases: Dict[str, str] = {}

    def _add(key: str, name: str) -> None:
        if not key:
            return
        existing = aliases.get(key)
        if existing is None:
            aliases[key] = name
            return
        # Prefer the more-specific match if a stem collides across refs.
        key_lc = key.lower()
        if key_lc in name.lower() and key_lc not in existing.lower():
            aliases[key] = name

    refs = list_references(vsnp3_path)
    for ref in refs:
        name = ref.get("name")
        base = Path(ref.get("path", ""))
        if not name or not base.exists():
            continue
        for ext in (".fa", ".fasta", ".fna", ".fas"):
            for fasta in base.rglob(f"*{ext}"):
                stem = fasta.stem
                _add(stem, name)
                # NCBI version-suffix variant: `NZ_LS483305.1` -> `NZ_LS483305`.
                stripped = re.sub(r"\.\d+$", "", stem)
                if stripped != stem:
                    _add(stripped, name)
                # Fully extensionless variant for multi-dot stems like
                # `MTBC0_v1.1` -> would still resolve via the previous rule,
                # but covers anything else weird.
                root = fasta.name.split(".", 1)[0]
                if root and root != stem:
                    _add(root, name)
            if name in aliases.values():
                break
    return aliases


def _refs_match(a: str, b: str, allow_fuzzy: bool) -> bool:
    if allow_fuzzy:
        ca = _canonical_ref_key(a)
        cb = _canonical_ref_key(b)
        return ca == cb or ca.startswith(cb) or cb.startswith(ca)
    return a == b


def _canonical_ref_key(ref: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]", "", ref.lower())


def _unique_target(base_dir: Path, filename: str) -> Path:
    stem = Path(filename).stem
    suffix = "".join(Path(filename).suffixes)
    idx = 1
    while True:
        candidate = base_dir / f"{stem}_import{idx}{suffix}"
        if not candidate.exists():
            return candidate
        idx += 1


def _source_prefix(vcf: Path, source_roots: List[Path]) -> str:
    for root in source_roots:
        try:
            vcf.relative_to(root)
            return root.name or "source"
        except ValueError:
            continue
    return "source"


def _sample_from_vcf(vcf: Path) -> str:
    name = vcf.name
    for suffix in ("_zc.vcf.gz", "_zc.vcf"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return vcf.stem


# ---------------------------------------------------------------------------
# Serve frontend static files (added for OOD deployment)
# ---------------------------------------------------------------------------
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse as _FileResponse

_frontend_dist = Path(__file__).parent.parent.parent / "frontend" / "dist"

if _frontend_dist.exists():
    @app.get("/", include_in_schema=False)
    async def _serve_root():
        # The entry document points at content-hashed bundles and must be
        # revalidated after an update. In particular, Safari may otherwise keep
        # an already-open GUI tab on the previous bundle indefinitely.
        return _FileResponse(
            _frontend_dist / "index.html",
            headers={"Cache-Control": "no-store"},
        )

    app.mount("/assets", StaticFiles(directory=str(_frontend_dist / "assets")), name="static_assets")

    # Serve other static files at root level (favicon, etc.)
    for _f in _frontend_dist.iterdir():
        if _f.is_file() and _f.name != "index.html":
            _fname = _f.name
            @app.get(f"/{_fname}", include_in_schema=False)
            async def _serve_static(fname=_fname):
                return _FileResponse(_frontend_dist / fname)


def _posthoc_lock_path(group_dir: Path, tool: str) -> Path:
    return group_dir / "posthoc" / f".{tool}.lock"


def _posthoc_clear_stale_lock(lock_path: Path) -> None:
    if not lock_path.exists():
        return
    job_id = lock_path.read_text(encoding="utf-8").strip()
    if not job_id:
        lock_path.unlink()
        return
    job = job_manager.get_job(job_id)
    if not job or job.get("status") in {"succeeded", "failed", "cancelled"}:
        lock_path.unlink()


def _posthoc_stub_command(cfg: Dict, stats_path: Path, group: str, tool: str) -> str:
    code = (
        "import json, sys, time\n"
        "from pathlib import Path\n"
        "out=Path(sys.argv[1])\n"
        "payload={\n"
        "    'tool': sys.argv[2],\n"
        "    'group': sys.argv[3],\n"
        "    'status': 'stub',\n"
        "    'generated_at': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())\n"
        "}\n"
        "out.write_text(json.dumps(payload, indent=2), encoding='utf-8')\n"
    )
    cmd_list = conda_python_cmd(cfg, code, [str(stats_path), tool, group])
    return " ".join(shlex.quote(part) for part in cmd_list)


def _posthoc_snp_analysis_command(group_dir: Path, group_name: str, out_dir: Path, tool_bin: str, scope: str) -> str:
    snp_dists_path = "snp-dists"
    if tool_bin:
        candidate = Path(tool_bin) / "snp-dists"
        if candidate.exists():
            snp_dists_path = str(candidate)
    cmd_parts = [
        sys.executable,
        "-m",
        "app.posthoc.snp_analysis",
        "--group-dir",
        str(group_dir),
        "--group-name",
        group_name,
        "--out-dir",
        str(out_dir),
        "--snp-dists",
        snp_dists_path,
        "--scope",
        scope,
    ]
    return " ".join(shlex.quote(part) for part in cmd_parts)
