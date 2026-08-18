"""The streaming renderer must match the full one — colour above all.

Every colour in a vSNP3 SNP table comes from CONDITIONAL FORMATTING; the cells
carry no direct fill whatsoever. openpyxl's read-only worksheets do not expose
conditional formatting, so the first version of the streaming renderer produced
a perfectly laid out, completely colourless table — and, because the IGV links
are gated on a cell having a background, no links either. It looked fine in a
size measurement and was useless on screen.

So this compares the two renderers cell-for-cell on a workbook built the way
vSNP3 builds them, and pins the output size bound that keeps a dense table from
rendering to tens of megabytes.

Run directly:  python test_xlsx_render.py
"""

from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

import xlsx_html


def assert_eq(actual, expected, label):
    if actual != expected:
        raise AssertionError(f"{label}: expected {expected!r}, got {actual!r}")
    print(f"  OK  {label} = {actual!r}")


def assert_true(condition, label):
    if not condition:
        raise AssertionError(f"{label}: expected truthy, got falsy")
    print(f"  OK  {label}")


def make_cascade_like(path: Path, rows: int, cols: int) -> None:
    """A workbook shaped like a vSNP3 cascade table.

    Row 1 is contig:pos loci, column 1 is sample names, and the variant cells
    are coloured by a CONDITIONAL FORMATTING rule rather than a direct fill —
    which is the whole point of the test.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for c in range(2, cols + 1):
        ws.cell(row=1, column=c, value=f"MTBC0:{1000 + c}")
    for r in range(2, rows + 1):
        ws.cell(row=r, column=1, value=f"SAMPLE{r - 1}")
        for c in range(2, cols + 1):
            # "A" matches the CF rule below; "-" does not.
            ws.cell(row=r, column=c, value="A" if (r + c) % 2 == 0 else "-")
    rng = f"B2:{openpyxl.utils.get_column_letter(cols)}{rows}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"A"'],
        fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000",
                         fill_type="solid")))
    wb.save(path)
    wb.close()


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="xlsx_render_"))
    try:
        book = tmp / "cascade.xlsx"
        make_cascade_like(book, rows=30, cols=40)
        samples = {f"SAMPLE{i}" for i in range(1, 30)}

        print("\n[the two renderers agree — colour comes from conditional formatting]")
        full = xlsx_html.xlsx_to_html(
            book, project="p", samples_with_bams=samples, samples_with_vcfs=set())
        window = xlsx_html.render_window(
            book, 30, 40, None, "p", samples, set(),
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        stream = xlsx_html.compose_page(window)

        # The renderers express colour differently — the full one inline per
        # cell, the streaming one as a class per distinct style — so compare
        # what is actually equivalent: how many cells are coloured, and how
        # many are clickable IGV targets.
        full_colour = full.count("background-color")
        assert_true(full_colour > 0, "the full renderer finds CF colour at all")
        assert_true(window["style_css"].count("background-color") > 0,
                    "streaming emits background colours in its style palette")
        # Count the anchors themselves — the class name also appears in the
        # stylesheet, which silently inflated a whole-page count by 3.
        full_clickable = full.count('<a class="xlsx-igv-cell"')
        stream_coloured = sum(r.count("xlsx-variant") for r in window["rows"])
        assert_eq(stream_coloured, full_clickable,
                  "streaming marks the same cells clickable as the full renderer")
        assert_true(stream_coloured > 0,
                    "variant cells are produced (they hang off the CF colour)")
        assert_true("xlsxTable" in stream, "the page carries the delegated-click table")

        print("\n[a dense sheet is bounded in BYTES, not just cells]")
        # 120 x 900 = 108,000 cells, under the cell cap, but nearly all of them
        # coloured with an IGV anchor — the shape that rendered to 48 MB when
        # only cells were counted.
        dense = tmp / "dense.xlsx"
        make_cascade_like(dense, rows=120, cols=900)
        many = {f"SAMPLE{i}" for i in range(1, 120)}
        budget = 512 * 1024
        w = xlsx_html.render_window(
            dense, 120, 900, None, "p", many, set(),
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS,
            max_table_bytes=budget)
        rendered = sum(len(r) for r in w["rows"])
        assert_true(rendered < budget * 1.6,
                    f"window ({rendered/1024:.0f} KB) respects a {budget/1024:.0f} KB budget")
        assert_true(w["shown_cols"] < 900, "the window was narrowed to fit")
        page = xlsx_html.compose_page(w)
        assert_true("Showing" in page, "the trim is disclosed on the page")
        assert_true(sum(r.count("xlsx-variant") for r in w["rows"]) > 0,
                    "trimming did not cost the colour")

        print("\n[the page ships a prefix and can serve the rest in batches]")
        tall = tmp / "tall.xlsx"
        make_cascade_like(tall, rows=500, cols=30)
        w2 = xlsx_html.render_window(
            tall, 500, 30, None, "p", {f"SAMPLE{i}" for i in range(1, 500)}, set(),
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        assert_eq(len(w2["rows"]), 500, "every row of the window is rendered")
        first = xlsx_html.compose_page(w2, initial_rows=200)
        assert_eq(first.count("<tr>"), 200, "only the first 200 rows are inlined")
        assert_true("rows_from" in first, "the page knows how to ask for the rest")
        # What the rows endpoint slices out — the same strings, no re-render.
        batch = "".join(w2["rows"][200:400])
        assert_eq(batch.count("<tr>"), 200, "a scroll batch is 200 rows")
        assert_true(len(w2["row_samples"]) == 500,
                    "a sample name per row, for resolving clicks after lazy loads")

        print("\n[a small sheet keeps the untouched full-fidelity path]")
        small = tmp / "small.xlsx"
        make_cascade_like(small, rows=5, cols=6)
        page = xlsx_html.xlsx_to_html(small, project="p",
                                      samples_with_bams={"SAMPLE1"}, samples_with_vcfs=set())
        assert_true("Showing the first" not in page, "no truncation notice")
        assert_true(page.count("background-color") > 0, "small sheet is coloured")

        print("\n[the reference-match rule (equal B$2) resolves while streaming]")
        # The REAL top-priority rule in a vSNP3 table is not a literal: it
        # paints a cell near-white when it equals the reference row — an
        # anchor-row reference (B$2). A streaming pass can honour it because
        # row 2 goes by before any row that needs it; without that capture,
        # reference-matching calls fell through to the per-base colour rules
        # and a streamed table painted every cell bright.
        refbook = tmp / "refrule.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        rows, cols = 12, 10
        for c in range(2, cols + 1):
            ws.cell(row=1, column=c, value=f"MTBC0:{1000 + c}")
        ws.cell(row=2, column=1, value="root")
        for c in range(2, cols + 1):
            ws.cell(row=2, column=c, value="C")
        for r in range(3, rows + 1):
            ws.cell(row=r, column=1, value=f"SAMPLE{r - 2}")
            for c in range(2, cols + 1):
                # Alternate reference-matching 'C' with variant 'A'.
                ws.cell(row=r, column=c, value="C" if (r + c) % 2 == 0 else "A")
        rng = f"B3:{openpyxl.utils.get_column_letter(cols)}{rows}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=["B$2"],
            fill=PatternFill(start_color="FFFDFEFE", end_color="FFFDFEFE",
                             fill_type="solid")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"A"'],
            fill=PatternFill(start_color="FF58FA82", end_color="FF58FA82",
                             fill_type="solid")))
        wb.save(refbook)
        wb.close()
        full = xlsx_html.xlsx_to_html(refbook, project="p",
                                      samples_with_bams=None, samples_with_vcfs=None)
        wref = xlsx_html.render_window(
            refbook, rows, cols, None, "p", None, None,
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        full_white = full.count("#fdfefe")
        assert_true(full_white > 0, "the full renderer resolves B$2 at all")
        # In the streamed output the near-white lives once in the palette;
        # count the cells wearing its class.
        import re as _re
        white_classes = _re.findall(r"\.(k\d+)\{[^}]*#fdfefe", wref["style_css"])
        assert_true(len(white_classes) >= 1, "streaming palette carries the near-white")
        stream_white = sum(
            sum(len(_re.findall(rf'\b{cls}\b', row)) for row in wref["rows"])
            for cls in white_classes)
        assert_eq(stream_white, full_white,
                  "streaming paints the same reference-matching cells near-white")
        green_classes = _re.findall(r"\.(k\d+)\{[^}]*#58fa82", wref["style_css"])
        stream_green = sum(
            sum(len(_re.findall(rf'\b{cls}\b', row)) for row in wref["rows"])
            for cls in green_classes)
        assert_eq(stream_green, full.count("#58fa82"),
                  "…and the same variant cells green")

        print("\n[a clade download holds the clade, and only the clade]")
        # The bug: "Download xlsx" from a clade view handed back the whole
        # group's table, because the download branch returned the file before
        # the selection was resolved. The export must now be the SAME subset the
        # page is showing — same rows, same columns, same colours.
        clade = tmp / "clade.xlsx"
        shutil.copy2(refbook, clade)
        rows, cols = 12, 10
        picked = ["SAMPLE2", "SAMPLE5", "SAMPLE7"]
        wf = xlsx_html.render_filtered_window(
            clade, rows, cols, None, "p", None, None, picked,
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        assert_eq(wf["filter"]["matched"], 3, "the three picked samples matched")
        assert_true(bool(wf["kept_rows"]) and bool(wf["kept_cols"]),
                    "the window records the subset in sheet coordinates")
        # Header + reference + the three samples, and nothing else in this book.
        assert_eq(wf["kept_rows"][:2], [1, 2], "header and reference row are kept")
        assert_eq(len(wf["kept_rows"]), 5, "no other rows come along")
        assert_eq(wf["kept_cols"][0], 1, "the sample-name column is always kept")

        out = tmp / "clade_subset.xlsx"
        xlsx_html.write_filtered_xlsx(clade, out, wf["kept_rows"], wf["kept_cols"],
                                      sheet_title=wf["sheet"])
        assert_true(out.exists() and out.stat().st_size < clade.stat().st_size,
                    "the subset is a smaller file than the table it came from")
        got = openpyxl.load_workbook(out)
        gws = got.active
        assert_eq(gws.max_column, len(wf["kept_cols"]),
                  "the export has exactly the kept columns")
        # Every value must be the one at the corresponding SOURCE coordinate: an
        # off-by-one in the row or column mapping would put a sample's calls on
        # another sample's row, which is the one failure that would not look
        # like a failure.
        srcws = openpyxl.load_workbook(clade).active
        mismatched = 0
        for out_r, src_r in enumerate(wf["kept_rows"], start=1):
            for out_c, src_c in enumerate(wf["kept_cols"], start=1):
                if (gws.cell(row=out_r, column=out_c).value
                        != srcws.cell(row=src_r, column=src_c).value):
                    mismatched += 1
        assert_eq(mismatched, 0, "every exported cell holds its source value")
        # Colour: CF is resolved into static fills, so the export carries the
        # same palette the page does, and the near-white reference-match rule
        # still beats the per-base rules.
        fills = {}
        for row in gws.iter_rows(min_row=1, max_row=gws.max_row):
            for c in row:
                if c.fill is not None and c.fill.patternType:
                    rgb = c.fill.fgColor.rgb
                    if isinstance(rgb, str):
                        fills[rgb[-6:].lower()] = fills.get(rgb[-6:].lower(), 0) + 1
        assert_true("fdfefe" in fills, "reference-matching cells are near-white")
        assert_true("58fa82" in fills, "variant cells are green")
        # And the same counts the page shows for those rows/columns.
        page_white = sum(
            sum(len(_re.findall(rf'\b{cls}\b', row)) for row in wf["rows"])
            for cls in _re.findall(r"\.(k\d+)\{[^}]*#fdfefe", wf["style_css"]))
        assert_eq(fills["fdfefe"], page_white,
                  "the export and the page colour the same number of cells white")
        assert_eq(gws.freeze_panes, None,
                  "no panes to translate when the source froze none")

        print("\n[the subset re-anchors frozen panes]")
        # The source freezes at B3 (one name column, two header rows). Both
        # survive the filter here, so the export must freeze at B3 too — and it
        # has to be set before the first row is streamed out, or openpyxl drops
        # it without a word.
        frozen = tmp / "frozen.xlsx"
        fwb = openpyxl.load_workbook(clade)
        fwb.active.freeze_panes = "B3"
        fwb.save(frozen)
        fwb.close()
        wfr = xlsx_html.render_filtered_window(
            frozen, rows, cols, None, "p", None, None, picked,
            xlsx_html.DEFAULT_MAX_CELLS, xlsx_html.DEFAULT_MAX_ROWS)
        fout = tmp / "frozen_subset.xlsx"
        xlsx_html.write_filtered_xlsx(frozen, fout, wfr["kept_rows"], wfr["kept_cols"])
        assert_eq(openpyxl.load_workbook(fout).active.freeze_panes, "B3",
                  "the export freezes the header rows and the name column")

        print("\nAll xlsx render tests passed.")
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
